# tabs/dashboard_report_tab.py
import tkinter
from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk
import time, os, re, json, subprocess
from datetime import datetime

# --- EXCEL IMPORT ---

# Selenium Imports

# PDF & Image Imports
from fpdf import FPDF
from PIL import Image, ImageDraw, ImageFont 
from utils import resource_path, get_logger
from .base_tab import BaseAutomationTab
from .autocomplete_widget import AutocompleteEntry
import config

logger = get_logger()

class DashboardReportTab(BaseAutomationTab):
    def __init__(self, parent, app_instance):
        # Lazy imports
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException
        super().__init__(parent, app_instance, automation_key="dashboard_report")
        
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1) 
        
        # Columns to scrape (Strictly 4 columns)
        self.report_headers = [
            "S No.", "Project Name with code", "E-MR No.", "DateFrom-DateTo"
        ]
        
        self._create_widgets()
        self.load_inputs()

    def _create_widgets(self):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        import openpyxl
        from selenium import webdriver

        controls_frame = ctk.CTkFrame(self)
        controls_frame.grid(row=0, column=0, sticky="new", padx=10, pady=10)
        controls_frame.grid_columnconfigure(1, weight=1)

        # --- Input Fields ---
        ctk.CTkLabel(controls_frame, text="State:").grid(row=0, column=0, sticky='w', padx=15, pady=(15, 5))
        self.state_entry = AutocompleteEntry(controls_frame, 
                                             suggestions_list=self.app.history_manager.get_suggestions("dashboard_state"),
                                             app_instance=self.app, history_key="dashboard_state")
        self.state_entry.grid(row=0, column=1, sticky='ew', padx=15, pady=(15, 5))

        ctk.CTkLabel(controls_frame, text="District:").grid(row=1, column=0, sticky='w', padx=15, pady=5)
        self.district_entry = AutocompleteEntry(controls_frame, 
                                                suggestions_list=self.app.history_manager.get_suggestions("dashboard_district"),
                                                app_instance=self.app, history_key="dashboard_district")
        self.district_entry.grid(row=1, column=1, sticky='ew', padx=15, pady=5)

        ctk.CTkLabel(controls_frame, text="Block:").grid(row=2, column=0, sticky='w', padx=15, pady=5)
        self.block_entry = AutocompleteEntry(controls_frame, 
                                             suggestions_list=self.app.history_manager.get_suggestions("dashboard_block"),
                                             app_instance=self.app, history_key="dashboard_block")
        self.block_entry.grid(row=2, column=1, sticky='ew', padx=15, pady=5)

        ctk.CTkLabel(controls_frame, text="Panchayat:").grid(row=3, column=0, sticky='w', padx=15, pady=5)
        self.panchayat_entry = AutocompleteEntry(controls_frame, 
                                                 suggestions_list=self.app.history_manager.get_suggestions("dashboard_panchayat"),
                                                 app_instance=self.app, history_key="dashboard_panchayat")
        self.panchayat_entry.grid(row=3, column=1, sticky='ew', padx=15, pady=5)

        ctk.CTkLabel(controls_frame, text="Delay Column:").grid(row=4, column=0, sticky='w', padx=15, pady=5)
        self.delay_column_options = [
            "Attendance not filled in T+2 days",
            "Measurement Book not filled in T+5 days",
            "Wagelist not Sent in T+6 days",
            "Pending for I sig FTO in T+7 days",
            "Pending for II sig FTO in T+8 days"
        ]
        self.delay_column_entry = ctk.CTkComboBox(controls_frame, values=self.delay_column_options)
        self.delay_column_entry.grid(row=4, column=1, sticky='ew', padx=15, pady=5)
        if self.delay_column_options:
            self.delay_column_entry.set(self.delay_column_options[0])

        action_frame = self._create_action_buttons(parent_frame=controls_frame)
        action_frame.grid(row=5, column=0, columnspan=2, pady=10)

        # --- Output Tabs ---
        notebook = ctk.CTkTabview(self)
        notebook.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 10))
        workcode_tab = notebook.add("Workcode List")
        results_tab = notebook.add("Results Table")
        self._create_log_and_status_area(parent_notebook=notebook)

        # 1. Workcode List Tab
        workcode_tab.grid_columnconfigure(0, weight=1)
        workcode_tab.grid_rowconfigure(1, weight=1)
        
        copy_frame = ctk.CTkFrame(workcode_tab, fg_color="transparent")
        copy_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 0))
        self.copy_wc_button = ctk.CTkButton(copy_frame, text="Copy Workcodes", command=self._copy_workcodes)
        self.copy_wc_button.pack(side="left")

        self.run_mr_fill_button = ctk.CTkButton(copy_frame, text="Run MR Fill", command=self._run_mr_fill,
                                                  fg_color=config.COLORS["green_dashboard"], hover_color=config.COLORS["green_dashboard_hover"])
        self.run_mr_fill_button.pack_forget()

        self.workcode_textbox = ctk.CTkTextbox(workcode_tab, state="disabled")
        self.workcode_textbox.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)

        # 2. Results Tab
        results_tab.grid_columnconfigure(0, weight=1)
        results_tab.grid_rowconfigure(1, weight=1)
        
        export_frame = ctk.CTkFrame(results_tab, fg_color="transparent")
        export_frame.grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.export_button = ctk.CTkButton(export_frame, text="Export Report", command=self.export_report)
        self.export_button.pack(side="left")
        
        self.export_format_menu = ctk.CTkOptionMenu(export_frame, values=["Excel (.xlsx)", "PDF (.pdf)", "PNG (.png)"])
        self.export_format_menu.pack(side="left", padx=5)

        # --- Treeview Config ---
        self.results_tree = ttk.Treeview(results_tab, columns=self.report_headers, show='headings')
        for col in self.report_headers: 
            self.results_tree.heading(col, text=col)
            
        self.results_tree.column("S No.", width=50, anchor='center')
        self.results_tree.column("Project Name with code", width=450)
        self.results_tree.column("E-MR No.", width=120, anchor='center')
        self.results_tree.column("DateFrom-DateTo", width=180, anchor='center')

        self.results_tree.grid(row=1, column=0, sticky='nsew', padx=5, pady=5)
        scrollbar = ctk.CTkScrollbar(results_tab, command=self.results_tree.yview)
        self.results_tree.configure(yscroll=scrollbar.set); scrollbar.grid(row=1, column=1, sticky='ns')
        self.style_treeview(self.results_tree)

    def set_ui_state(self, running: bool):
        if not self._is_alive():
            return
        self.set_common_ui_state(running)
        state = "disabled" if running else "normal"
        self.state_entry.configure(state=state)
        self.district_entry.configure(state=state)
        self.block_entry.configure(state=state)
        self.panchayat_entry.configure(state=state)
        self.delay_column_entry.configure(state=state)
        self.run_mr_fill_button.configure(state=state)

    def reset_ui(self):
        pass
        
    def start_automation(self):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        self.run_mr_fill_button.pack_forget()
        for item in self.results_tree.get_children(): self.results_tree.delete(item)
        self._update_workcode_textbox("") 
        
        inputs = {
            'state': self.state_entry.get().strip(), 
            'district': self.district_entry.get().strip(), 
            'block': self.block_entry.get().strip(),
            'panchayat': self.panchayat_entry.get().strip(),
            'delay_column': self.delay_column_entry.get().strip()
        }
        
        if not all([inputs['state'], inputs['district'], inputs['block'], inputs['panchayat'], inputs['delay_column']]):
            messagebox.showwarning("Input Error", "All fields are required."); return
        
        self.save_inputs(inputs)
        self.app.update_history("dashboard_state", inputs['state'])
        self.app.update_history("dashboard_district", inputs['district'])
        self.app.update_history("dashboard_block", inputs['block'])
        self.app.update_history("dashboard_panchayat", inputs['panchayat'])
        
        self.app.start_automation_thread(self.automation_key, self.run_automation_logic, args=(inputs,))

    def _solve_captcha(self, driver, wait):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        self.app.log_message(self.log_display, "Attempting to solve CAPTCHA...")
        try:
            captcha_element = wait.until(EC.presence_of_element_located((By.ID, "ContentPlaceHolder1_lblStopSpam")))
            captcha_text = captcha_element.text
            match = re.search(r'(\d+)\s*([+\-*])\s*(\d+)', captcha_text)
            if not match: raise ValueError("Could not parse CAPTCHA.")
            num1, operator, num2 = match.groups(); num1, num2 = int(num1), int(num2)
            result = num1 + num2 if operator == '+' else (num1 - num2 if operator == '-' else num1 * num2)
            driver.find_element(By.ID, "ContentPlaceHolder1_txtCaptcha").send_keys(str(result))
            driver.find_element(By.ID, "ContentPlaceHolder1_btnLogin").click()
            time.sleep(1.0)  # Short wait after click
            if "Invalid Captcha Code" in driver.page_source: raise ValueError("CAPTCHA failed.")
            return True
        except TimeoutException:
            self.app.log_message(self.log_display, "CAPTCHA not found, skipping.", "info")
            return True
        except ValueError as e:
            self.app.log_message(self.log_display, f"CAPTCHA Error: {e}", "error")
            raise

    def run_automation_logic(self, inputs, retries=1):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        self.app.after(0, self.set_ui_state, True)
        self.app.after(0, self.app.set_status, "Starting Dashboard Report...") 
        self.app.after(0, self.update_status, "Initializing...", 0.0) 
        self.app.clear_log(self.log_display)

        try:
            driver = self.app.get_driver()
            if not driver: return 
            wait = WebDriverWait(driver, 20)

            # --- STANDARD FLOW ONLY (Direct Link Removed) ---
            self.app.log_message(self.log_display, "Navigating to Home Page...")
            driver.get(config.MIS_REPORTS_CONFIG["base_url"])
            self._solve_captcha(driver, wait)

            self.update_status("Selecting State...", 0.15)
            Select(wait.until(EC.element_to_be_clickable((By.ID, "ContentPlaceHolder1_ddl_States")))).select_by_visible_text(inputs['state'].upper())
            
            self.update_status("Opening Dashboard...", 0.2)
            report_link = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Dashboard for Delay Monitoring System")))
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", report_link)
            time.sleep(1); report_link.click()

            self.update_status("Selecting District...", 0.25)
            wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, inputs['district'].upper()))).click()

            self.update_status("Selecting Block...", 0.3)
            wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, inputs['block'].upper()))).click()

            self.update_status("Finding Panchayat...", 0.35)
            main_table_xpath = "//table[.//b[text()='S No.'] and .//b[text()='Panchayat']]"
            wait.until(EC.presence_of_element_located((By.XPATH, f"{main_table_xpath}//tr[1]/td/b[text()='Panchayat']")))
            panchayat_row = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, f"{main_table_xpath}//tr[td[2][normalize-space()='{inputs['panchayat']}']]")))

            self.update_status("Finding Column...", 0.4)
            header_cells = driver.find_elements(By.XPATH, f"{main_table_xpath}//tr[.//b[contains(text(), 'T+2')]]/td/b")
            target_col_index = -1
            for i, th_b in enumerate(header_cells):
                if ' '.join(inputs['delay_column'].split()).lower().strip() == ' '.join(th_b.text.split()).lower().strip():
                    target_col_index = i + 2
                    break

            if target_col_index == -1: raise ValueError(f"Column '{inputs['delay_column']}' not found.")

            row_cells = panchayat_row.find_elements(By.TAG_NAME, "td")
            target_cell = row_cells[target_col_index]

            try:
                target_link = target_cell.find_element(By.TAG_NAME, "a")
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", target_link)
                time.sleep(0.5); target_link.click()
            except NoSuchElementException:
                 if target_cell.text.strip() == '0':
                    messagebox.showinfo("No Data", f"No records found for {inputs['delay_column']} in {inputs['panchayat']}.")
                    self.success_message = None; return
                 else: raise ValueError("Target cell is not a clickable link.")

            self.update_status("Loading Final Report...", 0.5)
            FINAL_TABLE_XPATH = "//table[@bordercolor='green' and .//b[contains(text(), 'E-MR No.')]]"
            table = wait.until(EC.presence_of_element_located((By.XPATH, FINAL_TABLE_XPATH)))
            rows = table.find_elements(By.XPATH, ".//tr[position()>1]") 

            if not rows:
                self.app.log_message(self.log_display, "Table is empty.", "warning")
                return

            workcode_list = []
            pending_mr_count = 0

            for i, row in enumerate(rows):
                if self.app.stop_events[self.automation_key].is_set(): break
                self.update_status(f"Processing row {i+1}/{len(rows)}", 0.5 + ((i+1)/len(rows))*0.45)

                cells = row.find_elements(By.TAG_NAME, "td")
                # Original Table Indices: 0:SNo, 1:Dist, 2:Blk, 3:GP, 4:Agency, 5:Project, 6:EMR, 7:Date
                if len(cells) < 8: continue

                # --- EXTRACT DATA FOR NEW COLUMNS ---
                s_no = cells[0].text.strip()
                project_name = cells[5].text.strip()
                emr_no = cells[6].text.strip()
                dates = cells[7].text.strip()
                
                # Extract workcode
                wc_match = re.search(r'\(([^)]+)\)$', project_name)
                if wc_match: workcode_list.append(wc_match.group(1).strip())

                pending_mr_count += 1
                row_data = (s_no, project_name, emr_no, dates)
                self.app.after(0, lambda data=row_data: self.results_tree.insert("", "end", values=data))

            if self.app.stop_events[self.automation_key].is_set(): return
            self.app.after(0, self._update_workcode_textbox, "\n".join(workcode_list)) 
            self.success_message = f"Done.\n{pending_mr_count} Pending items found."

        except Exception as e:
            if "Session Expired" in str(e) and retries > 0:
                self.run_automation_logic(inputs, retries - 1)
                return
            self.app.log_message(self.log_display, f"Error: {e}", "error")
            self.success_message = None
        finally:
            self.app.after(0, self.set_ui_state, False)
            self.app.after(0, self.app.set_status, "Ready")
            self.app.after(0, self.update_status, "Ready", 0.0)
            if hasattr(self, 'success_message') and self.success_message:
                self.app.after(100, lambda: messagebox.showinfo("Complete", self.success_message))
                if inputs['delay_column'] == "Attendance not filled in T+2 days":
                    self.app.after(0, lambda: self.run_mr_fill_button.pack(side="left", padx=(10, 0)))

    def _update_workcode_textbox(self, text):
        self.workcode_textbox.configure(state="normal")
        self.workcode_textbox.delete("1.0", tkinter.END)
        self.workcode_textbox.insert("1.0", text)
        self.workcode_textbox.configure(state="disabled")

    def _copy_workcodes(self):
        text = self.workcode_textbox.get("1.0", tkinter.END).strip()
        if text:
            self.app.clipboard_clear(); self.app.clipboard_append(text)
            messagebox.showinfo("Copied", f"Copied to clipboard.")
        else: messagebox.showwarning("Empty", "No workcodes.")

    def _run_mr_fill(self):
        wc = self.workcode_textbox.get("1.0", tkinter.END).strip()
        gp = self.panchayat_entry.get().strip()
        if wc and gp: self.app.switch_to_mr_fill_with_data(wc, gp)
        else: messagebox.showwarning("Error", "Missing Data.")

    # =========================================================================
    # ==================== FINAL EXPORT LOGIC (FIXED) =========================
    # =========================================================================

    def export_report(self):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        if not self.results_tree.get_children():
            messagebox.showinfo("No Data", "No results to export.")
            return

        # Basic Info
        state = self.state_entry.get().strip()
        district = self.district_entry.get().strip()
        block = self.block_entry.get().strip()
        panchayat = self.panchayat_entry.get().strip()
        safe_panchayat = re.sub(r'[\\/*?:"<>|]', '_', panchayat or "Report") 
        
        export_format = self.export_format_menu.get()
        current_year = datetime.now().strftime("%Y")
        current_date_str = datetime.now().strftime("%d-%b-%Y")

        headers = self.report_headers # ["S No.", "Project Name...", "E-MR No.", "Date..."]
        data = [self.results_tree.item(item, 'values') for item in self.results_tree.get_children()]

        # Titles & Headers
        delay_type = self.delay_column_entry.get()
        if "Attendance" in delay_type: report_type = "Attendance Pending Report"
        elif "Measurement" in delay_type: report_type = "Measurement Book Pending Report"
        elif "Wagelist" in delay_type: report_type = "Wagelist Pending Report"
        elif "FTO" in delay_type: report_type = "FTO Pending Report"
        else: report_type = "Delay Compensation Report"

        main_title = f"{report_type.upper()}"
        sub_title = f"District: {district}  |  Block: {block}  |  Panchayat: {panchayat}"
        
        # Directory
        target_dir = os.path.join(self.app.get_user_downloads_path(), f"Reports {current_year}", safe_panchayat)
        try: os.makedirs(target_dir, exist_ok=True)
        except Exception as e: logger.debug("Failed to create dirs: %s", e)

        # --- EXCEL EXPORT ---
        if "Excel" in export_format:
            file_path = filedialog.asksaveasfilename(
                initialdir=target_dir, initialfile=f"{report_type.replace(' ','_')}_{safe_panchayat}_{current_date_str}.xlsx", 
                defaultextension=".xlsx", filetypes=[("Excel Workbook", "*.xlsx")])
            if file_path:
                if self._save_to_excel(data, headers, main_title, sub_title, file_path):
                    messagebox.showinfo("Success", f"Excel report saved:\n{file_path}")

        # --- PDF EXPORT ---
        elif "PDF" in export_format:
            file_path = filedialog.asksaveasfilename(
                initialdir=target_dir, initialfile=f"{report_type.replace(' ','_')}_{safe_panchayat}_{current_date_str}.pdf", 
                defaultextension=".pdf", filetypes=[("PDF Document", "*.pdf")])
            if file_path:
                # 4 Columns Widths (Normalized for A4 Landscape)
                col_widths = [15, 177, 35, 50] # Total ~277mm
                
                if self.generate_report_pdf(data, headers, col_widths, main_title, sub_title, file_path):
                    messagebox.showinfo("Success", f"PDF report saved:\n{file_path}")

        # --- PNG EXPORT ---
        elif "PNG" in export_format:
            file_path = filedialog.asksaveasfilename(
                initialdir=target_dir, initialfile=f"{report_type.replace(' ','_')}_{safe_panchayat}_{current_date_str}.png", 
                defaultextension=".png", filetypes=[("PNG Image", "*.png")])
            if file_path:
                if self._save_to_png(data, headers, main_title, sub_title, file_path):
                    messagebox.showinfo("Success", f"PNG report saved:\n{file_path}")

    def _save_to_excel(self, data, headers, title, subtitle, file_path):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        try:
            wb = openpyxl.Workbook(); ws = wb.active
            ws.title = "Report"
            
            # Use fixed number of columns (4) to prevent bleed into 'E'
            num_cols = 4 
            last_col_char = get_column_letter(num_cols) # Should be 'D'

            # Styles
            header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") 
            gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            center = Alignment(horizontal="center", vertical="center")
            
            # Row 1: Title
            ws.merge_cells(f'A1:{last_col_char}1')
            ws['A1'] = title
            ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
            ws['A1'].fill = header_fill
            ws['A1'].alignment = center

            # Row 2: Location Subtitle
            ws.merge_cells(f'A2:{last_col_char}2')
            ws['A2'] = subtitle
            ws['A2'].font = Font(size=11, bold=True)
            ws['A2'].alignment = center
            ws['A2'].fill = PatternFill(start_color="DCE6F1", fill_type="solid")

            # Row 3: Generated Date & Promo
            ws.merge_cells(f'A3:{last_col_char}3')
            ws['A3'] = f"Generated: {datetime.now().strftime('%d-%m-%Y %I:%M %p')} | Visit NregaBot.com"
            ws['A3'].font = Font(italic=True, size=9)
            ws['A3'].alignment = center

            # Row 5: Headers
            for col_idx, text in enumerate(headers, 1):
                c = ws.cell(row=5, column=col_idx, value=text)
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = header_fill
                c.alignment = center
                c.border = border

            # Data
            for r_idx, row_data in enumerate(data, 6):
                fill = gray_fill if r_idx % 2 == 0 else white_fill
                for c_idx, val in enumerate(row_data, 1):
                    # SAFETY: Ensure we don't write beyond column 4
                    if c_idx > num_cols: break 
                    c = ws.cell(row=r_idx, column=c_idx, value=val)
                    c.fill = fill
                    c.border = border
                    # Center align SNo(1), EMR(3), Date(4). Left align Project(2)
                    c.alignment = Alignment(horizontal="left", vertical="center") if c_idx == 2 else center

            # Widths (S No, Project, EMR, Date)
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 65
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 25

            wb.save(file_path)
            try: os.startfile(file_path) if os.name == 'nt' else subprocess.call(['open', file_path])
            except Exception as e: logger.debug("Failed to open file: %s", e)
            return True
        except Exception as e:
            messagebox.showerror("Export Error", f"{e}"); return False

    def generate_report_pdf(self, data, headers, col_widths, title, subtitle, file_path):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        class ProPDF(FPDF):
            def footer(self):
                self.set_y(-15); self.set_font('Arial', 'I', 8)
                self.cell(0, 10, f'Page {self.page_no()} - Generated by NregaBot.com', 0, 0, 'C')

        try:
            pdf = ProPDF(orientation="L", unit="mm", format="A4")
            pdf.set_auto_page_break(auto=True, margin=15); pdf.add_page()
            
            # Font handling
            try:
                font_reg = resource_path("assets/fonts/NotoSansDevanagari-Regular.ttf")
                font_bold = resource_path("assets/fonts/NotoSansDevanagari-Bold.ttf")
                pdf.add_font("Custom", "", font_reg, uni=True)
                pdf.add_font("Custom", "B", font_bold, uni=True)
                f_name = "Custom"
            except: f_name = "Arial"

            # Title Block
            pdf.set_font(f_name, "B", 16)
            pdf.cell(0, 10, title, 0, 1, "C")
            
            # Subtitle (Location)
            pdf.set_font(f_name, "B", 10)
            pdf.set_fill_color(220, 230, 241) # Light Blue Bar
            pdf.cell(0, 8, subtitle, 0, 1, "C", fill=True)
            
            # Date
            pdf.set_font(f_name, "", 8) # Regular font (Not Italic to avoid crash)
            pdf.cell(0, 6, f"Date: {datetime.now().strftime('%d-%b-%Y')}", 0, 1, "R")
            pdf.ln(2)

            # Table Headers
            pdf.set_font(f_name, "B", 9)
            pdf.set_fill_color(31, 73, 125) # Dark Blue
            pdf.set_text_color(255, 255, 255) # White Text
            for i, h in enumerate(headers):
                pdf.cell(col_widths[i], 10, h, 1, 0, "C", fill=True)
            pdf.ln()

            # Table Data (Improved Row Height Calculation)
            pdf.set_text_color(0, 0, 0)
            pdf.set_font(f_name, "", 8)
            fill = False
            
            for row in data:
                pdf.set_fill_color(242, 242, 242) if fill else pdf.set_fill_color(255, 255, 255)
                
                # 1. Calculate Maximum Row Height First
                max_row_height = 0
                cell_lines = [] # Store line counts to avoid re-calculating
                for i, txt in enumerate(row):
                    # Get number of lines this cell will take
                    # FPDF multi_cell simply prints, so we use string width to estimate
                    # OR we use a temporary multi_cell approach
                    
                    # Robust method: split_only=True returns the lines
                    lines = pdf.multi_cell(col_widths[i], 5, str(txt), border=0, split_only=True)
                    num_lines = len(lines)
                    max_row_height = max(max_row_height, num_lines * 5)
                    cell_lines.append(num_lines)

                # Ensure minimum height
                max_row_height = max(max_row_height, 6)

                # 2. Check Page Break
                if pdf.get_y() + max_row_height > 190:
                    pdf.add_page()
                    # Re-print headers
                    pdf.set_font(f_name, "B", 9)
                    pdf.set_fill_color(31, 73, 125); pdf.set_text_color(255, 255, 255)
                    for i, h in enumerate(headers): pdf.cell(col_widths[i], 10, h, 1, 0, "C", fill=True)
                    pdf.ln(); pdf.set_text_color(0, 0, 0); pdf.set_font(f_name, "", 8)

                # 3. Print Cells
                current_x = pdf.get_x()
                current_y = pdf.get_y()
                
                for i, txt in enumerate(row):
                    # Save x position
                    pdf.set_xy(current_x, current_y)
                    
                    # Print MultiCell
                    # align='L' for Project Name (index 1), 'C' for others
                    align = 'L' if i == 1 else 'C'
                    pdf.multi_cell(col_widths[i], 5, str(txt), border=1, align=align, fill=True)
                    
                    # Move X pointer to next column
                    current_x += col_widths[i]
                
                # 4. Move Y pointer to next row start
                pdf.set_xy(10, current_y + max_row_height) # 10 is left margin
                
                fill = not fill

            pdf.output(file_path)
            return True
        except Exception as e:
            messagebox.showerror("PDF Error", f"{e}"); return False

    def _save_to_png(self, data, headers, title, subtitle, file_path):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        try:
            # Font Setup
            try:
                fr = resource_path("assets/fonts/NotoSansDevanagari-Regular.ttf")
                fb = resource_path("assets/fonts/NotoSansDevanagari-Bold.ttf")
                font_title = ImageFont.truetype(fb, 40)
                font_sub = ImageFont.truetype(fb, 30)
                font_head = ImageFont.truetype(fb, 28)
                font_body = ImageFont.truetype(fr, 24)
            except:
                font_title = ImageFont.load_default()
                font_sub = font_head = font_body = font_title

            # Layout Config
            W = 2400; margin = 50
            # Width Ratios: SNo(0.05), Proj(0.60), EMR(0.15), Date(0.20)
            col_ws = [W*0.05, W*0.60, W*0.15, W*0.20 - (2*margin)] 
            col_ws = [int(x) for x in col_ws]
            
            # Colors
            c_head_bg = (31, 73, 125); c_head_txt = (255, 255, 255)
            c_row_odd = (242, 242, 242); c_row_even = (255, 255, 255)
            
            # Start Image
            H = 2000 # Dynamic
            img = Image.new("RGB", (W, H), (255, 255, 255))
            draw = ImageDraw.Draw(img)
            y = margin

            # Draw Title
            w_text = font_title.getlength(title)
            draw.text(((W-w_text)/2, y), title, font=font_title, fill=(0,0,0))
            y += 60

            # Draw Subtitle Box
            draw.rectangle([margin, y, W-margin, y+50], fill=(220, 230, 241))
            w_sub = font_sub.getlength(subtitle)
            draw.text(((W-w_sub)/2, y+8), subtitle, font=font_sub, fill=(0,0,0))
            y += 80

            # Draw Headers
            x = margin; h_height = 60
            for i, h in enumerate(headers):
                draw.rectangle([x, y, x+col_ws[i], y+h_height], fill=c_head_bg, outline=(0,0,0))
                th_w = font_head.getlength(h)
                draw.text((x + (col_ws[i]-th_w)/2, y+15), h, font=font_head, fill=c_head_txt)
                x += col_ws[i]
            y += h_height

            # Draw Rows
            line_h = 40
            for idx, row in enumerate(data):
                bg = c_row_odd if idx%2==0 else c_row_even
                
                wrapped_cells = []
                max_lines = 1
                for i, txt in enumerate(row):
                    lines = self._wrap_text(str(txt), font_body, col_ws[i]-20) 
                    wrapped_cells.append(lines)
                    max_lines = max(max_lines, len(lines))
                
                row_h = max_lines * line_h + 30 
                
                if y + row_h + margin + 50 > img.height:
                    new_img = Image.new("RGB", (W, img.height + 2500), (255, 255, 255))
                    new_img.paste(img, (0,0))
                    img = new_img; draw = ImageDraw.Draw(img)

                x = margin
                for i, lines in enumerate(wrapped_cells):
                    draw.rectangle([x, y, x+col_ws[i], y+row_h], fill=bg, outline=(100,100,100))
                    
                    ty = y + 15
                    for line in lines:
                        tw = font_body.getlength(line)
                        tx = x + 15 if i == 1 else x + (col_ws[i]-tw)/2
                        draw.text((tx, ty), line, font=font_body, fill=(0,0,0))
                        ty += line_h
                    x += col_ws[i]
                y += row_h

            # Footer
            y += 20
            draw.text((margin, y), "Report Generated by NregaBot.com", font=font_body, fill=(100,100,100))
            y += 50

            # Crop
            final_img = img.crop((0, 0, W, y))
            final_img.save(file_path)
            return True
        except Exception as e:
            messagebox.showerror("PNG Error", f"{e}"); return False

    def save_inputs(self, inputs):
        d = {k: inputs.get(k) for k in ('state', 'district', 'block', 'panchayat')}
        try:
            with open(self.app.get_data_path("dashboard_report_inputs.json"), 'w') as f: json.dump(d, f)
        except Exception as e: logger.debug("Dashboard: Could not save inputs: %s", e)

    def load_inputs(self):
        try:
            with open(self.app.get_data_path("dashboard_report_inputs.json"), 'r') as f: data = json.load(f)
            self.state_entry.delete(0, 'end'); self.state_entry.insert(0, data.get('state', ''))
            self.district_entry.delete(0, 'end'); self.district_entry.insert(0, data.get('district', ''))
            self.block_entry.delete(0, 'end'); self.block_entry.insert(0, data.get('block', ''))
            self.panchayat_entry.delete(0, 'end'); self.panchayat_entry.insert(0, data.get('panchayat', ''))
        except Exception as e: logger.debug("Dashboard: Could not load inputs: %s", e)