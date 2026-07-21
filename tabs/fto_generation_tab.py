# tabs/fto_generation_tab.py
import tkinter
from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk
import time, json, os, re
import threading
from datetime import datetime

import config
from .base_tab import BaseAutomationTab

class FtoGenerationTab(BaseAutomationTab):
    def __init__(self, parent, app_instance):
        # Lazy imports
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait, Select
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoAlertPresentException, StaleElementReferenceException
        super().__init__(parent, app_instance, automation_key="fto_gen")
        self.automation_has_run = False 
        self.stored_location_data = {} 
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        
        # Default common paths for Old Firefox
        self.default_paths = [
            r"C:\Program Files (x86)\Mozilla Firefox\firefox.exe",
            r"C:\Program Files\Mozilla Firefox\firefox.exe"
        ]
        
        self._create_widgets()
        self._load_saved_path()

    def _create_widgets(self):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.common.exceptions import NoAlertPresentException
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        import openpyxl
        from selenium import webdriver

        # Main container
        controls_frame = ctk.CTkFrame(self)
        controls_frame.grid(row=0, column=0, sticky="ew", pady=(0, 2)) 
        controls_frame.grid_columnconfigure(0, weight=1)

        note_text = "Instructions:\n1. Check/Set Old Firefox Path and Click 'Launch Old Firefox'.\n2. Log in manually, insert DSC Token, go to FTO page.\n3. Click 'Start' to sign pending FTOs or 'Delete' to remove."
        ctk.CTkLabel(controls_frame, text=note_text, justify="left").grid(row=0, column=0, sticky='w', padx=15, pady=(5, 2))
        
        # --- NEW: Browser Setup Frame ---
        setup_frame = ctk.CTkFrame(controls_frame, fg_color="transparent")
        setup_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=(2, 2))
        setup_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(setup_frame, text="Old Firefox Path:", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, sticky='w', padx=5, pady=5)
        
        self.ff_path_entry = ctk.CTkEntry(setup_frame, placeholder_text="e.g. C:\\Program Files (x86)\\Mozilla Firefox\\firefox.exe")
        self.ff_path_entry.grid(row=0, column=1, sticky='ew', padx=5, pady=5)

        self.browse_btn = ctk.CTkButton(setup_frame, text="Browse", width=70, command=self._browse_firefox)
        self.browse_btn.grid(row=0, column=2, padx=5, pady=5)

        self.check_install_btn = ctk.CTkButton(setup_frame, text="Check Install", width=100, fg_color="#D97706", hover_color="#B45309", command=self._check_installation)
        self.check_install_btn.grid(row=0, column=3, padx=5, pady=5)

        self.launch_btn = ctk.CTkButton(setup_frame, text="Launch Old Firefox", fg_color="#108842", hover_color="#1A994C", command=self._launch_firefox)
        self.launch_btn.grid(row=0, column=4, padx=5, pady=5)

        # --- Row 2: Main Action Buttons (Start/Stop/Reset) ---
        action_frame_wrapper = self._create_action_buttons(parent_frame=controls_frame)
        action_frame_wrapper.grid(row=2, column=0, sticky='ew', pady=(2, 2), padx=15)

        # Access inner container to add separator and Delete button
        inner_container = action_frame_wrapper.winfo_children()[0] 
        
        # Separator (Visual Gap)
        ctk.CTkFrame(inner_container, width=2, height=20, fg_color="gray60").pack(side="left", padx=15)

        # Delete Button (Separated from Reset)
        self.delete_btn = ctk.CTkButton(
            inner_container, 
            text="🗑 Delete FTOs", 
            command=self.start_delete_automation,
            fg_color="#9B2C2C", 
            hover_color="#7F1D1D",
            width=110,
            font=ctk.CTkFont(size=13, weight="bold")
        )
        self.delete_btn.pack(side="left", padx=(0, 0))

        # --- Row 3: ABPS Check Button Container ---
        self.abps_container = ctk.CTkFrame(controls_frame, fg_color="transparent", height=0)
        self.abps_container.grid(row=3, column=0, sticky="ew", pady=(0, 0))
        self.abps_container.grid_remove() # Hide initially
        
        self.check_abps_button = ctk.CTkButton(
            self.abps_container, 
            text="Check Pending ABPS Labour", 
            command=self._go_to_mr_tracking,
            width=200,
            height=32,
            fg_color="#3B82F6",
            hover_color="#2563EB",
            font=ctk.CTkFont(size=13, weight="bold")
        )
        
        # --- Results Area ---
        notebook = ctk.CTkTabview(self)
        notebook.grid(row=1, column=0, sticky="nsew", pady=0)
        
        self._create_log_and_status_area(parent_notebook=notebook)
        results_frame = notebook.add("Results")

        results_frame.grid_columnconfigure(0, weight=1)
        results_frame.grid_rowconfigure(0, weight=1)
        
        cols = ("Type", "Status", "Info", "Timestamp")
        self.results_tree = ttk.Treeview(results_frame, columns=cols, show='headings')
        for col in cols: self.results_tree.heading(col, text=col)
        self.results_tree.column("Type", width=150)
        self.results_tree.column("Status", width=100)
        self.results_tree.column("Info", width=300)
        self.results_tree.column("Timestamp", width=100, anchor='center')
        self.results_tree.grid(row=0, column=0, sticky='nsew')
        
        scrollbar = ctk.CTkScrollbar(results_frame, command=self.results_tree.yview)
        self.results_tree.configure(yscroll=scrollbar.set)
        scrollbar.grid(row=0, column=1, sticky='ns')
        self.style_treeview(self.results_tree)

    # ============================================================================
    # NEW OLD FIREFOX SETUP LOGIC
    # ============================================================================

    def _browse_firefox(self):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.common.exceptions import NoAlertPresentException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        path = filedialog.askopenfilename(
            title="Select Old Firefox Executable",
            filetypes=[("Executable Files", "*.exe")]
        )
        if path:
            self.ff_path_entry.delete(0, tkinter.END)
            self.ff_path_entry.insert(0, path)
            self._save_path(path)

    def _check_installation(self):
        path = self.ff_path_entry.get().strip()
        if not path:
            for p in self.default_paths:
                if os.path.exists(p):
                    path = p
                    self.ff_path_entry.insert(0, path)
                    break
        
        if path and os.path.exists(path):
            messagebox.showinfo("Success", f"Old Firefox found at:\n{path}")
            self._save_path(path)
        else:
            messagebox.showwarning("Not Found", "Old Firefox not found! Please browse and select 'firefox.exe' manually.")

    def _launch_firefox(self):
        path = self.ff_path_entry.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showerror("Error", "Valid Firefox path is required!")
            return
        
        self.app.log_message(self.log_display, "Launching Old Firefox...")
        self.launch_btn.configure(state="disabled", text="Launching...")
        
        url = "https://nregade4.nic.in/netnrega/Login.aspx?&level=HomePO&state_code=34"

        def _thread():
            # Yeh call aapke browser_manager.py ka naya function hit karegi
            success, msg = self.app.browser_manager.launch_old_firefox(path, url)
            if success:
                self.app.after(0, lambda: self.app.log_message(self.log_display, "Browser Launched. Please login manually, plug in your DSC token, then click Start.", "success"))
                self.app.after(0, lambda: messagebox.showinfo("Browser Ready", "Old Firefox is open.\n\n1. Login to NREGA.\n2. Go to FTO page.\n3. Return here and click 'Start'."))
            else:
                self.app.after(0, lambda: self.app.log_message(self.log_display, msg, "error"))
                self.app.after(0, lambda: messagebox.showerror("Browser Error", msg))
            
            self.app.after(0, lambda: self.launch_btn.configure(state="normal", text="Launch Old Firefox"))

        threading.Thread(target=_thread, daemon=True).start()

    def _save_path(self, path):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.common.exceptions import NoAlertPresentException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        self.app.update_history("old_firefox_path", path)

    def _load_saved_path(self):
        saved = self.app.history_manager.get_suggestions("old_firefox_path")
        if saved:
            self.ff_path_entry.insert(0, saved[0])

    def set_ui_state(self, running: bool):
        self.set_common_ui_state(running)
        self.delete_btn.configure(state="disabled" if running else "normal")
        self.ff_path_entry.configure(state="disabled" if running else "normal")
        self.browse_btn.configure(state="disabled" if running else "normal")
        self.check_install_btn.configure(state="disabled" if running else "normal")
        self.launch_btn.configure(state="disabled" if running else "normal")
        
        if running:
            self.check_abps_button.pack_forget()
            self.abps_container.grid_remove() 
        elif self.automation_has_run:
            self.abps_container.grid() 
            self.check_abps_button.pack(anchor="center", pady=(5, 5)) 
        else:
             self.check_abps_button.pack_forget()
             self.abps_container.grid_remove() 

    def start_automation(self):
        self.automation_has_run = False 
        self.stored_location_data = {} 
        self.check_abps_button.pack_forget()
        self.abps_container.grid_remove() 
        self.app.start_automation_thread(self.automation_key, self.run_generation_logic)

    def start_delete_automation(self):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.common.exceptions import NoAlertPresentException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        if not messagebox.askyesno("Confirm Delete", "This will delete the FIRST FTO in the dropdown.\n\nEnsure you want to proceed."):
            return
        self.app.start_automation_thread(self.automation_key + "_del", self.run_delete_logic)

    def _log_result(self, r_type, status, info):
        self.app.after(0, lambda: self.results_tree.insert("", "end", values=(r_type, status, info, datetime.now().strftime("%H:%M:%S"))))

    # ============================================================================
    # GENERATION LOGIC (WITH FIXED SCRAPING) - UNTOUCHED
    # ============================================================================

    def run_generation_logic(self):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.common.exceptions import NoAlertPresentException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        self.app.after(0, self.set_ui_state, True)
        self.app.clear_log(self.log_display)
        self.app.after(0, lambda: [self.results_tree.delete(item) for item in self.results_tree.get_children()])
        self.update_status("Starting Generation...", 0)
        
        try:
            driver = self.app.get_driver()
            if not driver: return
            cfg = config.FTO_GEN_CONFIG
            wait = WebDriverWait(driver, 15)
            
            self.update_status("Processing Aadhaar FTO...", 0.25)
            self._process_verification_page(driver, wait, cfg["aadhaar_fto_url"], "Aadhaar Gen")

            self.update_status("Processing Top-Up FTO...", 0.75)
            self._process_verification_page(driver, wait, cfg["top_up_fto_url"], "Top-Up Gen")
            
            self.app.log_message(self.log_display, "Generation complete.")
            messagebox.showinfo("Success", "FTO Generation Process Completed.")

        except Exception as e:
            self.app.log_message(self.log_display, f"Error: {e}", "error")
        finally:
            self.automation_has_run = True
            self.app.after(0, self.set_ui_state, False)
            self.app.after(0, self.update_status, "Finished", 1.0)
            self.app.after(0, self.app.set_status, "Ready")

    def _process_verification_page(self, driver, wait, url, name):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.common.exceptions import NoAlertPresentException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        try:
            self.app.log_message(self.log_display, f"Navigating to {name}...")
            driver.get(url)
            
            # --- SCRAPE LOCATION DATA IMMEDIATELY ---
            if not self.stored_location_data: 
                self._scrape_location_from_page(driver)
            # ----------------------------------------

            wait.until(EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_wage_list_verify")))
            
            if not driver.find_elements(By.XPATH, "//input[contains(@id, '_auth')]"):
                self.app.log_message(self.log_display, "No records found.", "warning")
                self._log_result(name, "Skipped", "No records")
                return

            driver.execute_script("document.querySelectorAll('input[id*=\"_auth\"]').forEach(radio => radio.click());")
            
            submit_btn = wait.until(EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_ch_verified")))
            driver.execute_script("arguments[0].click();", submit_btn)
            
            auth_btn = wait.until(EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_btn")))
            driver.execute_script("arguments[0].click();", auth_btn)
            
            alert = wait.until(EC.alert_is_present())
            fto_match = re.search(r'FTO No : \((.*?)\)', alert.text)
            fto = fto_match.group(1) if fto_match else "Generated"
            
            self.app.log_message(self.log_display, f"Success: {fto}", "success")
            self._log_result(name, "Success", fto)
            alert.accept()
        except TimeoutException:
            self._log_result(name, "Failed", "Page load/Login error")
        except Exception as e:
            self._log_result(name, "Error", str(e))

    def _scrape_location_from_page(self, driver):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.common.exceptions import NoAlertPresentException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        """Helper to scrape District, Block, Panchayat from TABLE CELLS (Not IDs)."""
        try:
            self.app.log_message(self.log_display, "Capturing location info...")
            
            def get_text_by_xpath(label):
                # ---- Lazy imports ----
                from selenium.webdriver.common.by import By
                from selenium.webdriver.support.ui import Select, WebDriverWait
                from selenium.webdriver.support import expected_conditions as EC
                from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
                from selenium.common.exceptions import NoAlertPresentException
                from selenium import webdriver
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                from openpyxl.utils import get_column_letter
                from openpyxl.worksheet.page import PageMargins
                from openpyxl.drawing.image import Image as XLImage
                try:
                    # Finds <td> containing "District" and gets text
                    elem = driver.find_element(By.XPATH, f"//td[contains(text(), '{label}')]")
                    text = elem.text 
                    if ":" in text:
                        return text.split(":")[1].strip()
                    return text.strip()
                except:
                    return None

            # 1. District
            dist = get_text_by_xpath("District")
            if dist: self.stored_location_data['district'] = dist

            # 2. Block
            blk = get_text_by_xpath("Block")
            if blk: self.stored_location_data['block'] = blk

            # 3. Panchayat
            panch = get_text_by_xpath("Panch")
            if panch:
                self.stored_location_data['panchayat'] = panch

            if self.stored_location_data:
                self.app.log_message(self.log_display, f"Location captured: {self.stored_location_data}")
            else:
                self.app.log_message(self.log_display, "Could not capture location data using XPath.", "warning")

        except Exception as e:
            print(f"Scrape Error: {e}")

    # ============================================================================
    # DELETION LOGIC (WITH SCROLL FIX) - UNTOUCHED
    # ============================================================================
    
    def run_delete_logic(self):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.common.exceptions import NoAlertPresentException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        self.app.after(0, self.set_ui_state, True)
        self.app.clear_log(self.log_display)
        self.update_status("Starting Deletion...", 0)
        
        try:
            driver = self.app.get_driver()
            if not driver: return
            cfg = config.FTO_GEN_CONFIG
            wait = WebDriverWait(driver, 10)

            self.update_status("Deleting from Link 1...", 0.2)
            self._process_deletion_page(driver, wait, cfg["delete_url_1"], "Delete (Type 1)")

            self.update_status("Deleting from Link 2...", 0.6)
            self._process_deletion_page(driver, wait, cfg["delete_url_2"], "Delete (Type 2)")
            
            self.app.log_message(self.log_display, "Deletion process finished.")
            messagebox.showinfo("Finished", "FTO Deletion check complete.")

        except Exception as e:
            self.app.log_message(self.log_display, f"Critical Error: {e}", "error")
        finally:
            self.app.after(0, self.set_ui_state, False)
            self.app.after(0, self.update_status, "Ready", 1.0)

    def _process_deletion_page(self, driver, wait, url, log_name):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.common.exceptions import NoAlertPresentException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        try:
            self.app.log_message(self.log_display, f"Navigating to {log_name}...")
            driver.get(url)
            
            try:
                dd_elem = wait.until(EC.presence_of_element_located((By.TAG_NAME, "select")))
                select = Select(dd_elem)
            except:
                self.app.log_message(self.log_display, "Dropdown not found.", "warning")
                self._log_result(log_name, "Skipped", "Dropdown not found")
                return

            if len(select.options) <= 1:
                self.app.log_message(self.log_display, "No FTOs available.", "info")
                self._log_result(log_name, "Skipped", "No FTOs")
                return

            fto_text = select.options[1].text
            self.app.log_message(self.log_display, f"Selecting: {fto_text}")
            select.select_by_index(1)

            self.app.log_message(self.log_display, "Waiting for reload...")
            try:
                WebDriverWait(driver, 10).until(
                    lambda d: d.execute_script('return document.readyState') == 'complete'
                )
            except TimeoutException:
                pass
            
            self.app.log_message(self.log_display, "Scrolling to bottom...")
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.ID, 'ctl00_ContentPlaceHolder1_'))
                )
            except (TimeoutException, NoSuchElementException):
                pass

            max_retries = 3
            radio_found = False
            for i in range(max_retries):
                try:
                    no_radio = wait.until(EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_rb_chkfto_n")))
                    driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", no_radio)
                    try:
                        WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.ID, 'ctl00_ContentPlaceHolder1_'))
                        )
                    except (TimeoutException, NoSuchElementException):
                        pass
                    driver.execute_script("arguments[0].click();", no_radio)
                    self.app.log_message(self.log_display, "Selected 'No'.")
                    radio_found = True
                    break
                except StaleElementReferenceException:
                    try:
                        WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.ID, 'ctl00_ContentPlaceHolder1_'))
                        )
                    except (TimeoutException, NoSuchElementException):
                        pass
                except Exception:
                    try:
                        WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.ID, 'ctl00_ContentPlaceHolder1_'))
                        )
                    except (TimeoutException, NoSuchElementException):
                        pass

            if not radio_found:
                self.app.log_message(self.log_display, "Could not find 'No' radio button (Check filters or scroll).", "error")
                return

            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.ID, 'ctl00_ContentPlaceHolder1_'))
                )
            except (TimeoutException, NoSuchElementException):
                pass
            try:
                sign_btn = driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_btn")
                driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", sign_btn)
                time.sleep(0.5)
                driver.execute_script("arguments[0].click();", sign_btn)
                self.app.log_message(self.log_display, "Clicked Signature.")
                
                try:
                    alert = wait.until(EC.alert_is_present())
                    alert.accept()
                    self._log_result(log_name, "Success", f"Deleted {fto_text}")
                except TimeoutException:
                    self._log_result(log_name, "Check", f"Attempted {fto_text}")
            except Exception as e:
                self.app.log_message(self.log_display, f"Sign Error: {e}", "error")

        except Exception as e:
            self._log_result(log_name, "Error", str(e))

    def _go_to_mr_tracking(self):
        """Uses data scraped DURING generation to switch tabs."""
        print(f"DEBUG: Using Stored Data: {self.stored_location_data}")
        self.app.switch_to_mr_tracking_for_abps(self.stored_location_data)