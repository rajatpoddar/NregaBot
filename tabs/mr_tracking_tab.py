# tabs/mr_tracking_tab.py
import subprocess
import json
import tkinter
from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk
import time, os, re
from datetime import datetime
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# --- Imports jo add kiye gaye hain ---
from selenium import webdriver
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
# --- End Imports ---

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException

from fpdf import FPDF
from PIL import Image, ImageDraw, ImageFont # Import Pillow
from utils import resource_path
from .base_tab import BaseAutomationTab
from .autocomplete_widget import AutocompleteEntry
import config  # <-- Make sure config is imported

class MrTrackingTab(BaseAutomationTab):
    def __init__(self, parent, app_instance):
        super().__init__(parent, app_instance, automation_key="mr_tracking")
        
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1) # Main notebook takes up space
        
        self.report_headers = [
            "SNo.", "Panchayat Name", "Muster roll number", "Technical Staff (Designation)", 
            "Record No.", "Work Code", "date of Closure of Muster Roll", "Muster Roll Filling Status", 
            "date of generation of Wage List", "Wagelist No.", "FTO No.", "Date of generation of FTO", 
            "Date of 1st sign", "Date of 2nd sign"
        ]
        
        # --- NEW: Headers for the ABPS results tab ---
        self.abps_report_headers = [
            "Panchayat Name", "Muster Roll No.", "Work Code", "Wagelist Number", "Labour Name", "Jobcard Number"
        ]
        
        # --- Naya Badlaav: Is tab ka apna driver hoga ---
        self.driver = None
        
        self._create_widgets()
        self.load_inputs()

    def _create_widgets(self):
        # Frame for all user input controls
        controls_frame = ctk.CTkFrame(self)
        controls_frame.grid(row=0, column=0, sticky="new", padx=10, pady=10)
        controls_frame.grid_columnconfigure(1, weight=1)

        # --- Input Fields ---
        ctk.CTkLabel(controls_frame, text="State:").grid(row=0, column=0, sticky='w', padx=15, pady=(15, 5))
        self.state_entry = AutocompleteEntry(controls_frame, 
                                             suggestions_list=self.app.history_manager.get_suggestions("mr_track_state"),
                                             app_instance=self.app, history_key="mr_track_state")
        self.state_entry.grid(row=0, column=1, sticky='ew', padx=15, pady=(15, 5))

        ctk.CTkLabel(controls_frame, text="District:").grid(row=1, column=0, sticky='w', padx=15, pady=5)
        self.district_entry = AutocompleteEntry(controls_frame, 
                                                suggestions_list=self.app.history_manager.get_suggestions("mr_track_district"),
                                                app_instance=self.app, history_key="mr_track_district")
        self.district_entry.grid(row=1, column=1, sticky='ew', padx=15, pady=5)

        ctk.CTkLabel(controls_frame, text="Block:").grid(row=2, column=0, sticky='w', padx=15, pady=5)
        self.block_entry = AutocompleteEntry(controls_frame, 
                                             suggestions_list=self.app.history_manager.get_suggestions("mr_track_block"),
                                             app_instance=self.app, history_key="mr_track_block")
        self.block_entry.grid(row=2, column=1, sticky='ew', padx=15, pady=5)

        ctk.CTkLabel(controls_frame, text="Panchayat:").grid(row=3, column=0, sticky='w', padx=15, pady=5)
        self.panchayat_entry = AutocompleteEntry(controls_frame, 
                                                 suggestions_list=self.app.history_manager.get_suggestions("mr_track_panchayat"),
                                                 app_instance=self.app, history_key="mr_track_panchayat")
        self.panchayat_entry.grid(row=3, column=1, sticky='ew', padx=15, pady=5)

        # --- Filter Checkboxes ---
        filter_frame = ctk.CTkFrame(controls_frame, fg_color="transparent")
        filter_frame.grid(row=4, column=1, sticky="w", padx=15, pady=10)

        self.pending_only_var = tkinter.IntVar(value=0)
        self.pending_only_check = ctk.CTkCheckBox(filter_frame, 
                                                  text="Show only 'Pending for filling'", 
                                                  variable=self.pending_only_var,
                                                  command=self._on_filter_check_changed)
        self.pending_only_check.pack(side="left")

        # --- NEW Zero MR Checkbox ---
        self.zero_mr_filter_var = tkinter.IntVar(value=0)
        self.zero_mr_filter_check = ctk.CTkCheckBox(filter_frame,
                                                    text="Find T+8 to T+15 (for Zero MR)",
                                                    variable=self.zero_mr_filter_var,
                                                    command=self._on_filter_check_changed)
        self.zero_mr_filter_check.pack(side="left", padx=(20, 0))

        self.abps_pending_var = tkinter.IntVar(value=0)
        self.abps_pending_check = ctk.CTkCheckBox(filter_frame, 
                                                  text="Show only 'Pending for ABPS'", 
                                                  variable=self.abps_pending_var,
                                                  command=self._on_filter_check_changed)
        self.abps_pending_check.pack(side="left", padx=(20, 0))

        action_frame = self._create_action_buttons(parent_frame=controls_frame)
        action_frame.grid(row=5, column=0, columnspan=2, pady=10)

        # --- Output Tabs ---
        notebook = ctk.CTkTabview(self)
        notebook.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 10))
        workcode_tab = notebook.add("Workcode List")
        results_tab = notebook.add("Results Table")
        abps_results_tab = notebook.add("ABPS Pendency Results") 
        self._create_log_and_status_area(parent_notebook=notebook)

        # 1. Workcode List Tab
        workcode_tab.grid_columnconfigure(0, weight=1)
        workcode_tab.grid_rowconfigure(1, weight=1)
        
        copy_frame = ctk.CTkFrame(workcode_tab, fg_color="transparent")
        copy_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 0))
        
        self.copy_wc_button = ctk.CTkButton(copy_frame, text="Copy Workcodes", command=self._copy_workcodes)
        self.copy_wc_button.pack(side="left")

        self.run_mr_payment_button = ctk.CTkButton(copy_frame, 
                                                   text="Run MR Payment", 
                                                   command=self._run_mr_payment, 
                                                   fg_color="#108842", 
                                                   hover_color="#1A994C")
        self.run_mr_payment_button.pack_forget() 
        
        self.run_emb_entry_button = ctk.CTkButton(copy_frame,
                                                    text="Run eMB Entry",
                                                    command=self._run_emb_entry,
                                                    fg_color="#0A708C", 
                                                    hover_color="#0E95BA")
        self.run_emb_entry_button.pack_forget() 

        self.run_zero_mr_button = ctk.CTkButton(copy_frame,
                                                  text="Forward to Zero MR",
                                                  command=self._run_zero_mr,
                                                  fg_color="#D9534F", 
                                                  hover_color="#C9302C")
        self.run_zero_mr_button.pack_forget()

        self.workcode_textbox = ctk.CTkTextbox(workcode_tab, state="disabled")
        self.workcode_textbox.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)

        # 2. Results Tab (Table)
        results_tab.grid_columnconfigure(0, weight=1)
        results_tab.grid_rowconfigure(1, weight=1)
        
        export_frame = ctk.CTkFrame(results_tab, fg_color="transparent")
        export_frame.grid(row=0, column=0, sticky="w", padx=5, pady=5)
        
        self.export_button = ctk.CTkButton(export_frame, text="Export Report", command=self.export_report)
        self.export_button.pack(side="left")
        
        self.export_format_menu = ctk.CTkOptionMenu(export_frame, values=["Excel (.xlsx)", "PDF (.pdf)", "PNG (.png)"])
        self.export_format_menu.pack(side="left", padx=5)

        # --- NEW BUTTON FOR PENDENCY REPORT ---
        self.generate_pendency_btn = ctk.CTkButton(
            export_frame, 
            text="Generate Pendency Report (T0-T8)", 
            command=self._open_pendency_report_window,
            fg_color="#B45309", hover_color="#92400E"
        )
        self.generate_pendency_btn.pack(side="left", padx=15)
        # --------------------------------------

        self.results_tree = ttk.Treeview(results_tab, columns=self.report_headers, show='headings')
        for col in self.report_headers: self.results_tree.heading(col, text=col)
        self.results_tree.column("SNo.", width=40, anchor='center')
        self.results_tree.column("Muster roll number", width=80)
        self.results_tree.column("Work Code", width=250)
        self.results_tree.column("Muster Roll Filling Status", width=250)
        self.results_tree.column("Technical Staff (Designation)", width=200)

        self.results_tree.grid(row=1, column=0, sticky='nsew', padx=5, pady=5)
        scrollbar = ctk.CTkScrollbar(results_tab, command=self.results_tree.yview)
        self.results_tree.configure(yscroll=scrollbar.set); scrollbar.grid(row=1, column=1, sticky='ns')
        self.style_treeview(self.results_tree)
        
        # 3. ABPS Pendency Results Tab
        abps_results_tab.grid_columnconfigure(0, weight=1)
        abps_results_tab.grid_rowconfigure(1, weight=1)
        
        abps_export_frame = ctk.CTkFrame(abps_results_tab, fg_color="transparent")
        abps_export_frame.grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.abps_export_button = ctk.CTkButton(abps_export_frame, text="Export ABPS Report", command=self._export_abps_report)
        self.abps_export_button.pack(side="left")
        self.abps_export_format_menu = ctk.CTkOptionMenu(abps_export_frame, values=["Excel (.xlsx)"])
        self.abps_export_format_menu.pack(side="left", padx=5)

        self.abps_results_tree = ttk.Treeview(abps_results_tab, columns=self.abps_report_headers, show='headings')
        for col in self.abps_report_headers: self.abps_results_tree.heading(col, text=col)
        self.abps_results_tree.column("Panchayat Name", width=120)
        self.abps_results_tree.column("Muster Roll No.", width=100)
        self.abps_results_tree.column("Work Code", width=250)
        self.abps_results_tree.column("Wagelist Number", width=150)
        self.abps_results_tree.column("Labour Name", width=150)
        self.abps_results_tree.column("Jobcard Number", width=150)
        
        self.abps_results_tree.grid(row=1, column=0, sticky='nsew', padx=5, pady=5)
        abps_scrollbar = ctk.CTkScrollbar(abps_results_tab, command=self.abps_results_tree.yview)
        self.abps_results_tree.configure(yscroll=abps_scrollbar.set); abps_scrollbar.grid(row=1, column=1, sticky='ns')
        self.style_treeview(self.abps_results_tree)

    def _on_filter_check_changed(self):
        if self.zero_mr_filter_var.get() == 1:
            self.pending_only_check.configure(state="disabled")
            self.abps_pending_check.configure(state="disabled")
            self.pending_only_var.set(0)
            self.abps_pending_var.set(0)
        elif self.abps_pending_var.get() == 1:
            self.pending_only_check.configure(state="disabled")
            self.zero_mr_filter_check.configure(state="disabled")
            self.pending_only_var.set(0)
            self.zero_mr_filter_var.set(0)
        elif self.pending_only_var.get() == 1:
            self.abps_pending_check.configure(state="disabled")
            self.zero_mr_filter_check.configure(state="disabled")
            self.abps_pending_var.set(0)
            self.zero_mr_filter_var.set(0)
        else:
            self.pending_only_check.configure(state="normal")
            self.abps_pending_check.configure(state="normal")
            self.zero_mr_filter_check.configure(state="normal")

    def set_ui_state(self, running: bool):
        self.set_common_ui_state(running)
        state = "disabled" if running else "normal"
        
        self.state_entry.configure(state=state)
        self.district_entry.configure(state=state)
        self.block_entry.configure(state=state)
        self.panchayat_entry.configure(state=state)
        
        self.pending_only_check.configure(state=state)
        self.abps_pending_check.configure(state=state)
        self.zero_mr_filter_check.configure(state=state)
        
        if state == "normal":
            self._on_filter_check_changed()
        
        self.run_mr_payment_button.configure(state=state)
        self.run_emb_entry_button.configure(state=state)
        self.run_zero_mr_button.configure(state=state)
        self.generate_pendency_btn.configure(state=state) # Control new button state
        
        self.abps_export_button.configure(state=state)
        self.abps_export_format_menu.configure(state=state)

    def set_for_abps_check(self):
        self.reset_ui(reset_all_filters=False) 
        self.abps_pending_var.set(1)
        self._on_filter_check_changed()

    def reset_ui(self, reset_all_filters=True):
        if reset_all_filters:
            self.pending_only_var.set(0)
            self.abps_pending_var.set(0)
            self.zero_mr_filter_var.set(0)
            self._on_filter_check_changed()
        
        for item in self.results_tree.get_children(): self.results_tree.delete(item)
        for item in self.abps_results_tree.get_children(): self.abps_results_tree.delete(item)
        self._update_workcode_textbox("")
        
        self.app.log_message(self.log_display, "Form has been reset.")
        self.update_status("Ready", 0.0)
        
    def _get_new_driver(self):
        self.app.log_message(self.log_display, "Naya Headless Chrome browser shuru kar raha hoon...", "info")
        try:
            chrome_options = ChromeOptions()
            chrome_options.add_argument("--headless=new")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--disable-software-rasterizer")
            chrome_options.add_argument("--log-level=3")
            chrome_options.add_argument("--silent")
            chrome_options.add_argument("--window-size=1920,1080") 
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"]) 
            
            service = ChromeService(ChromeDriverManager().install())
            service.creation_flags = subprocess.CREATE_NO_WINDOW if config.OS_SYSTEM == "Windows" else 0
            
            driver = webdriver.Chrome(service=service, options=chrome_options)
            
            self.app.log_message(self.log_display, "Headless browser safaltapoorvak shuru ho gaya.", "info")
            return driver
        except Exception as e:
            self.app.log_message(self.log_display, f"Headless browser shuru karne mein BADI GADBAD: {e}", "error")
            messagebox.showerror("Browser Error", f"Naya Headless Chrome browser shuru nahi ho saka.\n\nError: {e}\n\nKya Chrome installed hai?")
            return None
        
    def start_automation(self):
        self.run_mr_payment_button.pack_forget() 
        self.run_emb_entry_button.pack_forget() 
        self.run_zero_mr_button.pack_forget() 
        
        for item in self.results_tree.get_children(): self.results_tree.delete(item)
        for item in self.abps_results_tree.get_children(): self.abps_results_tree.delete(item) 
        self._update_workcode_textbox("") 
        
        inputs = {
            'state': self.state_entry.get().strip(), 
            'district': self.district_entry.get().strip(), 
            'block': self.block_entry.get().strip(),
            'panchayat': self.panchayat_entry.get().strip(),
            'pending_only': self.pending_only_var.get() == 1,
            'abps_pending': self.abps_pending_var.get() == 1,
            'zero_mr_filter': self.zero_mr_filter_var.get() == 1
        }
        
        if not all([inputs['state'], inputs['district'], inputs['block'], inputs['panchayat']]):
            messagebox.showwarning("Input Error", "State, District, Block, and Panchayat are required."); return
        
        self.save_inputs(inputs)
        self.app.update_history("mr_track_state", inputs['state'])
        self.app.update_history("mr_track_district", inputs['district'])
        self.app.update_history("mr_track_block", inputs['block'])
        self.app.update_history("mr_track_panchayat", inputs['panchayat'])
        
        if self.driver:
            self.app.log_message(self.log_display, "Ek automation pehle se chal raha hai. Rukiye...", "warning")
            messagebox.showwarning("Automation Jaari Hai", "Report generation pehle se chal raha hai.")
            return
        
        self.driver = self._get_new_driver()
        if not self.driver:
            self.app.log_message(self.log_display, "ERROR: WebDriver nahi mila. Automation ruka.", "error")
            return
        
        self.app.after(0, self.set_ui_state, True) 
        self.app.start_automation_thread(self.automation_key, self.run_automation_logic, args=(inputs,))

    def run_automation_logic(self, inputs):
        self.app.after(0, self.app.set_status, "Starting MR Tracking...") 
        self.app.after(0, self.update_status, "Initializing...", 0.0) 
        self.app.clear_log(self.log_display)
        self.app.log_message(self.log_display, "Starting MR Tracking automation...")
        
        self.zero_mr_data = [] 
        
        try:
            driver = self.driver 
            if not driver:
                self.app.log_message(self.log_display, "ERROR: Browser driver not found.", "error")
                return 
                
            wait = WebDriverWait(driver, 20)
            
            url = config.MR_TRACKING_CONFIG["url"]
            self.app.after(0, self.app.set_status, "Navigating to MR Tracking...")
            self.app.after(0, self.update_status, "Navigating...", 0.1)
            self.app.log_message(self.log_display, f"Navigating to MR Tracking page...")
            driver.get(url)
            
            main_window_handle = driver.current_window_handle 
            
            STATE_ID = "ctl00_ContentPlaceHolder1_ddl_state"
            DIST_ID = "ctl00_ContentPlaceHolder1_ddl_dist"
            BLOCK_ID = "ctl00_ContentPlaceHolder1_ddl_blk"
            PANCH_ID = "ctl00_ContentPlaceHolder1_ddl_pan"
            RADIO_PAYMENT_PENDING_ID = "ctl00_ContentPlaceHolder1_Rbtn_pay_1"
            RADIO_T8_T15_ID = "ctl00_ContentPlaceHolder1_Rbtn_pay_2" 
            SUBMIT_BTN_ID = "ctl00_ContentPlaceHolder1_Button1"
            TABLE_XPATH = "//table[@bordercolor='#EBEBEB' and .//b[text()='SNo.']]"

            def wait_for_dropdown(dropdown_id, step_name, progress):
                self.app.after(0, self.app.set_status, f"Waiting for {step_name}...")
                self.app.after(0, self.update_status, f"Waiting for {step_name}...", progress)
                self.app.log_message(self.log_display, f"Waiting for dropdown {dropdown_id} to populate...")
                wait.until(
                    EC.presence_of_element_located((By.XPATH, f"//select[@id='{dropdown_id}']/option[position()>1]"))
                )
                self.app.log_message(self.log_display, "Dropdown populated.")
                time.sleep(0.5) 

            self.app.after(0, self.app.set_status, f"Selecting State: {inputs['state']}")
            self.app.after(0, self.update_status, "Selecting State...", 0.15)
            self.app.log_message(self.log_display, f"Selecting State: {inputs['state']}")
            state_select = Select(wait.until(EC.element_to_be_clickable((By.ID, STATE_ID))))
            state_select.select_by_visible_text(inputs['state'].upper())
            wait_for_dropdown(DIST_ID, "Districts", 0.2)

            self.app.after(0, self.app.set_status, f"Selecting District: {inputs['district']}")
            self.app.after(0, self.update_status, "Selecting District...", 0.25)
            self.app.log_message(self.log_display, f"Selecting District: {inputs['district']}")
            dist_select = Select(wait.until(EC.element_to_be_clickable((By.ID, DIST_ID))))
            dist_select.select_by_visible_text(inputs['district'].upper())
            wait_for_dropdown(BLOCK_ID, "Blocks", 0.3)

            self.app.after(0, self.app.set_status, f"Selecting Block: {inputs['block']}")
            self.app.after(0, self.update_status, "Selecting Block...", 0.35)
            self.app.log_message(self.log_display, f"Selecting Block: {inputs['block']}")
            Select(wait.until(EC.element_to_be_clickable((By.ID, BLOCK_ID)))).select_by_visible_text(inputs['block'])
            
            self.app.after(0, self.app.set_status, f"Selecting Panchayat: {inputs['panchayat']}")
            self.app.after(0, self.update_status, "Selecting Panchayat...", 0.45)
            self.app.log_message(self.log_display, f"Selecting Panchayat: {inputs['panchayat']}")
            Select(wait.until(EC.element_to_be_clickable((By.ID, PANCH_ID)))).select_by_visible_text(inputs['panchayat'])
            
            self.app.after(0, self.app.set_status, "Setting filter...")
            self.app.after(0, self.update_status, "Setting filter...", 0.5)
            
            if inputs['zero_mr_filter']:
                self.app.log_message(self.log_display, "Selecting '...T+8 and T+15'")
                wait.until(EC.element_to_be_clickable((By.ID, RADIO_T8_T15_ID))).click()
            else:
                self.app.log_message(self.log_display, "Selecting 'Where payment is pending'")
                wait.until(EC.element_to_be_clickable((By.ID, RADIO_PAYMENT_PENDING_ID))).click()
            
            self.app.after(0, self.app.set_status, "Submitting form...")
            self.app.after(0, self.update_status, "Submitting form...", 0.55)
            self.app.log_message(self.log_display, "Submitting form...")
            wait.until(EC.element_to_be_clickable((By.ID, SUBMIT_BTN_ID))).click()
            
            self.app.after(0, self.app.set_status, "Waiting for report...")
            self.app.after(0, self.update_status, "Waiting for report...", 0.6)
            self.app.log_message(self.log_display, "Waiting for report table...")
            table = wait.until(EC.presence_of_element_located((By.XPATH, TABLE_XPATH)))
            rows = table.find_elements(By.XPATH, ".//tr[position()>1]") 
            
            total_rows = len(rows)
            if total_rows == 0:
                self.app.log_message(self.log_display, "No records found for the selected criteria.", "warning")
                messagebox.showinfo("No Data", "No records found for the selected criteria.")
                self.success_message = None
                return

            self.app.log_message(self.log_display, f"Found {total_rows} records. Processing...")
            
            workcode_list = []
            displayed_rows = 0
            abps_pending_count = 0
            pending_filling_count = 0
            abps_pending_mrs = [] 
            
            for i, row in enumerate(rows):
                if self.app.stop_events[self.automation_key].is_set():
                    self.app.log_message(self.log_display, "Stop signal received.", "warning")
                    break
                
                progress = 0.6 + ( (i + 1) / total_rows ) * 0.2
                status_msg = f"Processing row {i+1}/{total_rows}"
                self.app.after(0, self.app.set_status, status_msg)
                self.app.after(0, self.update_status, status_msg, progress)
                
                cells = row.find_elements(By.TAG_NAME, "td")
                if not cells or len(cells) < len(self.report_headers):
                    continue
                    
                row_data = [cell.text.strip() for cell in cells]
                
                panchayat_name = row_data[1] 
                muster_roll_no = row_data[2] 
                work_code = row_data[5]
                muster_status = row_data[7]
                wagelist_no = row_data[9]
                fto_no = row_data[10]
                fto_date = row_data[11]
                first_sign_date = row_data[12]

                is_abps_pending = "Pending for signature of 1st Signatory" in first_sign_date and not fto_no and not fto_date
                is_pending_filling = "Pending for filling" in muster_status

                if inputs['abps_pending']:
                    if not is_abps_pending:
                        continue 
                elif inputs['pending_only']:
                    if not is_pending_filling:
                        continue 

                    if "since 0 days" in muster_status or "since 1 days" in muster_status or "since 1 Day" in muster_status:
                        self.app.log_message(self.log_display, f"Skipping MR {muster_roll_no} (0/1 days pending).", "info")
                        continue 
                    
                    
                elif inputs['zero_mr_filter']:
                    self.zero_mr_data.append({
                        "panchayat": panchayat_name,
                        "work_code": work_code,
                        "msr_no": muster_roll_no
                    })
                
                self.app.after(0, lambda data=tuple(row_data): self.results_tree.insert("", "end", values=data))
                displayed_rows += 1
                
                if work_code:
                    workcode_list.append(work_code)
                
                if is_abps_pending:
                    abps_pending_count += 1
                    abps_pending_mrs.append({
                        "panchayat": row_data[1],
                        "mr_no": row_data[2],
                        "work_code": work_code,
                        "wagelist_no": wagelist_no
                    })
                if is_pending_filling:
                    pending_filling_count += 1

            if self.app.stop_events[self.automation_key].is_set():
                 self.app.log_message(self.log_display, "Automation stopped by user.", "warning")
                 self.success_message = None 
                 return 

            self.app.after(0, self._update_workcode_textbox, "\n".join(workcode_list))
            
            if inputs['abps_pending'] and abps_pending_mrs:
                self.app.log_message(self.log_display, f"Found {abps_pending_count} MRs pending for ABPS. Now finding workers...")
                
                wagelists_to_search = {}
                for mr in abps_pending_mrs:
                  wl = mr["wagelist_no"]
                  if not wl: 
                      self.app.log_message(self.log_display, f"Skipping MR {mr['mr_no']} (Workcode: {mr['work_code']}) as Wagelist No. is blank.", "warning")
                      continue
                  if wl not in wagelists_to_search:
                    wagelists_to_search[wl] = []
                  wagelists_to_search[wl].append(mr)
                
                self.app.log_message(self.log_display, f"Found {len(wagelists_to_search)} unique wagelists to scan.")
                
                total_wl = len(wagelists_to_search)
                for i, (wagelist_no, mr_list) in enumerate(wagelists_to_search.items()):
                    if self.app.stop_events[self.automation_key].is_set(): break
                    
                    progress = 0.8 + ( (i + 1) / total_wl ) * 0.2
                    status_msg = f"Scanning Wagelist {i+1}/{total_wl} ({wagelist_no})"
                    self.app.after(0, self.app.set_status, status_msg)
                    self.app.after(0, self.update_status, status_msg, progress)
                    
                    self._search_wagelist_for_pending_abps(driver, wait, inputs, wagelist_no, mr_list, main_window_handle)

                if driver.current_window_handle != main_window_handle:
                    driver.switch_to.window(main_window_handle)

            if inputs['abps_pending']:
                self.success_message = f"MR Tracking complete. Found {abps_pending_count} MRs pending for ABPS."
            elif inputs['pending_only']:
                self.success_message = f"MR Tracking complete. Found {pending_filling_count} MRs pending for filling."
            elif inputs['zero_mr_filter']:
                self.success_message = f"MR Tracking complete. Found {len(self.zero_mr_data)} MRs for Zero MR processing."
            else:
                self.success_message = f"MR Tracking complete. Displayed {displayed_rows} total records."
            
            self.app.log_message(self.log_display, f"Processing complete. {self.success_message.replace('MR Tracking complete. ', '')}", "success")
            
        except (TimeoutException, NoSuchElementException, StaleElementReferenceException) as e:
            if driver and "Session Expired" in driver.page_source:
                self.app.log_message(self.log_display, "Session expired, par yeh headless hai, isliye retry nahi kar rahe.", "warning")
                error_msg = "Session Expired during headless operation. Please try again."
                self.app.log_message(self.log_display, error_msg, "error")
                messagebox.showerror("Automation Error", error_msg)
            else:
                error_msg = f"A browser error occurred: {str(e).splitlines()[0]}"
                self.app.log_message(self.log_display, error_msg, "error")
                messagebox.showerror("Automation Error", error_msg)
                
            self.app.after(0, self.app.set_status, "Browser Error")
            self.success_message = None
        except Exception as e:
            self.app.log_message(self.log_display, f"An unexpected error occurred: {e}", "error")
            messagebox.showerror("Critical Error", f"An unexpected error occurred: {e}")
            self.app.after(0, self.app.set_status, "Unexpected Error")
            self.success_message = None
        finally:
            if self.driver: 
                try:
                    self.driver.quit()
                    self.app.after(0, self.app.log_message, self.log_display, "Automation ne browser ko band kar diya hai.", "info")
                except Exception as e:
                    self.app.after(0, self.app.log_message, self.log_display, f"Browser band karne mein error: {e}", "warning")
            
            self.driver = None 
            
            self.app.after(0, self.set_ui_state, False) 
            
            final_app_status = "Automation Stopped" if self.app.stop_events[self.automation_key].is_set() else \
                              ("Automation Finished" if hasattr(self, 'success_message') and self.success_message else "Automation Failed")
            final_tab_status = "Stopped" if self.app.stop_events[self.automation_key].is_set() else \
                              ("Finished" if hasattr(self, 'success_message') and self.success_message else "Failed")

            self.app.after(0, self.app.set_status, final_app_status)
            self.app.after(0, self.update_status, final_tab_status, 1.0)

            if not self.app.stop_events[self.automation_key].is_set():
                 self.app.after(5000, lambda: self.app.set_status("Ready")) 
                 self.app.after(5000, lambda: self.update_status("Ready", 0.0)) 

            if hasattr(self, 'success_message') and self.success_message and not self.app.stop_events[self.automation_key].is_set():
                self.app.after(100, lambda: messagebox.showinfo("Complete", self.success_message))
                
                if inputs.get('zero_mr_filter', False):
                    self.app.after(0, lambda: self.run_zero_mr_button.pack(side="left", padx=(10, 0)))
                else:
                    self.app.after(0, lambda: self.run_mr_payment_button.pack(side="left", padx=(10, 0)))
                    self.app.after(0, lambda: self.run_emb_entry_button.pack(side="left", padx=(10, 0)))

    def _search_wagelist_for_pending_abps(self, driver, wait, inputs, wagelist_no, mr_list, main_window_handle):
        try:
            self.app.log_message(self.log_display, f"   Opening homesearch tab for {wagelist_no}...")
            driver.execute_script("window.open(arguments[0], '_blank');", "https://mnregaweb4.nic.in/netnrega/homesearch.htm")
            time.sleep(1) 
            
            popup_handle = [handle for handle in driver.window_handles if handle != main_window_handle][-1]
            driver.switch_to.window(popup_handle)

            self.app.log_message(self.log_display, "   Waiting for iframe...")
            wait.until(EC.frame_to_be_available_and_switch_to_it((By.TAG_NAME, "iframe")))
            self.app.log_message(self.log_display, "   ...Switched to iframe.")
            
            self.app.log_message(self.log_display, "   Selecting 'WageList' from dropdown...")
            Select(wait.until(EC.element_to_be_clickable((By.ID, "ddl_search")))).select_by_value("WageList")
            
            self.app.log_message(self.log_display, "   Waiting for State dropdown to populate (Postback 1)...")
            wait.until(EC.presence_of_element_located((By.XPATH, "//select[@id='ddl_state']/option[text()='ANDAMAN AND NICOBAR']")))
            self.app.log_message(self.log_display, "   ...State dropdown populated.")
            
            self.app.log_message(self.log_display, f"   Selecting State: {inputs['state'].upper()}...")
            state_select = Select(wait.until(EC.element_to_be_clickable((By.ID, "ddl_state"))))
            state_select.select_by_visible_text(inputs['state'].upper())
            
            self.app.log_message(self.log_display, "   Waiting for District dropdown to populate (Postback 2)...")
            wait.until(EC.presence_of_element_located((By.XPATH, f"//select[@id='ddl_district']/option[text()='{inputs['district'].upper()}']")))
            self.app.log_message(self.log_display, "   ...District dropdown populated.")
            
            self.app.log_message(self.log_display, f"   Selecting District: {inputs['district'].upper()}...")
            dist_select = Select(driver.find_element(By.ID, "ddl_district"))
            dist_select.select_by_visible_text(inputs['district'].upper()) 
            
            self.app.log_message(self.log_display, "   Waiting for final postback (2 sec)...")
            time.sleep(2)
            self.app.log_message(self.log_display, "   ...Wait complete.")

            self.app.log_message(self.log_display, f"   Entering Wagelist No: {wagelist_no}...")
            keyword_box = wait.until(EC.element_to_be_clickable((By.ID, "txt_keyword2")))
            keyword_box.send_keys(wagelist_no)
            
            self.app.log_message(self.log_display, "   Clicking 'GO'...")
            driver.find_element(By.XPATH, "//input[@value='GO']").click()

            self.app.log_message(self.log_display, "   Waiting for search result popup...")
            wait.until(EC.number_of_windows_to_be(3))
            self.app.log_message(self.log_display, "   ...Search result popup appeared.")
            
            wagelist_search_popup_handle = [h for h in driver.window_handles if h != main_window_handle and h != popup_handle][0]
            driver.switch_to.window(wagelist_search_popup_handle)

            self.app.log_message(self.log_display, "   Clicking wagelist link in popup...")
            wl_link = wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, wagelist_no)))
            wl_link.click()

            self.app.log_message(self.log_display, "   Waiting for wagelist details page...")
            wait.until(EC.presence_of_element_located((By.ID, "lb_main")))
            self.app.log_message(self.log_display, "   ...Wagelist details page loaded.")
            
            self.app.log_message(self.log_display, f"   Scanning {wagelist_no} for pending workers...")
            details_table = wait.until(EC.presence_of_element_located((By.XPATH, "//span[@id='lb_main']/ancestor::center/table[1]")))
            worker_rows = details_table.find_elements(By.XPATH, ".//tr[position() > 1]") 
            
            found_workers = set() 
            
            for row in worker_rows:
                cells = row.find_elements(By.TAG_NAME, "td")
                
                if len(cells) < 15: continue 
                
                jobcard_no = cells[8].text.strip()
                applicant_name = cells[9].text.strip()
                fto_no = cells[12].text.strip()
                
                if not fto_no and (jobcard_no, applicant_name) not in found_workers:
                    found_workers.add((jobcard_no, applicant_name))
                    self.app.log_message(self.log_display, f"      > Found pending: {applicant_name} ({jobcard_no})")
                    for mr in mr_list:
                        result_data = (mr["panchayat"], mr["mr_no"], mr["work_code"], wagelist_no, applicant_name, jobcard_no)
                        self.app.after(0, lambda data=result_data: self.abps_results_tree.insert("", "end", values=data))
            
            if not found_workers:
                 self.app.log_message(self.log_display, f"   No pending workers found in {wagelist_no}.")

        except Exception as e:
            self.app.log_message(self.log_display, f"   ERROR scanning wagelist {wagelist_no}: {type(e).__name__} {str(e).splitlines()[0]}", "error")
        finally:
            self.app.log_message(self.log_display, "   Closing popup windows...")
            for handle in driver.window_handles:
                if handle != main_window_handle:
                    driver.switch_to.window(handle)
                    driver.close()
            driver.switch_to.window(main_window_handle)
            self.app.log_message(self.log_display, "   ...Finished wagelist scan.")
            time.sleep(0.5) 

    # --- PENDENCY REPORT FEATURE (T0 to T8+) ---

    def _open_pendency_report_window(self):
        """Calculates and displays the Pendency Report based on current table data."""
        items = self.results_tree.get_children()
        if not items:
            messagebox.showinfo("No Data", "Please run the MR Tracking automation first to get data.")
            return

        # --- Calculate Data ---
        summary_data = self._process_pendency_data(items)
        if not summary_data:
            messagebox.showinfo("No Pendency Data", "Could not find any 'since X days' text in the current results.")
            return

        # --- Create Popup Window ---
        win = ctk.CTkToplevel(self)
        win.title("Pendency Report (T0 - T8)")
        win.geometry("900x500")
        win.transient(self) 
        
        # Grid Configuration
        win.grid_columnconfigure(0, weight=1)
        win.grid_rowconfigure(1, weight=1)

        # Header
        header_frame = ctk.CTkFrame(win, fg_color="transparent")
        header_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=10)
        ctk.CTkLabel(header_frame, text="Panchayat-wise Pendency Analysis", font=("Arial", 16, "bold")).pack(side="left")
        
        export_btn = ctk.CTkButton(header_frame, text="Export Excel", 
                                   command=lambda: self._export_pendency_excel(summary_data),
                                   fg_color="#108842", hover_color="#1A994C")
        export_btn.pack(side="right")

        # Table
        cols = ["Panchayat", "Total MRs", "T0", "T1", "T2", "T3", "T4", "T5", "T6", "T7", "T8+"]
        tree = ttk.Treeview(win, columns=cols, show='headings')
        
        for col in cols:
            tree.heading(col, text=col)
            width = 150 if col == "Panchayat" else 60
            tree.column(col, width=width, anchor="center")
            
        tree.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 10))
        scrollbar = ctk.CTkScrollbar(win, command=tree.yview)
        tree.configure(yscroll=scrollbar.set)
        scrollbar.grid(row=1, column=1, sticky="ns", pady=(0, 10))
        
        # Populate Table
        sorted_panchayats = sorted(summary_data.keys())
        for panch in sorted_panchayats:
            row = summary_data[panch]
            # Calculate total
            total_mrs = sum(row[f"T{i}"] for i in range(9))
            
            values = (
                panch, total_mrs,
                row["T0"], row["T1"], row["T2"], row["T3"],
                row["T4"], row["T5"], row["T6"], row["T7"], row["T8"]
            )
            tree.insert("", "end", values=values)
            
        self.style_treeview(tree)

    def _process_pendency_data(self, tree_items):
        """
        Parses tree items to count days pending.
        Constraints:
        1. Look for 'since X days' in text.
        2. Unique MR per Panchayat (don't count same MR twice).
        """
        summary = {} # { "PanchayatName": {T0:0, T1:0... seen_mrs: set()} }
        
        # Regex to find number of days
        # Matches: "since 5 days", "since 1 day", "since 5 Day" (Case Insensitive)
        regex = re.compile(r'since\s+(\d+)\s*(?:days|day)', re.IGNORECASE)

        for item_id in tree_items:
            values = self.results_tree.item(item_id, 'values')
            if not values: continue
            
            # Extract relevant columns
            # Index 1: Panchayat, Index 2: MR No
            # Index 7: Status, Index 12: 1st Sign, Index 13: 2nd Sign
            panchayat = values[1]
            mr_no = values[2]
            
            # Combine text fields to search
            full_text = f"{values[7]} {values[12]} {values[13]}"
            
            match = regex.search(full_text)
            if not match:
                continue # Skip if no "since X days" found
                
            days_pending = int(match.group(1))
            
            # Initialize Panchayat Data
            if panchayat not in summary:
                summary[panchayat] = {f"T{i}": 0 for i in range(9)}
                summary[panchayat]["seen_mrs"] = set()
            
            # Check Duplicate MR (Constraint: Same MR count 1 hi hoga)
            if mr_no in summary[panchayat]["seen_mrs"]:
                continue
            
            # Add to seen
            summary[panchayat]["seen_mrs"].add(mr_no)
            
            # Bucket Allocation
            if days_pending >= 8:
                summary[panchayat]["T8"] += 1
            else:
                summary[panchayat][f"T{days_pending}"] += 1
                
        return summary

    def _export_pendency_excel(self, summary_data):
        if not summary_data: return
        
        # Prepare Data
        export_list = []
        for panch in sorted(summary_data.keys()):
            row = summary_data[panch]
            total = sum(row[f"T{i}"] for i in range(9))
            export_list.append([
                panch, total, 
                row["T0"], row["T1"], row["T2"], row["T3"],
                row["T4"], row["T5"], row["T6"], row["T7"], row["T8"]
            ])
            
        columns = ["Panchayat", "Total Pending", "T0", "T1", "T2", "T3", "T4", "T5", "T6", "T7", "T8+"]
        
        # File Dialog
        filename = f"Pendency_Report_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", 
            initialfile=filename,
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if not save_path: return
        
        try:
            df = pd.DataFrame(export_list, columns=columns)
            
            # Formatting with OpenPyXL
            with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                sheet_name = 'Pendency'
                # Row 1-4 reserved for Title/Subtitle, Data starts from Row 5
                df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=4)
                
                wb = writer.book
                ws = writer.sheets[sheet_name]
                
                # --- Styles Definition ---
                # Headers
                header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark Blue
                title_font = Font(size=14, bold=True, color="FFFFFF")
                subtitle_font = Font(italic=True, size=9)
                
                # Table Columns
                col_header_fill = PatternFill(start_color="DCE6F1", fill_type="solid") # Light Blue
                col_header_font = Font(bold=True)
                
                # Data Rows
                white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
                # Alignment & Border
                center_align = Alignment(horizontal="center", vertical="center")
                left_align = Alignment(horizontal="left", vertical="center")
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                     top=Side(style='thin'), bottom=Side(style='thin'))

                # --- 1. Main Title (Row 1) ---
                ws.merge_cells('A1:K1')
                cell = ws['A1']
                cell.value = f"PENDENCY REPORT (T0 - T8+)"
                cell.font = title_font
                cell.fill = header_fill
                cell.alignment = center_align
                
                # --- 2. Sub Title (Row 2) ---
                ws.merge_cells('A2:K2')
                cell = ws['A2']
                cell.value = f"Generated by NregaBot | Date: {datetime.now().strftime('%d-%m-%Y %I:%M %p')}"
                cell.font = subtitle_font
                cell.alignment = center_align

                # --- 3. Table Headers (Row 5) ---
                # Note: Pandas wrote headers at Row 5 because startrow=4
                for cell in ws[5]:
                    cell.font = col_header_font
                    cell.fill = col_header_fill
                    cell.alignment = center_align
                    cell.border = thin_border

                # --- 4. Data Rows Styling (Row 6 onwards) ---
                data_start_row = 6
                for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row, max_row=data_start_row + len(export_list) - 1, min_col=1, max_col=11)):
                    # Zebra Striping Logic
                    fill = gray_fill if row_idx % 2 == 0 else white_fill
                    
                    for cell in row:
                        cell.fill = fill
                        cell.border = thin_border
                        
                        # Formatting: Panchayat Left Aligned, Numbers Center Aligned
                        if cell.column == 1: # Column A
                            cell.alignment = left_align
                        else:
                            cell.alignment = center_align

                # --- 5. Column Widths ---
                ws.column_dimensions['A'].width = 25 # Panchayat Name Width
                for col_num in range(2, 12): # Columns B to K
                    col_letter = get_column_letter(col_num)
                    ws.column_dimensions[col_letter].width = 12
                    
            messagebox.showinfo("Success", f"Professional Pendency Report saved to:\n{save_path}")
            
            # Auto Open File
            try:
                if os.name == 'nt': os.startfile(save_path)
                else: subprocess.call(['open', save_path])
            except: pass
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to save Excel:\n{e}")

    # --- END PENDENCY REPORT FEATURE ---

    def _run_mr_payment(self):
        """Called when the 'Run MR Payment' button is clicked."""
        workcodes_raw = self.workcode_textbox.get("1.0", tkinter.END).strip()
        panchayat_name = self.panchayat_entry.get().strip()

        if not workcodes_raw:
            messagebox.showwarning("No Data", "There are no workcodes to send to the MR Payment tab.", parent=self)
            return
        
        if not panchayat_name or panchayat_name.upper() == "ALL":
            messagebox.showwarning("Invalid Panchayat", "A specific Panchayat name must be selected to run MR Payment.", parent=self)
            return

        # --- UPDATED LOGIC: Extract last 6 digits ---
        processed_list = []
        for code in workcodes_raw.splitlines():
            code = code.strip()
            if code:
                # Split by '/', take last part, take last 6 chars
                short_code = code.split('/')[-1][-6:]
                processed_list.append(short_code)
        
        final_workcodes = "\n".join(processed_list)
        # --- END UPDATED LOGIC ---

        self.app.switch_to_msr_tab_with_data(final_workcodes, panchayat_name)

    def _run_emb_entry(self):
        """Called when the 'Run eMB Entry' button is clicked."""
        workcodes_raw = self.workcode_textbox.get("1.0", tkinter.END).strip()
        panchayat_name = self.panchayat_entry.get().strip()

        if not workcodes_raw:
            messagebox.showwarning("No Data", "There are no workcodes to send to the eMB Entry tab.", parent=self)
            return
        
        if not panchayat_name or panchayat_name.upper() == "ALL":
            messagebox.showwarning("Invalid Panchayat", "A specific Panchayat name must be selected to run eMB Entry.", parent=self)
            return

        # --- UPDATED LOGIC: Extract last 6 digits ---
        processed_list = []
        for code in workcodes_raw.splitlines():
            code = code.strip()
            if code:
                # Split by '/', take last part, take last 6 chars
                short_code = code.split('/')[-1][-6:]
                processed_list.append(short_code)
        
        final_workcodes = "\n".join(processed_list)
        # --- END UPDATED LOGIC ---

        self.app.switch_to_emb_entry_with_data(final_workcodes, panchayat_name)

    def _run_zero_mr(self):
        """Called when the 'Forward to Zero MR' button is clicked."""
        if not hasattr(self, 'zero_mr_data') or not self.zero_mr_data:
            messagebox.showwarning("No Data", "No Workcode/MSR data found to forward.", parent=self)
            return

        self.app.log_message(self.log_display, f"Sending {len(self.zero_mr_data)} MRs to Zero MR tab...")
        
        # --- UPDATED LOGIC: Extract last 6 digits for Zero MR Data ---
        processed_data = []
        for item in self.zero_mr_data:
            original_wc = item['work_code']
            # Process WC: Split by '/', take last part, take last 6 chars
            short_wc = original_wc.split('/')[-1][-6:]
            
            processed_data.append({
                "panchayat": item['panchayat'],
                "work_code": short_wc, # Updated workcode
                "msr_no": item['msr_no']
            })
        # --- END UPDATED LOGIC ---
        
        self.app.switch_to_zero_mr_tab_with_data(processed_data)

    def _update_workcode_textbox(self, text):
        self.workcode_textbox.configure(state="normal")
        self.workcode_textbox.delete("1.0", tkinter.END)
        self.workcode_textbox.insert("1.0", text)
        self.workcode_textbox.configure(state="disabled")

    def _copy_workcodes(self):
        text = self.workcode_textbox.get("1.0", tkinter.END).strip()
        if text:
            self.app.clipboard_clear()
            self.app.clipboard_append(text)
            messagebox.showinfo("Copied", f"{len(text.splitlines())} workcodes copied to clipboard.", parent=self)
        else:
            messagebox.showwarning("Empty", "There are no workcodes to copy.", parent=self)

    def export_report(self):
        if not self.results_tree.get_children():
            messagebox.showinfo("No Data", "There are no results to export.")
            return
            
        panchayat = self.panchayat_entry.get().strip() or "Report"
        safe_panchayat = re.sub(r'[\\/*?:"<>|]', '_', panchayat) 
        export_format = self.export_format_menu.get()
        
        current_year = datetime.now().strftime("%Y")
        current_date_str = datetime.now().strftime("%d-%b-%Y") 
        
        headers = self.results_tree['columns']
        data = [self.results_tree.item(item, 'values') for item in self.results_tree.get_children()]
        
        title = f"MR Tracking Report Panchayat - {panchayat}"
        date_str = f"Date - {datetime.now().strftime('%d-%m-%Y')}"
        
        downloads_path = self.app.get_user_downloads_path() 
        target_dir = os.path.join(downloads_path, "NregaBot", f"Reports {current_year}", safe_panchayat) 
        try:
            os.makedirs(target_dir, exist_ok=True)
        except OSError as e:
             messagebox.showerror("Folder Error", f"Could not create report directory:\n{target_dir}\nError: {e}")
             return

        if "Excel" in export_format:
            ext = ".xlsx"
            file_type_tuple = ("Excel Workbook", "*.xlsx")
            default_filename = f"MR_Tracking_{safe_panchayat}-{current_date_str}{ext}"
            
            file_path = filedialog.asksaveasfilename(
                initialdir=target_dir,
                initialfile=default_filename,
                defaultextension=ext,
                filetypes=[file_type_tuple, ("All Files", "*.*")],
                title="Save Report As"
            )
            if not file_path: return
            success = self._save_to_excel(data, headers, f"{title} {date_str}", file_path)
            if success:
                messagebox.showinfo("Success", f"Excel report saved successfully to:\n{file_path}")

        elif "PDF" in export_format:
            ext = ".pdf"
            file_type_tuple = ("PDF Document", "*.pdf")
            default_filename = f"MR_Tracking_{safe_panchayat}-{current_date_str}{ext}"

            file_path = filedialog.asksaveasfilename(
                initialdir=target_dir,
                initialfile=default_filename,
                defaultextension=ext,
                filetypes=[file_type_tuple, ("All Files", "*.*")],
                title="Save Report As"
            )
            if not file_path: return

            col_widths = [10, 20, 25, 30, 15, 45, 20, 45, 20, 25, 25, 20, 20, 20] 
            total_width_ratio = sum(col_widths)
            effective_page_width = 297 - 20 
            actual_col_widths = [(w / total_width_ratio) * effective_page_width for w in col_widths]
            
            success = self.generate_report_pdf(data, headers, actual_col_widths, title, date_str, file_path)
            if success:
                messagebox.showinfo("Success", f"PDF report saved successfully to:\n{file_path}")
        
        elif "PNG" in export_format:
            ext = ".png"
            file_type_tuple = ("PNG Image", "*.png")
            default_filename = f"MR_Tracking_{safe_panchayat}-{current_date_str}{ext}"
            
            file_path = filedialog.asksaveasfilename(
                initialdir=target_dir,
                initialfile=default_filename,
                defaultextension=ext,
                filetypes=[file_type_tuple, ("All Files", "*.*")],
                title="Save Report As"
            )
            if not file_path: return
            success = self._save_to_png(data, headers, title, date_str, file_path)
            if success:
                messagebox.showinfo("Success", f"PNG report saved successfully to:\n{file_path}")

    def _export_abps_report(self):
        if not self.abps_results_tree.get_children():
            messagebox.showinfo("No Data", "There are no ABPS results to export.")
            return
            
        panchayat = self.panchayat_entry.get().strip() or "Report"
        safe_panchayat = re.sub(r'[\\/*?:"<>|]', '_', panchayat) 
        
        current_year = datetime.now().strftime("%Y")
        current_date_str = datetime.now().strftime("%d-%b-%Y")
        
        headers = self.abps_report_headers
        data = [self.abps_results_tree.item(item, 'values') for item in self.abps_results_tree.get_children()]
        
        title = f"ABPS Pendency Report - {panchayat}"
        date_str = f"Date - {datetime.now().strftime('%d-%m-%Y')}"
        
        downloads_path = self.app.get_user_downloads_path() 
        target_dir = os.path.join(downloads_path, "NregaBot", f"Reports {current_year}", safe_panchayat) 
        try:
            os.makedirs(target_dir, exist_ok=True)
        except OSError as e:
             messagebox.showerror("Folder Error", f"Could not create report directory:\n{target_dir}\nError: {e}")
             return

        ext = ".xlsx"
        file_type_tuple = ("Excel Workbook", "*.xlsx")
        default_filename = f"ABPS_Pendency_{safe_panchayat}-{current_date_str}{ext}"
        
        file_path = filedialog.asksaveasfilename(
            initialdir=target_dir,
            initialfile=default_filename,
            defaultextension=ext,
            filetypes=[file_type_tuple, ("All Files", "*.*")],
            title="Save ABPS Report As"
        )
        if not file_path: return
        success = self._save_to_excel(data, headers, f"{title} {date_str}", file_path)
        if success:
            messagebox.showinfo("Success", f"ABPS Excel report saved successfully to:\n{file_path}")

    def _save_to_excel(self, data, headers, full_title, file_path):
        try:
            df = pd.DataFrame(data, columns=headers)
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                sheet_name = 'Report'
                df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
                worksheet = writer.sheets[sheet_name]
                
                worksheet['A1'] = full_title
                worksheet['A1'].font = Font(bold=True, size=14)
                worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
                worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
                
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
                for cell in worksheet["2:2"]: 
                    cell.font = header_font
                    cell.fill = header_fill

                for col_idx, col in enumerate(df.columns, 1):
                    column_letter = get_column_letter(col_idx)
                    try:
                        max_length = max(len(str(col)), df[col].astype(str).map(len).max())
                    except (TypeError, ValueError):
                         max_length = len(str(col)) 
                    adjusted_width = min((max_length + 2), 50) 
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            return True
        except Exception as e:
            messagebox.showerror("Excel Export Error", f"Could not generate Excel report.\nError: {e}", parent=self)
            return False

    def generate_report_pdf(self, data, headers, col_widths, title, date_str, file_path):
        class PDFWithFooter(FPDF):
            def footer(self):
                self.set_y(-15) 
                try:
                    self.set_font(font_name, '', 8) 
                except NameError: 
                    self.set_font('Helvetica', '', 8) 
                self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')
                self.set_xy(self.l_margin, -15)
                self.cell(0, 10, 'Report Generated by NregaBot.com', 0, 0, 'L')

        try:
            pdf = PDFWithFooter(orientation="L", unit="mm", format="A4")
            pdf.set_auto_page_break(auto=True, margin=15) 
            pdf.add_page()
            
            try:
                font_path_regular = resource_path("assets/fonts/NotoSansDevanagari-Regular.ttf")
                font_path_bold = resource_path("assets/fonts/NotoSansDevanagari-Bold.ttf")
                pdf.add_font("NotoSansDevanagari", "", font_path_regular, uni=True)
                pdf.add_font("NotoSansDevanagari", "B", font_path_bold, uni=True)
                font_name = "NotoSansDevanagari"
            except RuntimeError:
                font_name = "Helvetica" 

            pdf.set_font(font_name, "B", 14) 
            pdf.cell(0, 10, title, 0, 1, "C")
            pdf.set_font(font_name, "", 10) 
            pdf.cell(0, 8, date_str, 0, 1, "R") 
            pdf.ln(4) 

            pdf.set_font(font_name, "B", 7) 
            pdf.set_fill_color(200, 220, 255)
            header_height = 8 
            
            if len(col_widths) != len(headers):
                self.app.log_message(self.log_display, "PDF Export Warning: Column width count mismatch.", "warning")
                col_widths = [(pdf.w - 2 * pdf.l_margin) / len(headers)] * len(headers)
                
            for i, header in enumerate(headers):
                pdf.cell(col_widths[i], header_height, header, 1, 0, "C", fill=True) 
            pdf.ln()

            pdf.set_font(font_name, "", 6) 
            line_height = 4 
            
            for row_data in data:
                if len(row_data) != len(headers):
                    continue

                max_lines = 1
                for i, cell_text in enumerate(row_data):
                    lines = pdf.multi_cell(col_widths[i], line_height, str(cell_text), border=0, align='L', split_only=True)
                    current_lines = len(lines) if lines else 1 
                    if current_lines > max_lines: max_lines = current_lines
                
                row_height = line_height * max_lines
                
                if pdf.get_y() + row_height > pdf.page_break_trigger:
                    pdf.add_page()
                    pdf.set_font(font_name, "B", 7)
                    for i, header in enumerate(headers):
                         pdf.cell(col_widths[i], header_height, header, 1, 0, "C", fill=True)
                    pdf.ln()
                    pdf.set_font(font_name, "", 6) 

                y_start = pdf.get_y()
                x_start = pdf.l_margin 
                
                for i, cell_text in enumerate(row_data):
                    col_width = col_widths[i]
                    x_current = x_start + sum(col_widths[:i]) 
                    pdf.set_xy(x_current, y_start) 
                    pdf.multi_cell(col_width, line_height, str(cell_text), border=1, align='L', max_line_height=line_height) 
                
                pdf.set_y(y_start + row_height) 

            pdf.output(file_path)
            return True
        except Exception as e:
            messagebox.showerror("PDF Export Error", f"Could not generate PDF report.\nError: {e}", parent=self)
            return False

    def _save_to_png(self, data, headers, title, date_str, file_path):
        try:
            try:
                font_path_regular = resource_path("assets/fonts/NotoSansDevanagari-Regular.ttf")
                font_path_bold = resource_path("assets/fonts/NotoSansDevanagari-Bold.ttf")
                font_title = ImageFont.truetype(font_path_bold, 28)
                font_date = ImageFont.truetype(font_path_regular, 18)
                font_header = ImageFont.truetype(font_path_bold, 16)
                font_body = ImageFont.truetype(font_path_regular, 14)
            except IOError:
                self.app.log_message(self.log_display, "Warning: NotoSansDevanagari fonts not found. Using default PIL fonts. Ensure 'assets/fonts' exist.", "warning")
                font_title = ImageFont.load_default(size=28)
                font_date = ImageFont.load_default(size=18)
                font_header = ImageFont.load_default(size=16)
                font_body = ImageFont.load_default(size=14)

            img_width = 2400  
            margin_x = 80
            margin_y = 60
            
            header_bg_color = (220, 235, 255) 
            row_even_bg_color = (255, 255, 255) 
            row_odd_bg_color = (245, 245, 245)  
            text_color = (0, 0, 0) 
            border_color = (180, 180, 180) 

            draw_start_y = margin_y 

            col_widths_pixels = []
            min_col_width = 100 
            
            for i, header in enumerate(headers):
                text_width = font_header.getlength(header)
                col_widths_pixels.append(max(min_col_width, text_width + 40)) 

            for row_data in data:
                for i, cell_text in enumerate(row_data):
                    lines = self._wrap_text(str(cell_text), font_body, col_widths_pixels[i] - 20) 
                    max_text_width = 0
                    for line in lines:
                        max_text_width = max(max_text_width, font_body.getlength(line))
                    col_widths_pixels[i] = max(col_widths_pixels[i], max_text_width + 40) 

            current_total_width = sum(col_widths_pixels)
            available_width = img_width - 2 * margin_x
            if current_total_width < available_width:
                extra_space_per_col = (available_width - current_total_width) / len(col_widths_pixels)
                col_widths_pixels = [w + extra_space_per_col for w in col_widths_pixels]

            current_total_width = sum(col_widths_pixels) 
            
            estimated_row_height = font_body.getbbox("Tg")[3] - font_body.getbbox("Tg")[1] + 20 
            estimated_header_height = font_header.getbbox("Tg")[3] - font_header.getbbox("Tg")[1] + 20 

            title_bbox = font_title.getbbox(title)
            date_bbox = font_date.getbbox(date_str)
            title_height = title_bbox[3] - title_bbox[1]
            date_height = date_bbox[3] - date_bbox[1]

            total_estimated_height = margin_y * 2 + title_height + date_height + 20 + estimated_header_height + (len(data) * estimated_row_height)
            
            img = Image.new("RGB", (img_width, int(total_estimated_height)), (255, 255, 255))
            draw = ImageDraw.Draw(img)

            current_y = margin_y
            
            title_text_width = font_title.getlength(title)
            title_x = (img_width - title_text_width) / 2
            draw.text((title_x, current_y), title, font=font_title, fill=text_color)
            current_y += title_height + 5

            date_text_width = font_date.getlength(date_str)
            date_x = img_width - margin_x - date_text_width
            draw.text((date_x, current_y), date_str, font=font_date, fill=text_color)
            current_y += date_height + 20 

            header_y_start = current_y
            header_height = 0
            for i, header in enumerate(headers):
                wrapped_header = self._wrap_text(header, font_header, col_widths_pixels[i] - 10)
                line_height = font_header.getbbox("Tg")[3] - font_header.getbbox("Tg")[1] 
                header_height = max(header_height, len(wrapped_header) * line_height + 10) 
            
            current_x = margin_x
            for i, header in enumerate(headers):
                draw.rectangle([current_x, header_y_start, current_x + col_widths_pixels[i], header_y_start + header_height], fill=header_bg_color, outline=border_color, width=1)
                
                wrapped_header = self._wrap_text(header, font_header, col_widths_pixels[i] - 20)
                text_y = header_y_start + (header_height - len(wrapped_header) * font_header.getbbox("Tg")[3] - font_header.getbbox("Tg")[1] ) / 2 
                
                for line in wrapped_header:
                    line_width = font_header.getlength(line)
                    draw.text((current_x + (col_widths_pixels[i] - line_width) / 2, text_y), line, font=font_header, fill=text_color)
                    text_y += font_header.getbbox("Tg")[3] - font_header.getbbox("Tg")[1] 
                current_x += col_widths_pixels[i]
            current_y += header_height

            for row_idx, row_data in enumerate(data):
                row_bg_color = row_even_bg_color if row_idx % 2 == 0 else row_odd_bg_color

                max_row_text_height = 0
                temp_wrapped_cells = []
                for i, cell_text in enumerate(row_data):
                    wrapped_lines = self._wrap_text(str(cell_text), font_body, col_widths_pixels[i] - 20) 
                    temp_wrapped_cells.append(wrapped_lines)
                    line_height = font_body.getbbox("Tg")[3] - font_body.getbbox("Tg")[1] 
                    max_row_text_height = max(max_row_text_height, len(wrapped_lines) * line_height)

                row_data_height = max_row_text_height + 10 

                if current_y + row_data_height + margin_y > img.height:
                    new_height = int(img.height + (row_data_height + margin_y) * 1.5) 
                    new_img = Image.new("RGB", (img_width, new_height), (255, 255, 255))
                    new_img.paste(img, (0, 0))
                    img = new_img
                    draw = ImageDraw.Draw(img) 

                current_x = margin_x
                for i, cell_text in enumerate(row_data):
                    draw.rectangle([current_x, current_y, current_x + col_widths_pixels[i], current_y + row_data_height], fill=row_bg_color, outline=border_color, width=1)
                    
                    wrapped_lines = temp_wrapped_cells[i]
                    text_y = current_y + 5 
                    for line in wrapped_lines:
                        draw.text((current_x + 10, text_y), line, font=font_body, fill=text_color) 
                        text_y += font_body.getbbox("Tg")[3] - font_body.getbbox("Tg")[1] 
                    current_x += col_widths_pixels[i]
                current_y += row_data_height

            final_img = img.crop((0, 0, img_width, current_y + margin_y))
            final_img.save(file_path, "PNG", dpi=(300, 300)) 
            return True
        except Exception as e:
            messagebox.showerror("PNG Export Error", f"Could not generate PNG report.\nError: {e}", parent=self)
            return False

    def _wrap_text(self, text, font, max_width):
        """Helper to wrap text for Pillow."""
        if not text:
            return [""]
        words = text.split(' ')
        lines = []
        current_line = []
        for word in words:
            if font.getlength(' '.join(current_line + [word])) <= max_width:
                current_line.append(word)
            else:
                lines.append(' '.join(current_line))
                current_line = [word]
        lines.append(' '.join(current_line))
        return lines

    def save_inputs(self, inputs):
        """Saves non-sensitive inputs for this tab."""
        save_data = {
            'state': inputs.get('state'),
            'district': inputs.get('district'),
            'block': inputs.get('block'),
            'panchayat': inputs.get('panchayat')
        }
        try:
            config_file = self.app.get_data_path("mr_tracking_inputs.json")
            with open(config_file, 'w') as f:
                json.dump(save_data, f, indent=4)
        except Exception as e:
            print(f"Error saving MR Tracking inputs: {e}")

    def load_inputs(self):
        """Loads saved inputs for this tab."""
        try:
            config_file = self.app.get_data_path("mr_tracking_inputs.json")
            if not os.path.exists(config_file): return
            
            with open(config_file, 'r') as f:
                data = json.load(f)
            
            self.state_entry.delete(0, tkinter.END)
            self.state_entry.insert(0, data.get('state', ''))
            self.district_entry.delete(0, tkinter.END)
            self.district_entry.insert(0, data.get('district', ''))
            self.block_entry.delete(0, tkinter.END)
            self.block_entry.insert(0, data.get('block', ''))
            self.panchayat_entry.delete(0, tkinter.END)
            self.panchayat_entry.insert(0, data.get('panchayat', ''))
        except Exception as e:
            print(f"Error loading MR Tracking inputs: {e}")