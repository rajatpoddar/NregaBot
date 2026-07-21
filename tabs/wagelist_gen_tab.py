# tabs/wagelist_gen_tab.py
import tkinter
from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk
import time, os, sys, subprocess
import re  
import base64 
from datetime import datetime
from urllib.parse import urlparse, parse_qs
import config
from .base_tab import BaseAutomationTab
from .autocomplete_widget import AutocompleteEntry

class WagelistGenTab(BaseAutomationTab):
    def __init__(self, parent, app_instance):
        # Lazy imports
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException
        from selenium.webdriver.common.print_page_options import PrintOptions
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException
        from selenium.webdriver.common.print_page_options import PrintOptions
        super().__init__(parent, app_instance, automation_key="gen")
        self.grid_columnconfigure(0, weight=1); self.grid_rowconfigure(1, weight=1)
        self._create_widgets()

    def _create_widgets(self):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.webdriver.common.print_page_options import PrintOptions
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        import openpyxl
        from selenium import webdriver

        # Configure Main Grid (Full Expansion)
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # --- Main Notebook (Settings | Results | Log) ---
        notebook = ctk.CTkTabview(self)
        notebook.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

        settings_tab = notebook.add("Settings")
        results_tab = notebook.add("Results")
        # Log/Status Area added to Notebook
        self._create_log_and_status_area(parent_notebook=notebook)

        # ================== SETTINGS TAB ==================
        settings_tab.grid_columnconfigure(0, weight=1)
        settings_tab.grid_rowconfigure(0, weight=1) 
        
        # Controls Container (Scrollable or Frame)
        controls_frame = ctk.CTkFrame(settings_tab)
        controls_frame.pack(fill="both", expand=True, padx=10, pady=10)
        controls_frame.grid_columnconfigure(1, weight=1)
        
        # --- 1. Agency Entry ---
        ctk.CTkLabel(controls_frame, text=f"Agency Name ({config.AGENCY_PREFIX}...):").grid(row=0, column=0, sticky='w', padx=15, pady=(15,0))
        self.agency_entry = AutocompleteEntry(
            controls_frame, 
            suggestions_list=self.app.history_manager.get_suggestions("panchayat_name"),
            app_instance=self.app, 
            history_key="panchayat_name"
        )
        self.agency_entry.grid(row=0, column=1, sticky='ew', padx=15, pady=(15,0))
        
        # Note for Macro usage
        ctk.CTkLabel(controls_frame, text="Note: Use 'Macro Manager' tab for bulk processing multiple Panchayats.", text_color="gray60", font=ctk.CTkFont(size=11)).grid(row=1, column=1, sticky='w', padx=15, pady=(5,10))
        
        # --- 2. Settings Checkboxes ---
        self.save_pdf_var = ctk.StringVar(value="off")
        self.save_pdf_checkbox = ctk.CTkCheckBox(
            controls_frame, text="Save generated wagelist page as PDF",
            variable=self.save_pdf_var, onvalue="on", offvalue="off"
        )
        self.save_pdf_checkbox.grid(row=2, column=0, columnspan=2, sticky='w', padx=15, pady=(10, 0))

        self.send_to_sender_var = ctk.StringVar(value="on")
        self.send_to_sender_checkbox = ctk.CTkCheckBox(
            controls_frame, text="✓ Auto-start 'Send Wagelist' automation after generation completes",
            variable=self.send_to_sender_var, onvalue="on", offvalue="off"
        )
        self.send_to_sender_checkbox.grid(row=3, column=0, columnspan=2, sticky='w', padx=15, pady=10)

        # Action Buttons (Start/Stop)
        action_frame = self._create_action_buttons(parent_frame=controls_frame)
        action_frame.grid(row=4, column=0, columnspan=2, sticky='ew', pady=(20, 15))

        # ================== RESULTS TAB ==================
        results_tab.grid_columnconfigure(0, weight=1)
        results_tab.grid_rowconfigure(0, weight=1)
        results_tab.grid_rowconfigure(1, weight=0)

        # Treeview (Full Height)
        cols = ("Timestamp", "Work Code", "Status", "Wagelist No.", "Job Card No.", "Applicant Name")
        self.results_tree = ttk.Treeview(results_tab, columns=cols, show='headings')
        for col in cols: self.results_tree.heading(col, text=col)
        
        self.results_tree.column("Timestamp", width=80, anchor='center')
        self.results_tree.column("Work Code", width=180)
        self.results_tree.column("Status", width=120)
        self.results_tree.column("Wagelist No.", width=150)
        self.results_tree.column("Job Card No.", width=180)
        self.results_tree.column("Applicant Name", width=150)
        
        self.results_tree.grid(row=0, column=0, sticky='nsew', padx=5, pady=5)
        
        # Scrollbar linked to Treeview
        scrollbar = ctk.CTkScrollbar(results_tab, command=self.results_tree.yview)
        self.results_tree.configure(yscroll=scrollbar.set)
        scrollbar.grid(row=0, column=1, sticky='ns', pady=5)
        
        self.style_treeview(self.results_tree)
        self._setup_treeview_sorting(self.results_tree)

        # Results Tab Buttons (Bottom)
        results_action_frame = ctk.CTkFrame(results_tab, fg_color="transparent")
        results_action_frame.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(5, 10), padx=5)
        
        export_controls_frame = ctk.CTkFrame(results_action_frame, fg_color="transparent")
        export_controls_frame.pack(side='right', padx=(10, 0))
        
        self.export_button = ctk.CTkButton(export_controls_frame, text="Export Report", command=self.export_report)
        self.export_button.pack(side='left')
        
        self.export_format_menu = ctk.CTkOptionMenu(export_controls_frame, width=130, values=["PDF (.pdf)", "CSV (.csv)"], command=self._on_format_change)
        self.export_format_menu.pack(side='left', padx=5)
        
        self.export_filter_menu = ctk.CTkOptionMenu(export_controls_frame, width=150, values=["Export All", "Success Only", "Failed Only"])
        self.export_filter_menu.pack(side='left', padx=(0, 5))

    def _on_format_change(self, selected_format):
        if "CSV" in selected_format: self.export_filter_menu.configure(state="disabled")
        else: self.export_filter_menu.configure(state="normal")

    def set_ui_state(self, running: bool):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.webdriver.common.print_page_options import PrintOptions
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        self.set_common_ui_state(running)
        state = "disabled" if running else "normal"
        self.agency_entry.configure(state=state)
        self.save_pdf_checkbox.configure(state=state) 
        self.send_to_sender_checkbox.configure(state=state)
        self.export_button.configure(state=state)
        self.export_format_menu.configure(state=state)
        self.export_filter_menu.configure(state=state)
        if state == "normal": self._on_format_change(self.export_format_menu.get())

    def reset_ui(self):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.webdriver.common.print_page_options import PrintOptions
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        if messagebox.askokcancel("Reset Form?", "Are you sure?"):
            self.agency_entry.delete(0, tkinter.END)
            self.save_pdf_var.set("off") 
            self.send_to_sender_var.set("on")
            for item in self.results_tree.get_children(): self.results_tree.delete(item)
            self.app.clear_log(self.log_display)
            self.update_status("Ready", 0.0)
            self.app.log_message(self.log_display, "Form has been reset.")
            self.app.after(0, self.app.set_status, "Ready")

    def start_automation(self):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.webdriver.common.print_page_options import PrintOptions
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        agency = self.agency_entry.get().strip()
        
        if not agency:
            messagebox.showwarning("Input Error", "Please enter an Agency/Panchayat name.")
            return

        self.app.update_history("panchayat_name", agency)
        
        # We pass it as a list [agency] so the looping logic in run_automation_logic handles it correctly
        # This keeps it compatible with both manual run and macro run.
        self.app.start_automation_thread(self.automation_key, self.run_automation_logic, args=([agency],))

    def retry_logic_handler(self):
        """
        Retry Logic for Wagelist Gen.
        Since this works on a 'Pending List' from the website, 
        'Retrying' simply means running the automation again for the same panchayat.
        """
        if messagebox.askyesno("Retry", "Retrying will check for any remaining items in the list.\nContinue?"):
            self.start_automation()

    def run_automation_logic(self, agency_input):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.webdriver.common.print_page_options import PrintOptions
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        """
        Logic: Processes the provided agency/panchayat.
        agency_input: Can be a single string or a list of strings.
        """
        self.app.after(0, self.set_ui_state, True)
        self.app.clear_log(self.log_display)
        # Clear tree only at start
        self.app.after(0, lambda: [self.results_tree.delete(item) for item in self.results_tree.get_children()])
        
        # Normalize input to a list to support the loop
        agency_list = agency_input if isinstance(agency_input, list) else [agency_input]
        
        total_panchayats = len(agency_list)
        all_generated_wagelists = [] 

        try:
            driver = self.app.get_driver()
            if not driver: return
            wait = WebDriverWait(driver, 30)
            
            # --- LOOP THROUGH EACH PANCHAYAT ---
            for p_index, agency_name_part in enumerate(agency_list):
                if self.app.stop_events[self.automation_key].is_set(): break

                # UI Update for current item
                self.app.log_message(self.log_display, f"=== Processing '{agency_name_part}' ({p_index+1}/{total_panchayats}) ===", "info")
                self.app.after(0, self.app.set_status, f"Processing: {agency_name_part}")
                self.app.after(0, self.update_status, f"{agency_name_part}", (p_index / total_panchayats))
                
                # --- SETUP OUTPUT DIR (Per Panchayat) ---
                output_dir = None
                if self.save_pdf_var.get() == "on":
                    try:
                        safe_agency_name = "".join(c for c in agency_name_part if c.isalnum() or c in (' ', '_')).rstrip()
                        folder_name = config.WAGELIST_GEN_CONFIG.get('output_folder_name', 'NREGABot_WL_Output')
                        # Create a subfolder for this specific panchayat
                        output_dir = os.path.join(self.app.get_user_downloads_path(), folder_name, datetime.now().strftime('%Y-%m-%d'), safe_agency_name)
                        os.makedirs(output_dir, exist_ok=True)
                        self.app.log_message(self.log_display, f"   PDFs will be saved to: {output_dir}", "info")
                    except Exception:
                        output_dir = None

                total_errors_to_skip = 0
                panchayat_wagelists = []

                # --- INNER LOOP (Existing Logic for one Panchayat) ---
                while not self.app.stop_events[self.automation_key].is_set():
                    self.app.after(0, self.app.set_status, f"[{agency_name_part}] Processing item {total_errors_to_skip + 1}...")
                    
                    try:
                        # A. Load Page
                        loaded = False
                        for _ in range(3):
                            try:
                                driver.get(config.WAGELIST_GEN_CONFIG["base_url"])
                                loaded = True; break
                            except Exception: time.sleep(2)
                        
                        if not loaded: 
                            self.app.log_message(self.log_display, f"   Failed to load URL for {agency_name_part}. Skipping...", "error")
                            break # Move to next panchayat

                        # B. Select Agency
                        try:
                            agency_select = wait.until(EC.presence_of_element_located((By.ID, 'ctl00_ContentPlaceHolder1_exe_agency')))
                            select = Select(agency_select)
                            full_agency_name = config.AGENCY_PREFIX + agency_name_part
                            
                            match_text = next((opt.text for opt in select.options if opt.text.strip() == full_agency_name), None)
                            if not match_text:
                                self.app.log_message(self.log_display, f"   No pending wagelists or Agency not found: '{full_agency_name}'.", "warning")
                                break # Move to next panchayat
                            
                            select.select_by_visible_text(match_text)
                        except Exception as e:
                            self.app.log_message(self.log_display, f"   Error selecting agency: {e}", "error")
                            break 
                        
                        # C. Click Proceed
                        try:
                            proceed_btn = wait.until(EC.element_to_be_clickable((By.ID, 'ctl00_ContentPlaceHolder1_go')))
                            driver.execute_script("arguments[0].click();", proceed_btn)
                        except:
                            break

                        # D. Wait for Table & Find Row
                        try:
                            wagelist_table = wait.until(EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_wagelist_msr")))
                            rows = wagelist_table.find_elements(By.XPATH, ".//tr[td]")
                        except:
                            self.app.log_message(self.log_display, f"   No wagelist table found for {agency_name_part}.", "info")
                            break

                        if not rows or total_errors_to_skip >= len(rows): 
                            self.app.log_message(self.log_display, f"   Done with {agency_name_part}.", "success")
                            break
                            
                        row_to_process = rows[total_errors_to_skip]
                        
                        # E. Extract Data
                        try: 
                            checkbox = row_to_process.find_element(By.XPATH, ".//input[@type='checkbox']")
                            tds = row_to_process.find_elements(By.TAG_NAME, "td")
                            work_code = tds[2].get_attribute("innerText").strip()
                        except NoSuchElementException: 
                            break 

                        self.app.log_message(self.log_display, f"   Generating: {work_code}")
                        
                        # F. Click Checkbox & Generate
                        if not checkbox.is_selected():
                            driver.execute_script("arguments[0].click();", checkbox)
                        
                        gen_btn = wait.until(EC.element_to_be_clickable((By.ID, 'ctl00_ContentPlaceHolder1_btn_go')))
                        driver.execute_script("arguments[0].click();", gen_btn)
                        
                        # G. Wait for Outcome
                        try:
                            def check_outcome(d):
                                # ---- Lazy imports ----
                                from selenium.webdriver.common.by import By
                                from selenium.webdriver.support.ui import Select, WebDriverWait
                                from selenium.webdriver.support import expected_conditions as EC
                                from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
                                from selenium.webdriver.common.print_page_options import PrintOptions
                                from selenium import webdriver
                                import openpyxl
                                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                                from openpyxl.utils import get_column_letter
                                from openpyxl.worksheet.page import PageMargins
                                from openpyxl.drawing.image import Image as XLImage
                                if "view_wagelist.aspx" in d.current_url: return True
                                try:
                                    msg = d.find_element(By.ID, "ctl00_ContentPlaceHolder1_lblmsg")
                                    if msg.text.strip(): return True
                                except: pass
                                return False
                            WebDriverWait(driver, 45).until(check_outcome)
                        except TimeoutException:
                            self.app.log_message(self.log_display, "   Timeout: Page slow.", "error")
                            total_errors_to_skip += 1
                            continue

                        # H. Handle Result
                        if "view_wagelist.aspx" in driver.current_url:
                            parsed_url = urlparse(driver.current_url)
                            query_params = parse_qs(parsed_url.query)
                            wagelist_no = query_params.get('Wage_Listno', ['N/A'])[0]
                            
                            if wagelist_no != 'N/A': 
                                panchayat_wagelists.append(wagelist_no)
                                all_generated_wagelists.append(wagelist_no)
                            
                            pdf_info = ""
                            if output_dir and wagelist_no != 'N/A':
                                pdf_path = self._save_page_as_pdf(driver, wagelist_no, work_code, output_dir)
                                pdf_info = " (PDF Saved)" if pdf_path else " (PDF Failed)"

                            self.app.log_message(self.log_display, f"   SUCCESS: {wagelist_no}{pdf_info}", "success")
                            self._log_result(work_code, "Success", wagelist_no, "", "")
                            # Stay at index 0 because processed item is gone
                            pass 

                        else:
                            try:
                                err_elem = driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_lblmsg")
                                err_text = err_elem.get_attribute("innerText").strip()
                            except: err_text = "Unknown Error"
                            
                            self.app.log_message(self.log_display, f"   Failed: {err_text}", "error")
                            self._log_result(work_code, f"Failed ({err_text[:20]})", "N/A", "", "")
                            total_errors_to_skip += 1 # Skip this failed item

                    except Exception as e:
                        self.app.log_message(self.log_display, f"   Row Error: {e}", "error")
                        total_errors_to_skip += 1
                
                # End of While Loop (One Panchayat done)
                time.sleep(1) # Breathe before next panchayat

            # End of For Loop (Batch done)

            if not self.app.stop_events[self.automation_key].is_set():
                if all_generated_wagelists:
                    messagebox.showinfo("Complete", f"Processed {total_panchayats} Panchayat(s).\nGenerated {len(all_generated_wagelists)} Wagelists.")
                else:
                    messagebox.showinfo("Complete", "Process finished. No wagelists were generated.")

        except Exception as e: 
            self.app.log_message(self.log_display, f"Critical Error: {e}", level="error")
        finally:
            self.app.after(0, self.set_ui_state, False)
            self.app.after(0, self.update_status, "Finished", 1.0)
            self.app.after(0, self.app.set_status, "Finished")
            
            # Send ALL collected wagelists data if enabled
            if self.send_to_sender_var.get() == "on" and all_generated_wagelists and not self.app.stop_events[self.automation_key].is_set():
                self.app.after(0, self.app.set_status, "Auto-starting Send Wagelist...")
                
                # Pass auto_start=True to the workflow manager
                self.app.after(1000, lambda: self.app.send_wagelist_data_and_switch_tab(
                    all_generated_wagelists[0], 
                    all_generated_wagelists[-1],
                    auto_start=True
                ))
            else:
                self.app.after(3000, lambda: self.app.set_status("Ready"))

    def _log_result(self, work_code, status, wagelist_no, job_card, applicant_name):
        timestamp = datetime.now().strftime("%H:%M:%S")
        tags = ('success',) if 'success' in status.lower() else ('failed',)
        self.app.after(0, lambda: self.results_tree.insert("", "end", values=(timestamp, work_code, status, wagelist_no, job_card, applicant_name), tags=tags))

    # --- NEW METHOD: Save Page as PDF ---
    def _save_page_as_pdf(self, driver, wagelist_no, work_code, output_dir):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.webdriver.common.print_page_options import PrintOptions
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        """Saves the current page as a PDF."""
        try:
            # Create a safe filename
            safe_work_code = work_code.split('/')[-1][-6:] if '/' in work_code else work_code[-6:]
            base_filename = f"WL_{wagelist_no.replace('/', '-')}_{safe_work_code}"
            extension = ".pdf"
            counter = 1
            pdf_filename = f"{base_filename}{extension}"
            save_path = os.path.join(output_dir, pdf_filename)

            # Ensure filename is unique
            while os.path.exists(save_path):
                pdf_filename = f"{base_filename} ({counter}){extension}"
                save_path = os.path.join(output_dir, pdf_filename)
                counter += 1

            pdf_data_base64 = None
            
            # Use browser-specific commands to print to PDF
            if self.app.active_browser == 'firefox':
                self.app.log_message(self.log_display, "   - Using Firefox's print command to save PDF...", "info")
                print_options = PrintOptions()
                print_options.orientation = "landscape"
                print_options.scale = 0.7
                pdf_data_base64 = driver.print_page(print_options)

            elif self.app.active_browser == 'chrome':
                self.app.log_message(self.log_display, "   - Using Chrome's advanced print command (CDP) to save PDF...", "info")
                print_options = {
                    "landscape": True, 
                    "displayHeaderFooter": False, 
                    "printBackground": True, 
                    "scale": 0.7, 
                    "marginTop": 0.4, "marginBottom": 0.4,
                    "marginLeft": 0.4, "marginRight": 0.4,
                    "paperWidth": 8.27, # A4 width in inches
                    "paperHeight": 11.69 # A4 height in inches
                }
                result = driver.execute_cdp_cmd("Page.printToPDF", print_options)
                pdf_data_base64 = result['data']

            if pdf_data_base64:
                pdf_data = base64.b64decode(pdf_data_base64)
                with open(save_path, 'wb') as f:
                    f.write(pdf_data)
                return save_path
            else:
                self.app.log_message(self.log_display, f"Error: PDF data was not generated for {wagelist_no}.", "error")
                return None

        except Exception as e:
            self.app.log_message(self.log_display, f"Error saving PDF for {wagelist_no}: {e}", "error")
            return None

    def export_report(self):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.webdriver.common.print_page_options import PrintOptions
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        export_format = self.export_format_menu.get()
        if "CSV" in export_format:
            # For CSV, we export all columns as is
            self.export_treeview_to_csv(self.results_tree, "wagelist_gen_results.csv")
            return
            
        data, file_path = self._get_filtered_data_and_filepath(export_format)
        if not data: return

        # For PDF and Image, we consolidate columns for a cleaner report
        report_data, report_headers, col_widths = self._prepare_report_data(data)

        if "PDF" in export_format:
            self._handle_pdf_export(report_data, report_headers, col_widths, file_path)

    def _get_filtered_data_and_filepath(self, export_format):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.webdriver.common.print_page_options import PrintOptions
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        all_items = self.results_tree.get_children()
        if not all_items: messagebox.showinfo("No Data", "There are no results to export."); return None, None
        agency_name = self.agency_entry.get().strip()
        if not agency_name: messagebox.showwarning("Input Needed", "Please enter an Agency Name for the report title."); return None, None

        filter_option = self.export_filter_menu.get()
        data_to_export = []
        for item_id in all_items:
            row_values = self.results_tree.item(item_id)['values']
            status = row_values[2].upper() # Status is the third column
            if filter_option == "Export All": data_to_export.append(row_values)
            elif filter_option == "Success Only" and "SUCCESS" in status: data_to_export.append(row_values)
            elif filter_option == "Failed Only" and "SUCCESS" not in status: data_to_export.append(row_values)
        if not data_to_export: messagebox.showinfo("No Data", f"No records found for filter '{filter_option}'."); return None, None

        safe_name = "".join(c for c in agency_name if c.isalnum() or c in (' ', '_')).rstrip()
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        file_details = {"Image (.jpg)": { "ext": ".jpg", "types": [("JPEG Image", "*.jpg")]}, "PDF (.pdf)": { "ext": ".pdf", "types": [("PDF Document", "*.pdf")]}}
        details = file_details.get(export_format, {"ext": ".txt", "types": [("Text File", "*.txt")]}) # Fallback
        filename = f"WagelistGen_Report_{safe_name}_{timestamp}{details['ext']}"

        file_path = filedialog.asksaveasfilename(defaultextension=details['ext'], filetypes=details['types'], initialdir=self.app.get_user_downloads_path(), initialfile=filename, title="Save Report")
        return (data_to_export, file_path) if file_path else (None, None)

    def _prepare_report_data(self, raw_data):
        report_data, report_headers = [], ["Work Code", "Status", "Details", "Timestamp"]
        col_widths = [70, 45, 130, 25]
        for row in raw_data:
            timestamp, work_code, status, wagelist, jc, name = row
            details = f"Wagelist: {wagelist}"
            if jc: details += f" | Unfrozen JC: {jc} ({name})"
            report_data.append([work_code, status, details, timestamp])
        return report_data, report_headers, col_widths

    
    def _handle_pdf_export(self, data, headers, col_widths, file_path):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.webdriver.common.print_page_options import PrintOptions
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        title = f"Wagelist Generation Report: {self.agency_entry.get().strip()}"
        report_date = datetime.now().strftime('%d %b %Y')
        success = self.generate_report_pdf(data, headers, col_widths, title, report_date, file_path)
        if success:
            if messagebox.askyesno("Success", f"PDF Report saved to:\n{file_path}\n\nDo you want to open it?"):
                if sys.platform == "win32": os.startfile(file_path)
                else: subprocess.call(['open', file_path])