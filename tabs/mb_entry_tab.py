# tabs/mb_entry_tab.py
import tkinter
from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk
import time, os, json, sys, subprocess, random
import re
from datetime import datetime, date

# Selenium Imports
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    UnexpectedAlertPresentException, 
    NoSuchElementException, 
    TimeoutException
)

# Excel Imports
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# PDF Imports (ReportLab)
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

import config
from .base_tab import BaseAutomationTab
from .autocomplete_widget import AutocompleteEntry

class MbEntryTab(BaseAutomationTab):
    def __init__(self, parent, app_instance):
        """Initializes the eMB Entry tab."""
        super().__init__(parent, app_instance, automation_key="mb_entry")
        
        # Path to save/load form inputs
        self.config_file = self.app.get_data_path("mb_entry_inputs.json")

        self.mapping_file = self.app.get_data_path("mb_panchayat_mate_map.json")
        self.mapping_data = {}
        self._load_mapping_data()
        
        # Dictionary to hold form field variables
        self.config_vars = {}
        
        # Variable for the "Auto MB No." checkbox
        self.auto_mb_no_var = ctk.BooleanVar(value=True)
        
        # --- Panchayat-dependent mate name logic ---
        self.panchayat_after_id = None 
        self.notebook = None 
        # ---

        # Configure grid layout
        self.grid_columnconfigure(0, weight=1); self.grid_rowconfigure(1, weight=1)
        
        # Create and load UI elements
        self._create_widgets(); self._load_inputs()
        self._toggle_mb_no_entry() 

    def _create_widgets(self):
        """Creates and places all UI elements for this tab."""
        
        # --- Top Frame for Configuration ---
        top_frame = ctk.CTkFrame(self)
        top_frame.grid(row=0, column=0, sticky='ew', padx=10, pady=(10,0))
        
        config_frame = ctk.CTkFrame(top_frame)
        config_frame.pack(pady=(0, 10), fill='x')
        
        config_frame.grid_columnconfigure((1, 3), weight=1)
        
        # --- Form Fields ---
        self.panchayat_entry = self._create_autocomplete_field(config_frame, "panchayat_name", "Panchayat Name", 0, 0)
        self.panchayat_entry.bind("<KeyRelease>", self._on_panchayat_change_debounced)
        
        # --- MB No. with Auto Checkbox ---
        ctk.CTkLabel(config_frame, text="MB No.").grid(row=1, column=0, sticky='w', padx=15, pady=5)
        mb_frame = ctk.CTkFrame(config_frame, fg_color="transparent")
        mb_frame.grid(row=1, column=1, sticky='ew', padx=15, pady=5)
        mb_frame.grid_columnconfigure(0, weight=1)
        
        mb_var = ctk.StringVar()
        self.config_vars["measurement_book_no"] = mb_var
        self.mb_no_entry = ctk.CTkEntry(mb_frame, textvariable=mb_var)
        self.mb_no_entry.grid(row=0, column=0, sticky='ew')

        self.auto_mb_no_checkbox = ctk.CTkCheckBox(
            mb_frame, text="Auto", variable=self.auto_mb_no_var,
            command=self._toggle_mb_no_entry
        )
        self.auto_mb_no_checkbox.grid(row=0, column=1, padx=(10, 0))
        # --- End MB No. Section ---

        self.page_no_entry = self._create_field(config_frame, "page_no", "Page No.", 1, 2)
        self.unit_cost_entry = self._create_field(config_frame, "unit_cost", "Unit Cost (₹)", 2, 0)
        self.pit_count_entry = self._create_field(config_frame, "default_pit_count", "Pit Count", 2, 2)
        
        self.mate_name_entry = self._create_autocomplete_field(config_frame, "mate_name", "Mate Names (comma-separated)", 3, 0)
        self._on_panchayat_change()

        # Note for user
        note = ctk.CTkLabel(config_frame, text="ℹ️ Note: Use this emb automation only for single activity works.", text_color="#E53E3E", wraplength=450)
        note.grid(row=4, column=0, columnspan=4, sticky='w', padx=15, pady=(10, 15))

        # --- Action Buttons (Start, Stop, Reset) ---
        action_frame_container = ctk.CTkFrame(top_frame)
        action_frame_container.pack(pady=10, fill='x')
        action_frame = self._create_action_buttons(parent_frame=action_frame_container)
        action_frame.pack(expand=True, fill='x')
        
        # --- Tab View for Work Codes, Results, Logs ---
        self.notebook = ctk.CTkTabview(self)
        self.notebook.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0,10))
        work_codes_frame = self.notebook.add("Work Codes")
        results_frame = self.notebook.add("Results")
        self._create_log_and_status_area(parent_notebook=self.notebook) 

        # --- Work Codes Tab ---
        work_codes_frame.grid_columnconfigure(0, weight=1); work_codes_frame.grid_rowconfigure(1, weight=1)
        wc_controls_frame = ctk.CTkFrame(work_codes_frame, fg_color="transparent")
        wc_controls_frame.grid(row=0, column=0, sticky='ew')
        
        clear_button = ctk.CTkButton(wc_controls_frame, text="Clear", width=80, command=lambda: self.work_codes_text.delete("1.0", tkinter.END))
        clear_button.pack(side='right', pady=(5,0), padx=(0,5))
        
        extract_button = ctk.CTkButton(wc_controls_frame, text="Extract from Text", width=120,
                                       command=lambda: self._extract_and_update_workcodes(self.work_codes_text))
        extract_button.pack(side='right', pady=(5,0), padx=(0, 5))
        
        self.work_codes_text = ctk.CTkTextbox(work_codes_frame, wrap=tkinter.WORD)
        self.work_codes_text.grid(row=1, column=0, sticky='nsew', padx=5, pady=5)
        
        # --- Results Tab ---
        results_frame.grid_columnconfigure(0, weight=1); results_frame.grid_rowconfigure(1, weight=1)
        results_action_frame = ctk.CTkFrame(results_frame, fg_color="transparent")
        results_action_frame.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(5, 10), padx=5)
        
        export_controls_frame = ctk.CTkFrame(results_action_frame, fg_color="transparent")
        export_controls_frame.pack(side='right', padx=(10, 0))
        
        # Updated Export Menu
        self.export_button = ctk.CTkButton(export_controls_frame, text="Download Report", command=self.export_report, fg_color="#107C10")
        self.export_button.pack(side='left')
        
        self.export_format_menu = ctk.CTkOptionMenu(
            export_controls_frame, 
            width=160, 
            values=["Excel Professional (.xlsx)", "PDF Professional (.pdf)", "CSV (.csv)"], 
            command=self._on_format_change
        )
        self.export_format_menu.set("Excel Professional (.xlsx)") # Default
        self.export_format_menu.pack(side='left', padx=5)
        
        self.export_filter_menu = ctk.CTkOptionMenu(export_controls_frame, width=150, values=["Export All", "Success Only", "Failed Only"])
        self.export_filter_menu.pack(side='left', padx=(0, 5))

        # --- Results Treeview ---
        cols = ("Panchayat", "Work Code", "Work Name", "Muster Roll No", "MR Period", "Status", "Details", "Timestamp")
        self.results_tree = ttk.Treeview(results_frame, columns=cols, show='headings')
        
        for col in cols: 
            self.results_tree.heading(col, text=col)
        
        self.results_tree.column("Panchayat", width=100)
        self.results_tree.column("Work Code", width=120)
        self.results_tree.column("Work Name", width=250)
        self.results_tree.column("Muster Roll No", width=120)
        self.results_tree.column("MR Period", width=150)
        self.results_tree.column("Status", width=80, anchor='center')
        self.results_tree.column("Details", width=200)
        self.results_tree.column("Timestamp", width=80, anchor='center')
        
        self.results_tree.grid(row=1, column=0, sticky='nsew')
        scrollbar = ctk.CTkScrollbar(results_frame, command=self.results_tree.yview)
        self.results_tree.configure(yscroll=scrollbar.set); scrollbar.grid(row=1, column=1, sticky='ns')
        self.style_treeview(self.results_tree); self._setup_treeview_sorting(self.results_tree)

    def _toggle_mb_no_entry(self):
        """Enables or disables the MB No. entry based on the 'Auto' checkbox."""
        if self.auto_mb_no_var.get():
            self.mb_no_entry.configure(state="disabled")
            self.config_vars["measurement_book_no"].set("Auto from Workcode")
        else:
            self.mb_no_entry.configure(state="normal")
            saved_data = {}
            if os.path.exists(self.config_file):
                try:
                    with open(self.config_file, 'r') as f: saved_data = json.load(f)
                except (json.JSONDecodeError, IOError): pass
            self.config_vars["measurement_book_no"].set(saved_data.get("measurement_book_no", ""))

    def _on_format_change(self, selected_format):
        if "CSV" in selected_format: self.export_filter_menu.configure(state="disabled")
        else: self.export_filter_menu.configure(state="normal")

    def _create_field(self, parent, key, text, r, c):
        ctk.CTkLabel(parent, text=text).grid(row=r, column=c, sticky='w', padx=15, pady=5)
        var = ctk.StringVar(); self.config_vars[key] = var
        entry = ctk.CTkEntry(parent, textvariable=var)
        entry.grid(row=r, column=c+1, sticky='ew', padx=15, pady=5)
        return entry

    def _create_autocomplete_field(self, parent, key, text, r, c):
        ctk.CTkLabel(parent, text=text).grid(row=r, column=c, sticky='w', padx=15, pady=5)
        var = ctk.StringVar(); self.config_vars[key] = var
        initial_suggestions = self.app.history_manager.get_suggestions(key)
        entry = AutocompleteEntry(parent, textvariable=var, suggestions_list=initial_suggestions, app_instance=self.app, history_key=key)
        entry.grid(row=r, column=c+1, columnspan=3, sticky='ew', padx=15, pady=5)
        return entry

    # --- Panchayat-dependent mate name logic ---
    def _get_current_mate_key(self):
        panchayat_name = self.panchayat_entry.get().strip().lower()
        panchayat_safe_name = "".join(c for c in panchayat_name if c.isalnum() or c == '_').rstrip()
        if not panchayat_safe_name: return "mate_name_default"
        return f"mate_name_{panchayat_safe_name}"

    def _on_panchayat_change_debounced(self, event=None):
        if self.panchayat_after_id: self.after_cancel(self.panchayat_after_id)
        if event and event.keysym in ("Up", "Down", "Return", "Enter", "Tab"): return
        self.panchayat_after_id = self.after(300, self._on_panchayat_change)

    def _load_mapping_data(self):
        if os.path.exists(self.mapping_file):
            try:
                with open(self.mapping_file, 'r') as f: self.mapping_data = json.load(f)
            except Exception: self.mapping_data = {}

    def _save_mapping_pair(self, panchayat, mate_names):
        if not panchayat or not mate_names: return
        key = panchayat.strip().lower()
        self.mapping_data[key] = mate_names.strip()
        try:
            with open(self.mapping_file, 'w') as f: json.dump(self.mapping_data, f, indent=4)
        except Exception: pass

    def _on_panchayat_change(self):
        if self.panchayat_after_id: self.after_cancel(self.panchayat_after_id); self.panchayat_after_id = None
        mate_key = self._get_current_mate_key()
        new_suggestions = self.app.history_manager.get_suggestions(mate_key)
        if self.mate_name_entry:
            self.mate_name_entry.history_key = mate_key
            self.mate_name_entry.suggestions = new_suggestions
            current_panchayat = self.panchayat_entry.get().strip().lower()
            if current_panchayat in self.mapping_data:
                saved_mate = self.mapping_data[current_panchayat]
                if self.mate_name_entry.get().strip() != saved_mate:
                    self.mate_name_entry.delete(0, tkinter.END)
                    self.mate_name_entry.insert(0, saved_mate)

    def set_ui_state(self, running: bool):
        self.set_common_ui_state(running) 
        state = "disabled" if running else "normal"
        self.work_codes_text.configure(state=state)
        self.panchayat_entry.configure(state=state)
        self.page_no_entry.configure(state=state)
        self.unit_cost_entry.configure(state=state)
        self.mate_name_entry.configure(state=state)
        self.pit_count_entry.configure(state=state)
        self.export_button.configure(state=state)
        self.export_format_menu.configure(state=state)
        self.export_filter_menu.configure(state=state)
        self.auto_mb_no_checkbox.configure(state=state)
        if state == "normal":
            self._on_format_change(self.export_format_menu.get())
            self._toggle_mb_no_entry()
        else:
            self.mb_no_entry.configure(state="disabled")

    def reset_ui(self):
        if messagebox.askokcancel("Reset Form?", "Clear all inputs and logs?"):
            self._load_inputs()
            self.config_vars['panchayat_name'].set("") 
            self.work_codes_text.configure(state="normal")
            self.work_codes_text.delete("1.0", tkinter.END)
            self.work_codes_text.configure(state="disabled")
            for item in self.results_tree.get_children(): self.results_tree.delete(item)
            self.app.clear_log(self.log_display)
            self.update_status("Ready", 0.0)
            self.app.log_message(self.log_display, "Form has been reset.")
            self.app.after(0, self.app.set_status, "Ready")

    def start_automation(self):
        cfg = {key: var.get().strip() for key, var in self.config_vars.items()}
        if not self.auto_mb_no_var.get() and not cfg.get("measurement_book_no"):
            messagebox.showwarning("Input Error", "MB No. field is required when 'Auto' is unchecked.")
            return
        required_fields = ["panchayat_name", "page_no", "unit_cost", "default_pit_count", "mate_name"]
        if any(not cfg.get(key) for key in required_fields):
            messagebox.showwarning("Input Error", "All configuration fields must be filled out.")
            return
        work_codes_raw = [line.strip() for line in self.work_codes_text.get("1.0", tkinter.END).strip().splitlines() if line.strip()]
        if not work_codes_raw:
            messagebox.showwarning("Input Required", "Please paste at least one work code.")
            return
        self._save_mapping_pair(cfg['panchayat_name'], cfg['mate_name'])
        self._save_inputs(cfg)
        self.app.start_automation_thread(self.automation_key, self.run_automation_logic, args=(cfg, work_codes_raw))
    
    def _save_inputs(self, cfg):
        try:
            with open(self.config_file, 'w') as f: json.dump(cfg, f, indent=4)
        except Exception as e: self.app.log_message(self.log_display, f"Could not save inputs: {e}", "warning")

    def _load_inputs(self):
        saved_data = {}
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, 'r') as f: saved_data = json.load(f)
            except (json.JSONDecodeError, IOError) as e: self.app.log_message(self.log_display, f"Could not load inputs: {e}", "warning")
        for key, var in self.config_vars.items():
            default_value = config.MB_ENTRY_CONFIG["defaults"].get(key, "")
            var.set(saved_data.get(key, default_value))
        self.after(100, self._on_panchayat_change)

    def run_automation_logic(self, cfg, work_codes_raw):
        self.app.after(0, self.set_ui_state, True) 
        self.app.clear_log(self.log_display) 
        self.app.after(0, lambda: [self.results_tree.delete(item) for item in self.results_tree.get_children()])
        self.app.log_message(self.log_display, "Starting eMB Entry automation...")
        self.app.after(0, self.app.set_status, "Running eMB Entry...") 
        
        try:
            driver = self.app.get_driver()
            if not driver: return 

            mate_names_list = [name.strip() for name in cfg["mate_name"].split(',') if name.strip()]
            if not mate_names_list:
                messagebox.showerror("Input Error", "Please provide at least one Mate Name.")
                return

            if not self.app.stop_events[self.automation_key].is_set():
                self.app.update_history("panchayat_name", cfg['panchayat_name'])
                mate_key = self._get_current_mate_key()
                for mate in mate_names_list: self.app.update_history(mate_key, mate)
            
            processed_codes = set()
            total = len(work_codes_raw)
            self.app.after(0, self.app.set_status, f"Starting eMB Entry for {total} workcodes...")

            for i, work_code in enumerate(work_codes_raw):
                if self.app.stop_events[self.automation_key].is_set():
                    self.app.log_message(self.log_display, "Automation stopped.", "warning"); break
                
                self.app.after(0, self.update_status, f"Processing {i+1}/{total}: {work_code}", (i+1) / total)
                
                if work_code in processed_codes:
                    self._log_result(cfg, work_code, "Skipped", "Duplicate entry.")
                    continue
                
                self._process_single_work_code(driver, work_code, cfg, mate_names_list)
                processed_codes.add(work_code)

            final_msg = "Automation finished." if not self.app.stop_events[self.automation_key].is_set() else "Stopped."
            self.app.after(0, self.update_status, final_msg, 1.0)
            if not self.app.stop_events[self.automation_key].is_set(): 
                messagebox.showinfo("Complete", "e-MB Entry process has finished.")
        
        except Exception as e:
            self.app.log_message(self.log_display, f"A critical error occurred: {e}", "error")
            messagebox.showerror("Automation Error", f"An error occurred:\n\n{e}")
        finally:
            self.app.after(0, self.set_ui_state, False)
            self.app.after(0, self.app.set_status, "Automation Finished")

    def _log_result(self, cfg, work_code, status, details, work_name="-", mr_no="-", mr_period="-"):
        timestamp = datetime.now().strftime("%H:%M:%S")
        panchayat = cfg.get('panchayat_name', '-')
        tags = ('failed',) if 'success' not in status.lower() else ()
        values = (panchayat, work_code, work_name, mr_no, mr_period, status, details, timestamp)
        self.app.after(0, lambda: self.results_tree.insert("", "end", values=values, tags=tags))

    def _process_single_work_code(self, driver, work_code, cfg, mate_names_list):
        wait = WebDriverWait(driver, 25) 
        extracted_work_name = "-"; extracted_mr_no = "-"; extracted_mr_period = "-"
        
        try:
            self.app.after(0, self.app.set_status, f"Navigating for {work_code}...")
            if "MustorRoll/MeasurementBook.aspx" not in driver.current_url: driver.get(config.MB_ENTRY_CONFIG["url"])

            try:
                panchayat_dropdown = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, 'ctl00_ContentPlaceHolder1_ddl_panch')))
                selected_option = Select(panchayat_dropdown).first_selected_option
                if selected_option.text.strip() != cfg['panchayat_name']:
                    Select(panchayat_dropdown).select_by_visible_text(cfg['panchayat_name'])
                    wait.until(EC.staleness_of(panchayat_dropdown))
                    wait.until(EC.presence_of_element_located((By.ID, 'ctl00_ContentPlaceHolder1_ddl_panch')))
            except Exception: pass
            
            wait.until(EC.presence_of_element_located((By.ID, 'ctl00_ContentPlaceHolder1_txtMBNo')))

            mb_no_to_use = cfg["measurement_book_no"]
            if self.auto_mb_no_var.get() and len(work_code) >= 4: mb_no_to_use = work_code[-4:] 

            self.app.after(0, self.app.set_status, f"Searching {work_code}...")
            driver.execute_script(f"document.getElementById('ctl00_ContentPlaceHolder1_txtMBNo').value = '{mb_no_to_use}';")
            driver.execute_script(f"document.getElementById('ctl00_ContentPlaceHolder1_txtpageno').value = '{cfg['page_no']}';")
            driver.execute_script(f"document.getElementById('ctl00_ContentPlaceHolder1_txtWrkCode').value = '{work_code}';")
            
            work_dropdown_old = driver.find_element(By.ID, 'ctl00_ContentPlaceHolder1_ddlSelWrk')
            search_btn = driver.find_element(By.ID, 'ctl00_ContentPlaceHolder1_imgButtonSearch')
            driver.execute_script("arguments[0].click();", search_btn)
            
            self.app.after(0, self.app.set_status, "Waiting for search results...")
            try: wait.until(EC.staleness_of(work_dropdown_old))
            except TimeoutException: pass

            self.app.after(0, self.app.set_status, f"Selecting work details...")
            select_work_elem = wait.until(EC.presence_of_element_located((By.ID, 'ctl00_ContentPlaceHolder1_ddlSelWrk')))
            select_work = Select(select_work_elem)
            found_work = False; target_index = 1 
            clean_search_code = str(work_code).strip()
            
            for index, option in enumerate(select_work.options):
                if (clean_search_code in option.get_attribute("value")) or (clean_search_code in option.text):
                    target_index = index; found_work = True; break
            
            try: extracted_work_name = re.findall(r'\((.*?)\)', select_work.options[target_index].text)[-1]
            except: extracted_work_name = "Unknown"

            element_to_go_stale = select_work_elem 
            select_work.select_by_index(target_index)
            try: wait.until(EC.staleness_of(element_to_go_stale))
            except TimeoutException: pass

            self.app.log_message(self.log_display, "🔘 Clicking Radio Button...")
            radio_btn = wait.until(EC.element_to_be_clickable((By.ID, "ctl00_ContentPlaceHolder1_rddist_0")))
            driver.execute_script("arguments[0].click();", radio_btn)
            time.sleep(1) 

            self.app.log_message(self.log_display, "⏳ Waiting for Period Dropdown...")
            period_dropdown_elem = wait.until(EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_ddlSelMPeriod")))
            period_dropdown = Select(period_dropdown_elem)
            if len(period_dropdown.options) <= 1: raise ValueError("No measurement period found.")
            extracted_mr_period = period_dropdown.options[1].text
            period_element_to_stale = period_dropdown_elem
            period_dropdown.select_by_index(1)
            
            self.app.log_message(self.log_display, "⏳ Waiting for Refresh...")
            try: wait.until(EC.staleness_of(period_element_to_stale))
            except TimeoutException: pass

            wait.until(EC.presence_of_element_located((By.ID, 'ctl00_ContentPlaceHolder1_lbl_person_days')))
            wait.until(lambda d: d.find_element(By.ID, 'ctl00_ContentPlaceHolder1_lbl_person_days').get_attribute('value') != '')

            try: extracted_mr_no = driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_lbl_msr").text
            except: extracted_mr_no = "-"

            pd_elem = driver.find_element(By.ID, 'ctl00_ContentPlaceHolder1_lbl_person_days')
            total_persondays = int(pd_elem.get_attribute('value') or 0)
            if total_persondays == 0: raise ValueError("0 Persondays / eMB already Booked")

            self.app.after(0, self.app.set_status, f"Filling activity details...")
            prefix = self._find_activity_prefix(driver) 
            total_cost = total_persondays * int(cfg["unit_cost"])
            
            driver.execute_script(f"document.getElementsByName('{prefix}$qty')[0].value = '{total_persondays}';")
            driver.execute_script(f"document.getElementsByName('{prefix}$unitcost')[0].value = '{cfg['unit_cost']}';")
            self.app.log_message(self.log_display, "⚙️ Triggering Auto-Calculation (check)...")
            driver.execute_script("if(typeof check === 'function') { check(); }")
            driver.execute_script(f"document.getElementsByName('{prefix}$labcomp')[0].value = '{total_cost}';")
            self.app.log_message(self.log_display, "⚙️ Triggering Validation (checkLabCom)...")
            driver.execute_script("if(typeof checkLabCom === 'function') { checkLabCom(); }")
            
            try: driver.execute_script(f"document.getElementById('ctl00_ContentPlaceHolder1_txtpit').value = '{cfg['default_pit_count']}';")
            except: pass 
            random_mate = random.choice(mate_names_list)
            driver.execute_script(f"document.getElementById('ctl00_ContentPlaceHolder1_txt_mat_name').value = '{random_mate}';")

            self.app.after(0, self.app.set_status, f"Saving...")
            save_btn = driver.find_element(By.XPATH, '//input[@value="Save"]')
            driver.execute_script("arguments[0].click();", save_btn)
            
            try:
                alert = wait.until(EC.alert_is_present())
                alert_text = alert.text
                alert.accept()
                status = "Success" if "success" in alert_text.lower() or "saved" in alert_text.lower() else "Failed"
                self._log_result(cfg, work_code, status, alert_text, extracted_work_name, extracted_mr_no, extracted_mr_period)
            except TimeoutException:
                self._log_result(cfg, work_code, "Failed", "No Alert Received", extracted_work_name, extracted_mr_no, extracted_mr_period)
        
        except Exception as e:
            err_msg = str(e).splitlines()[0]
            self.app.log_message(self.log_display, f"Error on {work_code}: {err_msg}", "error")
            self._log_result(cfg, work_code, "Failed", "Script Error", extracted_work_name, extracted_mr_no, extracted_mr_period)

    def _find_activity_prefix(self, driver):
        self.app.log_message(self.log_display, "Searching for 'Earth work' activity...")
        for i in range(1, 61): 
            try:
                activity_id = f"ctl00_ContentPlaceHolder1_activity_ctl{str(i).zfill(2)}_act_name"
                element = driver.find_element(By.ID, activity_id)
                text = element.get_attribute("innerText").lower()
                if "earth work" in text:
                    self.app.log_message(self.log_display, f"✅ Found 'Earth work' in row #{i}.", "success")
                    return f"ctl00$ContentPlaceHolder1$activity$ctl{str(i).zfill(2)}"
            except NoSuchElementException: continue 
        self.app.log_message(self.log_display, "⚠️ 'Earth work' not found, defaulting to first row (ctl01).", "warning")
        return "ctl00$ContentPlaceHolder1$activity$ctl01"
    
    def export_report(self):
        """Routes the export request to the correct handler."""
        export_format = self.export_format_menu.get()
        
        if "Excel" in export_format:
            self.export_professional_report()
        elif "CSV" in export_format:
            self.export_treeview_to_csv(self.results_tree, "mb_entry_results.csv")
        elif "PDF" in export_format:
            self.export_professional_pdf()

    def export_professional_report(self):
        """Generates a professional Excel report similar to eKYC Report."""
        all_items = self.results_tree.get_children()
        if not all_items:
            messagebox.showinfo("No Data", "No records to export."); return

        panchayat = self.panchayat_entry.get().strip()
        if not panchayat:
            messagebox.showwarning("Required", "Panchayat Name missing."); return

        # --- Filter Data ---
        filter_mode = self.export_filter_menu.get()
        data_export = []
        
        total_recs = 0
        success_count = 0
        failed_count = 0

        for item_id in all_items:
            vals = self.results_tree.item(item_id)['values']
            # Indexes: 0=Panch, 1=WC, 2=Name, 3=MR, 4=Period, 5=Status, 6=Detail, 7=Time
            status = vals[5].upper()
            
            if "SUCCESS" in status: success_count += 1
            else: failed_count += 1
            total_recs += 1

            if filter_mode == "Export All": data_export.append(vals)
            elif filter_mode == "Success Only" and "SUCCESS" in status: data_export.append(vals)
            elif filter_mode == "Failed Only" and "SUCCESS" not in status: data_export.append(vals)

        if not data_export:
            messagebox.showinfo("Empty", "No data matches the selected filter."); return

        # --- Setup Path (Downloads/NregaBot/MB Reports {Year}/{Panchayat}) ---
        year = date.today().year
        date_str = date.today().strftime("%d-%m-%Y")
        user_downloads = self.app.get_user_downloads_path()
        save_dir = os.path.join(user_downloads, "NregaBot", f"Reports {year}", "MB Report", panchayat)
        
        if not os.path.exists(save_dir): os.makedirs(save_dir)
        
        default_name = f"eMB_Report_{panchayat}_{date_str}.xlsx"
        filename = filedialog.asksaveasfilename(
            initialdir=save_dir, 
            initialfile=default_name, 
            defaultextension=".xlsx", 
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if not filename: return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "eMB Report"

            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            # Green Theme for MB (Payment related)
            header_fill = PatternFill(start_color="107C10", end_color="107C10", fill_type="solid") 
            white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            center = Alignment(horizontal="center", vertical="center")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # Main Header
            ws.merge_cells('A1:G1')
            ws['A1'] = f"e-MB ENTRY REPORT: {panchayat.upper()}"
            ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
            ws['A1'].fill = header_fill
            ws['A1'].alignment = center

            ws.merge_cells('A2:G2')
            ws['A2'] = f"Report Generated from NregaBot.com | Date: {datetime.now().strftime('%d-%m-%Y %I:%M %p')}"
            ws['A2'].font = Font(italic=True, size=9)
            ws['A2'].alignment = center

            # Summary Stats (Row 4 & 5)
            headers = ["Total Processed", "Success", "Failed"]
            vals = [total_recs, success_count, failed_count]
            
            # Place summary starting at Col C to E to center it roughly
            start_col = 3 
            for i, (h, v) in enumerate(zip(headers, vals)):
                col_idx = start_col + i
                c_h = ws.cell(row=4, column=col_idx, value=h)
                c_h.font = Font(bold=True)
                c_h.fill = PatternFill(start_color="DCE6F1", fill_type="solid")
                c_h.alignment = center
                c_h.border = border
                
                c_v = ws.cell(row=5, column=col_idx, value=v)
                c_v.font = Font(bold=True, size=11)
                c_v.alignment = center
                c_v.border = border
                if h == "Failed" and v > 0: c_v.font = Font(color="FF0000", bold=True)
                if h == "Success": c_v.font = Font(color="006100", bold=True)

            # Data Table (Row 7)
            # Tree Columns: 0=Panch, 1=WC, 2=Name, 3=MR, 4=Period, 5=Status, 6=Detail, 7=Time
            # Excel Cols: SNo, WorkCode, WorkName, MR No, Period, Status, Detail
            
            table_headers = ["S.No", "Work Code", "Work Name", "Muster Roll No", "Period", "Status", "Details"]
            t_row = 7

            for i, h in enumerate(table_headers, 1):
                c = ws.cell(row=t_row, column=i, value=h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = center
                c.border = border

            for idx, r in enumerate(data_export, 1):
                r_idx = t_row + idx
                fill = gray_fill if idx % 2 == 0 else white_fill
                
                # SNo
                c1 = ws.cell(row=r_idx, column=1, value=idx)
                c1.alignment = center; c1.fill = fill; c1.border = border
                
                # WC
                c2 = ws.cell(row=r_idx, column=2, value=r[1])
                c2.fill = fill; c2.border = border
                
                # Name (Handles Hindi Automatically)
                c3 = ws.cell(row=r_idx, column=3, value=r[2])
                c3.fill = fill; c3.border = border
                
                # MR
                c4 = ws.cell(row=r_idx, column=4, value=r[3])
                c4.alignment = center; c4.fill = fill; c4.border = border
                
                # Period
                c5 = ws.cell(row=r_idx, column=5, value=r[4])
                c5.alignment = center; c5.fill = fill; c5.border = border

                # Status
                c6 = ws.cell(row=r_idx, column=6, value=r[5])
                c6.alignment = center; c6.fill = fill; c6.border = border
                if "SUCCESS" in r[5].upper(): c6.font = Font(color="006100", bold=True)
                else: c6.font = Font(color="FF0000", bold=True)

                # Details
                c7 = ws.cell(row=r_idx, column=7, value=r[6])
                c7.fill = fill; c7.border = border

            # Widths
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 40 # Wide for Name
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 30

            wb.save(filename)
            messagebox.showinfo("Success", f"Professional Report Saved!\n{filename}")
            try:
                if os.name == 'nt': os.startfile(filename)
                else: subprocess.call(['open', filename])
            except: pass

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to save Excel: {e}")

    def export_professional_pdf(self):
        """Generates a professional PDF report with Hindi font support."""
        # --- Check if Library Exists ---
        if not HAS_REPORTLAB:
            messagebox.showerror("Missing Library", "PDF generation requires 'reportlab'.\nPlease run in terminal: pip install reportlab")
            return

        all_items = self.results_tree.get_children()
        if not all_items:
            messagebox.showinfo("No Data", "No records to export."); return

        panchayat = self.panchayat_entry.get().strip()
        if not panchayat:
            messagebox.showwarning("Required", "Panchayat Name missing."); return

        # --- Filter Data ---
        filter_mode = self.export_filter_menu.get()
        data_to_export = []
        
        total_recs = 0
        success_count = 0
        failed_count = 0

        for item_id in all_items:
            vals = self.results_tree.item(item_id)['values']
            status = vals[5].upper()
            
            if "SUCCESS" in status: success_count += 1
            else: failed_count += 1
            total_recs += 1

            if filter_mode == "Export All": data_to_export.append(vals)
            elif filter_mode == "Success Only" and "SUCCESS" in status: data_to_export.append(vals)
            elif filter_mode == "Failed Only" and "SUCCESS" not in status: data_to_export.append(vals)

        if not data_to_export:
            messagebox.showinfo("Empty", "No data matches the selected filter."); return

        # --- Path Setup ---
        year = date.today().year
        date_str = date.today().strftime("%d-%m-%Y")
        user_downloads = self.app.get_user_downloads_path()
        save_dir = os.path.join(user_downloads, "NregaBot", f"Reports {year}", "MB Report", panchayat)
        if not os.path.exists(save_dir): os.makedirs(save_dir)
        
        default_name = f"eMB_Report_{panchayat}_{date_str}.pdf"
        filename = filedialog.asksaveasfilename(
            initialdir=save_dir, 
            initialfile=default_name, 
            defaultextension=".pdf", 
            filetypes=[("PDF Document", "*.pdf")]
        )
        if not filename: return

        try:
            # --- 1. Font Registration (Fix for Hindi/Question Marks) ---
            font_name = "Helvetica" # Default fallback
            
            try:
                # Common Fonts paths
                font_paths = [
                    "C:\\Windows\\Fonts\\arial.ttf", 
                    "C:\\Windows\\Fonts\\Nirmala.ttf", 
                    "C:\\Windows\\Fonts\\mangal.ttf",
                    "/Library/Fonts/Arial Unicode.ttf", # Mac
                    "/System/Library/Fonts/Supplemental/Arial.ttf" # Mac
                ]
                
                selected_font = None
                for fp in font_paths:
                    if os.path.exists(fp):
                        selected_font = fp
                        break
                
                if selected_font:
                    pdfmetrics.registerFont(TTFont('HindiFont', selected_font))
                    font_name = 'HindiFont'
            except Exception as e:
                print(f"Font registration failed: {e}")

            # --- 2. Build PDF Elements ---
            doc = SimpleDocTemplate(filename, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'TitleStyle', 
                parent=styles['Heading1'], 
                fontName=font_name, 
                fontSize=16, 
                alignment=1, 
                spaceAfter=10,
                textColor=colors.white,
                backColor=colors.HexColor("#107C10"),
                borderPadding=5
            )
            elements.append(Paragraph(f"e-MB ENTRY REPORT: {panchayat.upper()}", title_style))
            
            # Subtitle
            sub_style = ParagraphStyle('SubStyle', parent=styles['Normal'], fontName=font_name, fontSize=9, alignment=1, spaceAfter=20)
            elements.append(Paragraph(f"Generated by NregaBot.com | Date: {datetime.now().strftime('%d-%m-%Y %I:%M %p')}", sub_style))

            # Summary Table
            summary_data = [
                ["Total Processed", "Success", "Failed"],
                [str(total_recs), str(success_count), str(failed_count)]
            ]
            
            sum_table = Table(summary_data, colWidths=[120, 100, 100])
            sum_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#DCE6F1")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('TEXTCOLOR', (2, 1), (2, 1), colors.red), 
                ('TEXTCOLOR', (1, 1), (1, 1), colors.green),
                ('FONTSIZE', (0, 1), (-1, 1), 12),
            ]))
            elements.append(sum_table)
            elements.append(Spacer(1, 20))

            # Main Data Table
            table_header = ["S.No", "Work Code", "Work Name", "Muster Roll", "Period", "Status", "Details"]
            table_data = [table_header]

            row_style = ParagraphStyle('RowStyle', fontName=font_name, fontSize=9)
            
            for idx, item in enumerate(data_to_export, 1):
                row = [
                    str(idx),
                    Paragraph(str(item[1]), row_style),
                    Paragraph(str(item[2]), row_style),
                    str(item[3]),
                    str(item[4]),
                    str(item[5]),
                    Paragraph(str(item[6]), row_style)
                ]
                table_data.append(row)

            col_widths = [40, 100, 200, 80, 110, 70, 160]
            
            main_table = Table(table_data, colWidths=col_widths, repeatRows=1)
            main_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#107C10")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            
            elements.append(main_table)
            doc.build(elements)
            
            messagebox.showinfo("Success", f"Professional PDF Saved!\n{filename}")
            try:
                if os.name == 'nt': os.startfile(filename)
                else: subprocess.call(['open', filename])
            except: pass

        except Exception as e:
            messagebox.showerror("PDF Error", f"Failed to generate PDF: {e}")
    
    def load_data_from_mr_tracking(self, workcodes: str, panchayat_name: str):
        self.panchayat_entry.delete(0, tkinter.END)
        self.panchayat_entry.insert(0, panchayat_name)
        self._on_panchayat_change()
        self.work_codes_text.configure(state="normal")
        self.work_codes_text.delete("1.0", tkinter.END)
        self.work_codes_text.insert("1.0", workcodes)
        self.work_codes_text.configure(state="disabled")
        if self.notebook: self.notebook.set("Work Codes")