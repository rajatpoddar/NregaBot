# tabs/nmms_attendance_tab.py
# NMMS Daily Attendance Viewer & Report Generator
# New approach: "Scrape Current Page" workflow — reads whatever is already visible in browser

import tkinter
from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk
import time, os, re, json, requests, threading
from datetime import datetime
from urllib.parse import urlencode, urlparse, parse_qs

import pandas as pd  # type: ignore
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore

from selenium.webdriver.common.by import By  # type: ignore
from selenium.webdriver.support.ui import WebDriverWait  # type: ignore
from selenium.webdriver.support import expected_conditions as EC  # type: ignore
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException  # type: ignore

from .base_tab import BaseAutomationTab

# ---------------------------------------------------------------------------
NMMS_BASE_URL = "https://vbgramgrep.dord.gov.in/vbgramg/NMMS_DailyAttendance.aspx"


class NmmsAttendanceTab(BaseAutomationTab):
    """
    NMMS Daily Attendance scraper.
    Flow:
      1. 'Open NMMS Page' → opens portal in connected browser
      2. User selects State/Date/Block in browser → panchayat list appears
      3. 'Scrape Current Page' → reads the panchayat table already on screen
      4. User selects panchayats → Start
      5. For each panchayat: click MR link → scrape MR list → scrape each MR detail
      6. Export professional Excel report
    """

    SUMMARY_HEADERS = [
        "S No.", "Panchayat", "Work Code", "Msr No.", "Work Name",
        "No. of Workers", "Taken By", "Designation",
        "Photo-1 Taken", "Photo-1 Uploaded", "Geo Coordinates (P1)",
        "Photo-2 Taken", "Photo-2 Uploaded", "Geo Coordinates (P2)",
        "Photo-1 Saved", "Photo-2 Saved",
    ]
    WORKER_HEADERS = [
        "S No.", "Panchayat", "Work Code", "Msr No.",
        "Job Card No.", "Worker Name", "Gender", "Attendance Date", "Status",
    ]
    PAN_OVERVIEW_HEADERS = [
        "S No.", "Panchayat", "No. of Works", "No. of Muster Rolls", "Persondays Generated",
    ]

    def __init__(self, parent, app_instance):
        super().__init__(parent, app_instance, automation_key="nmms_attendance")
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self._panchayat_data: list = []
        self._config_file = self.app.get_data_path("nmms_inputs.json")
        self._create_widgets()
        self._load_inputs()

    # -----------------------------------------------------------------------
    # UI
    # -----------------------------------------------------------------------
    def _create_widgets(self):
        top = ctk.CTkFrame(self)
        top.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 5))
        top.grid_columnconfigure(0, weight=1)

        # Instructions
        instr = (
            "STEPS:  1. Click 'Open NMMS Page' — portal opens in your browser.  "
            "2. In browser: select State, Attendance Date, Block and click Go.  "
            "3. Come back here and click 'Scrape Current Page'.  "
            "4. Select panchayats and click ▶ Start."
        )
        ctk.CTkLabel(top, text=instr, justify="left", wraplength=950,
                     fg_color=("gray90", "#2A2A2A"), corner_radius=8,
                     padx=10, pady=8).grid(row=0, column=0, sticky="ew", padx=8, pady=(8, 4))

        # Button row
        btn_row = ctk.CTkFrame(top, fg_color="transparent")
        btn_row.grid(row=1, column=0, sticky="ew", padx=8, pady=(0, 4))

        self._open_btn = ctk.CTkButton(
            btn_row, text="🌐 Open NMMS Page", width=175,
            fg_color="#1565C0", hover_color="#0D47A1", command=self._open_nmms_page)
        self._open_btn.pack(side="left", padx=(0, 8))

        self._scrape_btn = ctk.CTkButton(
            btn_row, text="🔍 Scrape Current Page", width=190,
            fg_color="#2E7D32", hover_color="#1B5E20", command=self._scrape_current_page_thread)
        self._scrape_btn.pack(side="left", padx=(0, 16))

        self._save_photos_var = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(btn_row, text="Download Group Photos",
                        variable=self._save_photos_var).pack(side="left")

        # Panchayat selection
        pan_outer = ctk.CTkFrame(top, fg_color="transparent")
        pan_outer.grid(row=2, column=0, sticky="ew", padx=8, pady=(4, 2))
        pan_outer.grid_columnconfigure(0, weight=1)

        pan_hdr = ctk.CTkFrame(pan_outer, fg_color="transparent")
        pan_hdr.grid(row=0, column=0, sticky="ew")
        ctk.CTkLabel(pan_hdr, text="Select Panchayats:", font=ctk.CTkFont(weight="bold")).pack(side="left")
        ctk.CTkButton(pan_hdr, text="Select All", width=90, command=self._select_all).pack(side="right", padx=(4, 0))
        ctk.CTkButton(pan_hdr, text="Clear All", width=90, command=self._clear_all).pack(side="right")

        self._pan_scroll = ctk.CTkScrollableFrame(pan_outer, height=120)
        self._pan_scroll.grid(row=1, column=0, sticky="ew", pady=(4, 0))
        self._pan_checkboxes: dict = {}

        self._pan_info_lbl = ctk.CTkLabel(
            pan_outer, text_color="gray50", justify="left", wraplength=950,
            text="No panchayats loaded. Open the NMMS page, navigate to panchayat list, then click 'Scrape Current Page'.")
        self._pan_info_lbl.grid(row=2, column=0, sticky="w", pady=(2, 2))

        # Start / Stop / Retry / Reset buttons
        self._create_action_buttons(parent_frame=top).grid(row=3, column=0, pady=(8, 10))

        # Bottom notebook
        nb = ctk.CTkTabview(self)
        nb.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 10))
        self._build_summary_tab(nb.add("MR Summary"))
        self._build_workers_tab(nb.add("Workers Detail"))
        self._create_log_and_status_area(parent_notebook=nb)

    def _build_summary_tab(self, tab):
        tab.grid_columnconfigure(0, weight=1)
        tab.grid_rowconfigure(1, weight=1)

        toolbar = ctk.CTkFrame(tab, fg_color="transparent")
        toolbar.grid(row=0, column=0, columnspan=2, sticky="w", padx=5, pady=5)
        ctk.CTkButton(toolbar, text="📊 Export Excel Report",
                      fg_color="#2E7D32", hover_color="#1B5E20",
                      command=self._export_excel).pack(side="left", padx=(0, 8))
        ctk.CTkButton(toolbar, text="Clear Results", width=100,
                      fg_color=("gray70", "#4A4A4A"), text_color="white",
                      command=self._clear_results).pack(side="left")

        self.results_tree = ttk.Treeview(tab, columns=self.SUMMARY_HEADERS, show="headings")
        _w = {"S No.": 45, "Panchayat": 100, "Work Code": 195, "Msr No.": 55,
              "Work Name": 190, "No. of Workers": 80, "Taken By": 130, "Designation": 100,
              "Photo-1 Taken": 140, "Photo-1 Uploaded": 140, "Geo Coordinates (P1)": 150,
              "Photo-2 Taken": 140, "Photo-2 Uploaded": 140, "Geo Coordinates (P2)": 150,
              "Photo-1 Saved": 85, "Photo-2 Saved": 85}
        for col in self.SUMMARY_HEADERS:
            self.results_tree.heading(col, text=col)
            self.results_tree.column(col, width=_w.get(col, 100), minwidth=40)

        self.results_tree.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
        vsb = ctk.CTkScrollbar(tab, command=self.results_tree.yview)
        hsb = ctk.CTkScrollbar(tab, orientation="horizontal", command=self.results_tree.xview)
        self.results_tree.configure(yscroll=vsb.set, xscroll=hsb.set)
        vsb.grid(row=1, column=1, sticky="ns")
        hsb.grid(row=2, column=0, sticky="ew")
        self.style_treeview(self.results_tree)

    def _build_workers_tab(self, tab):
        tab.grid_columnconfigure(0, weight=1)
        tab.grid_rowconfigure(1, weight=1)

        toolbar = ctk.CTkFrame(tab, fg_color="transparent")
        toolbar.grid(row=0, column=0, columnspan=2, sticky="w", padx=5, pady=5)
        ctk.CTkButton(toolbar, text="📋 Export Workers CSV",
                      command=lambda: self.export_treeview_to_csv(
                          self.workers_tree, "NMMS_Workers_Detail.csv")).pack(side="left")

        self.workers_tree = ttk.Treeview(tab, columns=self.WORKER_HEADERS, show="headings")
        _ww = {"S No.": 45, "Panchayat": 100, "Work Code": 180, "Msr No.": 55,
               "Job Card No.": 160, "Worker Name": 180, "Gender": 55,
               "Attendance Date": 100, "Status": 70}
        for col in self.WORKER_HEADERS:
            self.workers_tree.heading(col, text=col)
            self.workers_tree.column(col, width=_ww.get(col, 100), minwidth=30)

        self.workers_tree.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
        vsb2 = ctk.CTkScrollbar(tab, command=self.workers_tree.yview)
        self.workers_tree.configure(yscroll=vsb2.set)
        vsb2.grid(row=1, column=1, sticky="ns")
        self.style_treeview(self.workers_tree)

    # -----------------------------------------------------------------------
    # CONFIG / HELPERS
    # -----------------------------------------------------------------------
    def _save_inputs(self):
        try:
            with open(self._config_file, "w", encoding="utf-8") as f:
                json.dump({"save_photos": self._save_photos_var.get()}, f, indent=4)
        except Exception:
            pass

    def _load_inputs(self):
        try:
            if os.path.exists(self._config_file):
                with open(self._config_file, "r", encoding="utf-8") as f:
                    self._save_photos_var.set(json.load(f).get("save_photos", True))
        except Exception:
            pass

    def _select_all(self):
        for v in self._pan_checkboxes.values(): v.set(True)

    def _clear_all(self):
        for v in self._pan_checkboxes.values(): v.set(False)

    def _clear_results(self):
        for i in self.results_tree.get_children(): self.results_tree.delete(i)
        for i in self.workers_tree.get_children(): self.workers_tree.delete(i)
        self.update_status("Cleared.", 0)

    def set_ui_state(self, running: bool):
        self.set_common_ui_state(running)
        s = "disabled" if running else "normal"
        self._open_btn.configure(state=s)
        self._scrape_btn.configure(state=s)

    def reset_ui(self):
        self._clear_results()
        self.update_status("Ready", 0)
        self.app.log_message(self.log_display, "Reset complete.")

    def _get_driver(self):
        driver = self.app.get_driver()
        if not driver:
            messagebox.showwarning(
                "Browser Not Connected",
                "No browser found.\n\nPlease launch Chrome/Edge from the app and log in to NREGA portal first.")
        return driver

    # -----------------------------------------------------------------------
    # PHASE 1 — OPEN PAGE
    # -----------------------------------------------------------------------
    def _open_nmms_page(self):
        driver = self._get_driver()
        if not driver:
            return
        try:
            driver.get(NMMS_BASE_URL)
            self.app.log_message(
                self.log_display,
                "NMMS page opened in browser.\n"
                "Please select State, Attendance Date and Block, then click Go.\n"
                "Once the panchayat list appears, come back and click 'Scrape Current Page'.",
                "info")
        except WebDriverException as e:
            messagebox.showerror("Browser Error", f"Could not open NMMS page:\n{e}")

    # -----------------------------------------------------------------------
    # PHASE 2 — SCRAPE CURRENT PAGE
    # -----------------------------------------------------------------------
    def _scrape_current_page_thread(self):
        driver = self._get_driver()
        if not driver:
            return
        self._scrape_btn.configure(state="disabled", text="Scraping...")
        threading.Thread(target=self._scrape_current_page_logic, args=(driver,), daemon=True).start()

    def _scrape_current_page_logic(self, driver):
        try:
            self.app.log_message(self.log_display, "Reading panchayat table from current browser page...", "info")
            self.app.log_message(self.log_display, f"  Current URL: {driver.current_url}")

            try:
                WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.XPATH, "//table//tr[td]")))
            except TimeoutException:
                body = driver.find_element(By.TAG_NAME, "body").text[:400]
                self.app.after(0, lambda t=body: messagebox.showwarning(
                    "No Table Found",
                    "No data table found on the current browser page.\n\n"
                    "Please navigate to the panchayat list in your browser first,\n"
                    "then click 'Scrape Current Page' again.\n\n"
                    f"Page preview:\n{t}"))
                return

            rows = driver.find_elements(By.XPATH, "//table//tr")
            data = []
            for row in rows:
                cells = row.find_elements(By.TAG_NAME, "td")
                if len(cells) < 4:
                    continue
                try:
                    sno = cells[0].text.strip()
                    if not sno.isdigit():
                        continue
                    pan = cells[1].text.strip()
                    if not pan or pan.lower() == "total":
                        continue
                    no_works   = cells[2].text.strip() if len(cells) > 2 else ""
                    no_mr      = cells[3].text.strip() if len(cells) > 3 else ""
                    persondays = cells[4].text.strip() if len(cells) > 4 else ""
                    mr_href = ""
                    try:
                        mr_href = cells[3].find_element(By.TAG_NAME, "a").get_attribute("href") or ""
                    except NoSuchElementException:
                        pass
                    data.append({"sno": sno, "name": pan, "no_works": no_works,
                                 "no_mr": no_mr, "persondays": persondays, "mr_href": mr_href})
                    self.app.log_message(self.log_display,
                        f"  Found: {pan} | MRs: {no_mr} | href: {(mr_href[:70] + '...') if len(mr_href) > 70 else mr_href}")
                except Exception:
                    continue

            self.app.after(0, lambda d=data: self._populate_panchayat_checkboxes(d))
        except Exception as e:
            self.app.log_message(self.log_display, f"Scrape error: {e}", "error")
            self.app.after(0, lambda err=str(e): messagebox.showerror("Scrape Error", f"Could not read the page:\n{err}"))
        finally:
            self.app.after(0, lambda: self._scrape_btn.configure(state="normal", text="🔍 Scrape Current Page"))

    def _populate_panchayat_checkboxes(self, data: list):
        for w in self._pan_scroll.winfo_children(): w.destroy()
        self._pan_checkboxes.clear()
        self._panchayat_data = data

        if not data:
            self._pan_info_lbl.configure(text="No panchayats found. Make sure panchayat list is visible in browser.")
            self.app.log_message(self.log_display, "No panchayats found on page.", "warning")
            return

        self._pan_scroll.grid_columnconfigure(0, weight=1)
        for i, item in enumerate(data):
            var = ctk.BooleanVar(value=True)
            label = f"{item['name']}  (Works: {item['no_works']} | MRs: {item['no_mr']} | Persondays: {item['persondays']})"
            ctk.CTkCheckBox(self._pan_scroll, text=label, variable=var).grid(row=i, column=0, sticky="w", padx=5, pady=2)
            self._pan_checkboxes[item["name"]] = var

        self._pan_info_lbl.configure(text=f"✅ {len(data)} panchayat(s) loaded. Select desired ones and click ▶ Start.")
        self.app.log_message(self.log_display, f"Scraped {len(data)} panchayats.", "success")

    # -----------------------------------------------------------------------
    # PHASE 3 — START AUTOMATION
    # -----------------------------------------------------------------------
    def start_automation(self):
        selected = [n for n, v in self._pan_checkboxes.items() if v.get()]
        if not selected:
            messagebox.showwarning("No Selection", "Please select at least one panchayat.")
            return
        self._save_inputs()
        self.app.log_message(self.log_display, f"Starting for {len(selected)} panchayat(s)...")
        self.app.start_automation_thread(self.automation_key, self._run_scrape_logic, args=(selected,))

    def _run_scrape_logic(self, selected_panchayats: list):
        self.app.after(0, self.set_ui_state, True)
        self.app.after(0, self.update_status, "Initializing...", 0.0)

        driver = self._get_driver()
        if not driver:
            self.app.after(0, self.set_ui_state, False)
            return

        date_safe = datetime.now().strftime("%d-%m-%Y")
        photos_dir = os.path.join(self.app.get_user_downloads_path(), "NregaBot", "NMMS_Attendance", date_safe, "Photos")
        os.makedirs(photos_dir, exist_ok=True)

        page1_url = driver.current_url
        summary_sno = len(self.results_tree.get_children())
        worker_sno  = len(self.workers_tree.get_children())
        total = len(selected_panchayats)

        try:
            for p_idx, pan_name in enumerate(selected_panchayats):
                if self.app.stop_events[self.automation_key].is_set():
                    self.app.log_message(self.log_display, "Stop signal received.", "warning")
                    break

                self.app.after(0, self.update_status, f"Panchayat {p_idx+1}/{total}: {pan_name}", p_idx / total)
                self.app.log_message(self.log_display, f"▶ {pan_name}")

                pan_info = next((d for d in self._panchayat_data if d["name"] == pan_name), None)
                if not pan_info:
                    self.app.log_message(self.log_display, f"  ⚠ No data for '{pan_name}'", "warning")
                    continue

                # Return to page 1 if needed
                if driver.current_url != page1_url:
                    driver.get(page1_url)
                    time.sleep(2)

                mr_rows = self._click_and_scrape_mr_list(pan_info, driver, pan_name)
                if not mr_rows:
                    self.app.log_message(self.log_display, f"  No MRs found for {pan_name}.", "warning")
                    continue

                self.app.log_message(self.log_display, f"  {len(mr_rows)} MR(s) found.")

                for mr_info in mr_rows:
                    if self.app.stop_events[self.automation_key].is_set():
                        break
                    summary_sno += 1
                    detail = self._scrape_mr_detail(mr_info, driver, pan_name, photos_dir)

                    srow = (summary_sno, pan_name,
                            mr_info.get("work_code",""), mr_info.get("msr_no",""),
                            detail.get("work_name",""), detail.get("worker_count","0"),
                            detail.get("taken_by",""), detail.get("designation",""),
                            detail.get("photo1_taken",""), detail.get("photo1_uploaded",""),
                            detail.get("photo1_geo",""),
                            detail.get("photo2_taken",""), detail.get("photo2_uploaded",""),
                            detail.get("photo2_geo",""),
                            detail.get("photo1_saved","N/A"), detail.get("photo2_saved","N/A"))
                    self.app.after(0, lambda r=srow: self.results_tree.insert("", "end", values=r))

                    for w in detail.get("workers", []):
                        worker_sno += 1
                        wrow = (worker_sno, pan_name,
                                mr_info.get("work_code",""), mr_info.get("msr_no",""),
                                w.get("jobcard",""), w.get("name",""),
                                w.get("gender",""), w.get("date",""), w.get("status",""))
                        self.app.after(0, lambda r=wrow: self.workers_tree.insert("", "end", values=r))

            self.app.after(0, self.update_status, "Completed!", 1.0)
            self.app.log_message(self.log_display, f"Done! {summary_sno} MR(s) scraped.", "success")
            cnt = summary_sno
            self.app.after(200, lambda: messagebox.showinfo("Complete",
                f"Scraping complete!\nTotal MRs: {cnt}\nPhotos: {photos_dir}\n\nClick 'Export Excel Report' to save."))

        except Exception as e:
            self.app.log_message(self.log_display, f"Critical error: {e}", "error")
            self.app.after(0, lambda err=str(e): messagebox.showerror("Error", f"Scraping failed:\n{err}"))
        finally:
            self.app.after(0, self.set_ui_state, False)
            self.app.after(0, self.app.set_status, "Ready")

    # -----------------------------------------------------------------------
    # SCRAPING HELPERS
    # -----------------------------------------------------------------------
    def _click_and_scrape_mr_list(self, pan_info: dict, driver, pan_name: str) -> list:
        """Navigate to MR list page using saved href, scrape MR rows."""
        mr_rows = []
        try:
            href = pan_info.get("mr_href", "")
            if href:
                self.app.log_message(self.log_display, f"  MR list: {href[:80]}")
                driver.get(href)
            else:
                # Fallback: find link by panchayat name on current page
                self.app.log_message(self.log_display, f"  No href, searching link for '{pan_name}'...", "warning")
                try:
                    link = driver.find_element(
                        By.XPATH,
                        f"//tr[td[normalize-space()='{pan_name}']]//td[4]//a"
                    )
                    link.click()
                except NoSuchElementException:
                    self.app.log_message(self.log_display, f"  ⚠ MR link not found for '{pan_name}'.", "warning")
                    return []

            time.sleep(2)

            try:
                WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.XPATH, "//table//tr[td]")))
            except TimeoutException:
                self.app.log_message(self.log_display, "  ⚠ Timeout on MR list page.", "warning")
                return []

            for row in driver.find_elements(By.XPATH, "//table//tr"):
                cells = row.find_elements(By.TAG_NAME, "td")
                if len(cells) < 6:
                    continue
                try:
                    if not cells[0].text.strip().isdigit():
                        continue
                    work_code  = cells[4].text.strip()
                    msr_cell   = cells[5]
                    msr_no     = msr_cell.text.strip()
                    persondays = cells[6].text.strip() if len(cells) > 6 else ""
                    detail_href = ""
                    try:
                        detail_href = msr_cell.find_element(By.TAG_NAME, "a").get_attribute("href") or ""
                    except NoSuchElementException:
                        pass
                    mr_rows.append({"work_code": work_code, "msr_no": msr_no,
                                    "persondays": persondays, "detail_href": detail_href})
                except Exception:
                    continue
        except Exception as e:
            self.app.log_message(self.log_display, f"  MR list error: {e}", "error")
        return mr_rows

    def _scrape_mr_detail(self, mr_info: dict, driver, pan_name: str, photos_dir: str) -> dict:
        """Navigate to MR detail page, scrape photo info + worker table."""
        detail = {
            "work_name": "",
            "photo1_taken": "", "photo1_uploaded": "", "photo1_geo": "",
            "photo2_taken": "", "photo2_uploaded": "", "photo2_geo": "",
            "taken_by": "", "designation": "",
            "worker_count": "0", "workers": [],
            "photo1_saved": "No", "photo2_saved": "No",
        }
        href = mr_info.get("detail_href", "")
        if not href:
            self.app.log_message(self.log_display,
                f"    ⚠ No detail href for MR {mr_info.get('msr_no','?')}", "warning")
            return detail

        try:
            self.app.log_message(self.log_display,
                f"    MR {mr_info['msr_no']} | {mr_info['work_code']}")
            driver.get(href)
            time.sleep(2)
            page_src = driver.page_source

            # Work Name
            try:
                wn = driver.find_element(By.XPATH,
                    "//*[contains(translate(text(),'WORKNAME ','workname '),'work name')]/following::td[1]")
                detail["work_name"] = wn.text.strip()
            except NoSuchElementException:
                detail["work_name"] = self._re(page_src, r"Work Name[^:]*:\s*([^\n<]{3,100})")

            # Photo info blocks
            self._extract_photo_info(driver, page_src, 1, detail)
            self._extract_photo_info(driver, page_src, 2, detail)

            # Photos download
            if self._save_photos_var.get():
                detail["photo1_saved"] = self._download_photo(1, pan_name, mr_info, photos_dir, driver)
                detail["photo2_saved"] = self._download_photo(2, pan_name, mr_info, photos_dir, driver)

            # Worker table
            workers = []
            # Try table with Job Card header first
            w_rows = driver.find_elements(By.XPATH,
                "//table[.//th[contains(text(),'Job Card')] or "
                ".//td[contains(text(),'Job Card No')]]//tr[position()>1]")
            if not w_rows:
                w_rows = driver.find_elements(By.XPATH, "//table//tr[count(td)>=5]")

            for wr in w_rows:
                wc = wr.find_elements(By.TAG_NAME, "td")
                if len(wc) < 5 or not wc[0].text.strip().isdigit():
                    continue
                try:
                    raw  = wc[2].text.strip()
                    gm   = re.search(r"\(([MFmf])\)", raw)
                    workers.append({
                        "jobcard": wc[1].text.strip(),
                        "name":    re.sub(r"\s*\([MFmf]\)\s*", "", raw).strip(),
                        "gender":  gm.group(1).upper() if gm else "",
                        "date":    wc[3].text.strip(),
                        "status":  wc[4].text.strip(),
                    })
                except Exception:
                    continue

            detail["workers"]      = workers
            detail["worker_count"] = str(len(workers))

        except Exception as e:
            self.app.log_message(self.log_display, f"    Detail error: {e}", "error")
        return detail

    def _extract_photo_info(self, driver, page_src: str, photo_no: int, detail: dict):
        """Extract timestamp / geo / taken-by for a photo block. Writes into detail dict."""
        pkey = f"photo{photo_no}_"
        try:
            block = driver.find_element(By.XPATH,
                f"//td[contains(.,'Timestamp for Photo-{photo_no}')]"
                f"/ancestor::table[1]")
            detail[f"{pkey}taken"]    = self._field(block, ["Taken"])
            detail[f"{pkey}uploaded"] = self._field(block, ["Uploaded"])
            detail[f"{pkey}geo"]      = self._field(block, ["Geo Co-ordinates", "Geo"])
            if photo_no == 1:
                detail["taken_by"]    = self._field(block, ["Taken by", "Taken By"])
                detail["designation"] = self._field(block, ["Designation"])
        except NoSuchElementException:
            if photo_no == 1:
                detail[f"{pkey}taken"]    = self._re(page_src, r"Taken\s*:\s*([^\|<\n]{5,45})")
                detail[f"{pkey}uploaded"] = self._re(page_src, r"Uploaded\s*:\s*([^\|<\n]{5,45})")
                detail[f"{pkey}geo"]      = self._re(page_src, r"Geo Co-ordinates\s*:\s*([\d.,\- ]+)")
                detail["taken_by"]        = self._re(page_src, r"Taken by\s*:\s*([A-Za-z ]{3,50})")
                detail["designation"]     = self._re(page_src, r"Designation\s*:\s*([^\n<]{3,60})")
            # Photo 2 not uploaded is normal — no warning

    def _field(self, block_elem, labels: list) -> str:
        """Extract a labelled field value from a table block element."""
        try:
            tds = block_elem.find_elements(By.TAG_NAME, "td")
            for i, td in enumerate(tds):
                txt = td.text.strip()
                if any(lb.lower() in txt.lower() for lb in labels):
                    if i + 1 < len(tds):
                        val = tds[i + 1].text.strip()
                        if val and not any(lb.lower() in val.lower()
                                           for lb in ["taken", "uploaded", "geo", "designation"]):
                            return val
                    if ":" in txt:
                        return txt.split(":", 1)[1].strip()
        except Exception:
            pass
        return ""

    def _re(self, html: str, pattern: str) -> str:
        m = re.search(pattern, html, re.IGNORECASE)
        return m.group(1).strip() if m else ""

    def _download_photo(self, photo_no: int, pan_name: str,
                        mr_info: dict, photos_dir: str, driver) -> str:
        """Download a group photo using the browser's session cookies."""
        try:
            # Find img by id pattern
            imgs = driver.find_elements(By.XPATH,
                f"//img[contains(@id,'Photo{photo_no}') or contains(@id,'photo{photo_no}') "
                f"or contains(@id,'Img{photo_no}') or contains(@id,'img{photo_no}')]")
            if not imgs:
                all_imgs = driver.find_elements(By.TAG_NAME, "img")
                cands = [i for i in all_imgs
                         if any(k in (i.get_attribute("src") or "").lower()
                                for k in ("photo", "grp", "group", "attendance"))]
                if photo_no <= len(cands):
                    imgs = [cands[photo_no - 1]]
                else:
                    return "Not Uploaded"

            src = imgs[0].get_attribute("src") if imgs else ""
            if not src or any(s in src.lower() for s in ("spacer", "blank", "1x1")):
                return "Not Uploaded"

            # Ensure absolute URL
            if src.startswith("//"):
                src = "https:" + src
            elif not src.startswith("http"):
                src = NMMS_BASE_URL.rsplit("/", 1)[0] + "/" + src.lstrip("/")

            safe_pan = re.sub(r'[\\/*?:"<>|]', "_", pan_name)
            safe_wc  = re.sub(r'[\\/*?:"<>|/]', "_", mr_info.get("work_code", "WC"))
            fname    = f"{safe_pan}_MR{mr_info.get('msr_no','0')}_{safe_wc}_Photo{photo_no}.jpg"
            path     = os.path.join(photos_dir, fname)

            cookies  = {c["name"]: c["value"] for c in driver.get_cookies()}
            resp     = requests.get(src, headers={"User-Agent": "Mozilla/5.0"},
                                    cookies=cookies, timeout=20)
            if resp.status_code == 200 and len(resp.content) > 500:
                with open(path, "wb") as f:
                    f.write(resp.content)
                self.app.log_message(self.log_display, f"    📷 Photo {photo_no}: {fname}")
                return "Yes"
            return "No"
        except Exception as e:
            self.app.log_message(self.log_display, f"    Photo {photo_no} failed: {e}", "warning")
            return "Error"

    # -----------------------------------------------------------------------
    # EXCEL EXPORT
    # -----------------------------------------------------------------------
    def _export_excel(self):
        if not self.results_tree.get_children():
            messagebox.showinfo("No Data", "No results to export. Run the scrape first.")
            return

        date_safe = datetime.now().strftime("%d-%m-%Y")
        target_dir = os.path.join(self.app.get_user_downloads_path(), "NregaBot", "NMMS_Attendance", date_safe)
        os.makedirs(target_dir, exist_ok=True)

        file_path = filedialog.asksaveasfilename(
            initialdir=target_dir,
            initialfile=f"NMMS_Attendance_Report_{date_safe}.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Save NMMS Attendance Report")
        if not file_path:
            return

        try:
            self._write_excel(file_path, date_safe)
            messagebox.showinfo("Exported", f"Report saved!\n\n{file_path}")
            if os.name == "nt":
                try: os.startfile(file_path)
                except Exception: pass
        except Exception as e:
            messagebox.showerror("Export Error", f"Could not save report:\n{e}")

    def _write_excel(self, path: str, date_str: str):
        summary_data = [self.results_tree.item(i, "values") for i in self.results_tree.get_children()]
        worker_data  = [self.workers_tree.item(i, "values")  for i in self.workers_tree.get_children()]

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            # Sheet 1 — MR Summary
            pd.DataFrame(summary_data, columns=self.SUMMARY_HEADERS).to_excel(
                writer, sheet_name="MR Summary", index=False, startrow=4)
            self._style_sheet(writer.sheets["MR Summary"],
                len(summary_data), self.SUMMARY_HEADERS,
                f"NMMS Daily Attendance — MR Summary",
                f"Date: {date_str}  |  Generated by NregaBot", "1565C0")

            # Sheet 2 — Workers Detail
            pd.DataFrame(worker_data, columns=self.WORKER_HEADERS).to_excel(
                writer, sheet_name="Workers Detail", index=False, startrow=4)
            self._style_sheet(writer.sheets["Workers Detail"],
                len(worker_data), self.WORKER_HEADERS,
                "NMMS Daily Attendance — Workers Detail",
                f"Date: {date_str}  |  Generated by NregaBot", "2E7D32")

            # Sheet 3 — Block Overview
            if self._panchayat_data:
                pan_rows = [[d["sno"], d["name"], d["no_works"], d["no_mr"], d["persondays"]]
                            for d in self._panchayat_data]
                pd.DataFrame(pan_rows, columns=self.PAN_OVERVIEW_HEADERS).to_excel(
                    writer, sheet_name="Block Overview", index=False, startrow=4)
                self._style_sheet(writer.sheets["Block Overview"],
                    len(pan_rows), self.PAN_OVERVIEW_HEADERS,
                    "NMMS Block Overview — Panchayat Summary",
                    f"Date: {date_str}  |  Generated by NregaBot", "6A1B9A")

    def _style_sheet(self, ws, n_data_rows: int, headers: list,
                     title: str, subtitle: str, hdr_color: str):
        """Apply title, subtitle, generated-by, header styling, and zebra striping."""
        n = len(headers)
        WHITE = "FFFFFF"
        thin  = Side(style="thin", color="BDBDBD")
        bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
        ctr   = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Row 1 — Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
        c1 = ws.cell(row=1, column=1, value=title)
        c1.font      = Font(size=14, bold=True, color=WHITE)
        c1.fill      = PatternFill(start_color=hdr_color, end_color=hdr_color, fill_type="solid")
        c1.alignment = ctr
        ws.row_dimensions[1].height = 26

        # Row 2 — Subtitle
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
        c2 = ws.cell(row=2, column=1, value=subtitle)
        c2.font = Font(italic=True, size=9); c2.alignment = ctr

        # Row 3 — Generated by
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n)
        c3 = ws.cell(row=3, column=1,
                     value=f"Generated by NregaBot  |  {datetime.now().strftime('%d-%m-%Y %I:%M %p')}")
        c3.font = Font(italic=True, size=8, color="808080"); c3.alignment = ctr

        # Row 5 — Column headers (pandas wrote them at row 5 due to startrow=4)
        hdr_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
        for cell in ws[5]:
            cell.font = Font(bold=True, size=10)
            cell.fill = hdr_fill
            cell.alignment = ctr
            cell.border = bdr

        # Rows 6+ — Data with zebra striping
        EVEN = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        ODD  = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
        for r_idx, row_cells in enumerate(
                ws.iter_rows(min_row=6, max_row=5 + n_data_rows, min_col=1, max_col=n)):
            fill = EVEN if r_idx % 2 == 0 else ODD
            for cell in row_cells:
                cell.fill   = fill
                cell.border = bdr
                cell.alignment = Alignment(vertical="center")

        # Auto-size columns
        for col_idx, col_name in enumerate(headers, start=1):
            letter = get_column_letter(col_idx)
            best = min(max(len(str(col_name)) + 3, 12), 55)
            ws.column_dimensions[letter].width = best
