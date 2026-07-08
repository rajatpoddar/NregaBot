# tabs/ekyc_report_tab.py
import time
import threading
import json
import os
import datetime
import subprocess
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException

# Excel Imports
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import config
from .base_tab import BaseAutomationTab
from .autocomplete_widget import AutocompleteEntry 

class EKycReportTab(BaseAutomationTab):
    def __init__(self, parent, app_instance):
        super().__init__(parent, app_instance, "ekyc_report")
        
        if self.automation_key not in self.app.stop_events:
            self.app.stop_events[self.automation_key] = threading.Event()
        
        self.all_scraped_data = []
        self.scraped_keys = set() # To track duplicates
        self._setup_ui()
        self.load_inputs()

    def _setup_ui(self):
        # --- 1. Input Section ---
        input_frame = ctk.CTkFrame(self)
        input_frame.pack(fill="x", padx=10, pady=10)

        # Panchayat Input (Autocomplete Linked to Global History)
        ctk.CTkLabel(input_frame, text="Panchayat:").grid(row=0, column=0, padx=(10, 5), pady=10, sticky="w")
        self.panchayat_entry = AutocompleteEntry(
            input_frame, 
            width=140, 
            placeholder_text="Leave empty for ALL",
            suggestions_list=self.app.history_manager.get_suggestions("panchayat_name"),
            app_instance=self.app,
            history_key="panchayat_name"
        )
        self.panchayat_entry.grid(row=0, column=1, padx=5, pady=10)

        # Village Input (Autocomplete Linked to Global History)
        ctk.CTkLabel(input_frame, text="Village:").grid(row=0, column=2, padx=(10, 5), pady=10, sticky="w")
        self.village_entry = AutocompleteEntry(
            input_frame, 
            width=140, 
            placeholder_text="Leave empty for ALL",
            suggestions_list=self.app.history_manager.get_suggestions("village_name"),
            app_instance=self.app,
            history_key="village_name"
        )
        self.village_entry.grid(row=0, column=3, padx=5, pady=10)

        # Filter Dropdown
        ctk.CTkLabel(input_frame, text="Filter:").grid(row=0, column=4, padx=(10, 5), pady=10, sticky="w")
        self.filter_var = ctk.StringVar(value="All")
        self.filter_cb = ctk.CTkComboBox(input_frame, variable=self.filter_var, 
                                         values=["All", "Verified (Yes)", "Not Verified (No)"], width=130,
                                         command=self.apply_filter_visuals)
        self.filter_cb.grid(row=0, column=5, padx=5, pady=10)

        # --- NEW: ABPS Pending Checkbox ---
        self.abps_check_var = ctk.BooleanVar(value=False)
        self.abps_check = ctk.CTkCheckBox(input_frame, text="Include ABPS Pending (Override)", 
                                          variable=self.abps_check_var, 
                                          command=self.apply_filter_visuals,
                                          font=("Arial", 11, "bold"), text_color="orange")
        self.abps_check.grid(row=1, column=0, columnspan=3, padx=10, pady=(5, 10), sticky="w")

        note_label = ctk.CTkLabel(self, text="ℹ️ Note: Leave Panchayat empty for ALL panchayats. Check 'Include ABPS Pending' to show rows with ABPS 'No' even if eKYC is 'Yes'.", 
                                  text_color=("gray40", "gray70"), font=("Arial", 11, "italic"))
        note_label.pack(anchor="w", padx=20, pady=(0, 5))

        # --- Stats Frame (per-panchayat summary) ---
        self.stats_frame = ctk.CTkFrame(self)
        self.stats_frame.pack(fill="x", padx=10, pady=(0, 5))
        self.stats_label_header = ctk.CTkLabel(self.stats_frame, text="📊 Panchayat-wise Summary:", font=("Arial", 11, "bold"))
        self.stats_label_header.pack(anchor="w", padx=10, pady=(5, 2))
        self.stats_text = ctk.CTkLabel(self.stats_frame, text="(No data yet)", font=("Arial", 10), justify="left", wraplength=900)
        self.stats_text.pack(anchor="w", padx=15, pady=(0, 5))

        # --- 2. Action Buttons ---
        self._create_action_buttons(self).pack(fill="x", padx=10, pady=5)
        
        self.export_btn = ctk.CTkButton(self, text="Download Professional Excel Report", command=self.export_professional_report, 
                                        state="disabled", fg_color="#107C10")
        self.export_btn.pack(pady=5)

        # --- 3. Tabs (Results & Logs) ---
        self.tab_view = ctk.CTkTabview(self)
        self.tab_view.pack(fill="both", expand=True, padx=10, pady=5)
        
        self.tab_view.add("Results")
        self._create_log_and_status_area(self.tab_view)

        # --- Results Table ---
        result_frame = self.tab_view.tab("Results")
        
        columns = ("sno", "panchayat", "village", "jobcard", "name", "abps_status", "ekyc_status")
        self.tree = ttk.Treeview(result_frame, columns=columns, show="headings", selectmode="browse")
        
        self.tree.heading("sno", text="S.No")
        self.tree.heading("panchayat", text="Panchayat")
        self.tree.heading("village", text="Village")
        self.tree.heading("jobcard", text="Job Card No")
        self.tree.heading("name", text="Applicant Name")
        self.tree.heading("abps_status", text="ABPS Enabled?")
        self.tree.heading("ekyc_status", text="eKYC Done?")
        
        self.tree.column("sno", width=50, anchor="center")
        self.tree.column("panchayat", width=130)
        self.tree.column("village", width=120)
        self.tree.column("jobcard", width=180)
        self.tree.column("name", width=200)
        self.tree.column("abps_status", width=100, anchor="center")
        self.tree.column("ekyc_status", width=100, anchor="center")

        self.style_treeview(self.tree)

        sb = ttk.Scrollbar(result_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

    def update_status(self, text, progress=None):
        self.status_label.configure(text=f"Status: {text}")
        if progress is not None: self.progress_bar.set(progress)
        try: self.app.set_status(f"eKYC Bot: {text}")
        except: pass
        self.update_idletasks()

    def start_automation(self):
        self.save_inputs()
        self.set_common_ui_state(running=True)
        self.export_btn.configure(state="disabled")
        
        self.all_scraped_data = []
        self.scraped_keys = set()
        for item in self.tree.get_children(): self.tree.delete(item)
        self.stats_text.configure(text="(Processing...)")

        threading.Thread(target=self.run_process, daemon=True).start()

    def _update_stats_display(self):
        """Scraped data se panchayat-wise stats calculate karke display karo"""
        if not self.all_scraped_data:
            self.stats_text.configure(text="(No data yet)")
            return
        
        from collections import defaultdict
        panchayat_stats = defaultdict(lambda: {"total": 0, "done": 0, "abps_pending": 0})
        
        for r in self.all_scraped_data:
            p = r.get("panchayat", "Unknown")
            panchayat_stats[p]["total"] += 1
            if "yes" in r["ekyc"].lower():
                panchayat_stats[p]["done"] += 1
            if "no" in r["abps"].lower():
                panchayat_stats[p]["abps_pending"] += 1
        
        grand_total = len(self.all_scraped_data)
        grand_done = sum(1 for r in self.all_scraped_data if "yes" in r["ekyc"].lower())
        grand_abps = sum(1 for r in self.all_scraped_data if "no" in r["abps"].lower())
        
        lines = []
        for p_name, s in sorted(panchayat_stats.items()):
            pending = s["total"] - s["done"]
            lines.append(
                f"📌 {p_name}:  Total={s['total']}  |  eKYC Done={s['done']}  |  Pending={pending}  |  ABPS Pending={s['abps_pending']}"
            )
        
        if len(panchayat_stats) > 1:
            lines.append(
                f"\n🔢 GRAND TOTAL:  Total={grand_total}  |  eKYC Done={grand_done}  |  Pending={grand_total - grand_done}  |  ABPS Pending={grand_abps}"
            )
        
        self.stats_text.configure(text="\n".join(lines))

    def run_process(self):
        try:
            panchayat_target = self.panchayat_entry.get().strip()
            village_target = self.village_entry.get().strip()

            self.tab_view.set("Logs & Status")
            driver = self.app.browser_manager.get_driver()
            wait = WebDriverWait(driver, 20)
            
            self.update_status("Opening Website...")
            driver.get(config.EKYC_REPORT_CONFIG["url"])

            # 1. Uncheck Pending (IMPROVED WAIT LOGIC)
            try:
                chk_locator = (By.ID, "ctl00_ContentPlaceHolder1_chbx_freshCase")
                chk = wait.until(EC.presence_of_element_located(chk_locator))
                is_checked = driver.execute_script("return arguments[0].checked;", chk)
                
                if is_checked:
                    self.update_status("Unchecking Pending Box...")
                    driver.execute_script("arguments[0].click();", chk)
                    try: wait.until(EC.staleness_of(chk))
                    except: time.sleep(2)
            except Exception as e:
                self.app.log_message(self.log_display, f"Warning in Uncheck Pending: {e}", "warning")

            # 2. Determine Panchayats to Process
            panchayats_to_process = []
            
            if panchayat_target:
                panchayats_to_process.append(panchayat_target)
                self.app.update_history("panchayat_name", panchayat_target)
            else:
                self.update_status("Fetching panchayat list...")
                try:
                    panchayat_dd_elem = wait.until(EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_DDL_panchayat")))
                    options = Select(panchayat_dd_elem).options
                    
                    for opt in options:
                        val = opt.get_attribute("value")
                        txt = opt.text.strip()
                        if val not in ["00", "99"] and txt != "---Select---" and "All" not in txt:
                            panchayats_to_process.append(txt)
                    
                    self.app.log_message(self.log_display, f"Found {len(panchayats_to_process)} panchayats to scan.", "info")
                except Exception as e:
                    raise Exception(f"Could not fetch panchayat list: {e}")

            # 3. Iterate over Panchayats
            total_panchayats = len(panchayats_to_process)
            
            for p_idx, p_name in enumerate(panchayats_to_process, 1):
                if self.app.stop_events[self.automation_key].is_set(): break
                
                self.update_status(f"Processing Panchayat {p_idx}/{total_panchayats}: {p_name}")
                self.app.log_message(self.log_display, f"\n{'='*50}\nSelecting Panchayat: {p_name}\n{'='*50}", "info")
                
                # Select Panchayat
                try:
                    old_html = driver.find_element(By.TAG_NAME, "html")
                    panchayat_elem = wait.until(EC.element_to_be_clickable((By.ID, "ctl00_ContentPlaceHolder1_DDL_panchayat")))
                    panchayat_dd = Select(panchayat_elem)
                    panchayat_dd.select_by_visible_text(p_name)
                    
                    try: wait.until(EC.staleness_of(old_html))
                    except: time.sleep(3)
                    
                except Exception as e:
                    self.app.log_message(self.log_display, f"Failed to select {p_name}: {e}", "error")
                    continue

                # 4. Determine Villages to Process for this Panchayat
                villages_to_process = []
                
                if village_target and ("All Village" in village_target or village_target == "99"):
                    village_target = ""

                if village_target:
                    villages_to_process.append(village_target)
                    self.app.update_history("village_name", village_target)
                else:
                    self.update_status(f"Fetching village list for {p_name}...")
                    try:
                        village_dd_elem = None
                        for _ in range(3):
                            try:
                                village_dd_elem = driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_DDL_Village")
                                break
                            except: time.sleep(1)
                        
                        if not village_dd_elem: raise Exception("Village Dropdown not found")

                        options = Select(village_dd_elem).options
                        for opt in options:
                            val = opt.get_attribute("value")
                            txt = opt.text.strip()
                            if val not in ["00", "99"] and txt != "---Select---" and txt != "--All Villages--":
                                villages_to_process.append(txt)
                        
                        self.app.log_message(self.log_display, f"Found {len(villages_to_process)} villages in {p_name}.", "info")
                    except Exception as e:
                        self.app.log_message(self.log_display, f"Could not fetch village list for {p_name}: {e}", "error")
                        continue

                # 5. Process Villages in this Panchayat
                total_villages = len(villages_to_process)
                
                for v_idx, v_name in enumerate(villages_to_process, 1):
                    if self.app.stop_events[self.automation_key].is_set(): break
                    
                    self.update_status(f"[{p_name}] Village {v_idx}/{total_villages}: {v_name}")
                    self.app.log_message(self.log_display, f"  Selecting Village: {v_name}", "info")
                    
                    # Selection Retry Logic
                    selection_success = False
                    for attempt in range(1, 4):
                        try:
                            old_html = driver.find_element(By.TAG_NAME, "html")
                            v_dd_elem = wait.until(EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_DDL_Village")))
                            v_dd = Select(v_dd_elem)
                            v_dd.select_by_visible_text(v_name)
                            
                            try: wait.until(EC.staleness_of(old_html))
                            except: time.sleep(2)
                            
                            selection_success = True
                            break 
                        except Exception as e:
                            self.app.log_message(self.log_display, f"    Retry {attempt} for {v_name}...", "warning")
                            time.sleep(2)
                    
                    if not selection_success:
                        self.app.log_message(self.log_display, f"    Skipping {v_name} (Selection Failed)", "error")
                        continue

                    self.scrape_current_table(driver, p_name, v_name)

            self.update_status("Completed")
            self._update_stats_display()
            self.export_btn.configure(state="normal")
            messagebox.showinfo("Success", f"Scan Complete.\nRecords Found: {len(self.all_scraped_data)}")

        except Exception as e:
            self.handle_error(e)
        finally:
            self.set_common_ui_state(running=False)
            self.update_status("Ready")

    def scrape_current_table(self, driver, panchayat_name, village_name):
        current_page_num = 1
        
        # Reset to Page 1
        try:
            page_one_link = driver.find_elements(By.XPATH, "//a[text()='1']")
            if page_one_link:
                self.app.log_message(self.log_display, f"Resetting to Page 1 for {village_name}...", "info")
                old_table = driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_gvData")
                driver.execute_script("arguments[0].click();", page_one_link[0])
                try: WebDriverWait(driver, 10).until(EC.staleness_of(old_table))
                except: time.sleep(2)
        except: pass

        while True:
            if self.app.stop_events[self.automation_key].is_set(): return

            if "No Record Found" in driver.page_source:
                self.app.log_message(self.log_display, f"No records in {village_name}.", "warning")
                break

            try:
                table = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_gvData")))
                rows = table.find_elements(By.TAG_NAME, "tr")
            except:
                self.app.log_message(self.log_display, "Table not found.", "error")
                break

            count_on_page = 0
            if len(rows) > 1:
                for row in rows[1:]:
                    cols = row.find_elements(By.TAG_NAME, "td")
                    if len(cols) < 5: continue 

                    try:
                        jc = cols[1].text.strip()
                        if "Job Card" in jc: continue 
                        
                        # --- FIX: GARBAGE/PAGINATION ROW FILTER ---
                        # "1", "2" jaise page numbers ko filter karne ke liye length check
                        if len(jc) < 5: continue 
                        # ------------------------------------------

                        name = cols[3].text.strip()
                        abps = cols[-2].text.strip()
                        ekyc = cols[-1].text.strip()
                        
                        # --- STRICT DUPLICATE CHECK ---
                        # Spaces hata kar aur lowercase karke check karenge
                        clean_jc = "".join(jc.split()).lower()
                        clean_name = "".join(name.split()).lower()
                        
                        unique_key = f"{panchayat_name}|{village_name}|{clean_jc}|{clean_name}"
                        
                        if unique_key in self.scraped_keys: continue
                        self.scraped_keys.add(unique_key)
                        # -----------------------

                        record = {
                            "panchayat": panchayat_name,
                            "village": village_name,
                            "jobcard": jc, "name": name, "abps": abps, "ekyc": ekyc
                        }
                        self.all_scraped_data.append(record)
                        self.check_and_insert_to_tree(record)
                        count_on_page += 1
                    except: continue

            self.app.log_message(self.log_display, f"  > Page {current_page_num}: {count_on_page} new records.", "info")

            next_page_num = current_page_num + 1
            try:
                next_link = driver.find_element(By.XPATH, f"//a[contains(@href, 'Page${next_page_num}')]")
                self.update_status(f"Loading {village_name} - Page {next_page_num}...")
                old_table = driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_gvData")
                
                driver.execute_script("arguments[0].click();", next_link)
                
                try: WebDriverWait(driver, 10).until(EC.staleness_of(old_table))
                except: time.sleep(3)
                current_page_num += 1
            except NoSuchElementException:
                break
            except Exception as e:
                self.app.log_message(self.log_display, f"Pagination error: {e}", "warning")
                break

    def _should_show_record(self, record):
        """Combined Logic for Filters + ABPS Checkbox"""
        filter_mode = self.filter_var.get()
        include_abps_pending = self.abps_check_var.get()
        
        ekyc_yes = "yes" in record['ekyc'].lower()
        abps_no = "no" in record['abps'].lower()
        
        # 1. Base Filter Logic
        show = False
        if filter_mode == "All": 
            show = True
        elif filter_mode == "Verified (Yes)" and ekyc_yes: 
            show = True
        elif filter_mode == "Not Verified (No)" and not ekyc_yes: 
            show = True
            
        # 2. Override Logic: Include ABPS Pending?
        if include_abps_pending and abps_no:
            show = True
            
        return show

    def check_and_insert_to_tree(self, record):
        if self._should_show_record(record):
            sno = len(self.tree.get_children()) + 1
            self.tree.insert("", "end", values=(sno, record['panchayat'], record['village'], record['jobcard'], record['name'], record['abps'], record['ekyc']))
            if sno % 10 == 0: self.tree.yview_moveto(1)

    def apply_filter_visuals(self, _=None):
        for item in self.tree.get_children(): self.tree.delete(item)
        for r in self.all_scraped_data: self.check_and_insert_to_tree(r)

    def export_professional_report(self):
        if not self.all_scraped_data: return

        # --- 1. Stats Calculation (per-Panchayat) ---
        from collections import defaultdict
        panchayat_stats = defaultdict(lambda: {"total": 0, "done": 0, "abps_pending": 0})
        
        for r in self.all_scraped_data:
            p = r.get("panchayat", "Unknown")
            panchayat_stats[p]["total"] += 1
            if "yes" in r["ekyc"].lower():
                panchayat_stats[p]["done"] += 1
            if "no" in r["abps"].lower():
                panchayat_stats[p]["abps_pending"] += 1
        
        all_data = self.all_scraped_data
        total = len(all_data)
        done = sum(1 for r in all_data if 'yes' in r['ekyc'].lower())
        pending = total - done
        abps_pending = sum(1 for r in all_data if 'no' in r['abps'].lower())

        # --- 2. Filter Rows for Table ---
        data_export = [r for r in all_data if self._should_show_record(r)]

        # File Setup
        panchayat = self.panchayat_entry.get()
        village_input = self.village_entry.get()
        
        if panchayat and village_input:
            file_part = f"{panchayat}_{village_input}"
            header_text = f"eKYC & ABPS REPORT: {village_input}, {panchayat.upper()}"
        elif panchayat:
            file_part = f"Panchayat - {panchayat}"
            header_text = f"eKYC & ABPS REPORT: Panchayat - {panchayat.upper()}"
        else:
            file_part = "All_Panchayats"
            header_text = "eKYC & ABPS REPORT: ALL PANCHAYATS"
        
        year = datetime.date.today().year
        date_str = datetime.date.today().strftime("%d-%m-%Y")
        
        user_downloads = self.app.get_user_downloads_path()
        save_dir = os.path.join(user_downloads, "NregaBot", f"Reports {year}", "eKYC Reports")
        if not os.path.exists(save_dir): os.makedirs(save_dir)
            
        default_name = f"ekyc_report_{file_part}_{date_str}.xlsx"
        filename = filedialog.asksaveasfilename(initialdir=save_dir, initialfile=default_name, defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])

        if not filename: return

        try:
            wb = openpyxl.Workbook()
            
            # --- SHEET 1: DETAILED DATA ---
            ws = wb.active
            ws.title = "Detailed Report"

            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
            white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            center = Alignment(horizontal="center", vertical="center")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # Header
            ws.merge_cells('A1:G1')
            ws['A1'] = header_text
            ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
            ws['A1'].fill = header_fill
            ws['A1'].alignment = center

            ws.merge_cells('A2:G2')
            ws['A2'] = f"Report Generated from NregaBot.com | Date: {datetime.datetime.now().strftime('%d-%m-%Y %I:%M %p')}"
            ws['A2'].font = Font(italic=True, size=9)
            ws['A2'].alignment = center

            # Grand Summary (Row 4 & 5)
            headers = ["Total Laborers", "eKYC Done", "eKYC Pending", "ABPS Pending (No)"]
            vals = [total, done, pending, abps_pending]
            
            for i, (h, v) in enumerate(zip(headers, vals), start=2):
                c_h = ws.cell(row=4, column=i, value=h)
                c_h.font = Font(bold=True)
                c_h.fill = PatternFill(start_color="DCE6F1", fill_type="solid")
                c_h.alignment = center
                c_h.border = border
                
                c_v = ws.cell(row=5, column=i, value=v)
                c_v.font = Font(bold=True, size=11)
                c_v.alignment = center
                c_v.border = border
                if i == 4 and v > 0: c_v.font = Font(color="FF0000", bold=True)

            # Data Table
            t_row = 7
            cols = ["S.No", "Panchayat", "Village", "Job Card No", "Applicant Name", "ABPS Enabled?", "eKYC Done?"]
            for i, h in enumerate(cols, 1):
                c = ws.cell(row=t_row, column=i, value=h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = center
                c.border = border

            for idx, r in enumerate(data_export, 1):
                r_idx = t_row + idx
                fill = gray_fill if idx % 2 == 0 else white_fill
                
                ws.cell(row=r_idx, column=1, value=idx).fill = fill
                ws.cell(row=r_idx, column=1).border = border
                ws.cell(row=r_idx, column=1).alignment = center
                
                ws.cell(row=r_idx, column=2, value=r['panchayat']).fill = fill
                ws.cell(row=r_idx, column=2).border = border
                
                ws.cell(row=r_idx, column=3, value=r['village']).fill = fill
                ws.cell(row=r_idx, column=3).border = border
                
                ws.cell(row=r_idx, column=4, value=r['jobcard']).fill = fill
                ws.cell(row=r_idx, column=4).border = border
                
                ws.cell(row=r_idx, column=5, value=r['name']).fill = fill
                ws.cell(row=r_idx, column=5).border = border
                
                c6 = ws.cell(row=r_idx, column=6, value=r['abps'])
                c6.alignment = center; c6.fill = fill; c6.border = border
                if "no" in r['abps'].lower(): c6.font = Font(color="FF0000", bold=True)
                else: c6.font = Font(color="006100", bold=True)
                
                c7 = ws.cell(row=r_idx, column=7, value=r['ekyc'])
                c7.alignment = center; c7.fill = fill; c7.border = border
                if "no" in r['ekyc'].lower(): c7.font = Font(color="FF0000", bold=True)
                else: c7.font = Font(color="006100", bold=True)

            # Widths
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 22
            ws.column_dimensions['E'].width = 30
            ws.column_dimensions['F'].width = 16
            ws.column_dimensions['G'].width = 13

            # --- SHEET 2: PANCHAYAT-WISE SUMMARY ---
            if len(panchayat_stats) > 1:
                ws2 = wb.create_sheet("Panchayat Summary")
                
                ws2.merge_cells('A1:E1')
                ws2['A1'] = "📊 PANCHAYAT-WISE SUMMARY"
                ws2['A1'].font = Font(size=13, bold=True, color="FFFFFF")
                ws2['A1'].fill = header_fill
                ws2['A1'].alignment = center
                
                # Header Row
                summary_cols = ["Panchayat", "Total", "eKYC Done", "eKYC Pending", "ABPS Pending"]
                for i, h in enumerate(summary_cols, 1):
                    c = ws2.cell(row=3, column=i, value=h)
                    c.font = header_font
                    c.fill = header_fill
                    c.alignment = center
                    c.border = border
                
                row_idx = 4
                for p_name, s in sorted(panchayat_stats.items()):
                    p_pending = s["total"] - s["done"]
                    fill = gray_fill if row_idx % 2 == 0 else white_fill
                    
                    ws2.cell(row=row_idx, column=1, value=p_name).fill = fill
                    ws2.cell(row=row_idx, column=1).border = border
                    
                    ws2.cell(row=row_idx, column=2, value=s["total"]).fill = fill
                    ws2.cell(row=row_idx, column=2).border = border
                    ws2.cell(row=row_idx, column=2).alignment = center
                    
                    ws2.cell(row=row_idx, column=3, value=s["done"]).fill = fill
                    ws2.cell(row=row_idx, column=3).border = border
                    ws2.cell(row=row_idx, column=3).alignment = center
                    ws2.cell(row=row_idx, column=3).font = Font(color="006100", bold=True)
                    
                    ws2.cell(row=row_idx, column=4, value=p_pending).fill = fill
                    ws2.cell(row=row_idx, column=4).border = border
                    ws2.cell(row=row_idx, column=4).alignment = center
                    if p_pending > 0:
                        ws2.cell(row=row_idx, column=4).font = Font(color="FF0000", bold=True)
                    
                    ws2.cell(row=row_idx, column=5, value=s["abps_pending"]).fill = fill
                    ws2.cell(row=row_idx, column=5).border = border
                    ws2.cell(row=row_idx, column=5).alignment = center
                    if s["abps_pending"] > 0:
                        ws2.cell(row=row_idx, column=5).font = Font(color="FF6600", bold=True)
                    
                    row_idx += 1
                
                # Grand Total Row
                ws2.cell(row=row_idx, column=1, value="GRAND TOTAL").fill = PatternFill(start_color="FFC000", fill_type="solid")
                ws2.cell(row=row_idx, column=1).border = border
                ws2.cell(row=row_idx, column=1).font = Font(bold=True)
                
                ws2.cell(row=row_idx, column=2, value=total).fill = PatternFill(start_color="FFC000", fill_type="solid")
                ws2.cell(row=row_idx, column=2).border = border
                ws2.cell(row=row_idx, column=2).alignment = center
                ws2.cell(row=row_idx, column=2).font = Font(bold=True)
                
                ws2.cell(row=row_idx, column=3, value=done).fill = PatternFill(start_color="FFC000", fill_type="solid")
                ws2.cell(row=row_idx, column=3).border = border
                ws2.cell(row=row_idx, column=3).alignment = center
                ws2.cell(row=row_idx, column=3).font = Font(bold=True, color="006100")
                
                ws2.cell(row=row_idx, column=4, value=pending).fill = PatternFill(start_color="FFC000", fill_type="solid")
                ws2.cell(row=row_idx, column=4).border = border
                ws2.cell(row=row_idx, column=4).alignment = center
                ws2.cell(row=row_idx, column=4).font = Font(bold=True, color="FF0000" if pending > 0 else "000000")
                
                ws2.cell(row=row_idx, column=5, value=abps_pending).fill = PatternFill(start_color="FFC000", fill_type="solid")
                ws2.cell(row=row_idx, column=5).border = border
                ws2.cell(row=row_idx, column=5).alignment = center
                ws2.cell(row=row_idx, column=5).font = Font(bold=True, color="FF6600" if abps_pending > 0 else "000000")
                
                # Widths
                ws2.column_dimensions['A'].width = 25
                ws2.column_dimensions['B'].width = 12
                ws2.column_dimensions['C'].width = 12
                ws2.column_dimensions['D'].width = 14
                ws2.column_dimensions['E'].width = 14

            wb.save(filename)
            messagebox.showinfo("Success", f"File saved!\n{filename}")
            
            try:
                if os.name == 'nt': os.startfile(filename)
                else: subprocess.call(['open', filename])
            except: pass

        except Exception as e:
            messagebox.showerror("Error", f"Save Failed: {e}")

    def save_inputs(self):
        data = {
            "panchayat": self.panchayat_entry.get().strip(),
            "village": self.village_entry.get().strip(),
            "filter": self.filter_var.get(),
            "abps_check": self.abps_check_var.get()
        }
        try:
            config_file = self.app.get_data_path("ekyc_inputs.json")
            with open(config_file, "w") as f: json.dump(data, f, indent=4)
        except: pass

    def load_inputs(self):
        config_file = self.app.get_data_path("ekyc_inputs.json")
        if os.path.exists(config_file):
            try:
                with open(config_file, "r") as f: data = json.load(f)
                self.panchayat_entry.delete(0, "end"); self.panchayat_entry.insert(0, data.get("panchayat", ""))
                self.village_entry.delete(0, "end"); self.village_entry.insert(0, data.get("village", ""))
                if data.get("filter"): self.filter_var.set(data.get("filter"))
                if data.get("abps_check"): self.abps_check_var.set(data.get("abps_check"))
            except: pass