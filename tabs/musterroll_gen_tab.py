# tabs/musterroll_gen_tab.py
import tkinter
from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk
import os, json, time, base64, sys, subprocess, requests, re, threading
from datetime import datetime
from pypdf import PdfWriter 
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, 
    NoSuchElementException, 
    StaleElementReferenceException, 
    UnexpectedAlertPresentException
)
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

import config
from .base_tab import BaseAutomationTab
from .autocomplete_widget import AutocompleteEntry

class MusterrollGenTab(BaseAutomationTab):
    def __init__(self, parent, app_instance):
        super().__init__(parent, app_instance, automation_key="muster")
        self.config_file = self.app.get_data_path("muster_roll_inputs.json")
        
        self.mapping_file = self.app.get_data_path("mr_panchayat_staff_map.json")
        self.mapping_data = {} 
        
        self.success_count = 0
        self.skipped_count = 0
        self.output_dir = ""
        self.current_session_files = []
        
        # --- NEW: Data collection for eKYC ---
        self.collected_mr_data = [] 
        
        self.panchayat_after_id = None 
        self._load_mapping_data() 
        
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)
        self._create_widgets()
        self.load_inputs()

    def _create_widgets(self):
        # This frame holds all the user input fields
        controls_frame = ctk.CTkFrame(self)
        controls_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=(10,0))
        controls_frame.grid_columnconfigure((1,3), weight=1)
        
        ctk.CTkLabel(controls_frame, text="Panchayat Name:").grid(row=0, column=0, sticky='w', padx=15, pady=(15,0))
        self.panchayat_entry = AutocompleteEntry(controls_frame, suggestions_list=self.app.history_manager.get_suggestions("panchayat_name"),
            app_instance=self.app,
            history_key="panchayat_name")
        self.panchayat_entry.grid(row=0, column=1, columnspan=3, sticky='ew', padx=15, pady=(15,0))
        self.panchayat_entry.bind("<KeyRelease>", self._on_panchayat_change_debounced, add="+")
        
        ctk.CTkLabel(controls_frame, text="Note: Must exactly match the name on the VB-G-RAM-G portal.", text_color="gray50").grid(row=1, column=1, columnspan=3, sticky='w', padx=15, pady=(0,10))
        
        # --- Start Date ---
        ctk.CTkLabel(controls_frame, text="तारीख से (DD/MM/YYYY):").grid(row=2, column=0, sticky='w', padx=15, pady=5)
        start_date_frame = ctk.CTkFrame(controls_frame, fg_color="transparent")
        start_date_frame.grid(row=2, column=1, sticky='ew', padx=(15,5), pady=5)
        self.start_date_entry = ctk.CTkEntry(start_date_frame, placeholder_text="DD/MM/YYYY")
        self.start_date_entry.pack(side="left", fill="x", expand=True)
        ctk.CTkButton(start_date_frame, text="📅", width=30, fg_color=("gray85", "gray25"), text_color=("black", "white"),
                    command=lambda: self.open_date_picker(lambda d: [self.start_date_entry.delete(0, "end"), self.start_date_entry.insert(0, d)])).pack(side="right", padx=(5,0))

        # --- End Date ---
        ctk.CTkLabel(controls_frame, text="तारीख को (DD/MM/YYYY):").grid(row=2, column=2, sticky='w', padx=10, pady=5)
        end_date_frame = ctk.CTkFrame(controls_frame, fg_color="transparent")
        end_date_frame.grid(row=2, column=3, sticky='ew', padx=(5,15), pady=5)
        self.end_date_entry = ctk.CTkEntry(end_date_frame, placeholder_text="DD/MM/YYYY")
        self.end_date_entry.pack(side="left", fill="x", expand=True)
        ctk.CTkButton(end_date_frame, text="📅", width=30, fg_color=("gray85", "gray25"), text_color=("black", "white"),
                    command=lambda: self.open_date_picker(lambda d: [self.end_date_entry.delete(0, "end"), self.end_date_entry.insert(0, d)])).pack(side="right", padx=(5,0))
        
        ctk.CTkLabel(controls_frame, text="Select Designation:").grid(row=3, column=0, sticky='w', padx=15, pady=5)
        designation_options = ["Junior Engineer--BP", "Assistant Engineer--BP", "Technical Assistant--BP", "Acrited Engineer(AE)--GP", "Junior Engineer--GP", "Technical Assistant--GP"]
        self.designation_combobox = ctk.CTkComboBox(controls_frame, values=designation_options)
        self.designation_combobox.grid(row=3, column=1, sticky='ew', padx=(15,5), pady=5)
        
        ctk.CTkLabel(controls_frame, text="Select Technical Staff:").grid(row=3, column=2, sticky='w', padx=10, pady=5)
        self.staff_entry = AutocompleteEntry(controls_frame, suggestions_list=self.app.history_manager.get_suggestions("staff_name"))
        self.staff_entry.grid(row=3, column=3, sticky='ew', padx=(5,15), pady=5)
        
        ctk.CTkLabel(controls_frame, text="Output Action:").grid(row=4, column=0, sticky='w', padx=15, pady=5)
        self.output_action_combobox = ctk.CTkComboBox(controls_frame, values=["Save as PDF", "Print"])
        self.output_action_combobox.set("Save as PDF")
        self.output_action_combobox.grid(row=4, column=1, sticky='ew', padx=(15,5), pady=5)
        
        self.save_to_cloud_var = tkinter.BooleanVar(value=True) 
        self.save_to_cloud_checkbox = ctk.CTkCheckBox(
            controls_frame, 
            text="Save generated PDF to Cloud", 
            variable=self.save_to_cloud_var
        )
        self.save_to_cloud_checkbox.grid(row=4, column=2, columnspan=2, sticky='w', padx=15, pady=5)

        ctk.CTkLabel(controls_frame, text="Orientation:").grid(row=5, column=0, sticky='w', padx=15, pady=5)
        self.orientation_var = ctk.StringVar(value="Landscape")
        self.orientation_segmented_button = ctk.CTkSegmentedButton(controls_frame, variable=self.orientation_var, values=["Landscape", "Portrait"])
        self.orientation_segmented_button.grid(row=5, column=1, sticky='ew', padx=(15,5), pady=5)

        ctk.CTkLabel(controls_frame, text="PDF Scale:").grid(row=5, column=2, sticky='w', padx=10, pady=5)
        scale_frame = ctk.CTkFrame(controls_frame, fg_color="transparent")
        scale_frame.grid(row=5, column=3, sticky="ew", padx=(5,15), pady=5)
        scale_frame.grid_columnconfigure(0, weight=1)
        self.scale_slider = ctk.CTkSlider(scale_frame, from_=50, to=100, number_of_steps=50, command=self._update_scale_label)
        self.scale_slider.set(75)
        self.scale_slider.grid(row=0, column=0, sticky="ew")
        self.scale_label = ctk.CTkLabel(scale_frame, text="75%", width=40)
        self.scale_label.grid(row=0, column=1, padx=(10, 0))
        
        # --- NEW: eKYC Checkbox ---
        self.run_ekyc_var = ctk.BooleanVar(value=False)
        self.run_ekyc_checkbox = ctk.CTkCheckBox(
            controls_frame, 
            text="Run eKYC Verification after Completion", 
            variable=self.run_ekyc_var,
            text_color="orange",
            font=("Arial", 11, "bold")
        )
        self.run_ekyc_checkbox.grid(row=6, column=0, columnspan=2, sticky='w', padx=15, pady=10)

        ctk.CTkLabel(controls_frame, text="ℹ️ Generated MRs are saved in 'Downloads/NregaBot/MR_Output'.", text_color="gray50").grid(row=6, column=2, columnspan=2, sticky='e', padx=15, pady=(10,15))
        
        action_frame_container = ctk.CTkFrame(self)
        action_frame_container.grid(row=1, column=0, sticky="ew", padx=10, pady=10)
        action_frame = self._create_action_buttons(parent_frame=action_frame_container)
        action_frame.pack(expand=True, fill='x')

        data_notebook = ctk.CTkTabview(self)
        data_notebook.grid(row=2, column=0, sticky="nsew", padx=10, pady=(0,10))
        work_codes_tab = data_notebook.add("Work Search Keys (or auto)")
        results_tab = data_notebook.add("Results")
        self._create_log_and_status_area(parent_notebook=data_notebook)
        
        work_codes_tab.grid_columnconfigure(0, weight=1)
        work_codes_tab.grid_rowconfigure(1, weight=1)
        wc_controls = ctk.CTkFrame(work_codes_tab, fg_color="transparent")
        wc_controls.grid(row=0, column=0, sticky='ew')
        clear_button = ctk.CTkButton(wc_controls, text="Clear", width=80, command=lambda: self.work_codes_text.delete("1.0", tkinter.END))
        clear_button.pack(side='right', pady=(5,0), padx=(0,5))
        
        extract_button = ctk.CTkButton(wc_controls, text="Extract from Text", width=120,
                                       command=lambda: self._extract_and_update_workcodes(self.work_codes_text))
        extract_button.pack(side='right', pady=(5,0), padx=(0, 5))
        
        self.work_codes_text = ctk.CTkTextbox(work_codes_tab, height=100)
        self.work_codes_text.grid(row=1, column=0, sticky='nsew', padx=5, pady=5)
        
        results_tab.grid_columnconfigure(0, weight=1); results_tab.grid_rowconfigure(2, weight=1)
        
        results_action_frame = ctk.CTkFrame(results_tab, fg_color="transparent")
        results_action_frame.grid(row=0, column=0, sticky="ew", pady=(5,10), padx=5)
        
        self.merge_pdfs_button = ctk.CTkButton(results_action_frame, text="Merge Saved PDFs", command=self.merge_saved_pdfs)
        self.merge_pdfs_button.pack(side='left', padx=(0, 10))

        export_controls_frame = ctk.CTkFrame(results_action_frame, fg_color="transparent")
        export_controls_frame.pack(side='right', padx=(10, 0))
        self.export_button = ctk.CTkButton(export_controls_frame, text="Export Report", command=self.export_report)
        self.export_button.pack(side='left')
        self.export_format_menu = ctk.CTkOptionMenu(export_controls_frame, width=130, values=["PDF (.pdf)", "CSV (.csv)"], command=self._on_format_change)
        self.export_format_menu.pack(side='left', padx=5)
        self.export_filter_menu = ctk.CTkOptionMenu(export_controls_frame, width=150, values=["Export All", "Success Only", "Failed Only"])
        self.export_filter_menu.pack(side='left', padx=(0, 5))
        
        summary_frame = ctk.CTkFrame(results_tab, fg_color="transparent")
        summary_frame.grid(row=1, column=0, sticky="ew", pady=(0, 10))
        summary_frame.grid_columnconfigure((0, 1), weight=1)
        self.success_label = ctk.CTkLabel(summary_frame, text="Success: 0", text_color="#2E8B57", font=ctk.CTkFont(weight="bold")); self.success_label.grid(row=0, column=0, sticky='w')
        self.skipped_label = ctk.CTkLabel(summary_frame, text="Skipped/Failed: 0", text_color="#DAA520", font=ctk.CTkFont(weight="bold")); self.skipped_label.grid(row=0, column=1, sticky='w')
        
        cols = ("Timestamp", "Work Code/Key", "Status", "Details"); self.results_tree = ttk.Treeview(results_tab, columns=cols, show='headings')
        for col in cols: self.results_tree.heading(col, text=col)
        self.results_tree.column("Timestamp", width=80, anchor='center'); self.results_tree.column("Work Code/Key", width=250); self.results_tree.column("Status", width=100, anchor='center'); self.results_tree.column("Details", width=400)
        self.results_tree.grid(row=2, column=0, sticky='nsew')
        scrollbar = ctk.CTkScrollbar(results_tab, command=self.results_tree.yview); self.results_tree.configure(yscroll=scrollbar.set); scrollbar.grid(row=2, column=1, sticky='ns')
        self.style_treeview(self.results_tree)
        self._setup_treeview_sorting(self.results_tree)

    def _on_format_change(self, selected_format):
        if "CSV" in selected_format: self.export_filter_menu.configure(state="disabled")
        else: self.export_filter_menu.configure(state="normal")

    def _update_scale_label(self, value):
        self.scale_label.configure(text=f"{int(value)}%")

    def set_ui_state(self, running: bool):
        self.set_common_ui_state(running)
        state = "disabled" if running else "normal"
        self.panchayat_entry.configure(state=state)
        self.start_date_entry.configure(state=state)
        self.end_date_entry.configure(state=state)
        self.staff_entry.configure(state=state)
        self.designation_combobox.configure(state=state)
        self.orientation_segmented_button.configure(state=state)
        self.scale_slider.configure(state=state)
        self.output_action_combobox.configure(state=state)
        self.work_codes_text.configure(state=state)
        self.save_to_cloud_checkbox.configure(state=state)
        self.run_ekyc_checkbox.configure(state=state) 
        
        self.export_button.configure(state=state)
        self.export_format_menu.configure(state=state)
        self.export_filter_menu.configure(state=state)
        self.merge_pdfs_button.configure(state=state)
        if state == "normal": self._on_format_change(self.export_format_menu.get())

    def _load_mapping_data(self):
        if os.path.exists(self.mapping_file):
            try:
                with open(self.mapping_file, 'r') as f:
                    self.mapping_data = json.load(f)
            except Exception:
                self.mapping_data = {}

    def _save_mapping_pair(self, panchayat, staff):
        if not panchayat or not staff: return
        key = panchayat.strip().lower()
        self.mapping_data[key] = staff.strip()
        try:
            with open(self.mapping_file, 'w') as f:
                json.dump(self.mapping_data, f, indent=4)
        except Exception as e:
            print(f"Error saving mapping: {e}")

    def _on_panchayat_change_debounced(self, event=None):
        if self.panchayat_after_id:
            self.after_cancel(self.panchayat_after_id)
        if event and event.keysym in ("Up", "Down", "Return", "Enter", "Tab"): return
        self.panchayat_after_id = self.after(300, self._auto_fill_staff)

    def _auto_fill_staff(self):
        current_panchayat = self.panchayat_entry.get().strip().lower()
        if current_panchayat in self.mapping_data:
            saved_staff = self.mapping_data[current_panchayat]
            if self.staff_entry.get().strip() != saved_staff:
                self.staff_entry.delete(0, tkinter.END)
                self.staff_entry.insert(0, saved_staff)
        
    def start_automation(self):
        for item in self.results_tree.get_children(): self.results_tree.delete(item)
        self.success_count, self.skipped_count = 0, 0
        self.current_session_files = [] 
        self.collected_mr_data = [] 
        self.success_label.configure(text="Success: 0")
        self.skipped_label.configure(text="Skipped/Failed: 0")
        
        inputs = {
            'panchayat': self.panchayat_entry.get().strip(), 
            'start_date': self.start_date_entry.get().strip(), 
            'end_date': self.end_date_entry.get().strip(), 
            'designation': self.designation_combobox.get().strip(), 
            'staff': self.staff_entry.get().strip(), 
            'orientation': self.orientation_var.get(),
            'scale': self.scale_slider.get(),
            'output_action': self.output_action_combobox.get(), 
            'work_codes_raw': self.work_codes_text.get("1.0", tkinter.END).strip(),
            'save_to_cloud': self.save_to_cloud_var.get(),
            'run_ekyc': self.run_ekyc_var.get()
        }

        if not all(inputs[k] for k in ['panchayat', 'start_date', 'end_date', 'designation', 'staff']):
            messagebox.showwarning("Input Error", "All fields are required (except Work Search Keys).")
            return
        self._save_mapping_pair(inputs['panchayat'], inputs['staff'])
        inputs['work_codes'] = [line.strip() for line in inputs['work_codes_raw'].split('\n') if line.strip()]
        inputs['auto_mode'] = not bool(inputs['work_codes'])
        self.save_inputs(inputs)
        self.app.start_automation_thread(self.automation_key, self.run_automation_logic, args=(inputs,))

    def retry_logic_handler(self):
        failed_items = []
        for item_id in self.results_tree.get_children():
            values = self.results_tree.item(item_id)['values']
            work_code = str(values[1])
            status = str(values[2]).lower()
            if "success" not in status:
                failed_items.append(work_code)
        
        if not failed_items:
            messagebox.showinfo("Retry", "No failed items found to retry.")
            return

        if not messagebox.askyesno("Retry Failed", f"Found {len(failed_items)} failed/skipped items.\nLoad them and retry?"):
            return

        self.work_codes_text.configure(state="normal")
        self.work_codes_text.delete("1.0", tkinter.END)
        self.work_codes_text.insert("1.0", "\n".join(failed_items))
        self.work_codes_text.configure(state="disabled")

        for item in self.results_tree.get_children(): 
            self.results_tree.delete(item)
            
        self.success_count = 0
        self.skipped_count = 0
        self.update_status("Retrying failed items...", 0.0)
        self.start_automation()
        
    def reset_ui(self):
        if messagebox.askokcancel("Reset Form?", "Clear all inputs and logs?"):
            self.panchayat_entry.delete(0, tkinter.END)
            self.start_date_entry.clear(); self.end_date_entry.clear()
            self.staff_entry.delete(0, tkinter.END)
            self.designation_combobox.set('')
            self.orientation_var.set('Landscape')
            self.scale_slider.set(75); self.scale_label.configure(text="75%")
            self.output_action_combobox.set('Save as PDF')
            self.work_codes_text.delete('1.0', tkinter.END)
            for item in self.results_tree.get_children(): self.results_tree.delete(item)
            self.app.clear_log(self.log_display)
            self.update_status("Ready", 0.0)
            self.success_label.configure(text="Success: 0"); self.skipped_label.configure(text="Skipped/Failed: 0")
            self.app.log_message(self.log_display, "Form has been reset.")
            self.app.after(0, self.app.set_status, "Ready")
            
    def save_inputs(self, inputs):
        try:
            inputs_to_save = inputs.copy()
            inputs_to_save.pop('work_codes_raw', None)
            inputs_to_save.pop('work_codes', None)
            inputs_to_save.pop('auto_mode', None)
            with open(self.config_file, 'w') as f:
                json.dump(inputs_to_save, f, indent=4)
        except Exception as e: print(f"Error saving inputs: {e}")
        
    def load_inputs(self):
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r') as f: data = json.load(f)
                self.panchayat_entry.insert(0, data.get('panchayat', ''))
                self.start_date_entry.delete(0, "end"); self.start_date_entry.insert(0, data.get('start_date', ''))
                self.end_date_entry.delete(0, "end"); self.end_date_entry.insert(0, data.get('end_date', ''))
                self.designation_combobox.set(data.get('designation', ''))
                self.staff_entry.insert(0, data.get('staff', ''))
                self.orientation_var.set(data.get('orientation', 'Landscape'))
                self.scale_slider.set(data.get('scale', 75)); self._update_scale_label(self.scale_slider.get())
                self.output_action_combobox.set(data.get('output_action', 'Save as PDF'))
                self.save_to_cloud_var.set(data.get('save_to_cloud', True))
                if data.get("run_ekyc"): self.run_ekyc_var.set(data.get("run_ekyc"))
        except Exception as e: print(f"Error loading inputs: {e}")

    def _print_file(self, file_path):
        try:
            if not os.path.exists(file_path):
                self.app.log_message(self.log_display, f"Print Error: File not found at {file_path}", "error")
                return
            if sys.platform == "win32": os.startfile(file_path, "print")
            else: subprocess.run(["lpr", file_path], check=True)
            self.app.log_message(self.log_display, f"Sent {os.path.basename(file_path)} to printer.")
            time.sleep(2)
        except Exception as e:
            error_msg = f"An unexpected error occurred while printing: {e}"
            self.app.log_message(self.log_display, error_msg, "error")
            self.app.after(0, lambda: messagebox.showwarning("Print Error", error_msg))

    def _get_output_dir(self, panchayat_name):
        try:
            safe_panchayat_name = "".join(c for c in panchayat_name if c.isalnum() or c in (' ', '_')).rstrip()
            if not safe_panchayat_name: safe_panchayat_name = "Unknown_Panchayat"
            date_str = datetime.now().strftime('%Y-%m-%d')
            output_dir = os.path.join(self.app.get_user_downloads_path(), "NregaBot", "MR_Output", safe_panchayat_name, date_str)
            os.makedirs(output_dir, exist_ok=True)
            return output_dir
        except Exception as e:
            self.app.log_message(self.log_display, f"Error creating output directory: {e}", "error")
            messagebox.showerror("Directory Error", f"Could not create output directory: {e}")
            return None

    def run_automation_logic(self, inputs):
        self.app.after(0, self.set_ui_state, True)
        self.app.clear_log(self.log_display)
        self.app.log_message(self.log_display, f"Starting MR generation for: {inputs['panchayat']}")
        self.app.after(0, self.app.set_status, "Running MR Generation...")
        
        self.output_dir = self._get_output_dir(inputs['panchayat'])
        if not self.output_dir:
            self.app.log_message(self.log_display, "Failed to create output directory. Aborting.", "error")
            self.app.after(0, self.set_ui_state, False)
            return
            
        try:
            driver = self.app.get_driver()
            if not driver: 
                self.app.after(0, self.set_ui_state, False)
                return
            wait = WebDriverWait(driver, 20)
            
            self.app.log_message(self.log_display, f"Output will be in: {self.output_dir}", "info")
            
            if not self._validate_panchayat(driver, wait, inputs['panchayat']):
                self.app.after(0, self.set_ui_state, False)
                return
            
            self.app.update_history("panchayat_name", inputs['panchayat'])
            self.app.update_history("staff_name", inputs['staff'])

            items_to_process = self._get_items_to_process(driver, wait, inputs)
            session_skip_list = set()
            total_items = len(items_to_process)

            for index, item in enumerate(items_to_process):
                if self.app.stop_events[self.automation_key].is_set(): 
                    self.app.log_message(self.log_display, "Stop signal received.", "warning")
                    break
                self.app.log_message(self.log_display, f"\n--- Processing item ({index+1}/{total_items}): {item} ---", "info")
                self.app.after(0, self.update_status, f"Processing {item}", (index+1)/total_items)
                
                self._process_single_item(driver, wait, inputs, item, self.output_dir, session_skip_list)
            
            # --- START eKYC LOGIC HERE ---
            if inputs.get('run_ekyc') and not self.app.stop_events[self.automation_key].is_set():
                if self.collected_mr_data:
                    self.app.log_message(self.log_display, f"\n--- Starting eKYC Verification ({len(self.collected_mr_data)} workers collected) ---", "info")
                    self.run_post_mr_ekyc_check(driver, wait, inputs['panchayat'])
                else:
                    self.app.log_message(self.log_display, "Warning: No worker data was collected during MR generation. Skipping eKYC.", "warning")
            # -----------------------------

        except Exception as e:
            self.app.log_message(self.log_display, f"A critical error occurred: {e}", "error")
            if "in str" not in str(e): 
                messagebox.showerror("Critical Error", f"An unexpected error stopped the automation.\n\nError: {e}")
        
        finally:
            self.app.after(0, self.set_ui_state, False)
            self.app.after(0, self.update_status, "Automation Finished.", 1.0)
            self.app.after(100, self._show_completion_dialog, self.output_dir)
            self.app.after(0, self.app.set_status, "Automation Finished")

    def _show_completion_dialog(self, output_dir):
        summary = f"Automation complete.\nSuccess: {self.success_count}\nSkipped/Failed: {self.skipped_count}"
        if "macro" in self.app.active_automations:
            self.app.log_message(self.log_display, f"Batch Finished. Output saved to: {output_dir}", "info")
            return

        if self.success_count > 0 and output_dir and os.path.exists(output_dir):
            if messagebox.askyesno("Task Finished", f"{summary}\n\nDo you want to open the output folder?"):
                self.app.open_folder(output_dir)
        else:
            messagebox.showinfo("Task Finished", summary)

    def _validate_panchayat(self, driver, wait, panchayat_name):
        try:
            self.app.log_message(self.log_display, "Validating Panchayat name...")
            driver.get(config.MUSTER_ROLL_CONFIG["base_url"])
            panchayat_dropdown = Select(wait.until(EC.presence_of_element_located((By.ID, "exe_agency"))))
            target_panchayat = config.AGENCY_PREFIX + panchayat_name
            if target_panchayat not in [opt.text for opt in panchayat_dropdown.options]:
                error_msg = f"Panchayat name '{panchayat_name}' not found on the website. Please check spelling."
                if "macro" in self.app.active_automations:
                    self.app.log_message(self.log_display, f"Skipping: {error_msg}", "error")
                    return False
                messagebox.showerror("Validation Error", error_msg)
                return False
            self.app.log_message(self.log_display, "Panchayat name is valid.", "success")
            return True
        except Exception as e:
            self.app.log_message(self.log_display, f"Validation failed: Error: {e}", "error")
            return False

    def _get_items_to_process(self, driver, wait, inputs):
        if inputs['auto_mode']:
            self.app.log_message(self.log_display, "Auto Mode: Fetching available work codes...")
            try:
                Select(driver.find_element(By.ID, "exe_agency")).select_by_visible_text(config.AGENCY_PREFIX + inputs['panchayat'])
                wait.until(lambda d: len(Select(d.find_element(By.ID, "ddlWorkCode")).options) > 1)
                items = [opt.text for opt in Select(driver.find_element(By.ID, "ddlWorkCode")).options if opt.get_attribute("value")]
                self.app.log_message(self.log_display, f"Found {len(items)} available work codes.")
                return items
            except Exception as e:
                self.app.log_message(self.log_display, f"Could not fetch work codes automatically. Error: {e}", "error")
                return []
        else:
            self.app.log_message(self.log_display, f"Processing {len(inputs['work_codes'])} provided work keys.")
            return inputs['work_codes']

    def _process_single_item(self, driver, wait, inputs, item, output_dir, session_skip_list):
        full_work_code_text = ""
        try:
            self.app.log_message(self.log_display, "   - Navigating to MR page...")
            driver.get(config.MUSTER_ROLL_CONFIG["base_url"])
            
            self.app.log_message(self.log_display, "   - Selecting Panchayat...")
            panchayat_dropdown = wait.until(EC.presence_of_element_located((By.ID, "exe_agency")))
            Select(panchayat_dropdown).select_by_visible_text(config.AGENCY_PREFIX + inputs['panchayat'])
            
            self.app.log_message(self.log_display, f"   - Selecting work code for '{item}'...")
            full_work_code_text = self._select_work_code(driver, wait, item, inputs['auto_mode'])
            
            if full_work_code_text in session_skip_list:
                self._log_result(item, "Skipped", "Already processed in this session.")
                return

            self.app.log_message(self.log_display, "   - Entering dates and staff details...")
            driver.execute_script(f"document.getElementById('txtDateFrom').value = '{inputs['start_date']}';")
            driver.execute_script(f"document.getElementById('txtDateTo').value = '{inputs['end_date']}';")
            
            designation_dropdown = wait.until(EC.presence_of_element_located((By.ID, "ddldesg")))
            Select(designation_dropdown).select_by_visible_text(inputs['designation'])

            self.app.log_message(self.log_display, "   - Waiting for Technical Staff list to populate...")
            wait.until(EC.presence_of_element_located((By.XPATH, "//select[@id='ddlstaff']/option[position()>1]")))
            
            staff_dropdown = Select(driver.find_element(By.ID, "ddlstaff"))
            staff_found = False
            for opt in staff_dropdown.options:
                if inputs['staff'].lower() == opt.text.lower():
                    staff_dropdown.select_by_visible_text(opt.text)
                    staff_found = True
                    break
            
            if not staff_found:
                raise ValueError(f"Staff name '{inputs['staff']}' not found. Check spelling.")
            
            self.app.log_message(self.log_display, "   - Submitting form...")
            body_element = driver.find_element(By.TAG_NAME, 'body')
            btn_proceed = wait.until(EC.presence_of_element_located((By.ID, "btnProceed")))
            driver.execute_script("arguments[0].click();", btn_proceed)
            
            try:
                WebDriverWait(driver, 5).until(EC.alert_is_present())
                alert = driver.switch_to.alert
                alert_text = alert.text
                alert.accept()
                self._log_result(item, "Failed", f"Server Alert: {alert_text}")
                return
            except TimeoutException: pass
            
            wait.until(EC.staleness_of(body_element))
            
            error_reason = self._check_for_page_errors(driver)
            if error_reason:
                self._log_result(item, "Skipped", error_reason)
                session_skip_list.add(full_work_code_text)
                return
            
            # --- UPDATED: Pass extracted code to function ---
            if inputs.get('run_ekyc'):
                try:
                    # Pass the correct work code directly
                    self._extract_html_data_for_ekyc(driver, full_work_code_text)
                except Exception as e:
                    self.app.log_message(self.log_display, f"   - Warning: Failed to extract data for eKYC: {e}", "warning")
            # --------------------------------------------

            self.app.log_message(self.log_display, "   - Muster Roll is valid. Generating output...")
            pdf_path = self._save_mr_as_pdf(driver, full_work_code_text, output_dir, inputs['orientation'], inputs['scale'])
            
            log_detail = f"Saved as {os.path.basename(pdf_path)}" if pdf_path else "PDF Save Failed"
            
            if pdf_path:
                self.current_session_files.append(pdf_path)
                if inputs.get('save_to_cloud'):
                    self.app.log_message(self.log_display, "   - Uploading to cloud...")
                    if self._upload_to_cloud(pdf_path, inputs['panchayat']):
                        log_detail += " & Cloud Uploaded"
                    else:
                        log_detail += " (Cloud Failed)"

            if inputs['output_action'] == "Print" and pdf_path:
                self._print_file(pdf_path)

            self._log_result(item, "Success" if pdf_path else "Failed", log_detail)
            session_skip_list.add(full_work_code_text)

        except TimeoutException:
            self.app.log_message(self.log_display, f"Error on '{item}': Timeout (Slow Network)", "error")
            self._log_result(item, "Failed", "Timeout - Slow Network")
        except Exception as e:
            error_msg = str(e).splitlines()[0] if str(e) else "Unknown Error"
            self.app.log_message(self.log_display, f"Error on '{item}': {error_msg}", "error")
            self._log_result(item, "Failed", error_msg)

    # --- NEW: HTML Data Extraction Logic (Smart Search) ---
    def _extract_html_data_for_ekyc(self, driver, known_work_code):
        """Extracts data and uses the KNOWN Work Code to avoid regex errors."""
        panchayat_name = "Unknown"
        work_name = "Unknown"
        
        # 1. Header Extraction
        try:
            body_text = driver.find_element(By.TAG_NAME, "body").text
            
            # Panchayat Name Cleanup
            p_match = re.search(r"(?:Panchayat|पंचायत)\s*[:\-]\s*([^\n\r]+)", body_text, re.IGNORECASE)
            if p_match:
                raw_name = p_match.group(1).strip()
                # Clean extra text like "Financial Year"
                separators = ["Financial", "वित्तीय", "Janpad", "जनपद", "District", "Zila", "जिला", "Block", "खंड"]
                for sep in separators:
                    if sep.lower() in raw_name.lower():
                        idx = raw_name.lower().find(sep.lower())
                        raw_name = raw_name[:idx].strip()
                panchayat_name = raw_name.rstrip(":-").strip()

            # Work Name Parsing (Work Name is usually safe to extract)
            wn_match = re.search(r"(?:Work Name|कार्य का नाम)\s*[:\-]\s*([^\n\r]+)", body_text, re.IGNORECASE)
            if wn_match:
                work_name = wn_match.group(1).strip()
                
        except Exception as e:
            print(f"Header extraction warning: {e}")

        # 2. Find ALL Worker Tables
        target_tables = []
        all_tables = driver.find_elements(By.TAG_NAME, "table")
        
        for tbl in all_tables:
            if "Applicant Name" in tbl.text or "आवेदक का नाम" in tbl.text:
                target_tables.append(tbl)
        
        if not target_tables:
            self.app.log_message(self.log_display, "   - eKYC Extraction: Could not find any worker table.", "warning")
            return

        extracted_count = 0
        
        for table in target_tables:
            rows = table.find_elements(By.TAG_NAME, "tr")
            for row in rows:
                cols = row.find_elements(By.TAG_NAME, "td")
                if len(cols) > 4:
                    try:
                        col1_text = cols[0].text.strip()
                        if "S.No" in col1_text or "क्र.सं." in col1_text: continue

                        raw_jc_text = cols[1].text.strip() 
                        applicant_name = cols[3].text.strip()
                        village = cols[4].text.strip()
                        
                        if applicant_name and village and "Applicant" not in applicant_name:
                            self.collected_mr_data.append({
                                'panchayat': panchayat_name,
                                'work_code': known_work_code, # Use the correct passed code
                                'work_name': work_name,
                                'village': village,
                                'jobcard': raw_jc_text, 
                                'name': applicant_name
                            })
                            extracted_count += 1
                    except: continue
        
        if extracted_count > 0:
            self.app.log_message(self.log_display, f"   - eKYC Data: Captured {extracted_count} workers (Panchayat: {panchayat_name}).", "info")
        else:
            self.app.log_message(self.log_display, "   - eKYC Data: Tables found but no workers extracted.", "warning")

    # --- NEW: eKYC Execution Logic ---
    def run_post_mr_ekyc_check(self, driver, wait, input_panchayat_name):
        self.update_status("Starting eKYC Check...", 0)
        
        if not self.collected_mr_data: return

        # Determine best Panchayat Name
        target_panchayat_name = input_panchayat_name
        for rec in self.collected_mr_data:
            if rec.get('panchayat') and rec.get('panchayat') != "Unknown":
                target_panchayat_name = rec['panchayat']
                break
        
        self.app.log_message(self.log_display, f"Using Panchayat name for eKYC: '{target_panchayat_name}'")

        unique_villages = list(set(item['village'] for item in self.collected_mr_data if item['village']))
        self.app.log_message(self.log_display, f"Found {len(unique_villages)} unique villages to check.", "info")
        
        if not unique_villages: return

        # Open eKYC Page
        driver.get("https://nregade4.nic.in/Netnrega/UID/AppABPSRpt.aspx")
        
        # Uncheck Pending
        try:
            chk = wait.until(EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_chbx_freshCase")))
            if driver.execute_script("return arguments[0].checked;", chk):
                driver.execute_script("arguments[0].click();", chk)
                try: wait.until(EC.staleness_of(chk))
                except: time.sleep(2)
        except: pass

        # Select Panchayat
        try:
            self.app.log_message(self.log_display, f"Selecting Panchayat...")
            panchayat_elem = wait.until(EC.element_to_be_clickable((By.ID, "ctl00_ContentPlaceHolder1_DDL_panchayat")))
            old_html = driver.find_element(By.TAG_NAME, "html")
            
            select = Select(panchayat_elem)
            try:
                select.select_by_visible_text(target_panchayat_name)
            except NoSuchElementException:
                self.app.log_message(self.log_display, "Exact name match failed, trying case-insensitive...", "warning")
                found_opt = False
                for opt in select.options:
                    if opt.text.lower() == target_panchayat_name.lower():
                        select.select_by_visible_text(opt.text)
                        found_opt = True
                        break
                if not found_opt: raise Exception(f"Panchayat '{target_panchayat_name}' not found.")

            try: wait.until(EC.staleness_of(old_html))
            except: time.sleep(3)
        except Exception as e:
            self.app.log_message(self.log_display, f"eKYC: Panchayat selection failed: {e}", "error")
            return

        # Scan Villages
        scraped_ekyc_data = {} 
        
        for idx, v_name in enumerate(unique_villages, 1):
            if self.app.stop_events[self.automation_key].is_set(): break
            
            self.update_status(f"eKYC Checking Village {idx}/{len(unique_villages)}: {v_name}")
            self.app.log_message(self.log_display, f"Scanning Village: {v_name}", "info")
            
            try:
                old_html = driver.find_element(By.TAG_NAME, "html")
                v_dd_elem = wait.until(EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_DDL_Village")))
                Select(v_dd_elem).select_by_visible_text(v_name)
                try: wait.until(EC.staleness_of(old_html))
                except: time.sleep(2)
            except:
                self.app.log_message(self.log_display, f"Skipping {v_name} (Selection Failed)", "error")
                continue

            # Force Reset to Page 1
            try:
                page_one_links = driver.find_elements(By.XPATH, "//a[contains(@href,'Page$1')]")
                if page_one_links:
                    old_table = driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_gvData")
                    driver.execute_script("arguments[0].click();", page_one_links[0])
                    try: wait.until(EC.staleness_of(old_table))
                    except: time.sleep(2)
            except: pass

            current_page = 1
            while True:
                try:
                    table = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_gvData")))
                    rows = table.find_elements(By.TAG_NAME, "tr")
                    
                    if len(rows) > 1:
                        for row in rows[1:]:
                            cols = row.find_elements(By.TAG_NAME, "td")
                            if len(cols) < 5: continue
                            try:
                                jc_text = cols[1].text.strip()
                                
                                # --- UPDATE: Capture ABPS Status ---
                                abps_stat = cols[-2].text.strip() # 2nd Last Column
                                ekyc_stat = cols[-1].text.strip() # Last Column
                                
                                clean_jc = "".join(jc_text.split()).lower()
                                key = (v_name, clean_jc)
                                
                                # Store both statuses
                                scraped_ekyc_data[key] = {
                                    'ekyc': ekyc_stat,
                                    'abps': abps_stat
                                }
                            except: continue
                    
                    next_btns = driver.find_elements(By.XPATH, f"//a[contains(@href, 'Page${current_page + 1}')]")
                    if next_btns:
                        self.update_status(f"Scanning {v_name} - Page {current_page + 1}...")
                        driver.execute_script("arguments[0].click();", next_btns[0])
                        time.sleep(2)
                        current_page += 1
                    else: break
                except: break

        self._generate_ekyc_verification_report(scraped_ekyc_data)

    def _generate_ekyc_verification_report(self, scraped_data):
        self.app.log_message(self.log_display, "Generating Professional Verification Report...", "info")
        
        failed_ekyc_records = []
        
        for record in self.collected_mr_data:
            v_name = record['village']
            raw_jc = record['jobcard']
            
            clean_mr_jc = "".join(raw_jc.split()).lower()
            
            # Default Status
            ekyc_status = "Not Found"
            abps_status = "Not Found"
            
            found = False
            for (sk_village, sk_jc), stats in scraped_data.items():
                if sk_village == v_name and (clean_mr_jc in sk_jc or sk_jc in clean_mr_jc):
                    ekyc_status = stats.get('ekyc', 'Unknown')
                    abps_status = stats.get('abps', 'Unknown')
                    found = True
                    break
            
            # --- LOGIC UPDATE: Add if eKYC is NO *OR* ABPS is NO ---
            is_ekyc_bad = "no" in ekyc_status.lower() or "not found" in ekyc_status.lower()
            is_abps_bad = "no" in abps_status.lower()
            
            if is_ekyc_bad or is_abps_bad:
                full_wc = record['work_code']
                short_wc = full_wc[-6:] if len(full_wc) >= 6 else full_wc
                
                failed_ekyc_records.append([
                    short_wc,
                    record['work_name'],
                    record['panchayat'],
                    record['village'],
                    record['jobcard'],
                    record['name'],
                    abps_status, # New Column Data
                    ekyc_status
                ])

        if not failed_ekyc_records:
            self.app.log_message(self.log_display, "Verification Complete: All laborers have eKYC & ABPS Verified.", "success")
            return

        # Save Professional Excel
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Verification Report"
            
            # --- STYLES ---
            header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark Blue
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            title_font = Font(size=16, bold=True, color="1F497D")
            subtitle_font = Font(size=10, italic=True, color="555555")
            
            border = Border(
                left=Side(style='thin'), right=Side(style='thin'), 
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            center_align = Alignment(horizontal="center", vertical="center")
            left_align = Alignment(horizontal="left", vertical="center")

            # --- HEADER SECTION ---
            # Title
            ws.merge_cells('A1:H1')
            ws['A1'] = f"ABPS & eKYC VERIFICATION REPORT: {self.panchayat_entry.get().upper()}"
            ws['A1'].font = title_font
            ws['A1'].alignment = center_align
            
            # Subtitle (Promotion & Date)
            ws.merge_cells('A2:H2')
            ws['A2'] = f"Generated by NregaBot.com | Date: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}"
            ws['A2'].font = subtitle_font
            ws['A2'].alignment = center_align

            # --- TABLE HEADERS ---
            headers = [
                "Work Code (Last 6)", "Work Name", "Panchayat", "Village", 
                "Job Card No", "Applicant Name", "ABPS/ NPCI", "eKYC Status"
            ]
            
            # Row 4 is for headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border

            # --- DATA ROWS ---
            start_row = 5
            for idx, row_data in enumerate(failed_ekyc_records, start=start_row):
                # (wc, work_name, panchayat, village, jc, name, abps, ekyc)
                
                # Standard Columns
                ws.cell(row=idx, column=1, value=row_data[0]).alignment = center_align
                ws.cell(row=idx, column=2, value=row_data[1]).alignment = left_align # Work Name Left
                ws.cell(row=idx, column=3, value=row_data[2]).alignment = center_align
                ws.cell(row=idx, column=4, value=row_data[3]).alignment = center_align
                ws.cell(row=idx, column=5, value=row_data[4]).alignment = center_align
                ws.cell(row=idx, column=6, value=row_data[5]).alignment = left_align # Name Left
                
                # Status Columns with Conditional Formatting (Red text if NO)
                abps_cell = ws.cell(row=idx, column=7, value=row_data[6])
                abps_cell.alignment = center_align
                if "no" in str(row_data[6]).lower():
                    abps_cell.font = Font(color="FF0000", bold=True)
                else:
                    abps_cell.font = Font(color="008000", bold=True) # Green for Yes

                ekyc_cell = ws.cell(row=idx, column=8, value=row_data[7])
                ekyc_cell.alignment = center_align
                if "no" in str(row_data[7]).lower() or "not" in str(row_data[7]).lower():
                    ekyc_cell.font = Font(color="FF0000", bold=True)
                else:
                    ekyc_cell.font = Font(color="008000", bold=True)

                # Apply Borders to all cells in row
                for col in range(1, 9):
                    ws.cell(row=idx, column=col).border = border

            # --- COLUMN WIDTHS ---
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 40 # Work Name wider
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 22
            ws.column_dimensions['F'].width = 25
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 15
            
            # Save Path
            filename = f"Verification_Report_{self.panchayat_entry.get()}_{datetime.now().strftime('%d%m%Y_%H%M')}.xlsx"
            save_path = os.path.join(self.output_dir, filename)
            
            wb.save(save_path)
            self.app.log_message(self.log_display, f"Report Saved: {save_path}", "success")
            
            # Note: Popup is handled by caller function (the completion dialog will open folder)
                
        except Exception as e:
            self.app.log_message(self.log_display, f"Failed to save verification report: {e}", "error")

    def _save_mr_as_pdf(self, driver, full_work_code, output_dir, orientation, scale):
        try:
            safe_work_code = full_work_code.split('/')[-1][-6:]
            base_filename = safe_work_code
            extension = ".pdf"
            counter = 1
            pdf_filename = f"{base_filename}{extension}"
            save_path = os.path.join(output_dir, pdf_filename)

            while os.path.exists(save_path):
                pdf_filename = f"{base_filename} ({counter}){extension}"
                save_path = os.path.join(output_dir, pdf_filename)
                counter += 1

            is_landscape = (orientation == "Landscape")
            pdf_scale = scale / 100.0
            pdf_data_base64 = None

            # --- CSS for Orientation ---
            if is_landscape:
                self.app.log_message(self.log_display, "   - Injecting CSS for landscape orientation...")
                driver.execute_script(
                    "var css = '@page { size: landscape; }';"
                    "var head = document.head || document.getElementsByTagName('head')[0];"
                    "var style = document.createElement('style');"
                    "style.type = 'text/css'; style.media = 'print';"
                    "if (style.styleSheet){ style.styleSheet.cssText = css; }"
                    "else { style.appendChild(document.createTextNode(css)); }"
                    "head.appendChild(style);"
                )

            # --- Fix: Remove blank page caused by website update ---
            self.app.log_message(self.log_display, "   - Removing blank page elements...")
            driver.execute_script("""
                // 1. Inject print CSS to suppress all known blank-page causes
                var styleTag = document.createElement('style');
                styleTag.innerHTML = `
                    @media print {
                        * { page-break-after: auto !important; page-break-before: auto !important; break-after: auto !important; break-before: auto !important; }
                        body::after { display: none !important; content: none !important; }
                    }
                `;
                document.head.appendChild(styleTag);

                // 2. Remove trailing empty block elements from body
                var bodyChildren = Array.from(document.body.children);
                for (var i = bodyChildren.length - 1; i >= 0; i--) {
                    var el = bodyChildren[i];
                    if (el.innerText.trim() === '' && el.querySelectorAll('img, input, table, iframe, canvas, video').length === 0) {
                        el.parentNode.removeChild(el);
                    } else {
                        break; // stop at first non-empty element from the end
                    }
                }

                // 3. Force remove page-break-after on ALL elements
                var allEls = document.querySelectorAll('*');
                allEls.forEach(function(el) {
                    el.style.pageBreakAfter = 'auto';
                    el.style.pageBreakBefore = 'auto';
                    el.style.breakAfter = 'auto';
                    el.style.breakBefore = 'auto';
                });
            """)

            if self.app.active_browser == 'firefox':
                # Firefox: Inject a fixed div using JavaScript
                footer_js = """
                var footer = document.createElement('div');
                footer.innerText = 'NregaBot.com';
                footer.style.position = 'fixed';
                footer.style.bottom = '0';
                footer.style.right = '0';
                footer.style.padding = '10px';
                footer.style.fontSize = '10px';
                footer.style.color = '#cccccc';  // Light Gray
                footer.style.fontFamily = 'Arial, sans-serif';
                footer.style.zIndex = '9999';
                document.body.appendChild(footer);
                """
                driver.execute_script(footer_js)

                self.app.log_message(self.log_display, "   - Using Firefox's print command...")
                self.app.log_message(self.log_display, "   - Note: PDF Scale setting is ignored for Firefox.", "warning")
                pdf_data_base64 = driver.print_page()

            elif self.app.active_browser == 'chrome':
                self.app.log_message(self.log_display, "   - Using Chrome's advanced print command (CDP)...")

                # Inject footer as a fixed-position element (avoids CDP footer causing extra blank page)
                driver.execute_script("""
                    var existing = document.getElementById('nregabot-footer');
                    if (!existing) {
                        var footer = document.createElement('div');
                        footer.id = 'nregabot-footer';
                        footer.innerText = 'NregaBot.com';
                        footer.style.position = 'fixed';
                        footer.style.bottom = '6px';
                        footer.style.right = '10px';
                        footer.style.fontSize = '9px';
                        footer.style.color = '#d3d3d3';
                        footer.style.fontFamily = 'Helvetica, sans-serif';
                        footer.style.zIndex = '9999';
                        document.body.appendChild(footer);
                    }
                """)

                print_options = {
                    "landscape": is_landscape,
                    "displayHeaderFooter": False,      # Disabled - footer injected via JS instead
                    "printBackground": False,
                    "scale": pdf_scale,
                    "marginTop": 0.4,
                    "marginBottom": 0.4,
                    "marginLeft": 0.4, "marginRight": 0.4
                }
                result = driver.execute_cdp_cmd("Page.printToPDF", print_options)
                pdf_data_base64 = result['data']

            if pdf_data_base64:
                pdf_data = base64.b64decode(pdf_data_base64)
                with open(save_path, 'wb') as f:
                    f.write(pdf_data)
                return save_path
            else:
                self.app.log_message(self.log_display, "Error: PDF data was not generated by the browser.", "error")
                return None

        except Exception as e:
            self.app.log_message(self.log_display, f"Error saving PDF: {e}", "error")
            return None

    def _select_work_code(self, driver, wait, item, is_auto_mode):
        """
        Selects the work code with Retry Logic to handle StaleElementExceptions.
        """
        work_code_dropdown_locator = (By.ID, "ddlWorkCode")
        
        for attempt in range(3): # Try 3 times
            try:
                if is_auto_mode:
                    # Auto Mode: Wait for options
                    wait.until(EC.presence_of_element_located(work_code_dropdown_locator))
                    wait.until(lambda d: len(Select(d.find_element(*work_code_dropdown_locator)).options) > 1)
                    
                    work_code_dropdown = Select(driver.find_element(*work_code_dropdown_locator))
                    
                    # Iterate to find matching text
                    found_option = next((opt for opt in work_code_dropdown.options if opt.text == item and opt.get_attribute("value")), None)

                    if found_option:
                        full_work_code_text = found_option.text
                        work_code_dropdown.select_by_visible_text(full_work_code_text)
                        self.app.log_message(self.log_display, f"   - Found and selected: {full_work_code_text}")
                        return full_work_code_text
                    else:
                        raise NoSuchElementException(f"Could not find a matching work for auto item '{item}'.")
                
                else:
                    # Manual Mode: Search box
                    search_key = item
                    search_box = wait.until(EC.presence_of_element_located((By.ID, "txtWork")))
                    driver.execute_script("arguments[0].value = arguments[1];", search_box, search_key)
                    
                    # Capture old dropdown to detect refresh
                    old_dropdown = driver.find_element(*work_code_dropdown_locator)
                    
                    search_btn = driver.find_element(By.ID, "imgButtonSearch")
                    driver.execute_script("arguments[0].click();", search_btn)
                    
                    # Wait for refresh
                    try: wait.until(EC.staleness_of(old_dropdown))
                    except TimeoutException: pass
                    
                    wait.until(EC.presence_of_element_located(work_code_dropdown_locator))
                    wait.until(lambda d: len(Select(d.find_element(*work_code_dropdown_locator)).options) > 1)
                    
                    work_code_dropdown = Select(driver.find_element(*work_code_dropdown_locator))
                    
                    found_option = next((opt for opt in work_code_dropdown.options if search_key in opt.text and opt.get_attribute("value")), None)
                    if found_option:
                        full_work_code_text = found_option.text
                        work_code_dropdown.select_by_visible_text(full_work_code_text)
                        self.app.log_message(self.log_display, f"   - Found and selected: {full_work_code_text}")
                        return full_work_code_text
                    else:
                        raise NoSuchElementException(f"Could not find a matching work for search key '{item}'.")
                        
            except StaleElementReferenceException:
                if attempt < 2:
                    self.app.log_message(self.log_display, "   - Stale element detected, retrying selection...", "warning")
                    time.sleep(2)
                    continue
                else:
                    raise
            except Exception as e:
                raise e

    def _check_for_page_errors(self, driver) -> str | None:
        """Checks for known error messages on the page. Returns the error string if found, else None."""
        page_source = driver.page_source.lower()
        if "geotag is not received" in page_source:
            return "Skipped: Geotag not received"
        if "greater than allowed limit" in page_source:
            return "Skipped: Greater than allowed limit"
        if "no worker available" in page_source:
            return "Skipped: No worker available"
        if "no muster roll available" in page_source:
            return "Skipped: No Muster Roll available"
        if "overlap that period" in page_source:
            return "Skipped: Date period overlaps with existing MR"
        return None

    def _log_result(self, item_key, status, details):
        timestamp = datetime.now().strftime("%H:%M:%S")
        values = (timestamp, item_key, status, details)
        
        # --- FIX: Explicit Success Tag ---
        # Pehle code sirf 'failed' check kar raha tha, ab 'success' bhi check karega
        tags = ()
        if 'success' in status.lower():
            tags = ('success',)
        else:
            tags = ('failed',)
        # ---------------------------------

        if status == "Success":
            self.success_count += 1
            self.app.after(0, lambda: self.success_label.configure(text=f"Success: {self.success_count}"))
        else:
            self.skipped_count += 1
            self.app.after(0, lambda: self.skipped_label.configure(text=f"Skipped/Failed: {self.skipped_count}"))
        
        self.app.after(0, lambda: self.results_tree.insert("", "end", values=values, tags=tags))

    def _upload_to_cloud(self, file_path, panchayat_name):
        """Uploads a given file to the user's cloud storage via the API."""
        if not self.app.license_info.get('key'):
            self.app.log_message(self.log_display, "   - Cloud Upload Skipped: No license key found.", "warning")
            return False
            
        headers = {'Authorization': f"Bearer {self.app.license_info['key']}"}
        url = f"{config.LICENSE_SERVER_URL}/files/api/upload"
        filename = os.path.basename(file_path)

        try:
            with open(file_path, 'rb') as f:
                files = {'file': (filename, f, 'application/pdf')}
                
                date_folder = datetime.now().strftime('%Y-%m-%d')
                safe_panchayat_name = "".join(c for c in panchayat_name if c.isalnum() or c in (' ', '_')).rstrip()
                
                relative_path = f'Muster_Rolls/{date_folder}/{safe_panchayat_name}/{filename}'
                
                data = {
                    'parent_id': '', 
                    'relative_path': relative_path
                }
                
                response = requests.post(url, headers=headers, files=files, data=data, timeout=120)

            if response.status_code == 201:
                return True
            else:
                self.app.log_message(self.log_display, f"   - Cloud upload failed with status {response.status_code}: {response.text}", "error")
                return False
        except requests.exceptions.RequestException as e:
            self.app.log_message(self.log_display, f"   - A connection error occurred during cloud upload: {e}", "error")
            return False
        except Exception as e:
            self.app.log_message(self.log_display, f"   - An unexpected error occurred during cloud upload: {e}", "error")
            return False

    def export_report(self):
        export_format = self.export_format_menu.get()
        if "CSV" in export_format:
            self.export_treeview_to_csv(self.results_tree, "muster_roll_gen_results.csv")
            return
            
        data, file_path = self._get_filtered_data_and_filepath(export_format)
        if not data: return

        report_data = [[row[1], row[2], row[3], row[0]] for row in data]
        report_headers = ["Work Code/Key", "Status", "Details", "Timestamp"]
        col_widths = [70, 35, 140, 25]

        if "PDF" in export_format:
            self._handle_pdf_export(report_data, report_headers, col_widths, file_path)

    def _get_filtered_data_and_filepath(self, export_format):
        if not self.results_tree.get_children(): messagebox.showinfo("No Data", "No results to export."); return None, None
        panchayat_name = self.panchayat_entry.get().strip()
        if not panchayat_name: messagebox.showwarning("Input Needed", "Panchayat Name is required for report title."); return None, None
        
        filter_option = self.export_filter_menu.get()
        data_to_export = []
        for item_id in self.results_tree.get_children():
            row_values = self.results_tree.item(item_id)['values']
            status = row_values[2].upper() 
            if filter_option == "Export All": data_to_export.append(row_values)
            elif filter_option == "Success Only" and "SUCCESS" in status: data_to_export.append(row_values)
            elif filter_option == "Failed Only" and "SUCCESS" not in status: data_to_export.append(row_values)
        if not data_to_export: messagebox.showinfo("No Data", f"No records found for filter '{filter_option}'."); return None, None

        safe_name = "".join(c for c in panchayat_name if c.isalnum() or c in (' ', '_')).rstrip()
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        details = {"Image (.jpg)": { "ext": ".jpg", "types": [("JPEG Image", "*.jpg")]}, "PDF (.pdf)": { "ext": ".pdf", "types": [("PDF Document", "*.pdf")]}}[export_format]
        filename = f"MR_Gen_Report_{safe_name}_{timestamp}{details['ext']}"
        file_path = filedialog.asksaveasfilename(defaultextension=details['ext'], filetypes=details['types'], initialdir=self.app.get_user_downloads_path(), initialfile=filename, title="Save Report")
        return (data_to_export, file_path) if file_path else (None, None)
    
    def _handle_pdf_export(self, data, headers, col_widths, file_path):
        title = f"Muster Roll Generation Report: {self.panchayat_entry.get().strip()}"
        report_date = datetime.now().strftime('%d %b %Y')
        success = self.generate_report_pdf(data, headers, col_widths, title, report_date, file_path)
        if success and messagebox.askyesno("Success", f"PDF Report saved to:\n{file_path}\n\nDo you want to open it?"):
            if sys.platform == "win32": os.startfile(file_path)
            else: subprocess.call(['open', file_path])

    # --- NEW MERGE PDFS METHOD ---
    def merge_saved_pdfs(self):
        self.app.log_message(self.log_display, "Starting PDF merge...")
        
        # Use ONLY the files generated in the current session
        pdf_files = self.current_session_files
        
        if not pdf_files:
            self.app.log_message(self.log_display, "No PDFs generated in this session to merge.", "warning")
            messagebox.showinfo("No Files", "No MRs have been successfully generated in this cycle yet.\nRun the automation first.", parent=self)
            return
            
        self.app.log_message(self.log_display, f"Merging {len(pdf_files)} files generated in this session.")

        # Get output file name from user
        dialog = ctk.CTkInputDialog(text="Enter a base name for the merged file:", title="Merge PDFs")
        base_name = dialog.get_input()
        
        if not base_name:
            self.app.log_message(self.log_display, "Merge cancelled by user.", "info")
            return

        # Get unique output path
        try:
            merge_output_dir = os.path.join(self.app.get_user_downloads_path(), "NregaBot", "Merged_Pdf_Output")
            os.makedirs(merge_output_dir, exist_ok=True)
            
            date_str = datetime.now().strftime("%d-%b-%Y")
            file_name = f"{base_name}_{date_str}.pdf"
            output_path = os.path.join(merge_output_dir, file_name)
            
            count = 1
            while os.path.exists(output_path):
                file_name = f"{base_name}_{date_str}({count}).pdf"
                output_path = os.path.join(merge_output_dir, file_name)
                count += 1
        except Exception as e:
            messagebox.showerror("Path Error", f"Could not create merge output path: {e}", parent=self)
            return

        # Run merge in a separate thread
        self.app.start_automation_thread(
            "pdf_merger_mr", 
            self._run_merge_logic, 
            args=(pdf_files, output_path)
        )

    def _run_merge_logic(self, file_list, output_path):
        """The actual PDF merging logic that runs in a thread."""
        self.app.after(0, self.set_ui_state, True)
        self.app.log_message(self.log_display, f"Merging {len(file_list)} files...")
        self.app.after(0, self.app.set_status, "Merging PDFs...")
        try:
            merger = PdfWriter()
            for pdf_path in file_list:
                if self.app.stop_events.get("pdf_merger_mr", threading.Event()).is_set():
                    self.app.log_message(self.log_display, "Merge cancelled.", "warning")
                    merger.close()
                    return
                merger.append(pdf_path)
            
            with open(output_path, "wb") as f_out:
                merger.write(f_out)
            merger.close()
            
            self.app.log_message(self.log_display, "Merge complete!", "success")
            messagebox.showinfo("Success", f"Successfully merged {len(file_list)} files into:\n{output_path}", parent=self)
            if messagebox.askyesno("Open Location?", "Open the Merged PDFs folder?", parent=self):
                self.app.open_folder(os.path.dirname(output_path))
                
        except Exception as e:
            self.app.log_message(self.log_display, f"Error during merge: {e}", "error")
            messagebox.showerror("Merge Error", f"An error occurred: {e}", parent=self)
        finally:
            self.app.after(0, self.set_ui_state, False)
            self.app.after(0, self.app.set_status, "Ready")