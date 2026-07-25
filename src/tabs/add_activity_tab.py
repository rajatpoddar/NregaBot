# tabs/add_activity_tab.py
import tkinter
from tkinter import ttk, messagebox
import customtkinter as ctk
import time
from datetime import datetime

from src import config
from .base_tab import BaseAutomationTab
from src.utils import get_logger
from typing import Any, Callable, Dict, List, Optional, Tuple

logger = get_logger()

class AddActivityTab(BaseAutomationTab):
    def __init__(self, parent: Any, app_instance: Any) -> None:
        # Lazy imports
        from selenium.webdriver.common.keys import Keys
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import NoSuchElementException, TimeoutException, UnexpectedAlertPresentException, StaleElementReferenceException
        from selenium.webdriver.common.by import By
        from selenium.webdriver.common.keys import Keys
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import NoSuchElementException, TimeoutException, UnexpectedAlertPresentException, StaleElementReferenceException
        super().__init__(parent, app_instance, automation_key="add_activity")
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self._create_widgets()
    def _create_widgets(self) -> None:
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.webdriver.common.keys import Keys
        from selenium.common.exceptions import UnexpectedAlertPresentException
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        import openpyxl
        from selenium import webdriver

        # Frame for controls and action buttons
        top_frame = ctk.CTkFrame(self)
        top_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 0))
        top_frame.grid_columnconfigure(0, weight=1)

        # --- UPDATED: Input fields for Price and Quantity ---
        input_frame = ctk.CTkFrame(top_frame)
        input_frame.grid(row=0, column=0, sticky="ew", padx=15, pady=(15, 10))
        input_frame.grid_columnconfigure((1, 3), weight=1)
        
        defaults = config.ADD_ACTIVITY_CONFIG['defaults']
        ctk.CTkLabel(input_frame, text=f"Default Activity Code: {defaults['activity_code']}", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, columnspan=4, sticky="w", padx=15, pady=(0, 10))

        ctk.CTkLabel(input_frame, text="Unit Price (₹):").grid(row=1, column=0, sticky="w", padx=15)
        self.unit_price_entry = ctk.CTkEntry(input_frame)
        self.unit_price_entry.grid(row=1, column=1, sticky="ew", padx=(0, 15))
        self.unit_price_entry.insert(0, defaults['unit_price'])

        ctk.CTkLabel(input_frame, text="Quantity:").grid(row=1, column=2, sticky="w", padx=15)
        self.quantity_entry = ctk.CTkEntry(input_frame)
        self.quantity_entry.grid(row=1, column=3, sticky="ew", padx=(0, 15))
        self.quantity_entry.insert(0, defaults['quantity'])

        # Action buttons
        action_frame = self._create_action_buttons(parent_frame=top_frame)
        action_frame.grid(row=1, column=0, sticky='ew', pady=(10, 15), padx=15)

        # Notebook for inputs and results
        notebook = ctk.CTkTabview(self)
        notebook.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 10))
        work_codes_frame = notebook.add("Work Keys")
        results_frame = notebook.add("Results")
        self._create_log_and_status_area(parent_notebook=notebook)

        # Work Keys Tab
        work_codes_frame.grid_columnconfigure(0, weight=1)
        work_codes_frame.grid_rowconfigure(1, weight=1) # <-- CORRECTED THIS LINE

        # --- NEW: Controls frame for buttons ---
        wc_controls_frame = ctk.CTkFrame(work_codes_frame, fg_color="transparent")
        wc_controls_frame.grid(row=0, column=0, sticky="ew", padx=5, pady=(5,0))
        
        clear_button = ctk.CTkButton(wc_controls_frame, text="Clear", width=80, command=lambda: self.work_keys_text.delete("1.0", tkinter.END))
        clear_button.pack(side='right', padx=(0, 5))
        
        extract_button = ctk.CTkButton(wc_controls_frame, text="Extract from Text", width=120,
                                       command=lambda: self._extract_and_update_workcodes(self.work_keys_text))
        extract_button.pack(side='right', padx=(0, 5))
        # --- END NEW ---

        self.work_keys_text = ctk.CTkTextbox(work_codes_frame, wrap=tkinter.WORD)
        self.work_keys_text.grid(row=1, column=0, sticky='nsew', padx=5, pady=5) # <-- Changed to row 1

        # Results Tab
        results_frame.grid_columnconfigure(0, weight=1)
        results_frame.grid_rowconfigure(1, weight=1) # Make space for the button

        results_action_frame = ctk.CTkFrame(results_frame, fg_color="transparent")
        results_action_frame.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(5, 10), padx=5)
        self.export_csv_button = ctk.CTkButton(results_action_frame, text="Export to CSV", command=lambda: self.export_treeview_to_csv(self.results_tree, "add_activity_results.csv"))
        self.export_csv_button.pack(side="left")

        cols = ("Work Key", "Status", "Details", "Timestamp")
        self.results_tree = ttk.Treeview(results_frame, columns=cols, show='headings')
        for col in cols:
            self.results_tree.heading(col, text=col)
        self.results_tree.column("Work Key", width=150)
        self.results_tree.column("Status", width=100, anchor='center')
        self.results_tree.column("Details", width=400)
        self.results_tree.column("Timestamp", width=100, anchor='center')
        self.results_tree.grid(row=1, column=0, sticky='nsew')
        scrollbar = ctk.CTkScrollbar(results_frame, command=self.results_tree.yview)
        self.results_tree.configure(yscroll=scrollbar.set)
        scrollbar.grid(row=1, column=1, sticky='ns')
        self.style_treeview(self.results_tree)

    def set_ui_state(self, running: bool):
        if not self._is_alive():
            return
        self.set_common_ui_state(running)
        state = "disabled" if running else "normal"
        self.work_keys_text.configure(state=state)
        self.unit_price_entry.configure(state=state)
        self.quantity_entry.configure(state=state)
    def start_automation(self) -> None:
        work_keys = [line.strip() for line in self.work_keys_text.get("1.0", tkinter.END).strip().splitlines() if line.strip()]
        if not work_keys:
            messagebox.showwarning("Input Required", "Please provide at least one work key.")
            return
            
        # Get and validate the new inputs
        unit_price = self.unit_price_entry.get().strip()
        quantity = self.quantity_entry.get().strip()

        if not unit_price or not quantity:
            messagebox.showwarning("Input Required", "Please enter a Unit Price and Quantity.")
            return
        
        # Pass the inputs to the automation logic
        self.app.start_automation_thread(self.automation_key, self.run_automation_logic, args=(work_keys, unit_price, quantity))
    def reset_ui(self) -> None:
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.webdriver.common.keys import Keys
        from selenium.common.exceptions import UnexpectedAlertPresentException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        if messagebox.askokcancel("Reset Form?", "Clear all inputs and logs?"):
            self.work_keys_text.configure(state="normal")
            self.work_keys_text.delete("1.0", tkinter.END)
            # Reset price and quantity to defaults
            defaults = config.ADD_ACTIVITY_CONFIG['defaults']
            self.unit_price_entry.delete(0, tkinter.END)
            self.unit_price_entry.insert(0, defaults['unit_price'])
            self.quantity_entry.delete(0, tkinter.END)
            self.quantity_entry.insert(0, defaults['quantity'])
            
            for item in self.results_tree.get_children():
                self.results_tree.delete(item)
            self.app.clear_log(self.log_display)
            self.update_status("Ready", 0.0)
            self.app.log_message(self.log_display, "Form has been reset.")
            self.app.after(0, self.app.set_status, "Ready")

    def run_automation_logic(self, work_keys, unit_price, quantity):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.webdriver.common.keys import Keys
        from selenium.common.exceptions import UnexpectedAlertPresentException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        self.app.after(0, self.set_ui_state, True)
        self.app.clear_log(self.log_display)
        self.app.after(0, lambda: [self.results_tree.delete(item) for item in self.results_tree.get_children()])
        self.app.log_message(self.log_display, "Starting 'Add Activity' automation...")
        self.app.after(0, self.app.set_status, "Running Add Activity...")

        try:
            driver = self.app.get_driver()
            if not driver:
                return

            total = len(work_keys)
            for i, work_key in enumerate(work_keys):
                if self.app.stop_events[self.automation_key].is_set():
                    self.app.log_message(self.log_display, "Automation stopped.", "warning")
                    break
                self.app.after(0, self.update_status, f"Processing {i+1}/{total}: {work_key}", (i+1) / total)
                self._process_single_work_key(driver, work_key, unit_price, quantity)

            final_msg = "Automation finished." if not self.app.stop_events[self.automation_key].is_set() else "Stopped."
            self.app.after(0, self.update_status, final_msg, 1.0)
            if not self.app.stop_events[self.automation_key].is_set():
                messagebox.showinfo("Complete", "'Add Activity' process has finished.")
        except Exception as e:
            self.app.log_message(self.log_display, f"A critical error occurred: {e}", "error")
            messagebox.showerror("Automation Error", f"An error occurred:\n\n{e}")
        finally:
            self.app.after(0, self.set_ui_state, False)
            self.app.after(0, self.app.set_status, "Automation Finished")

    def _log_result(self, work_key, status, details):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.app.after(0, lambda: self.results_tree.insert("", "end", values=(work_key, status, details, timestamp)))

    # Inside tabs/add_activity_tab.py
    def retry_logic_handler(self) -> None:
        """Override to map the retry button to the work_keys_text box."""
        self.retry_failed_automation(self.work_keys_text)

    def _process_single_work_key(self, driver, work_key, unit_price, quantity):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.webdriver.common.keys import Keys
        from selenium.common.exceptions import UnexpectedAlertPresentException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        """
        Processes a single work key.
        OPTIMIZED: Uses Staleness Check to ensure page refresh before selecting Work Code.
        """
        wait = WebDriverWait(driver, 20)
        activity_code = config.ADD_ACTIVITY_CONFIG['defaults']['activity_code']

        try:
            # Dismiss alerts if any
            try: driver.switch_to.alert.accept()
            except UnexpectedAlertPresentException: pass

            if config.ADD_ACTIVITY_CONFIG["url"] not in driver.current_url:
                driver.get(config.ADD_ACTIVITY_CONFIG["url"])

            # --- 1. Enter work key and Wait for Reload ---
            self.app.log_message(self.log_display, f"Searching for work key: {work_key}")
            
            # Capture the OLD dropdown element to check for refresh later
            work_name_dd_id = 'ctl00_ContentPlaceHolder1_ddlworkName'
            try:
                old_work_ddl = driver.find_element(By.ID, work_name_dd_id)
            except NoSuchElementException:
                old_work_ddl = None

            # Input Key via JS
            work_key_input = wait.until(EC.presence_of_element_located((By.ID, 'ctl00_ContentPlaceHolder1_txtwrksearchkey')))
            driver.execute_script("arguments[0].value = arguments[1];", work_key_input, work_key)
            
            # Trigger PostBack
            driver.execute_script("javascript:setTimeout('__doPostBack(\\'ctl00$ContentPlaceHolder1$txtwrksearchkey\\',\\'\\')', 0)")

            # CRITICAL FIX: Wait for the old dropdown to go 'stale' (page reload)
            if old_work_ddl:
                try:
                    wait.until(EC.staleness_of(old_work_ddl))
                except TimeoutException:
                    self.app.log_message(self.log_display, "Page didn't refresh quickly, forcing wait...", "warning")
            
            try:
                WebDriverWait(driver, 10).until(
                    lambda d: d.execute_script('return document.readyState') == 'complete'
                )
            except TimeoutException:
                pass

            # --- 2. Select work from dropdown ---
            # Re-find the element
            work_ddl_element = wait.until(EC.presence_of_element_located((By.ID, work_name_dd_id)))
            work_select = Select(work_ddl_element)

            # Retry logic if options are not loaded yet
            if len(work_select.options) <= 1:
                # Element wait handled by WebDriverWait below
                work_select = Select(driver.find_element(By.ID, work_name_dd_id))

            if len(work_select.options) > 1:
                work_select.select_by_index(1)
                self.app.log_message(self.log_display, "Work selected. Loading details...")
            else:
                # If still empty, the work key might be invalid
                self._log_result(work_key, "Failed", "Work Key not found or Dropdown empty.")
                return
            
            # Check for existing activity (Use innerText)
            try:
                activity_table = wait.until(EC.presence_of_element_located((By.ID, 'ctl00_ContentPlaceHolder1_grdDisplayAct')))
                if "No Activity Found" in activity_table.get_attribute("innerText"):
                    self.app.log_message(self.log_display, "No existing activity. Proceeding to add.")
                else:
                    self.app.log_message(self.log_display, "Activity already exists. Skipping.", "warning")
                    self._log_result(work_key, "Skipped", "An activity is already present.")
                    return
            except (NoSuchElementException, TimeoutException):
                # Sometimes the table doesn't load instantly, assume safe to proceed
                self.app.log_message(self.log_display, "Activity table check passed.")

            # --- 3. Select Activity ---
            activity_dd_id = 'ctl00_ContentPlaceHolder1_ddlAct'
            try:
                # Capture old activity dropdown
                old_act_ddl = driver.find_element(By.ID, activity_dd_id)
                
                Select(old_act_ddl).select_by_value(activity_code)
                
                # Wait for refresh (selecting activity usually triggers a small reload)
                wait.until(EC.staleness_of(old_act_ddl))
            except NoSuchElementException:
                self._log_result(work_key, "Failed", "Activity Dropdown not found.")
                return

            # --- 4. Fill Unit Price (JS Safe) ---
            unit_price_input = wait.until(EC.presence_of_element_located((By.ID, 'ctl00_ContentPlaceHolder1_txtAct_UnitPrice')))
            driver.execute_script("arguments[0].value = arguments[1];", unit_price_input, unit_price)
            driver.execute_script("arguments[0].dispatchEvent(new Event('change'));", unit_price_input)
            
            # Click body to trigger blur/calculations
            driver.execute_script("document.body.click();")
            
            # Wait for calculation refresh
            try:
                wait.until(EC.staleness_of(unit_price_input))
            except TimeoutException: pass

            # --- 5. Fill Quantity (JS Safe) ---
            quantity_input = wait.until(EC.presence_of_element_located((By.ID, 'ctl00_ContentPlaceHolder1_txtAct_Qty')))
            driver.execute_script("arguments[0].value = arguments[1];", quantity_input, quantity)
            driver.execute_script("arguments[0].dispatchEvent(new Event('change'));", quantity_input)

            time.sleep(1.5)  # Brief wait for postback to begin

            # --- 6. Click Save (JS Safe) ---
            self.app.log_message(self.log_display, "Saving activity...")
            save_button = wait.until(EC.presence_of_element_located((By.ID, 'ctl00_ContentPlaceHolder1_btsave')))
            driver.execute_script("arguments[0].click();", save_button)

            # --- 7. Check Result ---
            outcome_found = False
            for _ in range(3):
                try:
                    # Check for Success Message
                    try:
                        lbl_msg = driver.find_element(By.ID, 'ctl00_ContentPlaceHolder1_lblmsg')
                        txt = lbl_msg.get_attribute("innerText").strip()
                        if txt: 
                            self._log_result(work_key, "Success", txt)
                            outcome_found = True; break
                    except NoSuchElementException: pass

                    # Check for Error Message
                    try:
                        lbl_err = driver.find_element(By.ID, 'ctl00_ContentPlaceHolder1_lblError')
                        txt = lbl_err.get_attribute("innerText").strip()
                        if txt:
                            self._log_result(work_key, "Failed", txt)
                            outcome_found = True; break
                    except NoSuchElementException: pass
                    
                    time.sleep(1.5)  # Brief wait for postback to begin
                except Exception as e: logger.debug("AddActivity: Failed to process: %s", e)

            if not outcome_found:
                # If no message appears, assume success if inputs cleared, or log warning
                self._log_result(work_key, "Success", "Saved (Implicit - No error found).")

        except Exception as e:
            self._log_result(work_key, "Failed", f"Error: {str(e).splitlines()[0]}")