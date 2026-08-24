# tabs/demand_tab.py
import tkinter
from tkinter import ttk, messagebox, filedialog, Toplevel
import customtkinter as ctk
import os, csv, time, threading, json, re, requests
import unicodedata
from datetime import datetime

from src import config
from .base_tab import BaseAutomationTab
from src.utils import get_logger
from src.i18n import tr
from typing import Any, Callable, Dict, List, Optional, Tuple
from ._imports import By, Keys, Select, WebDriverWait, EC, NoSuchElementException, StaleElementReferenceException, TimeoutException  # noqa: F401


logger = get_logger()

# --- Cloud File Picker Toplevel Window ---
class CloudFilePicker(ctk.CTkToplevel):
    """
    A Toplevel window to select a file from the user's cloud storage.
    """
    def __init__(self, parent: Any, app_instance: Any) -> None:
        """
        Initializes the Toplevel window for the cloud file picker.
        """
        super().__init__(parent)
        self.app = app_instance
        self.selected_file = None # This will store the {'id': ..., 'filename': ...} dict
        self.current_folder_id = None
        self.current_path_str = "/"
        self.history = [] # Stack to store (folder_id, path_str)
        self.work_data = []

        self.title("Select File from Cloud")
        w, h = 400, 500
        x = (self.winfo_screenwidth() // 2) - (w // 2)
        y = (self.winfo_screenheight() // 2) - (h // 2)
        self.geometry(f'{w}x{h}+{x}+{y}')
        self.resizable(False, False)
        self.transient(parent)
        self.attributes("-topmost", True)
        self.grab_set()

        self.protocol("WM_DELETE_WINDOW", self._on_close)
        
        self.grid_rowconfigure(2, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # Frame for back button and path
        self.header_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.header_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 5))
        
        self.back_button = ctk.CTkButton(self.header_frame, text=tr("form.demand.back"), width=60, command=self._go_back, state="disabled")
        self.back_button.pack(side="left")

        self.path_label = ctk.CTkLabel(self.header_frame, text=self.current_path_str, anchor="w")
        self.path_label.pack(side="left", fill="x", expand=True, padx=10)

        # Status label (e.g., "Loading...")
        self.status_label = ctk.CTkLabel(self, text=tr("common.loading_files"), text_color="gray")
        self.status_label.grid(row=1, column=0, sticky="ew", padx=10, pady=5)
        
        # Scrollable frame for file/folder list
        self.file_frame = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self.file_frame.grid(row=2, column=0, sticky="nsew", padx=10, pady=(0, 10))

        # Start loading files from the root
        threading.Thread(target=self._load_files, args=(None,), daemon=True).start()

    def _load_files(self, folder_id):
        """
        Fetches the list of files and folders from the cloud server
        for a given folder_id (or root if None).
        """
        self.after(0, self.status_label.configure, {"text": "Loading..."})
        self.after(0, self._clear_list)
        
        token = self.app.license_info.get('key')
        if not token:
            self.after(0, self.status_label.configure, {"text": "Error: Not authenticated."})
            return

        headers = {'Authorization': f'Bearer {token}'}
        base_url = f"{config.LICENSE_SERVER_URL}/files/api/list"
        url = f"{base_url}/{folder_id}" if folder_id else base_url
        
        try:
            resp = self.app.http_session.get(url, headers=headers, timeout=10)
            if not resp.ok:
                raise Exception(f"Server error: {resp.status_code}")
                
            data = resp.json()
            if data.get('status') == 'success':
                files = data.get('files', [])
                # Filter for folders and report files (CSV / Excel)
                display_items = [f for f in files if f['is_folder'] or f['filename'].lower().endswith(('.csv', '.xlsx'))]
                self.after(0, self._populate_list, display_items)
            else:
                raise Exception(data.get('reason', 'Failed to list files.'))
        except Exception as e:
            self.after(0, self.status_label.configure, {"text": f"Error: {e}"})

    def _populate_list(self, files):
        """
        Populates the scrollable frame with buttons for each file/folder.
        """
        self._clear_list()
        self.status_label.configure(text=tr("form.demand.select_file_or_folder"))
        
        if not files:
            ctk.CTkLabel(self.file_frame, text=tr("form.demand.no_csv_xlsx"), text_color="gray").pack(pady=10)
            return

        # Sort: Folders first, then by name
        files.sort(key=lambda x: (not x['is_folder'], x['filename'].lower()))

        for file_data in files:
            icon = "📁" if file_data['is_folder'] else "📄"
            btn_text = f"{icon} {file_data['filename']}"
            
            btn = ctk.CTkButton(
                self.file_frame, 
                text=btn_text, 
                anchor="w",
                fg_color="transparent",
                text_color=("gray10", "gray90"), # Theme-aware text color
                hover_color=ctk.ThemeManager.theme["CTkButton"]["fg_color"],
                command=lambda f=file_data: self._on_item_click(f)
            )
            btn.pack(fill='x', padx=5, pady=2)

    def _clear_list(self):
        """
        Removes all widgets from the file list frame.
        """
        for widget in self.file_frame.winfo_children():
            widget.destroy()

    def _on_item_click(self, file_data):
        """
        Handles clicks on a file or folder.
        If folder, navigates into it. If file, selects it and closes.
        """
        if file_data['is_folder']:
            # Save current state to history
            self.history.append((self.current_folder_id, self.current_path_str))
            
            # Update current state
            self.current_folder_id = file_data['id']
            self.current_path_str = f"{self.current_path_str}{file_data['filename']}/"
            
            # Update UI
            self.path_label.configure(text=self.current_path_str)
            self.back_button.configure(state="normal")
            
            # Load files for the new folder
            threading.Thread(target=self._load_files, args=(self.current_folder_id,), daemon=True).start()
        else:
            # This is a file, select it and close
            self.selected_file = file_data
            self.grab_release()
            self.destroy()
            
    def _on_close(self):
        """Handles the window being closed via the 'X' button."""
        self.grab_release()
        self.destroy()

    def _go_back(self):
        """
        Navigates to the previous folder in the history.
        """
        if not self.history:
            return
            
        # Restore previous state from history
        prev_folder_id, prev_path_str = self.history.pop()
        
        self.current_folder_id = prev_folder_id
        self.current_path_str = prev_path_str
        
        # Update UI
        self.path_label.configure(text=self.current_path_str)
        if not self.history:
            self.back_button.configure(state="disabled")
            
        # Load files for the parent folder
        threading.Thread(target=self._load_files, args=(self.current_folder_id,), daemon=True).start()

# --- End of CloudFilePicker Class ---


class DemandTab(BaseAutomationTab):
    """
    The main class for the "Demand" automation tab.
    """
    # Portal messages that mean the demand for this period ALREADY exists
    # (e.g. 'Demand of ARMAN ANSARI for period 10/08/2026-23/08/2026 is
    # already there and work is allotted'). These are NOT a new success —
    # they get their own 'Already' status (warning colour) and are excluded
    # from auto-allocation. The bare 'already' is a catch-all for other
    # states' wordings ('already done', 'already allotted for this period',
    # ...) — a genuine success message never contains it.
    ALREADY_PHRASES = ("already there", "already exists", "is already",
                       "already been", "already")

    def __init__(self, parent: Any, app_instance: Any) -> None:
        """
        Initializes the Demand automation tab.
        """
        super().__init__(parent, app_instance, automation_key="demand")
        # self.worker_thread = None <-- This is now managed by main_app
        self.csv_path = None # Stores the path to the *processed* file (local or temp)
        self.config_file = self.app.get_data_path("demand_inputs.json")

        self.all_applicants_data = [] # Holds all data from CSV

        self._suppress_search_refresh = False  # keeps the result list stable while ticking

        # ── Smoothness caches (job-card selection lag fix) ──
        self._search_after_id = None   # debounce timer for the search box
        self._jc_index = {}            # 'Job card number' -> [applicant dicts]
        self._ordered_jcs = []         # JC order as in the report (follow-JC rows)
        self._jc_pos = {}              # JC -> position in _ordered_jcs
        self._jc_row_cache = {}        # JC -> row_frame in Selected panel (incremental updates)
        self._jc_more_label = None     # '+N more selected job cards' label widget

        self._current_panchayat = ""  # active group while automation runs (results display)
        self._current_village = ""

        # Retry chain ke pehle-wale successful labourers (run 1) yahan carry
        # hote hain — retry run start_automation() results tree CLEAR kar deta
        # hai, isliye finally-block unhe is set se merge karta hai taaki
        # auto-allocation me SABHI successful jobcards aa saken (sirf retried
        # nahi). _handoff_allocation._proceed() par clear ho jata hai.
        self._retry_prior_success_names = set()
        self._skip_retry_clear = False  # _retry_failed_applicants ise True karta hai

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self._create_widgets()
        self.load_inputs()
    def _create_widgets(self) -> None:

        """
        Creates all the UI elements (buttons, entries, frames) for the tab.
        """
        # Main tab view
        notebook = ctk.CTkTabview(self)
        notebook.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

        settings_tab = notebook.add("Settings")
        select_tab = notebook.add("Select Jobcard")
        results_tab = notebook.add("Results")
        self._create_log_and_status_area(notebook)

        # Open on the 'Select Jobcard' tab by default — otherwise the user
        # lands on 'Settings' and the job-card list is never visible until
        # they click the second tab (this is why the list 'did not show' in
        # the UI even though search was working).
        notebook.set("Select Jobcard")

        settings_tab.grid_columnconfigure(0, weight=1)

        # Jobcard selection gets its own tab — full width & height for the panels
        select_tab.grid_columnconfigure(0, weight=1)
        select_tab.grid_rowconfigure(0, weight=1)

        # ── Header card ──
        self._create_header_card(settings_tab, "📥", tr("tab.demand.title"), tr("tab.demand.subtitle"),
                                 icon_key="emoji_demand")

        # Wrap ALL settings content in a bordered scrollable card
        settings_scroll = ctk.CTkScrollableFrame(settings_tab, corner_radius=12,
                                                 border_width=1, border_color=("gray85", "gray30"))
        settings_scroll.grid(row=1, column=0, sticky="nsew", padx=12, pady=6)
        settings_scroll.grid_columnconfigure(0, weight=1)

        # Only the run-settings controls live in Settings; job card selection
        # has its own 'Select Jobcard' tab.
        content = settings_scroll  # alias for clarity

        # --- Settings Tab Widgets ---
        controls_frame = ctk.CTkFrame(content)
        controls_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=10)
        # 4 columns for compact layout
        controls_frame.grid_columnconfigure((1, 3), weight=1)
        controls_frame.grid_columnconfigure((0, 2), weight=0)

        # --- Row 0: State and Panchayat ---
        ctk.CTkLabel(controls_frame, text=tr("common.state_label")).grid(row=0, column=0, padx=(10, 5), pady=5, sticky="w")
        self.state_var = ctk.StringVar()
        # Only the user's own state(s) — saved from license/settings — are shown,
        # like the panchayat dropdowns in the other automations. Falls back to
        # all configured states when nothing is saved yet.
        self.state_options = self._get_state_options()
        self.state_combobox = ctk.CTkOptionMenu(controls_frame, variable=self.state_var, values=self.state_options)
        self.state_combobox.grid(row=0, column=1, padx=(0, 10), pady=5, sticky="ew")

        ctk.CTkLabel(controls_frame, text=tr("common.panchayat_label")).grid(row=0, column=2, padx=(0, 5), pady=5, sticky="w")
        # ALL panchayats saved in Settings > Location Data (merged across keys),
        # like every other tab — NOT just the 'location_panchayat' key. The
        # 'All Panchayats'/'My Saved Panchayats' labels are filtered out (they
        # are dropdown entries from other tabs, not real panchayats).
        p_vals = self._demand_panchayat_options() or [""]
        self.panchayat_var = ctk.StringVar()
        self.panchayat_menu = ctk.CTkOptionMenu(controls_frame, variable=self.panchayat_var, values=p_vals)
        self.panchayat_menu.grid(row=0, column=3, padx=5, pady=5, sticky="ew")

        # --- Row 1: Demand Date (From) ---
        ctk.CTkLabel(controls_frame, text=tr("form.demand.work_demand_from")).grid(row=1, column=0, padx=(10, 5), pady=5, sticky="w")

        # Demand Date Frame
        d_date_frame = ctk.CTkFrame(controls_frame, fg_color="transparent")
        d_date_frame.grid(row=1, column=1, padx=(0, 10), pady=5, sticky="ew")
        self.demand_date_entry = ctk.CTkEntry(d_date_frame, placeholder_text=tr("common.date_format"))
        self.demand_date_entry.pack(side="left", fill="x", expand=True)
        ctk.CTkButton(d_date_frame, text="📅", width=30, fg_color=("gray85", "gray25"), text_color=("black", "white"),
                    command=lambda: self.open_date_picker(lambda d: [self.demand_date_entry.delete(0, "end"), self.demand_date_entry.insert(0, d)])).pack(side="right", padx=(5,0))

        # --- Row 2: Days and No. of Labour ---
        
        # Days Input
        ctk.CTkLabel(controls_frame, text=tr("form.demand.days")).grid(row=2, column=0, padx=(10, 5), pady=5, sticky="w")
        self.days_entry = ctk.CTkEntry(controls_frame, validate="key", validatecommand=(self.register(lambda P: P.isdigit() or P == ""), '%P'))
        self.days_entry.grid(row=2, column=1, padx=(0, 10), pady=5, sticky="ew")
        self.days_entry.insert(0, self.app.history_manager.get_suggestions("demand_days")[0] if self.app.history_manager.get_suggestions("demand_days") else "14")
        
        # No. of Labour (Custom Selection)
        ctk.CTkLabel(controls_frame, text=tr("form.demand.no_of_labour")).grid(row=2, column=2, padx=(0, 5), pady=5, sticky="w")
        
        custom_select_frame = ctk.CTkFrame(controls_frame, fg_color="transparent")
        custom_select_frame.grid(row=2, column=3, sticky="ew", padx=5, pady=5)
        custom_select_frame.grid_columnconfigure(0, weight=1) 
        
        # UPDATED: Removed width=70
        self.custom_select_entry = ctk.CTkEntry(custom_select_frame, validate="key", validatecommand=(self.register(lambda P: P.isdigit() or P == ""), '%P'), placeholder_text=tr("common.count_col"))
        self.custom_select_entry.grid(row=0, column=0, sticky="w", padx=(0, 5))
        
        self.custom_select_button = ctk.CTkButton(custom_select_frame, text=tr("common.select"), command=self._select_custom_number, width=70)
        self.custom_select_button.grid(row=0, column=1, sticky="e")
        
        # --- Row 3: Work Key (typeable input) ---
        # User types ONE work key here — selected workers are allocated to it
        # automatically after the demand is submitted. If the report CSV has a
        # per-worker 'Allocation Work Code' column, those codes take priority.
        ctk.CTkLabel(controls_frame, text=tr("form.demand.work_key")).grid(row=3, column=0, padx=(10, 5), pady=5, sticky="w")
        
        work_key_frame = ctk.CTkFrame(controls_frame, fg_color="transparent")
        work_key_frame.grid(row=3, column=1, columnspan=3, padx=5, pady=5, sticky="ew")
        work_key_frame.grid_columnconfigure(0, weight=1)
        
        self.allocation_work_key_var = ctk.StringVar()
        self.allocation_work_key_entry = ctk.CTkEntry(
            work_key_frame,
            textvariable=self.allocation_work_key_var,
            placeholder_text=tr("form.demand.workkey_placeholder"))
        self.allocation_work_key_entry.grid(row=0, column=0, sticky="ew")
        
        # --- END Row 3 ---

        # Applicant selection frame — lives in the dedicated 'Select Jobcard' tab
        applicant_frame = ctk.CTkFrame(select_tab)
        applicant_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        applicant_frame.grid_columnconfigure(0, weight=1)
        applicant_frame.grid_rowconfigure(2, weight=1)

        # --- Row 0: File buttons + file label ---
        file_row = ctk.CTkFrame(applicant_frame, fg_color="transparent")
        file_row.grid(row=0, column=0, sticky="ew", padx=10, pady=(8, 2))
        file_row.grid_columnconfigure(4, weight=1)

        self.select_csv_button = ctk.CTkButton(file_row, text=tr("form.demand.upload_report"), command=self._select_csv_from_computer)
        self.select_csv_button.grid(row=0, column=0, padx=(0, 6))

        # Load a previously-uploaded eKYC report for the same panchayat (stored
        # locally by panchayat — no need to re-upload every time).
        self.previous_reports_button = ctk.CTkButton(
            file_row, text=tr("form.demand.previous"), width=100,
            command=self._select_report_from_history,
            fg_color=config.COLORS["teal_named"], hover_color=config.COLORS["teal_hover"])
        self.previous_reports_button.grid(row=0, column=1, padx=(0, 6))

        self.demo_csv_button = ctk.CTkButton(file_row, text=tr("common.demo_csv"), command=lambda: self.app.save_demo_csv("demand"), fg_color=config.COLORS["btn_start"], hover_color=config.COLORS["green_button_hover"], width=90)
        self.demo_csv_button.grid(row=0, column=2, padx=(0, 6))

        self.select_all_button = ctk.CTkButton(file_row, text=tr("common.select_all"), command=self._select_all_applicants, width=90)
        self.select_all_button.grid(row=0, column=3, padx=(0, 6))

        self.clear_selection_button = ctk.CTkButton(file_row, text=tr("common.clear"), command=self._clear_selection, fg_color="gray", hover_color="gray50", width=60)
        self.clear_selection_button.grid(row=0, column=4, sticky="w")

        self.file_label = ctk.CTkLabel(file_row, text=tr("errors.no_file_loaded"), text_color="gray", anchor="w")
        self.file_label.grid(row=1, column=0, columnspan=5, sticky="w", pady=(2, 0))

        self.selection_summary_label = ctk.CTkLabel(file_row, text=tr("form.demand.applicants_selected"), text_color="gray", anchor="w")
        self.selection_summary_label.grid(row=2, column=0, columnspan=5, sticky="w")

        # --- Row 1: Quick-select bar ---
        qs_frame = ctk.CTkFrame(applicant_frame, fg_color=("gray85", "gray18"), corner_radius=8)
        qs_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=(4, 4))
        qs_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(qs_frame, text=tr("form.demand.quick_select"), font=ctk.CTkFont(weight="bold")).grid(
            row=0, column=0, padx=(10, 6), pady=6, sticky="w")

        self.quick_select_entry = ctk.CTkEntry(
            qs_frame,
            placeholder_text=tr("form.demand.jc_suffixes"))
        self.quick_select_entry.grid(row=0, column=1, padx=(0, 6), pady=6, sticky="ew")
        self.quick_select_entry.bind("<Return>", lambda e: self._quick_select_jcs())

        ctk.CTkButton(qs_frame, text=tr("common.add"), width=60, command=self._quick_select_jcs).grid(
            row=0, column=2, padx=(0, 6), pady=6)

        # Search row (below quick-select)
        # StringVar-trace based (like home_tab) — fires on EVERY text change,
        # far more reliable than a <KeyRelease> binding in packaged builds.
        self.search_var = ctk.StringVar()
        # Debounced: har keystroke par poori list rebuild nahi hoti — typing
        # rukne ke ~180ms baad ek hi baar refresh hota hai (smooth feel).
        self.search_var.trace_add("write", lambda *_: self._schedule_search_refresh())
        self.search_entry = ctk.CTkEntry(qs_frame, textvariable=self.search_var,
                                         placeholder_text=tr("form.demand.search_placeholder"))
        self.search_entry.grid(row=1, column=0, columnspan=3, padx=10, pady=(0, 6), sticky="ew")

        # --- Row 2: Selected-JC summary table + search results panel ---
        # Panels stretch to fill the tab's full height for easier browsing.
        # Left summary card compact rahta hai (~300px) — search ko baki width.
        bottom_frame = ctk.CTkFrame(applicant_frame, fg_color="transparent")
        bottom_frame.grid(row=2, column=0, sticky="nsew", padx=10, pady=(0, 8))
        bottom_frame.grid_columnconfigure(0, weight=0, minsize=280)
        bottom_frame.grid_columnconfigure(1, weight=1)
        bottom_frame.grid_rowconfigure(0, weight=1)

        # Left: selected JCs summary (scrollable internally) — narrow + compact
        self.selected_jc_frame = ctk.CTkScrollableFrame(
            bottom_frame, label_text="✅ Selected Job Cards", width=300)
        self.selected_jc_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 4))
        self.selected_jc_frame.grid_columnconfigure(0, weight=1)

        # Right: search results (scrollable internally)
        self.search_results_frame = ctk.CTkScrollableFrame(
            bottom_frame, label_text="🔍 Search Results")
        self.search_results_frame.grid(row=0, column=1, sticky="nsew", padx=(4, 0))
        self.search_results_frame.grid_columnconfigure(0, weight=1)

        # --- Results Tab Widgets ---
        # Configure row weights
        results_tab.grid_rowconfigure(0, weight=1) # Treeview
        results_tab.grid_rowconfigure(1, weight=0) # Button frame
        
        # Configure column weights
        results_tab.grid_columnconfigure(0, weight=1) # Treeview
        results_tab.grid_columnconfigure(1, weight=0) # Scrollbar

        # Treeview
        cols = ("#", "Panchayat", "Village", "Job Card No", "Applicant Name", "Status")
        self.results_tree = ttk.Treeview(results_tab, columns=cols, show='headings')
        self.results_tree.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        
        # Scrollbar
        vsb = ttk.Scrollbar(results_tab, orient="vertical", command=self.results_tree.yview)
        vsb.grid(row=0, column=1, sticky='ns')
        self.results_tree.configure(yscrollcommand=vsb.set)
        
        # ... inside _create_widgets method ...

        # Button Frame for Results
        results_button_frame = ctk.CTkFrame(results_tab, fg_color="transparent")
        results_button_frame.grid(row=1, column=0, columnspan=2, sticky="ew", padx=5, pady=(0, 5))

        # Left Side: Retry Button
        self.retry_failed_button = ctk.CTkButton(results_button_frame, text=tr("form.demand.retry_failed"), command=self._retry_failed_applicants)
        self.retry_failed_button.pack(side="left", padx=5)

        # Right Side: Export Button
        export_controls_frame = ctk.CTkFrame(results_button_frame, fg_color="transparent")
        export_controls_frame.pack(side='right', padx=(10, 0))

        self.export_button = ctk.CTkButton(export_controls_frame, text=tr("common.export_excel"), command=self.export_report)
        self.export_button.pack(side='left')

        # ── Action buttons (OUTSIDE the card) ──
        action_buttons = self._create_action_buttons(parent_frame=settings_tab)
        action_buttons.grid(row=2, column=0, sticky="ew", padx=12, pady=6)

        self._setup_results_treeview()

    def _select_all_applicants(self):
        """Selects ALL valid (no *) applicants from the loaded report/CSV.

        A demand CSV contains exactly the job cards the user wants demand for,
        so Select All selects every row — no batch cap (the old 400 limit made
        big CSVs silently skip job cards).
        """
        if not self.all_applicants_data:
            return
        selected_count = 0
        for app_data in self.all_applicants_data:
            if "*" not in app_data.get('Name of Applicant', ''):
                app_data['_selected'] = True
                selected_count += 1
        self._refresh_selected_jc_panel()
        self._update_selection_summary()
        self.log_info(f"Selected {selected_count} valid applicants (of {len(self.all_applicants_data)} loaded).")
    def _select_custom_number(self):
        """
        Selects a custom number of applicants from the top of the list.
        """
        if not self.all_applicants_data:
            messagebox.showwarning(tr("dialogs.no_data"), tr("dialogs.load_csv_first"))
            return

        try:
            num_to_select = int(self.custom_select_entry.get().strip())
        except ValueError:
            messagebox.showwarning(tr("dialogs.invalid_input"), tr("dialogs.valid_number_select"))
            return

        if num_to_select <= 0:
            messagebox.showwarning(tr("dialogs.invalid_input"), tr("dialogs.number_greater_zero"))
            return
            
        if num_to_select > len(self.all_applicants_data):
            num_to_select = len(self.all_applicants_data)
            messagebox.showinfo(tr("dialogs.adjustment"), tr("dialogs.selecting_max_applicants", count=num_to_select))

        self._clear_selection() # Clear any existing selection first

        selected_count = 0
        
        # Iterate through the master list and select the first 'num_to_select' valid entries
        for i, applicant_data in enumerate(self.all_applicants_data):
            if selected_count >= num_to_select:
                break
            
            # Check if the applicant is valid (no '*')
            if "*" not in applicant_data.get('Name of Applicant', ''):
                applicant_data['_selected'] = True
                selected_count += 1
            
        self._refresh_selected_jc_panel()
        self._update_selection_summary()
        self.log_info(f"Selected first {selected_count} valid applicants.")
        self._update_jc_header_counters()

    def _clear_processed_selection(self):
        """
        Deselects ONLY successfully processed applicants.
        Keeps 'Failed' or 'Skipped' applicants selected for easy retry.
        """
        self.log_info("Updating selection based on results...")        
        # 1. Collect Successful JobCard+Name pairs from results
        successful_pairs = set()
        for item in self.results_tree.get_children():
            values = self.results_tree.item(item)['values']
            # values = (RowID, Panchayat, Village, JC, Name, Status)
            if len(values) >= 6:
                jc = str(values[3]).strip()
                name = str(values[4]).strip()
                status = str(values[5]).lower()
                
                # Agar status me Success ya Already hai, tabhi uncheck karein
                if "success" in status or "already" in status:
                    successful_pairs.add((jc, name))

        # 2. Update Master Data
        deselected_count = 0
        for app_data in self.all_applicants_data:
            jc_no = app_data.get('Job card number', '').strip()
            app_name = app_data.get('Name of Applicant', '').strip()
            
            # Agar ye pair successful list me hai, to deselect karo
            if (jc_no, app_name) in successful_pairs:
                app_data['_selected'] = False
                deselected_count += 1
            # Warna selected rehne do (agar pehle se selected tha)

        # 3. Refresh panels
        self._refresh_selected_jc_panel()
        self._refresh_search_results()
        self._update_selection_summary()
        self.log_info(f"Deselected {deselected_count} successful applicants. Failed items remain checked.")
        self._update_jc_header_counters()

    def _select_csv_from_computer(self):
        """Opens a file dialog to select an eKYC report or a simple Demand CSV
        (Excel/CSV) and processes it."""
        path = filedialog.askopenfilename(
            title=tr("form.demand.select_ekyc_report"),
            filetypes=[("eKYC Report / Demand CSV", "*.xlsx *.csv"),
                       ("Excel Workbook", "*.xlsx"),
                       ("CSV", "*.csv")]
        )
        if not path:
            return
        self._process_input_file(path)

    def _demand_panchayat_options(self) -> List[str]:
        """Saved panchayats for the Demand dropdown, minus the special
        'All Panchayats' / 'My Saved Panchayats' labels. Those labels are real
        dropdown entries in other tabs and can end up saved in history — they
        are NOT selectable panchayats and would fail as 'Panchayat Not Found'
        if chosen here."""
        return [p for p in (self._get_saved_panchayats() or [])
                if p and not self._is_panchayat_label(p)]

    def _process_input_file(self, path):
        """
        Reads an eKYC & ABPS report (Excel/CSV) exported by the eKYC Report tab
        and populates self.all_applicants_data.

        eKYC report columns:  sno, panchayat, village, jobcard, name,
        abps_status, ekyc_status  (4 title/summary rows before the header row).

        Legacy demand CSVs (Job Card Number / Name of Applicant / Allocation
        Work Code) are still accepted for backward compatibility.
        """
        self.csv_path = path
        self.file_label.configure(text=os.path.basename(path))
        self.all_applicants_data = []

        try:
            rows, headers = self._read_table_file(path)
            if not headers:
                raise ValueError("File is empty or could not be read.")

            norm_headers = [str(h).lower().replace(" ", "").replace("_", "").replace(".", "") for h in headers]

            def find_col(*names: str) -> int:
                for n in names:
                    if n in norm_headers:
                        return norm_headers.index(n)
                return -1

            jc_idx    = find_col("jobcard", "jobcardnumber", "jobcardno", "jccode")
            name_idx  = find_col("name", "nameofapplicant", "workername", "applicantname", "nameofworker")
            pan_idx   = find_col("panchayat", "panchayatname")
            vill_idx  = find_col("village", "villagename")
            # Work-Allocation exports carry the work key per labourer — reuse it
            # as the auto-allocation code for the next demand.
            alloc_idx = find_col("allocationworkcode", "workcode", "workkey",
                                 "selectedworkcode", "workname")
            status_idx = find_col("status")

            if jc_idx == -1 or name_idx == -1:
                raise ValueError(
                    "Headers not recognized. Expected eKYC report columns "
                    "('jobcard', 'name') or legacy 'Job Card Number' + 'Name of Applicant'."
                )

            def cell(i: int) -> str:
                if i == -1 or i >= len(row):
                    return ""
                v = row[i]
                return str(v).strip() if v is not None else ""

            for row_num, row in enumerate(rows, 1):
                if not row:
                    continue
                name, job_card = cell(name_idx), cell(jc_idx)
                if not name or not job_card:
                    continue
                # Work-Allocation exports: only keep rows that were allocated
                # (a 'Status' column exists only in that format).
                if status_idx != -1:
                    st = cell(status_idx).strip().lower()
                    if st and not any(k in st for k in ("success", "done", "allocated", "already")):
                        continue
                self.all_applicants_data.append({
                    'original_index': row_num,
                    'Name of Applicant': name,
                    'Job card number': job_card,
                    'panchayat': cell(pan_idx),
                    'village': cell(vill_idx),
                    'allocation_work_code': cell(alloc_idx),
                    '_selected': False,
                })

            loaded_count = len(self.all_applicants_data)
            self.log_info(f"Loaded {loaded_count} applicants from '{os.path.basename(path)}'.")

            # Precompute the JC index — search follow-JC expansion and the
            # Selected panel use it for O(1) lookups instead of rescanning the
            # whole dataset on every keystroke/tick.
            self._rebuild_jc_index()

            # ── '?' corruption check ──
            # Excel me 'CSV (Comma delimited)' (ANSI) me save karne se Devanagari/
            # regional names '?' ban jate hain — wo data file me hi corrupt hai,
            # app se recover nahi hota. User ko turant clear message do taaki wo
            # file ko 'CSV UTF-8' me re-save kare. Sath hi '?' wale names ko
            # skip/match na karne ki wajah samajh aaye.
            q_names = []
            for a in self.all_applicants_data:
                nm = (a.get('Name of Applicant') or '').strip()
                if nm and '?' in nm:
                    q_names.append(nm)
            if q_names:
                sample = ", ".join(sorted(set(q_names))[:3])
                self.log_warning(
                    f"⚠️ CSV me {len(q_names)} name(s) me '?' mila (e.g. {sample}). "
                    "Ye tab hota hai jab file Excel me 'CSV (Comma delimited)' (ANSI) "
                    "me save ki gayi ho aur names Hindi/regional hoon. File ko Excel me "
                    "'CSV UTF-8 (Comma delimited)' format me re-save karke dobara upload karein "
                    "— portal grid se name match nahi hoga jab tak names sahi na hoon.")
                self.app.after(0, lambda: messagebox.showwarning(
                    tr("demand.encoding_warn_title"),
                    tr("demand.encoding_warn_msg", count=len(q_names), sample=sample)))

            # Legacy demand CSVs carry a per-worker 'Allocation Work Code' column —
            # when present, auto-allocation after demand uses those specific codes.
            try:
                with_work_code = sum(1 for a in self.all_applicants_data
                                     if (a.get('allocation_work_code') or '').strip())
                if with_work_code:
                    self.log_info(f"{with_work_code} workers have their own work codes in the report — "
                                  "auto-allocation will use them after demand.")
            except Exception:
                pass

            # Auto-suggest panchayats present in the report — MERGED with the
            # user's saved panchayats so the dropdown never loses entries.
            pans = sorted({a['panchayat'] for a in self.all_applicants_data if a.get('panchayat')})
            if pans:
                try:
                    cur = self.panchayat_var.get().strip()
                    merged = list(dict.fromkeys([p for p in pans + self._demand_panchayat_options() if p]))
                    self.panchayat_menu.configure(values=merged or pans)
                    # Report ka panchayat auto-select karo — purana saved value
                    # (jaise MATIYARA) ko override karo jab report dusre panchayat
                    # ka ho. Agar user ka current value report ke panchayat se
                    # milta hai, use wahi rehne do.
                    if not cur or not any(str(p).strip().upper() == cur.upper()
                                          for p in pans if str(p).strip()):
                        self.panchayat_var.set(pans[0])
                except Exception:
                    pass

            # Auto-detect the state from the report's job card prefixes, so the
            # dropdown shows (and pre-selects) the user's own state only.
            try:
                detected = self._detect_state_from_report()
                if detected and detected in getattr(self, 'state_options', []):
                    if self.state_var.get().strip() != detected:
                        self.state_var.set(detected)
                        self.log_info(f"State auto-detected from report: {detected}")
                elif detected:
                    # Detected state is not in the dropdown yet — offer it
                    opts = list(getattr(self, 'state_options', []))
                    if detected not in opts:
                        opts.append(detected)
                    self.state_options = opts
                    self.state_combobox.configure(values=opts)
                    self.state_var.set(detected)
                    self.log_info(f"State auto-detected from report: {detected}")
            except Exception:
                pass

            self._store_report_in_history(path)
            self._update_applicant_display()

        except Exception as e:
            messagebox.showerror(tr("dialogs.error_reading_report"), tr("dialogs.could_not_read_report", error=e))
            self.csv_path = None
            self.all_applicants_data = []
            self._jc_index = {}
            self._ordered_jcs = []
            self._jc_pos = {}
            self._jc_row_cache.clear()
            self._jc_more_label = None
            self.file_label.configure(text=tr("errors.no_file"))
            self._update_applicant_display()
            self._update_selection_summary()

    def _read_table_file(self, path: str):
        """Reads an xlsx or csv file, returning (data_rows, header_row)."""
        ext = os.path.splitext(path)[1].lower()
        if ext == ".xlsx":
            try:
                from openpyxl import load_workbook
                wb = load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                raw_rows = []
                for r in ws.iter_rows(values_only=True):
                    if any(v is not None and str(v).strip() for v in r):
                        raw_rows.append(["" if v is None else str(v) for v in r])
                wb.close()
            except Exception as e:
                raise ValueError(f"Excel file could not be opened ({e}).")
        else:
            raw_rows = []
            # Encoding sniffing: BOM wali files (UTF-8 BOM / UTF-16) pehle;
            # phir UTF-8; phir Windows ANSI (cp1252) — Excel 'CSV (Comma
            # delimited)' yehi use karta hai; last me latin-1 (catch-all).
            # NOTE: agar file Excel ANSI me save hui hai aur names Devanagari
            # hain, to '?' file me hi aa chuke hote hain — wo recover nahi
            # hote, isliye _process_input_file me '?' warning diya jata hai.
            with open(path, "rb") as f:
                head = f.read(4)
            encodings = ["utf-8-sig", "utf-8", "cp1252", "latin-1"]
            if head[:2] in (b"\xff\xfe", b"\xfe\xff"):
                encodings = ["utf-16"] + encodings
            for enc in encodings:
                try:
                    with open(path, mode="r", encoding=enc, newline="") as f:
                        reader = csv.reader(f)
                        raw_rows = [r for r in reader if any(c.strip() for c in r)]
                    break
                except (UnicodeDecodeError, UnicodeError):
                    continue
        return self._extract_table(raw_rows)

    @staticmethod
    def _extract_table(raw_rows: List[List[str]]):
        """
        Finds the header row (the first row that has jobcard + name columns)
        and returns (data_rows, header_row). Handles the eKYC report's leading
        title/summary rows automatically.
        """
        for i, row in enumerate(raw_rows):
            norm = [str(c).lower().replace(" ", "").replace("_", "").replace(".", "") for c in row]
            has_jc = any(n in ("jobcard", "jobcardnumber", "jobcardno", "jccode") for n in norm)
            has_name = any(n in ("name", "nameofapplicant", "workername", "applicantname", "nameofworker") for n in norm)
            if has_jc and has_name:
                return raw_rows[i + 1:], row
        return [], (raw_rows[0] if raw_rows else [])

    # ------------------------------------------------------------------
    # Report history — upload once, reload anytime (per panchayat)
    # ------------------------------------------------------------------
    def _report_history_path(self) -> str:
        return self.app.get_data_path("demand_reports.json")

    def _load_report_history(self) -> Dict[str, Dict[str, str]]:
        try:
            with open(self._report_history_path(), "r", encoding="utf-8") as f:
                data = json.load(f)
                return data if isinstance(data, dict) else {}
        except Exception:
            return {}

    def _store_report_in_history(self, path: str) -> None:
        """Remembers the uploaded eKYC report for its panchayat, so the user can
        reload it next time via the 'Previous' button. A new upload for the
        same panchayat replaces the old entry (job-card list updates)."""
        try:
            pans = [a.get('panchayat') for a in self.all_applicants_data if a.get('panchayat')]
            if not pans:
                # Simple demand CSVs carry no panchayat column — fall back to
                # the panchayat the user picks in the UI dropdown (if any).
                cur = self.panchayat_var.get().strip()
                if cur:
                    pans = [cur]
                else:
                    return
            panchayat = max(set(pans), key=pans.count)
            try:
                date_str = datetime.fromtimestamp(os.path.getmtime(path)).strftime("%d-%m-%Y")
            except Exception:
                date_str = datetime.now().strftime("%d-%m-%Y")
            data = self._load_report_history()
            data[panchayat] = {
                "path": os.path.abspath(path),
                "date": date_str,
                "label": os.path.basename(path),
            }
            db_path = self._report_history_path()
            os.makedirs(os.path.dirname(db_path) or ".", exist_ok=True)
            tmp_path = db_path + ".tmp"
            with open(tmp_path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2)
            os.replace(tmp_path, db_path)
            self.log_info(f"Report saved in history for panchayat: {panchayat} ({date_str})")
        except Exception as e:
            self.log_info(f"Could not save report history: {e}")

    def _select_report_from_history(self) -> None:
        """Shows a picker of previously-uploaded reports (by panchayat)."""
        data = self._load_report_history()
        if not data:
            messagebox.showinfo(tr("demand.no_history_title"), tr("demand.no_history_msg"))
            return

        win = ctk.CTkToplevel(self)
        win.title("Select Previous Report")
        win.geometry("520x400")
        win.transient(self)
        win.grab_set()
        win.grid_columnconfigure(0, weight=1)
        win.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(win, text=tr("form.demand.previous_reports"),
                     font=ctk.CTkFont(size=15, weight="bold")).grid(
            row=0, column=0, padx=12, pady=(12, 4), sticky="w")

        scroll = ctk.CTkScrollableFrame(win)
        scroll.grid(row=1, column=0, sticky="nsew", padx=12, pady=6)
        scroll.grid_columnconfigure(0, weight=1)

        def _load(pan: str) -> None:
            entry = data.get(pan, {})
            win.destroy()
            if entry.get('path') and os.path.exists(entry['path']):
                self._process_input_file(entry['path'])
            else:
                messagebox.showerror(tr("demand.file_missing_title"), tr("demand.file_missing_msg", path=entry.get('path', '?')))

        for pan, entry in sorted(data.items()):
            label = entry.get('label', '') or ''
            date_str = entry.get('date', '') or ''
            btn = ctk.CTkButton(
                scroll,
                text=f"🏘️ {pan}   ({date_str})\n{label}",
                anchor="w",
                fg_color="transparent", text_color=("gray10", "gray90"),
                hover_color=ctk.ThemeManager.theme["CTkButton"]["fg_color"],
                command=lambda p=pan: _load(p))
            btn.grid(row=len(scroll.winfo_children()), column=0, sticky="ew", padx=4, pady=3)

        ctk.CTkButton(win, text=tr("common.cancel"), width=90,
                      command=win.destroy).grid(row=2, column=0, padx=12, pady=(4, 10))

    # APPLICANT SELECTION — new fast approach

    def _quick_select_jcs(self):
        """
        Parses the quick-select entry (e.g. '1/5, 12/44, 10/150') and selects
        all valid members of those job cards from the loaded CSV.
        Supports both 'suffix-only' (12/44) and full JC number matching.
        """
        raw = self.quick_select_entry.get().strip()
        if not raw:
            return
        if not self.all_applicants_data:
            messagebox.showwarning(tr("dialogs.no_data"), tr("dialogs.load_csv_first"))
            return

        tokens = [t.strip() for t in re.split(r'[,\s]+', raw) if t.strip()]
        matched_jcs = set()

        for token in tokens:
            token_lower = token.lower()
            for app_data in self.all_applicants_data:
                jc = app_data.get('Job card number', '')
                # Match full JC or suffix after '/'
                suffix = jc.split('/')[-1] if '/' in jc else jc
                if token_lower == jc.lower() or token_lower == suffix.lower():
                    matched_jcs.add(jc)

        if not matched_jcs:
            self.log_warning(f"No JCs matched: {raw}")
            return

        added = 0
        for app_data in self.all_applicants_data:
            if app_data.get('Job card number') in matched_jcs:
                if "*" not in app_data.get('Name of Applicant', ''):
                    app_data['_selected'] = True
                    added += 1

        self.quick_select_entry.delete(0, "end")
        self._refresh_selected_jc_panel()
        self._update_selection_summary()
        self.log_info(f"Quick-selected {added} applicants from {len(matched_jcs)} JC(s).")

    # Distinct pastel tints per village so results from different villages are
    # easy to tell apart (light-mode colour, dark-mode colour).
    VILLAGE_TINTS = [
        ("#DCE9F7", "#1C2A3A"),   # blue
        ("#DDF0DD", "#1D331D"),   # green
        ("#FDEBD2", "#3A2F1A"),   # orange
        ("#EDE3F6", "#2E2338"),   # purple
        ("#FBE3E8", "#3A2026"),   # pink
        ("#D9F0F0", "#173332"),   # teal
    ]
    VILLAGE_ACCENTS = [
        ("#1F6FBF", "#6FB7FF"),   # blue
        ("#2E7D32", "#8BC48F"),   # green
        ("#C45A00", "#FFB74D"),   # orange
        ("#7B1FA2", "#CE93D8"),   # purple
        ("#C2185B", "#F48FB1"),   # pink
        ("#00796B", "#4DB6AC"),   # teal
    ]

    def _village_color(self, village):
        """Deterministic colour index for a village (stable across refreshes)."""
        idx = 0
        if village:
            idx = sum(ord(c) for c in village) % len(self.VILLAGE_TINTS)
        return self.VILLAGE_TINTS[idx], self.VILLAGE_ACCENTS[idx]

    def _refresh_panels(self):
        """Rebuilds both selection panels and the summary label.

        Runs via after() AFTER a checkbox command returns, so the clicked
        checkbox is never destroyed in the middle of its own callback
        (avoids 'invalid command name' Tcl errors in the app).
        """
        self._refresh_search_results()
        self._refresh_selected_jc_panel()
        self._update_selection_summary()

    # ────────────────────────────────────────────────────────────────
    # Smoothness helpers (job-card selection lag fix)
    # ────────────────────────────────────────────────────────────────

    def _rebuild_jc_index(self):
        """Precomputes JC -> applicant rows + report order once per file load.

        Search follow-JC expansion and the Selected panel use these for O(1)
        lookups instead of rescanning the whole dataset on every keystroke/tick.
        """
        index = {}
        ordered = []
        seen = set()
        for a in self.all_applicants_data:
            jc = a.get('Job card number', '')
            if not jc:
                continue
            index.setdefault(jc, []).append(a)
            if jc not in seen:
                seen.add(jc)
                ordered.append(jc)
        self._jc_index = index
        self._ordered_jcs = ordered
        self._jc_pos = {jc: i for i, jc in enumerate(ordered)}

    def _schedule_search_refresh(self):
        """Debounced search: cancels any pending refresh and re-arms a timer.

        Fast typing rebuilds the results list only once the user pauses
        (~180ms), so a keystroke never triggers a full widget rebuild by
        itself. Jab tick search-box ko clear karta hai (suppressed), tab
        rebuild schedule NAHI hota — list stable rehti hai.
        """
        if getattr(self, '_suppress_search_refresh', False):
            return
        if self._search_after_id is not None:
            try:
                self.after_cancel(self._search_after_id)
            except Exception:
                pass
        self._search_after_id = self.safe_after(180, self._run_search_refresh)

    def _run_search_refresh(self):
        self._search_after_id = None
        if self._is_alive():
            self._refresh_search_results()

    def _refresh_search_results(self):
        """
        Shows matching applicants in the right-hand search results panel.
        - Searches by name or job card number (single character is enough,
          so typing just a suffix digit like "5" or "1" also works).
        - Each result row is tinted by its village (same village = same colour).
        - When searching by number/suffix, the 2-3 job cards that follow the match
          in the report are also listed — people often raise demand for the
          neighbouring labour cards too.
        """
        if getattr(self, '_suppress_search_refresh', False):
            # A checkbox tick is clearing the search box — keep the current list
            # visible so the user can tick the next nearby job card too.
            return
        for w in self.search_results_frame.winfo_children():
            w.destroy()

        query = self.search_entry.get().strip().lower()

        if not self.all_applicants_data:
            ctk.CTkLabel(self.search_results_frame,
                         text=tr("form.demand.no_report"),
                         text_color="gray", justify="left").pack(padx=10, pady=16)
            return

        # Empty search box → show ALL loaded job cards so the user can see
        # (and tick) everything right away. Typing filters the list.
        if len(query) < 1:
            matches = list(self.all_applicants_data)
            header = (f"📋 All {len(self.all_applicants_data)} applicant(s) — "
                      "type to filter")
            ctk.CTkLabel(self.search_results_frame,
                         text=header, text_color=("gray20", "gray80"),
                         font=ctk.CTkFont(size=13, weight="bold"),
                         justify="left").pack(anchor="w", padx=8, pady=(6, 2))
        else:
            matches = [r for r in self.all_applicants_data if
                       query in r.get('Job card number', '').lower() or
                       query in r.get('Name of Applicant', '').lower()]

            if not matches:
                # Helpful empty-state: show how many applicants are loaded so the
                # user can tell a bad query from an empty/small report. Demo
                # files may not contain every job-card number.
                ctk.CTkLabel(self.search_results_frame,
                             text=(f"No match for '{query}'.\n"
                                   f"{len(self.all_applicants_data)} applicant(s) loaded.\n"
                                   f"Clear the search to see the full list."),
                             text_color="gray", justify="left").pack(padx=10, pady=16)
                return

        # When the query matches a JC number/suffix, also pull in the next 2-3
        # job cards of the SAME village that follow the matched JC in the report
        # order — people often raise demand for neighbouring labour cards too.
        # These 'following' cards are shown WITHOUT the village tint (grey) and
        # with a ↳ marker, so the exact matches stay easy to spot.
        # (Capped at 30 hits so a broad single-character search on a big report
        # stays fast — the visible list is capped below anyway. Uses the
        # prebuilt _jc_index/_ordered_jcs so it never rescans the whole
        # dataset on every keystroke.)
        exact_ids = {id(x) for x in matches}
        jc_hits = ([] if len(query) < 1 else
                   [r for r in matches if query in r.get('Job card number', '').lower()])
        if jc_hits:
            extra = []
            for r in jc_hits[:30]:
                jc = r.get('Job card number', '')
                vill = (r.get('village') or '').strip()
                jpos = self._jc_pos.get(jc)
                if jpos is None:
                    continue
                for njc in self._ordered_jcs[jpos + 1:jpos + 4]:
                    for a in self._jc_index.get(njc, []):
                        if (a.get('village') or '').strip() == vill:
                            extra.append(a)
            seen_ids = set(exact_ids)
            for e in extra:
                if id(e) not in seen_ids:
                    matches.append(e)
                    seen_ids.add(id(e))

        # Empty search shows the whole report — cap it lower than a filtered
        # search so the initial render stays instant on big eKYC reports.
        # Typing narrows the list; ticking works from any visible row.
        limit = 80 if len(query) < 1 else 120
        light_mode = ctk.get_appearance_mode().lower() == "light"

        def pick(pair):
            """Theme-aware (light, dark) tuple → concrete colour for tk widgets."""
            return pair[0] if light_mode else pair[1]

        for row in matches[:limit]:
            name = row.get('Name of Applicant', '')
            jc   = row.get('Job card number', '')
            vill = row.get('village', '') or ''
            is_disabled = "*" in name
            is_follow = id(row) not in exact_ids   # 'next 2-3 cards' row
            if is_follow:
                # Following cards: neutral grey, no village tint, ↳ marker.
                tint, accent = ("gray87", "gray23"), ("gray60", "gray45")
                text = f"↳  {jc}  –  {name}"
            else:
                tint, accent = self._village_color(vill)
                text = f"{jc}  –  {name}"

            var = ctk.StringVar(value="on" if row.get('_selected') else "off")

            def _toggle(data=row, v=var):
                data['_selected'] = (v.get() == "on")
                # Tick ke baad SIRF input box clean hota hai — search results
                # list WAISI HI rehti hai (reset nahi hoti) taaki user aas-pados
                # ka agla jobcard bhi wahi se tick kar sake. Naya number type
                # karte hi list khud filter ho jati hai.
                # Pending debounced refresh bhi cancel karo — tick ke baad list
                # kabhi rebuild NAHI hoti (pehle ye ~200-800ms 'render' feel
                # deta tha).
                if self._search_after_id is not None:
                    try:
                        self.after_cancel(self._search_after_id)
                    except Exception:
                        pass
                    self._search_after_id = None
                self._suppress_search_refresh = True
                try:
                    self.search_entry.delete(0, "end")
                except Exception:
                    pass
                finally:
                    self._suppress_search_refresh = False
                # Left panel ko poora rebuild karne ki jagah SIRF is JC ki row
                # update hoti hai (add/count/remove) — tick per instant feel.
                jc = data.get('Job card number', '')
                self.after(0, lambda: self._sync_jc_row(jc))
                self.after(0, self._update_selection_summary)
                self.after(0, self.search_entry.focus_set)

            # NATIVE tk widgets — CTk widgets ~50x slower to create (2.7ms vs
            # 0.05ms per row); 100+ rows wali rebuild wahi 'render' feel deti
            # thi. Look same: village tint background + coloured check box.
            row_bg = pick(tint)
            row_frame = tkinter.Frame(self.search_results_frame, bg=row_bg)
            row_frame.pack(anchor="w", padx=4, pady=2, fill="x")

            cb = tkinter.Checkbutton(
                row_frame,
                text=text,
                variable=var, onvalue="on", offvalue="off",
                command=_toggle,
                bg=row_bg, activebackground=row_bg,
                selectcolor=pick(accent),
                fg=pick(("gray10", "gray90")),
                activeforeground=pick(("gray10", "gray90")),
                anchor="w", justify="left",
                relief="flat", bd=0, highlightthickness=0,
                padx=8, pady=4,
            )
            if is_disabled:
                cb.configure(state="disabled", fg="gray50")
            cb.pack(fill="x")

        if len(matches) > limit:
            ctk.CTkLabel(self.search_results_frame,
                         text=(f"... {len(matches) - limit} more — "
                               "type to filter or use Select All"),
                         text_color="gray").pack(anchor="w", padx=6, pady=2)

    # How many JC rows the Selected panel renders at most (a Select All on a
    # very large eKYC report can select 1000+ job cards — the selection itself
    # stays complete; only the summary panel is capped).
    RENDER_CAP = 150

    def _refresh_selected_jc_panel(self):
        """
        Full rebuild of the left panel showing all currently selected JCs
        with applicant count and a ✕ remove button.

        Used for discrete actions (Select All, Clear, file load, de-select).
        Checkbox ticks use the cheaper _sync_jc_row() incremental path.
        """
        for w in list(self._jc_row_cache.values()):
            try:
                w.destroy()
            except Exception:
                pass
        self._jc_row_cache.clear()
        self._jc_more_label = None
        for w in self.selected_jc_frame.winfo_children():
            w.destroy()

        # Group selected applicants by JC (report order)
        selected_by_jc = {}
        for app_data in self.all_applicants_data:
            if app_data.get('_selected'):
                jc = app_data.get('Job card number', '')
                selected_by_jc.setdefault(jc, []).append(app_data)

        if not selected_by_jc:
            ctk.CTkLabel(self.selected_jc_frame,
                         text=tr("form.demand.nothing_selected"),
                         text_color="gray", justify="left").pack(padx=10, pady=20)
            return

        for jc, members in list(selected_by_jc.items())[:self.RENDER_CAP]:
            self._create_jc_row(jc, members)

        if len(selected_by_jc) > self.RENDER_CAP:
            self._show_jc_more_label(len(selected_by_jc) - self.RENDER_CAP)

    def _create_jc_row(self, jc, members):
        """Builds one JC row in the Selected panel and caches it for updates.

        Native tk widgets — CTk rows are ~50x slower to create and this panel
        rebuilds on Select All / Clear / de-select / run-finish. Compact:
        narrow card, small paddings, font 10, capped names (see
        _update_jc_row_label).
        """
        light_mode = ctk.get_appearance_mode().lower() == "light"
        row_bg = ("gray88", "gray22")[0 if light_mode else 1]
        row_frame = tkinter.Frame(self.selected_jc_frame, bg=row_bg)
        row_frame.pack(fill="x", padx=3, pady=1)

        tkinter.Button(
            row_frame, text="✕", width=2, relief="flat", bd=0, cursor="hand2",
            bg=row_bg, fg=("gray30", "gray80")[0 if light_mode else 1],
            activebackground=("gray70", "gray40")[0 if light_mode else 1],
            activeforeground=("gray30", "gray80")[0 if light_mode else 1],
            command=lambda j=jc: self._deselect_jc(j)
        ).pack(side="right", padx=(0, 3), pady=2)

        label = tkinter.Label(
            row_frame, text="", anchor="w", justify="left", wraplength=240,
            bg=row_bg, fg=("gray10", "gray90")[0 if light_mode else 1],
            font=("Segoe UI", 10))
        label.pack(side="left", fill="x", expand=True, padx=(6, 2), pady=2)

        row_frame._jc = jc
        row_frame._label = label
        self._jc_row_cache[jc] = row_frame
        self._update_jc_row_label(row_frame, members)

    def _update_jc_row_label(self, row_frame, members):
        """Refreshes the text of an existing JC row — compact 2-line layout.

        Line 1: '/suffix (count) 🏘️ village'
        Line 2: sirf pehle 3 names + '+K more' — 20+ members wale JC par bhi
                row chhota rehta hai (pehle saare names dikhte the → bahut
                lambi row).
        """
        jc = row_frame._jc
        suffix = jc.split('/')[-1] if '/' in jc else jc
        vill = next((m.get('village') for m in members if m.get('village')), "")
        line1 = f"/{suffix}  ({len(members)} person{'s' if len(members) > 1 else ''})"
        if vill:
            line1 += f"  🏘️ {vill}"

        all_names = [m.get('Name of Applicant', '') for m in members]
        if all_names:
            shown = all_names[:3]
            line2 = ", ".join(n for n in shown if n)
            extra = len(all_names) - len(shown)
            if extra > 0:
                line2 += f"  +{extra} more"
        else:
            line2 = ""

        label_text = line1 if not line2 else f"{line1}\n{line2}"
        try:
            row_frame._label.configure(text=label_text)
        except Exception:
            pass

    def _selected_members(self, jc):
        """Selected applicants belonging to a job card (index lookup + fallback)."""
        rows = self._jc_index.get(jc)
        if rows is None:
            return [a for a in self.all_applicants_data
                    if a.get('Job card number') == jc and a.get('_selected')]
        return [a for a in rows if a.get('_selected')]

    def _sync_jc_row(self, jc):
        """
        Incremental left-panel update after a checkbox tick — add/update/remove
        only THIS job card's row instead of rebuilding the whole panel.
        """
        if not self._is_alive():
            return
        members = self._selected_members(jc)
        row = self._jc_row_cache.get(jc)
        if not members:
            if row is not None:
                try:
                    row.destroy()
                except Exception:
                    pass
                del self._jc_row_cache[jc]
            self._update_jc_more_label()
            return
        if row is not None:
            self._update_jc_row_label(row, members)
        elif len(self._jc_row_cache) < self.RENDER_CAP:
            self._create_jc_row(jc, members)
        self._update_jc_more_label()

    def _count_selected_jcs(self):
        seen = set()
        for a in self.all_applicants_data:
            if a.get('_selected'):
                jc = a.get('Job card number', '')
                if jc:
                    seen.add(jc)
        return len(seen)

    def _show_jc_more_label(self, extra):
        if self._jc_more_label is None:
            self._jc_more_label = ctk.CTkLabel(
                self.selected_jc_frame,
                text="", text_color="gray", justify="left")
            self._jc_more_label.pack(padx=10, pady=6)
        self._jc_more_label.configure(text=f"... +{extra} more selected job cards")

    def _update_jc_more_label(self):
        """Keeps the '+N more selected job cards' hint in sync after ticks."""
        if not self._is_alive():
            return
        total = self._count_selected_jcs()
        shown = len(self._jc_row_cache)
        if total > shown:
            if self._jc_more_label is None:
                self._jc_more_label = ctk.CTkLabel(
                    self.selected_jc_frame,
                    text="", text_color="gray", justify="left")
                self._jc_more_label.pack(padx=10, pady=6)
            self._jc_more_label.configure(
                text=f"... +{total - shown} more selected job cards")
        elif self._jc_more_label is not None:
            try:
                self._jc_more_label.destroy()
            except Exception:
                pass
            self._jc_more_label = None

    def _deselect_jc(self, jc):
        """Removes all selections for a given job card."""
        for app_data in self.all_applicants_data:
            if app_data.get('Job card number') == jc:
                app_data['_selected'] = False
        self._refresh_selected_jc_panel()
        self._refresh_search_results()   # keep search results in sync
        self._update_selection_summary()

    def _update_applicant_display(self, event=None):
        """Compatibility shim — just refreshes both panels."""
        self._refresh_selected_jc_panel()
        self._refresh_search_results()

    # kept for set_ui_state / _clear_processed_selection compatibility
    def _update_jc_header_counters(self):
        pass  # no longer needed with the new panel

    def _create_applicant_checkbox(self, row_data, is_next_jc=False, parent_jc_frame=None):
        pass  # no longer used

    def _on_applicant_select(self, applicant_data, new_state):
        """Updates master data and refreshes panels."""
        applicant_data['_selected'] = (new_state == "on")
        self._refresh_selected_jc_panel()
        self._update_selection_summary()

    def _update_selection_summary(self):
        """
        Updates the label showing the count of selected applicants and unique job cards.
        """
        selected = [r for r in self.all_applicants_data if r.get('_selected', False)]
        unique_jcs = len(set(r.get('Job card number') for r in selected))
        self.selection_summary_label.configure(text=f"{len(selected)} applicants / {unique_jcs} unique job cards")

    def set_ui_state(self, running: bool):
        if not self._is_alive():
            return
        """
        Enables or disables UI elements based on whether automation is running.
        """
        self.set_common_ui_state(running)
        state = "disabled" if running else "normal"
        
        self.state_combobox.configure(state=state)
        self.panchayat_menu.configure(state=state)
        self.days_entry.configure(state=state)
        self.select_csv_button.configure(state=state)
        self.previous_reports_button.configure(state=state)
        self.search_entry.configure(state=state)
        self.quick_select_entry.configure(state=state)
        self.demand_date_entry.configure(state=state)
        self.select_all_button.configure(state=state)
        self.clear_selection_button.configure(state=state)
        self.allocation_work_key_entry.configure(state=state)
        self.retry_failed_button.configure(state=state)
        
        # New Export Controls
        self.export_button.configure(state=state)
    def _get_state_options(self):
        """
        States available for Demand (STATE_DEMAND_CONFIG), filtered to the user's
        own state(s) saved from license activation / other automations (history
        key 'location_state', stored UPPERCASE). Falls back to ALL configured
        states when nothing has been saved yet.
        """
        demand_config = config.get_state_demand_config()
        upper_to_key = {k.upper(): k for k in demand_config.keys()}
        opts: List[str] = []
        try:
            for s in (self.app.history_manager.get_suggestions("location_state") or []):
                key = upper_to_key.get(str(s).strip().upper())
                if key and key not in opts:
                    opts.append(key)
        except Exception:
            pass
        return opts or list(demand_config.keys())

    def _detect_state_from_report(self):
        """
        Detects the state from the job card prefixes in the loaded report
        (e.g. 'JH-22-003-008-001/1' -> 'Jharkhand'). Returns a STATE_DEMAND_CONFIG
        key, or None.
        """
        prefixes = config.get_state_job_card_prefixes()
        demand_config = config.get_state_demand_config()
        for app_data in self.all_applicants_data:
            jc = (app_data.get('Job card number') or '').strip().upper()
            for prefix, state_key in prefixes.items():
                if jc.startswith(prefix.upper()) and state_key in demand_config:
                    return state_key
        return None

    def _get_village_code(self, job_card, state_logic_key):
        """
        Extracts the village code from a job card number based on state-specific logic.
        """
        try:
            jc = job_card.split('/')[0]
            if state_logic_key == "jh": return jc.split('-')[-1]
            elif state_logic_key == "rj": return jc[-3:]
            else: self.log_info(f"Warn: Unknown state logic '{state_logic_key}'."); return jc.split('-')[-1]
        except IndexError: return None

    # Add this method inside DemandTab class
    def load_csv_data(self, file_path):
        """
        Naya function: Ye specifically Macro ke liye hai taaki bina button dabaye
        CSV file load aur parse ho jaye.
        """
        if not file_path:
            return

        try:
            # 1. Variable Update
            self.selected_file_path = file_path
            
            # 2. UI Entry Update
            if hasattr(self, 'file_entry'):
                self.file_entry.delete(0, tkinter.END)
                self.file_entry.insert(0, file_path)

            # 3. Data Parse & Treeview Update
            # (Agar aapke paas Treeview hai to use clear karein)
            if hasattr(self, 'tree'):
                self.tree.delete(*self.tree.get_children())
            
            # Data store karne ke liye list
            self.work_data = [] 

            # Encoding sniffing — _read_table_file jaisa hi (utf-16 BOM,
            # UTF-8, cp1252/ANSI, latin-1). Nahi to ANSI-saved CSV me
            # Devanagari names silently '?' ban jate hain.
            # NOTE: work_data ko har attempt me fresh karo — ek encoding beech
            # me UnicodeDecodeError de to partial rows duplicate na ho jayein.
            with open(file_path, "rb") as f:
                _head = f.read(4)
            _encodings = ["utf-8-sig", "utf-8", "cp1252", "latin-1"]
            if _head[:2] in (b"\xff\xfe", b"\xfe\xff"):
                _encodings = ["utf-16"] + _encodings
            for _enc in _encodings:
                _batch = []
                try:
                    with open(file_path, newline='', encoding=_enc) as f:
                        reader = csv.DictReader(f)
                        for row in reader:
                            # Row clean karein
                            clean_row = {k.strip(): v.strip() for k, v in row.items() if k}
                            _batch.append(clean_row)
                            
                            # Treeview me dikhayein (Optional, visual confirmation ke liye)
                            if hasattr(self, 'tree'):
                                # Sirf values extract karke tree me dalein
                                self.tree.insert("", "end", values=list(clean_row.values()))
                    self.work_data = _batch
                    break
                except (UnicodeDecodeError, UnicodeError):
                    continue
            
            self.log_info(f"Loaded {len(self.work_data)} rows from {file_path} (macro)")

        except Exception as e:
            self.log_warning(f"Error loading CSV data: {e}")
            # Error aane par bhi kam se kam path variable set rakhein
            self.selected_file_path = file_path
    
    def set_automation_inputs(self, panchayat_name, file_path):
        """
        Updated: Loads CSV and selects applicants IMMEDIATELY (Synchronously).
        Removes timing delays to prevent 'Select applicants' error.
        """
        self.log_info(f"Auto-Setup: Panchayat={panchayat_name}, File={file_path}")
        
        # 1. Panchayat Name Set
        self.panchayat_var.set(panchayat_name)
        
        # 2. File Load & Immediate Selection
        if file_path and os.path.exists(file_path):
            # Step A: Load Data
            self._process_input_file(file_path)
            
            # Step B: Select All IMMEDIATELY (Do not use self.after)
            # यह सुनिश्चित करता है कि start_automation चलने से पहले डेटा तैयार है
            self._select_all_applicants()
            
            self.log_info(f"Auto-Setup: Loaded & Selected {len(self.all_applicants_data)} applicants.")
        else:
            self.log_error(f"Macro Error: File not found {file_path}")
    def start_automation(self) -> None:
        """
        Validates all user inputs and starts the main automation thread
        using the app's built-in thread manager (which plays sound).
        """
        # FRESH run par carry-over set clear karo (retry chain ka data agle
        # run me leak na ho). _retry_failed_applicants ise skip karwata hai.
        skip_retry_clear = getattr(self, '_skip_retry_clear', False)
        self._skip_retry_clear = False
        if not skip_retry_clear:
            self._retry_prior_success_names = set()

        # --- 1. Get and Validate Inputs ---
        state = self.state_var.get()
        if not state: messagebox.showerror(tr("errors.input_error"), tr("dialogs.select_state")); return
        try:
            cfg = config.get_state_demand_config()[state]
            logic_key = cfg.get("village_code_logic", "jh")
            # State-aware: base_url ko user ke state ke portal host par resolve
            # karo (Rajasthan → vbgramgde3). Same-state me koi change nahi.
            url = self.resolve_portal_url(cfg["base_url"])
        except KeyError:
            messagebox.showerror(tr("dialogs.config_error"), tr("dialogs.config_missing_state", state=state)); return

        selected = [r for r in self.all_applicants_data if r.get('_selected', False)]
        panchayat = self.panchayat_var.get().strip(); days_str = self.days_entry.get().strip()
        work_key_for_allocation = self.allocation_work_key_var.get().strip()
        
        demand_to_date_str = ""  # Override date feature removed

        try: 
            demand_dt_str = self.demand_date_entry.get()
            demand_dt = datetime.strptime(demand_dt_str, '%d/%m/%Y').date() 
            work_start = demand_dt.strftime('%d/%m/%Y') 
        except ValueError: messagebox.showerror(tr("dialogs.invalid_date"), tr("dialogs.use_dd_mm_yyyy")); return

        if demand_dt < datetime.now().date():
            messagebox.showerror(tr("dialogs.invalid_date"), tr("dialogs.no_past_date"))
            return

        if not days_str: messagebox.showerror(tr("dialogs.missing_info"), tr("dialogs.days_required")); return
        if not self.csv_path: messagebox.showerror(tr("dialogs.missing_info"), tr("dialogs.load_ekyc_first")); return
        if not selected: messagebox.showwarning(tr("dialogs.no_selection"), tr("dialogs.select_applicants")); return
        try: days_int = int(days_str); assert days_int > 0
        except (ValueError, AssertionError): messagebox.showerror(tr("dialogs.invalid_input"), tr("dialogs.days_positive")); return

        # --- 2. Setup UI for Running State ---
        # self.stop_event.clear(); <-- Handled by app.start_automation_thread
        self.app.clear_log(self.log_display)
        for i in self.results_tree.get_children(): self.results_tree.delete(i)
        self.log_info(f"Starting demand: {len(selected)} applicant(s), State: {state}...")
        if work_key_for_allocation:
            self.log_info(f"   -> Auto-allocation is ENABLED for Work Key: {work_key_for_allocation}")
        
        # self.app.set_status("Running..."); <-- Handled by app.start_automation_thread
        self.set_ui_state(running=True) # Disable UI elements

        # --- 3. Save History and Group Data ---
        # Labels (All/My Saved) history mein kabhi save nahi hote — demand ka
        # dropdown unhe filter karta hai, par saved data se bhi leak na ho.
        self._update_panchayat_history(panchayat)
        if state and state in getattr(self, 'state_options', []):
            self.app.history_manager.save_entry("location_state", state.upper())
        self.app.history_manager.save_entry("demand_days", days_str)
        self.save_inputs({
            "state": state, 
            "panchayat": panchayat, 
            "demand_date": demand_dt_str, 
            "days": days_str, 
            "work_key_for_allocation": work_key_for_allocation
        })

        # Simple demand CSVs carry no panchayat column — the panchayat is only
        # known once the user picks it in the UI, so store the uploaded report
        # in history NOW (run ke baad save) for the 'Previous' button. Reports
        # that already have their own panchayat column were stored at upload.
        try:
            if self.csv_path:
                self._store_report_in_history(self.csv_path)
        except Exception:
            pass

        # Group selected applicants by Panchayat -> Village -> Job Card.
        # Village names come straight from the eKYC report; legacy CSVs fall
        # back to the village code parsed from the job card number.
        grouped = {}; skipped_malformed = 0
        for app in selected:
            jc = app.get('Job card number', '').strip()
            if not jc: continue
            pan = (app.get('panchayat') or '').strip() or panchayat
            vill = (app.get('village') or '').strip()
            if not vill:
                vill = self._get_village_code(jc, logic_key) or ""
            if not vill:
                skipped_malformed += 1
                continue
            grouped.setdefault(pan, {}).setdefault(vill, {}).setdefault(jc, []).append(app)
        if skipped_malformed: self.log_warning(f"Warn: Skipped {skipped_malformed} Job Cards (no village info).")
        # --- 4. Start Worker Thread using the App's Method ---
        # This will play the sound and manage the thread
        args_tuple = (
            state, grouped, days_int, work_start, url, work_key_for_allocation
        )
        self.app.start_automation_thread(
            key=self.automation_key,
            target=self._process_demand,
            args=args_tuple
        )
    def reset_ui(self) -> None:
        """Resets all inputs, selections, and logs on the tab."""
        if not messagebox.askokcancel(tr("dialogs.reset_question"), tr("dialogs.clear_inputs_selections_logs")): return
        self.state_var.set("")
        self.panchayat_var.set("")
        self.days_entry.delete(0, 'end')
        self.search_entry.delete(0, 'end')
        self.quick_select_entry.delete(0, 'end')
        self.allocation_work_key_var.set("")
        self.demand_date_entry.delete(0, 'end')

        self.csv_path = None
        self.all_applicants_data.clear()
        if self._search_after_id is not None:
            try:
                self.after_cancel(self._search_after_id)
            except Exception:
                pass
            self._search_after_id = None
        self._jc_index = {}
        self._ordered_jcs = []
        self._jc_pos = {}
        for w in list(self._jc_row_cache.values()):
            try:
                w.destroy()
            except Exception:
                pass
        self._jc_row_cache.clear()
        self._jc_more_label = None
        self._current_panchayat = ""
        self._current_village = ""
        self.file_label.configure(text=tr("errors.no_file_loaded"), text_color="gray")

        self._refresh_selected_jc_panel()
        self._refresh_search_results()
        self._update_selection_summary()
        for i in self.results_tree.get_children(): self.results_tree.delete(i)
        self.app.clear_log(self.log_display)
        self.app.after(0, self.app.set_status, "Ready")
        self.log_info("Form reset.")
    def _setup_results_treeview(self):
        """
        Configures the columns and headings for the results table.
        """
        cols = ("#", "Panchayat", "Village", "Job Card No", "Applicant Name", "Status")
        self.results_tree["columns"] = cols
        self.results_tree.column("#0", width=0, stretch=tkinter.NO); self.results_tree.column("#", anchor='c', width=40)
        self.results_tree.column("Panchayat", anchor='w', width=130); self.results_tree.column("Village", anchor='w', width=110)
        self.results_tree.column("Job Card No", anchor='w', width=180); self.results_tree.column("Applicant Name", anchor='w', width=150)
        self.results_tree.column("Status", anchor='w', width=250)
        self.results_tree.heading("#0", text=""); self.results_tree.heading("#", text="#")
        self.results_tree.heading("Panchayat", text=tr("common.panchayat_col")); self.results_tree.heading("Village", text=tr("common.village_col"))
        self.results_tree.heading("Job Card No", text=tr("common.jobcard_no_col")); self.results_tree.heading("Applicant Name", text=tr("common.applicant_name_col"))
        self.results_tree.heading("Status", text=tr("common.status_col"))
        self.style_treeview(self.results_tree)

    def _process_demand(self, state, grouped, user_days, demand_from, base_url, work_key_for_allocation):
        """
        The main automation function that runs in a thread.

        Groups are structured as:  panchayat -> village -> job card -> [applicants].
        Villages are selected by NAME (from the eKYC report) with a village-code
        fallback for legacy CSVs. Every dropdown change triggers an ASP.NET
        postback — we always wait for the dependent list to populate before
        continuing (see _wait_dropdown_populated).
        """
        self._demand_error = ""   # AUDIT FIX: cleared at run start; set only by the FATAL handler
        driver = None
        try:
            driver = self.app.get_driver()
            if not driver:
                self.app.after(0, self.app.log_message, self.log_display, "ERROR: WebDriver unavailable.", "error")
                return
            wait = WebDriverWait(driver, 20)

            # Element IDs that vary slightly across state portals
            p_ids = ["ctl00_ContentPlaceHolder1_DDL_panchayat", "ctl00_ContentPlaceHolder1_ddlPanchayat"]
            v_ids = ["ctl00_ContentPlaceHolder1_DDL_Village", "ctl00_ContentPlaceHolder1_ddlvillage"]
            j_ids = ["ctl00_ContentPlaceHolder1_DDL_Registration", "ctl00_ContentPlaceHolder1_ddlJobcard"]
            days_worked_ids = ["ctl00_ContentPlaceHolder1_Lbldays"]
            grid_ids = ["ctl00_ContentPlaceHolder1_gvData", "ctl00_ContentPlaceHolder1_GridView1"]
            btn_ids = ["ctl00_ContentPlaceHolder1_btnProceed", "ctl00_ContentPlaceHolder1_btnSave"]

            total_p, p_idx = len(grouped), 0
            for panchayat, villages in grouped.items():
                p_idx += 1
                if self.is_stopped():
                    break
                try:
                    # Fresh page load per panchayat keeps dropdown state clean
                    driver.get(base_url)
                    self._current_panchayat = panchayat
                    self.app.after(0, self.app.log_message, self.log_display,
                                   f"State: {state} — Panchayat {p_idx}/{total_p}: {panchayat}")
                    self.app.after(0, self.app.set_status, f"P {p_idx}/{total_p}: {panchayat}")

                    # --- Panchayat: Block/PO login (dropdown) select karta hai;
                    # Panchayat/GP login (no dropdown) SKIP — central helper
                    # dono ko har tab me ek jaisa handle karta hai. ---
                    self.app.after(0, self.app.set_status, f"Selecting Panchayat: {panchayat}")
                    status, page_panchayat = self._select_panchayat_or_skip(
                        driver, wait, panchayat, p_ids, v_ids,
                        label_ids=["ctl00_ContentPlaceHolder1_panch"])
                    if status == "notfound":
                        self.app.after(0, self.app.log_message, self.log_display,
                                       f"ERROR: Panchayat '{panchayat}' not found in dropdown. Skipping this panchayat.", "error")
                        for vill, jcs in villages.items():
                            for jc, apps in jcs.items():
                                for a in apps:
                                    self.app.after(0, self._update_results_tree,
                                                   (jc, a.get('Name of Applicant'), "FAIL: Panchayat Not Found", panchayat, vill))
                        continue
                    if status == "missing":
                        self.app.after(0, self.app.log_message, self.log_display,
                                       "ERROR: Panchayat name required for Block Login.", "error")
                        continue
                    # GP login: panchayat ka naam page par text me hota hai — use
                    # record karo taaki settings > location data me save ho.
                    if status == "gp" and not panchayat and page_panchayat:
                        panchayat = page_panchayat
                        self._current_panchayat = panchayat
                    # Panchayat + villages Settings > Location Data me auto-add
                    # (GP users ke liye bhi — bina panchayat dropdown select
                    # kiye unka panchayat aur villages add ho jate hain).
                    self._save_panchayat_villages_to_settings(panchayat, list(villages.keys()))

                    # --- Loop through villages ---
                    total_v, v_idx = len(villages), 0
                    for village, jcs_in_v in villages.items():
                        v_idx += 1
                        if self.is_stopped():
                            break
                        self._current_village = village
                        try:
                            self.app.after(0, self.app.set_status,
                                           f"P {p_idx}/{total_p}, V {v_idx}/{total_v}: {village}")
                            self.app.after(0, self.app.log_message, self.log_display,
                                           f"--- Village {v_idx}/{total_v}: {village} ---")
                            if not self._select_village(driver, wait, village, v_ids):
                                raise NoSuchElementException(f"Village '{village}' not found in dropdown.")
                            self._wait_dropdown_populated(driver, wait, j_ids, "job cards after village selection")

                            # --- Loop through job cards in the village ---
                            total_jc, jc_idx = len(jcs_in_v), 0
                            for jc, apps in jcs_in_v.items():
                                jc_idx += 1
                                if self.is_stopped():
                                    break
                                progress = (p_idx - 1 + (v_idx - 1 + jc_idx / total_jc) / total_v) / total_p
                                self.app.after(0, self.update_status,
                                               f"P{p_idx}/{total_p} V{v_idx}/{total_v} JC{jc_idx}/{total_jc}", progress)
                                self.app.after(0, self.app.set_status,
                                               f"P{p_idx}/{total_p} V{v_idx}/{total_v} JC{jc_idx}/{total_jc}: {jc.split('/')[-1]}")
                                self._process_single_job_card(driver, wait, jc, apps,
                                                              user_days, demand_from, days_worked_ids,
                                                              j_ids, grid_ids, btn_ids,
                                                              p_ids, v_ids, base_url)
                        except Exception as e:
                            self.app.after(0, self.app.log_message, self.log_display,
                                           f"ERROR Village {village}: {type(e).__name__} - {e}. Skipping.", "error")
                            for jc_err, apps_err in jcs_in_v.items():
                                for app_data_err in apps_err:
                                    self.app.after(0, self._update_results_tree,
                                                   (jc_err, app_data_err.get('Name of Applicant'), "Skipped (Village Error)", panchayat, village))
                            continue

                except Exception as e:
                    self.app.after(0, self.app.log_message, self.log_display,
                                   f"ERROR Panchayat {panchayat}: {type(e).__name__} - {e}. Skipping.", "error")
                    for vill, jcs in villages.items():
                        for jc, apps in jcs.items():
                            for a in apps:
                                self.app.after(0, self._update_results_tree,
                                               (jc, a.get('Name of Applicant'), "Skipped (Panchayat Error)", panchayat, vill))
                    continue

            if not self.is_stopped():
                self.app.after(0, self.app.log_message, self.log_display, "✅ All processed.")

        except Exception as e:
            # AUDIT FIX (25 Aug 2026): record the fatal error on self — the old
            # `elif 'e' in locals():` check in the finally block was ALWAYS
            # False (Python deletes the except-name when the block exits), so
            # crashed runs reported "Finished" and still triggered the
            # allocation handoff.
            self._demand_error = type(e).__name__
            self.app.after(0, self.app.log_message, self.log_display, f"CRITICAL ERROR: {type(e).__name__} - {e}", "error")
            self.app.after(0, self.update_status, f"Error: {type(e).__name__}", 0.0)
            self.app.after(0, lambda: messagebox.showerror(tr("dialogs.error"), tr("dialogs.automation_stopped", error=e)))
        finally:
            final_status_text = "Finished"
            final_tab_status = "Finished"
            final_progress = 1.0

            if self.is_stopped():
                self.app.after(0, self.app.log_message, self.log_display, "Stopped by user.", "warning")
                final_status_text = "Stopped"
                final_tab_status = "Stopped"
            elif getattr(self, '_demand_error', ''):
                # AUDIT FIX (25 Aug 2026): was `elif 'e' in locals():` which is
                # ALWAYS False in Python 3 (the except-name is deleted when the
                # block exits) → crashed runs reported "Finished" and fell into
                # the handoff branch below. The explicit flag fixes both.
                final_status_text = f"Error: {getattr(self, '_demand_error', '')}"
                final_tab_status = f"Error: {getattr(self, '_demand_error', '')}"
                final_progress = 0.0
            else:
                # --- INTELLIGENT HANDOFF LOGIC (auto work allocation) ---
                if not self.is_stopped():
                    success_names, failed_names = set(), set()
                    for item_id in self.results_tree.get_children():
                        values = self.results_tree.item(item_id)['values']
                        status = str(values[5]).lower() if len(values) > 5 else ""
                        name = str(values[4]).strip() if len(values) > 4 else ""
                        if not name:
                            continue
                        if "success" in status:
                            success_names.add(name)
                        elif "fail" in status or "error" in status:
                            failed_names.add(name)

                    # Retry run: start_automation() ne tree clear kar diya tha,
                    # isliye run 1 ke successful names tree me nahi hain — unhe
                    # carry-over set se merge karo taaki auto-allocation me
                    # SABHI successful jobcards aa saken (sirf retried nahi).
                    if self._retry_prior_success_names:
                        carried = self._retry_prior_success_names - success_names
                        if carried:
                            success_names |= carried
                            self.log_info(
                                f"↻ Retry: {len(carried)} pehle-successful labourer(s) "
                                f"bhi allocation me shamil kiye.")

                    panchayat = getattr(self, '_current_panchayat', '') or ""

                    # Per-worker work codes (legacy CSV 'Allocation Work Code')
                    # take priority; workers without their own code go to the
                    # typed work key.
                    allocation_map = {}
                    has_specific_codes = False
                    for app_data in self.all_applicants_data:
                        name = app_data.get('Name of Applicant', '').strip()
                        if name in success_names:
                            specific_code = app_data.get('allocation_work_code', '').strip()
                            if specific_code:
                                has_specific_codes = True
                                allocation_map.setdefault(specific_code, []).append(name)
                            elif work_key_for_allocation:
                                allocation_map.setdefault(work_key_for_allocation, []).append(name)

                    # A typed global work key (no per-worker codes) gets its
                    # successful labourers listed under that key too, so the
                    # handoff ALWAYS passes the labourer names — Work Allocation
                    # selects ONLY those workers, never 'Allocate All'.
                    if not allocation_map:
                        self.log_info("📊 Demand automation finished (no successful labourers to allocate).")
                    else:
                        self._handoff_allocation(panchayat, failed_names, allocation_map)

                self.app.after(0, self._clear_processed_selection)

            self.app.after(0, self.set_ui_state, False)
            self.app.after(0, self.app.set_status, final_status_text)
            self.app.after(0, self.update_status, final_tab_status, final_progress)

            if not self.is_stopped() and not getattr(self, '_demand_error', ''):
                self.app.after(5000, lambda: self.app.set_status("Ready"))
                self.app.after(5000, lambda: self.update_status("Ready", 0.0))

    def _handoff_allocation(self, panchayat: str, failed_names: set,
                            allocation_map: Dict) -> None:
        """Triggers auto-allocation after demand. If any labourer's demand
        failed, asks the user first (retry vs proceed) before allocating.
        allocation_map (work_code -> labourer names) contains ONLY the workers
        whose demand succeeded, so Work Allocation selects exactly those —
        never 'Allocate All'. An empty map means nothing succeeded, so no
        allocation happens."""
        def _proceed() -> None:
            # Carry-over set ab consume ho chuka hai (allocation map usi se
            # bana) — clear karo taaki agla FRESH run ise dobara na uthaye.
            self._retry_prior_success_names = set()
            if allocation_map:
                count = sum(len(v) for v in allocation_map.values())
                self.app.after(0, self.app.log_message, self.log_display,
                               f"✅ Triggering Auto-Allocation for {count} laborers across {len(allocation_map)} work codes.")
                self.app.after(500, self.app.run_work_allocation_from_demand,
                               panchayat, allocation_map)
            else:
                # No labourer names are known (nothing succeeded) — never fall
                # back to 'Allocate All', which would allocate EVERY worker of
                # the panchayat even though none of the demands succeeded.
                self.app.after(0, self.app.log_message, self.log_display,
                               "ℹ️ No successful labourers to allocate — skipping allocation.")

        if failed_names:
            sample = "\n".join(sorted(failed_names)[:6])
            more = "\n..." if len(failed_names) > 6 else ""
            msg = tr("demand.failures_msg", count=len(failed_names), sample=sample, more=more)
            self.app.after(0, lambda: self._ask_failed_handoff(msg, _proceed))
        else:
            _proceed()

    def _ask_failed_handoff(self, msg: str, proceed: Callable) -> None:
        """Main-thread dialog: Yes = retry failed (skip allocation), No = allocate."""
        do_retry = messagebox.askyesno(
            tr("demand.failures_title"),
            msg + tr("demand.failures_suffix"))
        if do_retry:
            self.log_warning("Allocation skipped — failed labourers ka retry shuru...")
            # Demand thread ko finish hone de, phir failed applicants ka turant retry
            self.app.after(800, self._retry_failed_applicants)
        else:
            proceed()

    def _process_single_job_card(self, driver, wait, jc, apps_in_jc,
                                 user_days, demand_from, days_worked_ids, jc_ids,
                                 grid_ids, btn_ids, p_ids, v_ids, base_url):
        """
        Selects the job card, fills the demand grid rows for the target members
        and submits. Handles the ASP.NET postbacks:
          - JC dropdown change  -> async postback renders the grid (gvData)
          - grid date-field change (d3) -> async postback auto-fills dt_to
          - Proceed click -> full postback that saves the demand
        """

        def mark(jc_, name_, status_):
            # Snapshot the location NOW — the after() callback runs later on the
            # main thread, by which time the worker may have moved on.
            self.app.after(0, self._update_results_tree,
                           (jc_, name_, status_,
                            getattr(self, '_current_panchayat', ''),
                            getattr(self, '_current_village', '')))

        def grid_id_of():
            return driver.find_element(
                By.CSS_SELECTOR, ", ".join(f"table[id='{x}']" for x in grid_ids)
            ).get_attribute("id")

        def find_row_index(name_target):
            """Returns the grid row index (1-based) whose _job name matches, else -1."""
            try:
                rows = driver.find_elements(By.CSS_SELECTOR, f"table[id='{grid_id}'] > tbody > tr")
            except Exception:
                return -1
            grid_names = []
            for i, r in enumerate(rows):
                if i == 0:
                    continue
                try:
                    span = r.find_element(By.CSS_SELECTOR, "span[id*='_job']")
                    nm = (span.get_attribute("innerText") or "").strip()
                    grid_names.append(nm)
                    if self._names_equal(nm, name_target):
                        return i
                except (StaleElementReferenceException, NoSuchElementException):
                    continue
                except Exception:
                    continue
            # Match nahi hua — grid me actual names log karo. '?' wala CSV name
            # ya Devanagari-vs-English mismatch ki wajah se hota hai; user ko
            # dono side dikh jaye to fix karna aasan ho jata hai.
            if grid_names:
                self.app.after(0, self.app.log_message, self.log_display,
                               f"   ⚠️ Name '{name_target}' grid me nahi mila. "
                               f"Grid names: {grid_names[:6]}{'...' if len(grid_names) > 6 else ''}",
                               "warning")
            return -1

        try:
            jc_suffix = jc.split('/')[-1]
            self.app.after(0, self.app.log_message, self.log_display, f"Processing JC: {jc}")

            # ── 1. Select the job card in the Registration dropdown (POSTBACK) ──
            # First wait for the *target* JC to appear in the list — this guarantees
            # the village -> registration async postback has finished, so we never
            # match against the previous village's stale options. Then select it
            # (option values are composite: '{JC}:{reg_date}:{days}:...').
            if not self._wait_jc_option(driver, jc, jc_ids):
                self.app.after(0, self.app.log_message, self.log_display,
                               f"   FAIL: JC '{jc}' not found in the village list.", "error")
                for a in apps_in_jc:
                    mark(jc, a.get('Name of Applicant'), "FAIL: JC Not Found")
                return

            # Snapshot the grid BEFORE the JC postback. After a submit the page keeps
            # showing the previous JC's already-filled grid — if we read it before the
            # async postback replaces it, overlapping worker names across job cards
            # make every later JC wrongly report 'Already Correct' (nothing saved).
            old_grid = None
            try:
                old_grid = driver.find_element(
                    By.CSS_SELECTOR, ", ".join(f"table[id='{x}']" for x in grid_ids))
            except Exception:
                pass

            jc_selected = False
            for _attempt in range(3):
                if self.is_stopped():
                    return
                try:
                    jc_el = wait.until(EC.presence_of_element_located(
                        (By.CSS_SELECTOR, ", ".join(f"#{x}" for x in jc_ids))))
                    if self._select_jc_option(driver, Select(jc_el), jc):
                        jc_selected = True
                        break
                except (StaleElementReferenceException, NoSuchElementException):
                    time.sleep(0.5)
                time.sleep(0.15)
            if not jc_selected:
                self.app.after(0, self.app.log_message, self.log_display,
                               f"   FAIL: JC '{jc}' not found in the village list.", "error")
                for a in apps_in_jc:
                    mark(jc, a.get('Name of Applicant'), "FAIL: JC Not Found")
                return
            self.app.after(0, self.app.log_message, self.log_display,
                           f"   Selected JC '{jc}' from the list.")

            # Wait for the grid to refresh after the JC postback: the OLD grid (from
            # the previous JC) must go stale first — only then is the new JC's grid in
            # place. Without this, the still-visible filled rows of the previous JC get
            # misread and every later JC is wrongly reported as 'Already Correct'.
            grid_ready = old_grid is None
            if old_grid is not None:
                try:
                    WebDriverWait(driver, 8, poll_frequency=0.2).until(EC.staleness_of(old_grid))
                    grid_ready = True
                except TimeoutException:
                    grid_ready = False
            if not grid_ready:
                # Grid never replaced — re-fire the change event once.
                try:
                    jc_el = driver.find_element(By.CSS_SELECTOR, ", ".join(f"#{x}" for x in jc_ids))
                    self._select_jc_option(driver, Select(jc_el), jc)
                    WebDriverWait(driver, 10, poll_frequency=0.2).until(EC.staleness_of(old_grid))
                    grid_ready = True
                except Exception:
                    grid_ready = False
            if not grid_ready:
                # Might be a genuine re-run (same JC, grid already correct). Only fail
                # when none of this JC's workers are present in the grid at all.
                if not self._grid_has_any_worker(driver, grid_ids, apps_in_jc):
                    for a in apps_in_jc:
                        mark(jc, a.get('Name of Applicant'), "Failed (Grid did not refresh)")
                    return

            try:
                wait.until(EC.presence_of_element_located(
                    (By.CSS_SELECTOR, ", ".join(f"table[id='{x}']" for x in grid_ids))))
            except TimeoutException:
                msg = self._read_result_message(driver, attempts=2)
                status = "Skipped (JC Not Issued)" if "not yet issued" in msg.lower() else "Skipped (No grid)"
                for a in apps_in_jc:
                    mark(jc, a.get('Name of Applicant'), status)
                return
            grid_id = grid_id_of()

            # ── 2. "Not yet issued" check ──
            try:
                WebDriverWait(driver, 1.0).until(EC.presence_of_element_located(
                    (By.XPATH, "//font[contains(text(), 'not yet issued')]")))
                for a in apps_in_jc:
                    mark(jc, a.get('Name of Applicant'), "Skipped (JC Not Issued)")
                return
            except TimeoutException:
                pass

            # ── 3. Days availability (100-day rule) ──
            worked = self._get_worked_days(driver, days_worked_ids)
            avail = 100 - worked
            if avail <= 0:
                for a in apps_in_jc:
                    mark(jc, a.get('Name of Applicant'), "Skipped (100 days)")
                return

            adj_days = user_days
            total_needed = user_days * len(apps_in_jc)
            if total_needed > avail:
                adj_days = avail // len(apps_in_jc) if len(apps_in_jc) else avail
            elif user_days > avail:
                adj_days = avail

            today = datetime.now().strftime('%d/%m/%Y')

            # ── 3b. Clear stale dates/days on rows that are NOT selected ──
            # The portal (ASP.NET) carries previously-entered values into the
            # grid across postbacks. A row we are NOT filling can therefore
            # still hold old dates — submitting would create demand for that
            # labourer too (e.g. both members of a job card get demand even
            # though only one was selected).
            self._clear_stale_grid_rows(
                driver, grid_id,
                [a.get('Name of Applicant', '').strip() for a in apps_in_jc])

            # ── 4. Fill rows (single postback per row via d3) ──
            status_map = {}  # name -> 'filled' | 'noday' | 'error' | 'notfound'
            for i, a in enumerate(apps_in_jc):
                if self.is_stopped():
                    return
                name = a.get('Name of Applicant', '').strip()
                if not name:
                    continue
                days_val = adj_days if adj_days > 0 else (avail if i == 0 else 0)
                if days_val <= 0:
                    status_map[name] = "noday"
                    continue
                row_i = find_row_index(name)
                if row_i == -1:
                    status_map[name] = "notfound"
                    continue
                pfx = f"{grid_id}_ctl{row_i + 1:02d}_"
                status_map[name] = self._fill_grid_row(driver, wait, pfx, today, demand_from, days_val)

            for name, s in status_map.items():
                if s == "noday":
                    mark(jc, name, "Skipped (No days left)")
                elif s == "notfound":
                    mark(jc, name, "Failed (Not found in Table)")
                elif s == "error":
                    mark(jc, name, "Failed (Grid Error)")

            filled_names = [n for n, s in status_map.items() if s == "filled"]
            if not filled_names:
                return

            # The fill postbacks re-render the grid — clear once more so no
            # stale row can slip into the final submit.
            self._clear_stale_grid_rows(
                driver, grid_id,
                [a.get('Name of Applicant', '').strip() for a in apps_in_jc])

            # ── 5. Submit (POSTBACK) ──
            self.app.after(0, self.app.set_status, f"JC {jc_suffix}: Submitting...")
            try:
                btn = wait.until(EC.presence_of_element_located(
                    (By.CSS_SELECTOR, ", ".join(f"#{x}" for x in btn_ids))))
                driver.execute_script("arguments[0].click();", btn)
            except Exception:
                for n in filled_names:
                    mark(jc, n, "FAIL: Submit button missing")
                return

            res, ok = self._collect_submit_result(driver, btn)
            for n in filled_names:
                if ok:
                    mark(jc, n, "Success")
                elif any(p in (res or "").lower() for p in self.ALREADY_PHRASES):
                    # Demand already exists for this period — not a new success.
                    mark(jc, n, f"Already: {res}")
                else:
                    mark(jc, n, f"FAIL: {res}")

        except Exception as e:
            self.app.after(0, self.app.log_message, self.log_display,
                           f"CRITICAL ERROR processing {jc}: {e}", "error")
            for a in apps_in_jc:
                mark(jc, a.get('Name of Applicant'), f"FAIL: {type(e).__name__}")
            # Restore the page (panchayat + current village) so the remaining
            # job cards of the run can still be processed. The old behaviour —
            # just driver.get(base_url) — left the browser at its default state,
            # so every later JC failed with 'JC not found in the village list'.
            self._recover_to_village(driver, wait, base_url, p_ids, v_ids, jc_ids)

    def _recover_to_village(self, driver, wait, base_url, p_ids, v_ids, j_ids):
        """Reloads the demand page and re-selects the current panchayat and
        village.

        Called after an unexpected error on a job card. Restoring the dropdown
        state lets the remaining job cards of the run continue normally.
        """
        try:
            driver.get(base_url)
            time.sleep(1.0)
            if self.is_stopped():
                return
            panchayat = getattr(self, '_current_panchayat', '') or ''
            # Block-login portals show a panchayat dropdown; GP login does not
            # — central helper dono cases handle karta hai.
            self._select_panchayat_or_skip(
                driver, wait, panchayat, p_ids, v_ids, timeout=4)
            village = getattr(self, '_current_village', '') or ''
            if village:
                try:
                    if self._select_village(driver, wait, village, v_ids):
                        self._wait_dropdown_populated(driver, wait, j_ids,
                                                      "job cards (recovery)")
                except Exception:
                    pass
        except Exception as e:
            self.app.after(0, self.app.log_message, self.log_display,
                           f"   Recovery to village '{getattr(self, '_current_village', '')}' failed: {e}",
                           "warning")

    # ------------------------------------------------------------------
    # Postback-aware helpers for the Demand page
    # ------------------------------------------------------------------

    @staticmethod
    def _names_key(name):
        # NFKC: Devanagari matras/composed forms normalize karta hai — portal
        # ka naam aur CSV ka naam visually same par Unicode code points alag
        # hone par bhi match ho jaye. Whitespace/case hatao.
        return "".join(unicodedata.normalize("NFKC", (name or "")).lower().split())

    @staticmethod
    def _names_equal(a, b):
        """Compares two names ignoring case, whitespace and Unicode form."""
        return DemandTab._names_key(a) == DemandTab._names_key(b)

    def _clear_stale_grid_rows(self, driver, grid_id, keep_names):
        """Clears dates/days on every grid row whose worker is NOT one of the
        target applicants.

        The ASP.NET portal carries previously-entered values (dt_app, dt_from,
        d3, dt_to) into grid rows across postbacks. If a row we are not filling
        still holds dates, the final submit would create demand for that
        labourer too — e.g. both members of a job card get demand even though
        only one was selected. Values are cleared via JS (no change event), so
        no extra postback fires here; the next real postback (d3 change / the
        submit itself) carries the cleared state.
        """
        keep = {self._names_key(n) for n in keep_names if n}
        if not keep:
            return
        try:
            rows = driver.find_elements(By.CSS_SELECTOR,
                                        f"table[id='{grid_id}'] > tbody > tr")
        except Exception:
            return
        for i, r in enumerate(rows):
            if i == 0:
                continue  # header row
            try:
                span = r.find_element(By.CSS_SELECTOR, "span[id*='_job']")
                nm = (span.get_attribute("innerText") or "").strip()
            except (StaleElementReferenceException, NoSuchElementException):
                continue
            except Exception:
                continue
            if nm and self._names_key(nm) in keep:
                continue  # target worker — will be filled below
            pfx = f"{grid_id}_ctl{i + 1:02d}_"
            for fld in ("dt_app", "dt_from", "d3", "dt_to"):
                try:
                    el = r.find_element(By.ID, f"{pfx}{fld}")
                    if (el.get_attribute("value") or "").strip():
                        self._set_js_value(driver, el, "", fire_change=False)
                except (StaleElementReferenceException, NoSuchElementException):
                    continue
                except Exception:
                    continue

    def _grid_has_any_worker(self, driver, grid_ids, apps_in_jc):
        """True if any of the JC's workers appears in the current grid rows.
        Used to distinguish a stale grid (previous JC still shown) from a genuine
        re-run where the grid already shows this JC's data."""
        try:
            gid = driver.find_element(
                By.CSS_SELECTOR, ", ".join(f"table[id='{x}']" for x in grid_ids)
            ).get_attribute("id")
            rows = driver.find_elements(By.CSS_SELECTOR, f"table[id='{gid}'] > tbody > tr")
            for r in rows[1:]:
                try:
                    span = r.find_element(By.CSS_SELECTOR, "span[id*='_job']")
                    t = (span.get_attribute("innerText") or "").strip()
                    for a in apps_in_jc:
                        nm = (a.get('Name of Applicant', '') or '').replace('*', '').strip()
                        if nm and self._names_equal(t, nm):
                            return True
                except (StaleElementReferenceException, NoSuchElementException):
                    continue
                except Exception:
                    continue
        except Exception:
            pass
        return False

    @staticmethod
    def _jc_value_matches(val, jc):
        """True if a Registration option value matches the full job card number.

        Website values embed extra fields after the JC, e.g.
        'JH-22-003-008-005/127:10/20/2006:53:N:OT:Y' — the JC is everything
        before the first separator (':', '|' or ';').
        """
        v = (val or '').strip()
        if not v or not jc:
            return False
        if v == jc:
            return True
        for sep in (':', '|', ';'):
            if sep in v and v.split(sep)[0].strip() == jc:
                return True
        return False

    @staticmethod
    def _set_dropdown_value(driver, select_el, value):
        """Sets a <select> value via JS and fires 'change' (which triggers the
        ASP.NET onchange postback). Bypasses Selenium's XPath-based value
        selection, which can fail on composite values containing '/' and ':'."""
        driver.execute_script(
            "var s = arguments[0]; s.value = arguments[1];"
            " s.dispatchEvent(new Event('change', { bubbles: true }));",
            select_el, value)

    @staticmethod
    def _wait_jc_option(driver, jc, jc_ids, timeout=12):
        """Waits until the Registration dropdown contains the target job card
        (value match). Confirms the village postback has finished.

        Tolerates StaleElementReferenceException: an ASP.NET postback can
        replace the <option> elements WHILE we scan them, and reading a
        replaced option raises 'stale element reference'. That is treated as
        'not ready yet' and the wait keeps polling until the dropdown settles —
        instead of the exception aborting the whole job card.
        """
        jc = jc.strip()
        try:
            def _found(d):
                try:
                    for sel_id in jc_ids:
                        for opt in d.find_elements(By.XPATH, f"//select[@id='{sel_id}']/option"):
                            if DemandTab._jc_value_matches(opt.get_attribute('value'), jc):
                                return True
                except (StaleElementReferenceException, NoSuchElementException):
                    return False  # dropdown is mid-refresh — keep waiting
                return False
            WebDriverWait(driver, timeout).until(_found)
            return True
        except TimeoutException:
            return False

    @staticmethod
    def _select_jc_option(driver, sel, jc):
        """Selects the Registration option matching the full job card number.

        Option value format:  '{JC}:{reg_date}:{days}:{...}'
        Option text format:   '{serial}-{NAME}'
        Works for any state's job card format since the full JC is matched.
        """
        jc = jc.strip()
        if not jc:
            return False
        for opt in sel.options:
            if DemandTab._jc_value_matches(opt.get_attribute('value'), jc):
                DemandTab._set_dropdown_value(driver, sel._el, opt.get_attribute('value'))
                return True
        # Fallback: text prefix match on the serial suffix (covers any JC format)
        suffix = jc.split('/')[-1]
        for opt in sel.options:
            t = (opt.text or "").strip()
            if t.startswith(suffix + "-"):
                DemandTab._set_dropdown_value(driver, sel._el, opt.get_attribute('value'))
                return True
        return False

    @staticmethod
    def _set_js_value(driver, el, value, fire_change=False):
        """Sets an input's value via JS; optionally fires 'change' to trigger the
        ASP.NET onchange postback."""
        js = "arguments[0].focus(); arguments[0].value = arguments[1];"
        if fire_change:
            js += " arguments[0].dispatchEvent(new Event('change', { bubbles: true }));"
        driver.execute_script(js, el, value)

    def _fill_grid_row(self, driver, wait, pfx, today, demand_from, days_val):
        """
        Fills dt_app (today), dt_from (demand date) and d3 (days) on one grid row.

        dt_app / dt_from are set via JS WITHOUT firing events (no postback), then a
        single 'change' event is fired on d3 — that triggers the one async postback
        the server needs to auto-compute dt_to. Returns 'filled' | 'error'.

        NOTE: We never 'skip' a row whose fields already look correct. This portal
        (ASP.NET) carries the PREVIOUS job card's entered values into the new JC's
        grid rows across postbacks, so a freshly-selected JC can already show
        today / demand-from / days that were actually typed for the previous JC.
        Skipping on that made every JC after the first report 'Already Correct'
        without filling or submitting anything. We always force-set the values and
        fire the days postback so each JC is genuinely filled and submitted.
        """
        try:
            ids = {'app': f"{pfx}dt_app", 'from': f"{pfx}dt_from",
                   'days': f"{pfx}d3", 'till': f"{pfx}dt_to"}
            app_el = wait.until(EC.presence_of_element_located((By.ID, ids['app'])))
            from_el = driver.find_element(By.ID, ids['from'])
            days_el = driver.find_element(By.ID, ids['days'])
            till_el = driver.find_element(By.ID, ids['till'])

            cur_app = (app_el.get_attribute('value') or "").strip()
            cur_from = (from_el.get_attribute('value') or "").strip()
            cur_days = (days_el.get_attribute('value') or "").strip()

            if cur_app != today:
                self._set_js_value(driver, app_el, today, fire_change=False)
            if cur_from != demand_from:
                self._set_js_value(driver, from_el, demand_from, fire_change=False)
            # Always fire the d3 change postback — even when the value already
            # matches — so the server recomputes dt_to for THIS JC's rows.
            self._set_js_value(driver, days_el, str(days_val), fire_change=True)

            # Let the async postback replace the grid before reading dt_to back —
            # a carried-over value could otherwise make the wait below return
            # immediately on stale data and let us race the postback.
            time.sleep(0.2)

            # Wait for the server postback to auto-fill dt_to — this is also the
            # proof that the server received the JS-set dt_app/dt_from/d3 values.
            try:
                wait.until(lambda d: (d.find_element(By.ID, ids['till']).get_attribute("value") or "").strip() != "")
            except TimeoutException:
                return "error"
            return "filled"
        except Exception:
            return "error"

    def _get_worked_days(self, driver, days_worked_ids):
        try:
            el = WebDriverWait(driver, 1.0).until(EC.presence_of_element_located((By.ID, days_worked_ids[0])))
            t = (el.get_attribute("innerText") or "").strip()
            return int(t) if t.isdigit() else 0
        except Exception:
            return 0

    def _select_village(self, driver, wait, token, v_ids):
        """Selects the village dropdown option by name (case-insensitive), falling
        back to a value-suffix (village code) match for legacy inputs.

        Tolerates stale elements: an ASP.NET postback can refresh the dropdown's
        <option> elements WHILE we scan them, so the scan retries instead of
        aborting the whole village."""
        v_el = wait.until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR, ", ".join(f"#{x}" for x in v_ids))))
        for _attempt in range(3):
            if self.is_stopped():
                return False
            try:
                v_sel = Select(v_el)
                if self._select_by_text_case_insensitive(v_sel, token):
                    self.app.after(0, self.app.log_message, self.log_display,
                                   f"Selected Village '{token}'.")
                    return True
                for opt in v_sel.options:
                    val = (opt.get_attribute('value') or "").strip()
                    if val not in ("00", "99") and token and val.endswith(str(token)):
                        v_sel.select_by_value(val)
                        self.app.after(0, self.app.log_message, self.log_display,
                                       f"Selected Village '{opt.text}' (code ...{token}).")
                        return True
                break  # scan finished cleanly — option genuinely absent
            except (StaleElementReferenceException, NoSuchElementException):
                # Dropdown was mid-refresh — re-find it and retry.
                time.sleep(0.5)
                try:
                    v_el = wait.until(EC.element_to_be_clickable(
                        (By.CSS_SELECTOR, ", ".join(f"#{x}" for x in v_ids))))
                except TimeoutException:
                    return False
        try:
            avail = [o.text.strip() for o in v_sel.options
                     if o.text.strip() and o.text.strip() not in ("---Select---", "--All Villages--", "")]
            self.app.after(0, self.app.log_message, self.log_display,
                           f"   Available villages: {avail[:15]}{'...' if len(avail) > 15 else ''}", "warning")
        except Exception:
            pass
        return False

    def _wait_dropdown_populated(self, driver, wait, ids, label):
        """Waits for a dependent dropdown to gain options after an ASP.NET postback."""
        self.app.after(0, self.app.log_message, self.log_display, f"Waiting for {label}...")
        wait.until(EC.any_of(*[
            EC.presence_of_element_located((By.XPATH, f"//select[@id='{i}']/option[position()>1]"))
            for i in ids
        ]))
        time.sleep(0.4)  # let the partial postback settle

    def _read_result_message(self, driver, attempts=1):
        """Reads the post-submit message from the page (red fonts / message spans)."""
        xpaths = [
            "//font[@color='red']",
            "//span[contains(@id, '_lblmsg')]",
            "//span[@id='ctl00_ContentPlaceHolder1_Up_lit']",
            "//div[@id='divMesssge']",
        ]
        for _ in range(attempts):
            for xp in xpaths:
                try:
                    els = driver.find_elements(By.XPATH, xp)
                    for el in els:
                        t = (el.get_attribute("innerText") or "").strip()
                        if t and t.lower() != "updating msg..":
                            return t
                except Exception:
                    continue
            time.sleep(0.8)
        return ""

    def _collect_submit_result(self, driver, btn):
        """
        After clicking Proceed: handles the client-side confirm alert (older page
        versions), waits for the submit postback and returns (message, ok).
        """
        # 1. Client-side confirm alert?
        try:
            alert = WebDriverWait(driver, 3).until(EC.alert_is_present())
            text = (alert.text or "").strip()
            self.app.after(0, self.app.log_message, self.log_display, f"   ALERT: {text}")
            alert.accept()
            low = text.lower()
            if "please enter work demand" in low or "enter work demand for any one" in low:
                return text, False  # nothing was filled correctly
            if "do you want to submit" in low or "confirm" in low:
                time.sleep(0.8)
                # Some portals show a second alert after the confirm
                try:
                    a2 = WebDriverWait(driver, 4).until(EC.alert_is_present())
                    t2 = (a2.text or "").strip()
                    self.app.after(0, self.app.log_message, self.log_display, f"   ALERT 2: {t2}")
                    a2.accept()
                    low2 = t2.lower()
                    if any(x in low2 for x in ["error", "fail", "not saved", "problem", "please select"] + list(self.ALREADY_PHRASES)):
                        return t2, False
                    return t2, True
                except TimeoutException:
                    pass
                return "", True
            if any(x in low for x in ["error", "fail", "not saved", "problem", "please select"] + list(self.ALREADY_PHRASES)):
                return text, False
            return text, True
        except TimeoutException:
            pass

        # 2. No confirm alert — modern page: submit postback happened immediately.
        settled = False
        try:
            submit_wait = WebDriverWait(driver, 6)
            submit_wait.until(EC.staleness_of(btn))
            settled = True
        except TimeoutException:
            settled = False
        time.sleep(0.5)
        msg = self._read_result_message(driver, attempts=3)
        if msg:
            low = msg.lower()
            if any(x in low for x in ["error", "fail", "not saved", "problem", "please select", "enter work demand"] + list(self.ALREADY_PHRASES)):
                return msg, False
            return msg, True
        if settled:
            return "Submitted", True
        return "No response after submit", False
    def _update_results_tree(self, data):
        """
        Adds a new row to the results treeview with correct color tags.
        """
        jc, name, status = data[0], data[1], data[2]
        # Fallback values kabhi special labels nahi hote — agar panchayat_var
        # mein '🌐 All Panchayats'/'⭐ My Saved Panchayats' ho to uski jagah
        # '-' dikhao (label literal panchayat ki tarah result mein na aaye).
        fallback_panchayat = self._clean_panchayat_value(self.panchayat_var.get()) or "-"
        if len(data) >= 5:
            panchayat = data[3] or fallback_panchayat
            village = data[4] or "-"
        else:
            panchayat = getattr(self, '_current_panchayat', '') or fallback_panchayat
            village = getattr(self, '_current_village', '') or "-"
        row_id = len(self.results_tree.get_children()) + 1
        
        status_str = str(status)
        status_low = status_str.lower()
        tags = () # Default: No Color (White/Black)

        # 1. Failed Logic (Red)
        if any(e in status_low for e in ['fail', 'error', 'crash', 'not found', 'invalid', 'aadhaar', 'not saved', 'not issued']):
            tags = ('failed',)
            
        # 2. Warning Logic (Yellow) - 'already' (demand pehle se hai) + 'skipped' etc.
        elif any(w in status_low for w in ['already', 'skip', 'adjust', 'limit', '100 days']):
            tags = ('warning',)
            
        # 3. Success Logic (Green)
        elif any(s in status_low for s in ['success', 'saved', 'done']):
            tags = ('success',)

        # Display Text Truncation
        disp_status = (status_str[:100] + '...') if len(status_str) > 100 else status_str
        
        self.results_tree.insert("", "end", iid=row_id, values=(row_id, panchayat, village, jc, name, disp_status), tags=tags)
        self.results_tree.yview_moveto(1)

    def _retry_failed_applicants(self):
        """
        Simple retry: re-selects only the applicants whose demand FAILED and
        immediately re-runs the automation for them. No manual re-ticking,
        no waiting for the user to click Start again.
        """
        failed_items = self.results_tree.tag_has('failed')
        if not failed_items:
            messagebox.showinfo(tr("demand.retry_title"), tr("demand.retry_no_failed"))
            return

        # Clear all selections, then re-select ONLY the failed ones
        for app_data in self.all_applicants_data:
            app_data['_selected'] = False

        re_selected_count = 0
        for item_id in failed_items:
            try:
                values = self.results_tree.item(item_id, 'values')
                if not values:
                    continue
                jc_no, name = values[3], values[4]
                for app_data in self.all_applicants_data:
                    if (app_data.get('Job card number') == jc_no
                            and app_data.get('Name of Applicant') == name):
                        app_data['_selected'] = True
                        re_selected_count += 1
                        break
            except Exception:
                continue

        if not re_selected_count:
            messagebox.showwarning(
                tr("demand.retry_title"),
                tr("demand.retry_not_in_csv"))
            return

        # Retry run se PEHLE: abhi tree me jo successful hain (run 1 ke) unhe
        # carry-over set me accumulate karo. Retry ka start_automation() tree
        # clear kar dega — isliye ye names yahan preserve hote hain aur retry
        # run ke finally-block me merge hote hain. Isse auto-allocation me
        # run 1 + retry dono ke successful jobcards aate hain (sirf retried
        # wale nahi).
        try:
            for item_id in self.results_tree.get_children():
                values = self.results_tree.item(item_id, 'values')
                status = str(values[5]).lower() if len(values) > 5 else ""
                name = str(values[4]).strip() if len(values) > 4 else ""
                if name and "success" in status:
                    self._retry_prior_success_names.add(name)
        except Exception:
            pass

        self._refresh_selected_jc_panel()
        self._update_selection_summary()
        self._update_jc_header_counters()
        self.log_info(f"Re-selected {re_selected_count} failed applicants. Restarting demand automation...")

        # Turant automation re-run for just the failed applicants. Flag set
        # karo taaki start_automation() carry-over set clear na kare (usme
        # run 1 ke successful names hain jo retry ke baad allocation me
        # merge hone chahiye).
        self._skip_retry_clear = True
        self.start_automation()

    def export_results(self):
        """
        Exports the contents of the results treeview to a CSV file.
        """
        if not self.results_tree.get_children(): messagebox.showinfo(tr("dialogs.export"), tr("dialogs.no_results")); return
        p = self.panchayat_var.get().strip().replace(" ", "_") or "UnknownPanchayat"; s = self.state_var.get() or "UnknownState"
        fname = f"Demand_Report_{s}_{p}_{datetime.now():%Y%m%d_%H%M}.csv"; self.export_treeview_to_csv(self.results_tree, fname)

    def save_inputs(self, inputs):
        try:
            self.app.history_manager.save_tab_inputs_batch("demand", inputs)
        except Exception as e:
            self.log_warning(f"Could not save demand inputs: {e}")

    def load_inputs(self):
        """
        Loads the last saved inputs from the JSON file on tab startup.
        """
        today = datetime.now().strftime('%d/%m/%Y'); date_to_set = today
        days_to_set = self.app.history_manager.get_suggestions("demand_days")[0] if self.app.history_manager.get_suggestions("demand_days") else "14"
        work_key_to_set = ""
        
        data = self.app.history_manager.get_tab_inputs("demand")
        if data:
            saved_state = data.get('state', '')
            if saved_state and saved_state in getattr(self, 'state_options', []):
                self.state_var.set(saved_state)
            elif getattr(self, 'state_options', []):
                self.state_var.set(self.state_options[0])
            self.panchayat_var.set(self._clean_panchayat_value(data.get('panchayat')))
            days_to_set = data.get('days', days_to_set)
            work_key_to_set = data.get('work_key_for_allocation', '')
            loaded = data.get('demand_date', '');
            try:
                datetime.strptime(loaded, '%d/%m/%Y')
                date_to_set = loaded
            except ValueError:
                pass
            
        # --- FIX: Use delete/insert ---
        self.demand_date_entry.delete(0, "end")
        self.demand_date_entry.insert(0, date_to_set)
        
        self.days_entry.delete(0, 'end')
        self.days_entry.insert(0, days_to_set)
        
        self.allocation_work_key_var.set(work_key_to_set)

    def _clear_selection(self):
        """Clears the current selection of all applicants."""
        if not any(a.get('_selected', False) for a in self.all_applicants_data):
            self.log_info("No selection.")
            return
        for a in self.all_applicants_data:
            a['_selected'] = False
        self._refresh_selected_jc_panel()
        self._refresh_search_results()
        self._update_selection_summary()
        self.log_info("Selection cleared.")

    def _on_format_change(self, selected_format):
        """Disables the filter menu for CSV format as it exports all data."""
        if "CSV" in selected_format:
            self.export_filter_menu.configure(state="disabled")
        else:
            self.export_filter_menu.configure(state="normal")

    def export_report(self):
        """Export results to professional Excel."""
        panchayat_name = self.panchayat_var.get().strip()
        if not panchayat_name:
            messagebox.showwarning(tr("dialogs.input_needed"), tr("dialogs.enter_panchayat_filename"))
            return
        safe_p = "".join(c for c in panchayat_name if c.isalnum() or c in (' ', '_')).rstrip()
        self.export_treeview_to_excel(
            tree=self.results_tree,
            default_filename=f"Demand_Report_{safe_p}_{datetime.now():%Y%m%d_%H%M}.xlsx",
            filter_mode="Export All",
            title_prefix=f"Demand Report: {panchayat_name}"
        )