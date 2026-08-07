# tabs/base_tab.py
import csv
import tkinter
from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk
import os, sys, platform, re
from datetime import datetime
from typing import Any, Callable, Dict, List, Optional, Tuple
from PIL import Image, ImageDraw, ImageFont 

from src import config
from src.utils import resource_path, get_logger, truncate_workcode

logger = get_logger()

# Module-level imports for selenium and openpyxl (P4: moved from lazy imports in method bodies)
from selenium.common.exceptions import NoSuchWindowException, WebDriverException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException

# A2: Import extracted components from their own modules
from .date_picker_popup import DatePickerPopup
from .professional_pdf import ProfessionalPDF

class BaseAutomationTab(ctk.CTkFrame):
    def __init__(self, parent: Any, app_instance: Any, automation_key: str) -> None:
        super().__init__(parent, fg_color="transparent")
        self.app = app_instance
        self.automation_key = automation_key
        self.retry_btn = None # Placeholder for retry button
        self._tab_destroyed = False  # Flag: set True in destroy() to prevent UI updates on dead widgets
        
        # --- Activity Tracking for Logging & WhatsApp Notifications ---
        self.activity_start_time: Optional[float] = None
        self.activity_panchayat: str = ""
        self.activity_village: str = ""
        self.activity_details: str = ""  # JSON-like summary of results
        
        # --- AfterTracker for safe callback cleanup on tab destroy ---
        from src.ui_components import AfterTracker
        self._safe_after = AfterTracker(self)
        
    def _extract_activity_panchayat(self) -> str:
        """Auto-extract panchayat name from common widget patterns.
        Checks panchayat_var (StringVar), panchayat_menu/entry widgets,
        config_vars dict pattern (used by many tabs), and returns uppercase value or empty string."""
        try:
            # Priority 0: Check config_vars dict pattern (used by mb_entry, demand, etc.)
            cfg = getattr(self, 'config_vars', None)
            if cfg and isinstance(cfg, dict):
                for key in ['location_panchayat', 'panchayat_name', 'panchayat']:
                    var = cfg.get(key)
                    if var is not None and hasattr(var, 'get'):
                        val = var.get().strip().upper()
                        if val:
                            return val
            # Priority 1: StringVar named panchayat_var
            for attr in ['panchayat_var', 'panchayat']:
                v = getattr(self, attr, None)
                if v is not None and hasattr(v, 'get'):
                    val = v.get().strip().upper()
                    if val:
                        return val
            # Priority 2: CTkOptionMenu
            for attr in ['panchayat_menu', 'panchayat_dropdown', 'panchayat_entry']:
                w = getattr(self, attr, None)
                if w is not None:
                    if hasattr(w, 'cget'):
                        try:
                            var = w.cget('variable')
                            if var and hasattr(var, 'get'):
                                val = var.get().strip().upper()
                                if val:
                                    return val
                        except Exception:
                            pass
                    if hasattr(w, 'get'):
                        val = w.get().strip().upper()
                        if val:
                            return val
        except Exception:
            pass
        return ""
    
    def _extract_activity_village(self) -> str:
        """Auto-extract village name from common widget patterns."""
        try:
            # Check config_vars dict pattern first
            cfg = getattr(self, 'config_vars', None)
            if cfg and isinstance(cfg, dict):
                for key in ['location_village', 'village_name', 'village']:
                    var = cfg.get(key)
                    if var is not None and hasattr(var, 'get'):
                        val = var.get().strip().upper()
                        if val:
                            return val
            for attr in ['village_var', 'village']:
                v = getattr(self, attr, None)
                if v is not None and hasattr(v, 'get'):
                    val = v.get().strip().upper()
                    if val:
                        return val
            for attr in ['village_menu', 'village_entry']:
                w = getattr(self, attr, None)
                if w is not None and hasattr(w, 'get'):
                    val = w.get().strip().upper()
                    if val:
                        return val
        except Exception:
            pass
        return ""
    
    def _extract_activity_details(self) -> str:
        """Extract result summary from results_tree if available.
        
        Smart detection of the 'Status' column: checks column headings first,
        then falls back to common position patterns based on column count.
        Also includes work codes / keys for richer context.
        """
        try:
            tree = getattr(self, 'results_tree', None)
            if tree is None:
                return self.activity_details
            all_items = tree.get_children()
            if not all_items:
                return self.activity_details

            total = len(all_items)
            success = 0
            failed = 0
            skipped = 0

            # ── Find Status column index ──
            columns = tree["columns"]
            status_idx = 1  # default fallback
            work_code_idx = None  # optional: find a Work Code column
            if columns:
                col_list = list(columns)
                # Try to find "Status" by name
                for i, c in enumerate(col_list):
                    c_lower = c.lower().strip()
                    if c_lower == 'status':
                        status_idx = i
                        break
                else:
                    # No column named 'Status' — use heuristic based on count
                    if len(col_list) == 3:
                        # (Code, Status, Detail) → status at 1
                        status_idx = 1
                    elif len(col_list) == 4:
                        # (Time, Code, Status, Detail) or (Code, Status, Detail, Time) → status at 2
                        status_idx = 2
                    elif len(col_list) == 5:
                        # (Time, Panch, Code, Status, Detail) → status at 3
                        status_idx = 3
                    elif len(col_list) >= 8:
                        # Very wide tables (mb_entry=8, demand=10): status further left
                        status_idx = len(col_list) - 3
                    elif len(col_list) >= 6:
                        # Larger tables: status is usually second-to-last
                        status_idx = len(col_list) - 2
                
                # Try to find a "Work Code" or similar column for richer details
                for i, c in enumerate(col_list):
                    c_lower = c.lower().strip()
                    if any(kw in c_lower for kw in ['work code', 'work', 'jobcard', 'key', 'item']):
                        work_code_idx = i
                        break

            unique_codes = set()
            for item_id in all_items:
                values = tree.item(item_id)['values']
                if not values or len(values) <= status_idx:
                    continue

                status = str(values[status_idx]).lower()
                
                if 'success' in status or '✅' in status or 'verified' in status or 'saved' in status:
                    success += 1
                elif 'fail' in status or '❌' in status or 'error' in status or 'timeout' in status:
                    failed += 1
                elif 'skip' in status:
                    skipped += 1
                else:
                    skipped += 1

                # Collect unique work codes for richer context (last 6 digits only)
                if work_code_idx is not None and len(values) > work_code_idx:
                    code = str(values[work_code_idx]).strip()
                    if code and code != 'N/A' and code != '-':
                        unique_codes.add(truncate_workcode(code))

            # ── Build summary ──
            emoji_success = "✅" if success > 0 else ""
            emoji_failed = "❌" if failed > 0 else ""
            
            parts = [f"📊 Total: {total}"]
            if success > 0:
                parts.append(f"{emoji_success} OK: {success}")
            if failed > 0:
                parts.append(f"{emoji_failed} FAIL: {failed}")
            if skipped > 0:
                parts.append(f"⏭️ Skip: {skipped}")

            # Add work code count for context
            if len(unique_codes) > 0 and len(unique_codes) <= 5:
                codes_str = ", ".join(sorted(unique_codes)[:5])
                parts.append(f"📋 {codes_str}")
            elif len(unique_codes) > 5:
                parts.append(f"📋 {len(unique_codes)} codes")

            return " | ".join(parts)
        except Exception:
            return self.activity_details
    
    def _extract_tree_columns_rows(self) -> Tuple[List[str], List[List]]:
        """
        results_tree se raw columns + rows extract karta hai (cloud reports ke liye).

        Returns:
            (columns, rows) — rows list-of-lists, values stringified.
            Agar tree khali/nahi hai to ([], []).
        """
        try:
            tree = getattr(self, 'results_tree', None)
            if tree is None:
                return [], []
            columns = list(tree["columns"])
            rows: List[List] = []
            for item_id in tree.get_children():
                values = tree.item(item_id)['values']
                rows.append(["" if v is None else str(v) for v in values])
            return columns, rows
        except Exception:
            return [], []

    def _refresh_activity_data(self) -> None:
        """Call before/after automation to sync activity data from widgets."""
        self.activity_panchayat = self._extract_activity_panchayat()
        self.activity_village = self._extract_activity_village()
        self.activity_details = self._extract_activity_details()
        
    def show_automation_notification(self, status: str = "success", duration: int = 6000) -> None:
        """
        Show a professional toast notification when automation completes.
        
        Uses the upgraded ToastNotification with title + details support.
        Called automatically from AutomationMixin.on_automation_finished().
        
        Args:
            status: "success", "stopped", or "failed"
            duration: How long to show the notification (ms)
        """
        if not self._is_alive():
            return
        try:
            self._refresh_activity_data()
            summary = self.activity_details
            panchayat = self.activity_panchayat
            village = self.activity_village
            
            # Colour-coded by outcome: passed → green, failed → red,
            # stopped → amber.
            if status == "success":
                title = "✅ Automation Complete"
                kind = "success"  # GREEN = passed
            elif status == "stopped":
                title = "⏹ Automation Stopped"
                kind = "warning"  # AMBER = stopped
            else:
                title = "⚠️ Automation Failed"
                kind = "error"    # RED = failed
            
            # Build location string
            location_parts = []
            if panchayat:
                location_parts.append(f"📍 {panchayat}")
            if village:
                location_parts.append(f"🏘️ {village}")
            location_str = " | ".join(location_parts) if location_parts else ""
            
            key_display = self.automation_key.replace("_", " ").title()
            
            # Format details nicely
            detail_lines = []
            if summary:
                detail_lines.append(summary)
            if location_str and location_str not in (summary or ""):
                detail_lines.append(location_str)
            
            details_str = "\n".join(detail_lines) if detail_lines else "Check the 'Results' tab for full details"
            
            self.app.show_toast(
                message=f"📋 {key_display}",
                kind=kind,
                duration=duration,
                title=title,
                details=details_str
            )
        except Exception as e:
            logger.debug(f"Failed to show automation notification: {e}")
        
    def destroy(self) -> None:
        """
        Override destroy() to set a flag that prevents background threads
        from updating widgets on a destroyed tab.
        
        DO NOT call driver.quit() here — the automation thread may be using
        the same driver concurrently, causing a GIL fatal crash. The driver
        is cleaned up in start_automation_thread()'s wrapper() after the
        target() function returns.
        
        DO NOT set stop_event here — the new tab instance shares the same
        automation_key, which would replace the Event in stop_events[key],
        causing the old thread to read the new (unset) Event.
        """
        self._tab_destroyed = True
        super().destroy()
        
    def _is_alive(self) -> bool:
        """Returns True if the tab's widgets still exist and can be updated.
        
        Background threads call set_common_ui_state() / log_message() via
        self.app.after(0, ...). If the user navigated away before the callback
        fires, all widgets are destroyed and winfo_exists() returns False.
        This guard prevents TclError: invalid command name.
        """
        if self._tab_destroyed:
            return False
        try:
            return bool(self.winfo_exists())
        except Exception:
            return False
        
    def safe_after(self, ms: int, callback: Callable, *args: Any) -> str:
        """
        Tracked version of after() that auto-cancels when tab is destroyed.
        Prevents ghost callbacks from firing after tab is gone.
        Use this instead of self.after() for recurring callbacks.
        """
        return self._safe_after.after(ms, callback, *args)
        
    def open_date_picker(self, callback: Callable[[str], None]) -> None:
        """Opens the reusable DatePickerPopup."""
        DatePickerPopup(self, callback)

    def handle_error(self, e: Exception) -> None:
        """Centralized error handler.

        Dialogs are scheduled via app.after() so they always run on the Tk
        main thread — handle_error may be called from automation threads, and
        showing a Tk dialog directly from a worker thread can crash the app
        (e.g. when the user closed the browser tab mid-run).
        """
        error_msg = str(e).lower()
        if "no such window" in error_msg or "target window already closed" in error_msg or "web view not found" in error_msg:
            self.log_error("Automation Stopped: Browser tab/window was closed.")
            try:
                self.app.after(0, lambda: messagebox.showwarning("Browser Closed", "Automation stopped because the browser window was closed."))
            except Exception:
                pass
        elif "invalid session id" in error_msg:
            self.log_error("Error: Browser session lost.")
            try:
                self.app.after(0, lambda: messagebox.showwarning("Connection Lost", "Browser session was lost. Please restart the browser."))
            except Exception:
                pass
        else:
            self.log_error(f"Error: {e}")
            try:
                self.app.after(0, lambda: messagebox.showerror("Automation Error", f"An error occurred:\n\n{e}"))
            except Exception:
                pass

    def _get_wkhtml_path(self) -> str:
        os_type = platform.system()
    
        if hasattr(sys, '_MEIPASS'):
            base_path = sys._MEIPASS
            if os_type == "Windows":
                return os.path.join(base_path, 'wkhtmltoimage.exe')
            elif os_type == "Darwin":
                return os.path.join(base_path, 'wkhtmltoimage')
        else:
            base_path = os.path.abspath(".")
            if os_type == "Windows":
                return os.path.join(base_path, 'bin', 'win', 'wkhtmltoimage.exe')
            elif os_type == "Darwin":
                return os.path.join(base_path, 'bin', 'mac', 'wkhtmltoimage')
                
        return 'wkhtmltoimage'
        
    def generate_report_image(self, data: List[List[str]], headers: List[str], title: str, date_str: str, output_path: str) -> bool:
        try:
            try:
                font_path_regular = resource_path("assets/fonts/NotoSansDevanagari-Regular.ttf")
                font_path_bold = resource_path("assets/fonts/NotoSansDevanagari-Bold.ttf")
                font_title = ImageFont.truetype(font_path_bold, 28)
                font_date = ImageFont.truetype(font_path_regular, 18)
                font_header = ImageFont.truetype(font_path_bold, 16)
                font_body = ImageFont.truetype(font_path_regular, 14)
            except IOError:
                print("Warning: NotoSansDevanagari fonts not found. Using default PIL fonts.")
                font_title = ImageFont.load_default()
                font_date = ImageFont.load_default()
                font_header = ImageFont.load_default()
                font_body = ImageFont.load_default()

            img_width = 2400
            margin_x = 80
            margin_y = 60
            
            header_bg_color = (220, 235, 255)
            row_even_bg_color = (255, 255, 255)
            row_odd_bg_color = (245, 245, 245)
            text_color = (0, 0, 0)
            border_color = (180, 180, 180)

            num_cols = len(headers)
            col_widths_pixels = []
            if num_cols > 0:
                available_width = img_width - (2 * margin_x)
                default_width = available_width / num_cols
                col_widths_pixels = [default_width] * num_cols
                
                sno_index = -1
                if any(str(h).lower() in ["s no.", "sno.", "s.no"] for h in headers):
                    sno_index = next((i for i, h in enumerate(headers) if str(h).lower() in ["s no.", "sno.", "s.no"]), -1)
                    if sno_index != -1:
                        col_widths_pixels[sno_index] = max(80, default_width * 0.4)
                        
                non_sno_width = sum(col_widths_pixels[i] for i in range(num_cols) if i != sno_index)
                remaining_width = available_width - (col_widths_pixels[sno_index] if sno_index != -1 else 0)
                original_non_sno_total = sum(default_width for i in range(num_cols) if i != sno_index)
                
                if original_non_sno_total > 0:
                    scale_factor = remaining_width / original_non_sno_total
                    for i in range(num_cols):
                        if i != sno_index:
                            col_widths_pixels[i] *= scale_factor
            
            for i, header in enumerate(headers):
                header_width = font_header.getlength(str(header)) + 40
                if col_widths_pixels[i] < header_width:
                    col_widths_pixels[i] = header_width

            current_total_width = sum(col_widths_pixels)
            if current_total_width == 0: return False
                
            scale_factor = (img_width - 2 * margin_x) / current_total_width
            col_widths_pixels = [w * scale_factor for w in col_widths_pixels]

            initial_height = 1600
            img = Image.new("RGB", (img_width, initial_height), (255, 255, 255))
            draw = ImageDraw.Draw(img)

            current_y = margin_y
            
            title_bbox = font_title.getbbox(title)
            title_height = title_bbox[3] - title_bbox[1]
            title_text_width = font_title.getlength(title)
            title_x = (img_width - title_text_width) / 2
            draw.text((title_x, current_y), title, font=font_title, fill=text_color)
            current_y += title_height + 5

            date_bbox = font_date.getbbox(date_str)
            date_height = date_bbox[3] - date_bbox[1]
            date_text_width = font_date.getlength(date_str)
            date_x = img_width - margin_x - date_text_width
            draw.text((date_x, current_y), date_str, font=font_date, fill=text_color)
            current_y += date_height + 20

            header_y_start = current_y
            header_height = 0
            for i, header in enumerate(headers):
                wrapped_header = self._wrap_text(str(header), font_header, col_widths_pixels[i] - 10)
                line_height = (font_header.getbbox("Tg")[3] - font_header.getbbox("Tg")[1]) * 1.2
                header_height = max(header_height, len(wrapped_header) * line_height + 10)
            
            current_x = margin_x
            for i, header in enumerate(headers):
                draw.rectangle([current_x, header_y_start, current_x + col_widths_pixels[i], header_y_start + header_height], fill=header_bg_color, outline=border_color, width=1)
                
                wrapped_header = self._wrap_text(str(header), font_header, col_widths_pixels[i] - 20)
                line_height = (font_header.getbbox("Tg")[3] - font_header.getbbox("Tg")[1]) * 1.2
                total_text_height = len(wrapped_header) * line_height
                text_y = header_y_start + (header_height - total_text_height) / 2
                
                for line in wrapped_header:
                    line_width = font_header.getlength(line)
                    draw.text((current_x + (col_widths_pixels[i] - line_width) / 2, text_y), line, font=font_header, fill=text_color)
                    text_y += line_height
                current_x += col_widths_pixels[i]
            current_y += header_height

            line_height = (font_body.getbbox("Tg")[3] - font_body.getbbox("Tg")[1]) * 1.2
            for row_idx, row_data in enumerate(data):
                row_bg_color = row_even_bg_color if row_idx % 2 == 0 else row_odd_bg_color

                max_row_text_height = 0
                temp_wrapped_cells = []
                for i, cell_text in enumerate(row_data):
                    wrapped_lines = self._wrap_text(str(cell_text), font_body, col_widths_pixels[i] - 20)
                    temp_wrapped_cells.append(wrapped_lines)
                    max_row_text_height = max(max_row_text_height, len(wrapped_lines) * line_height)

                row_data_height = max_row_text_height + 10

                if current_y + row_data_height + margin_y > img.height:
                    new_height = int(img.height + (row_data_height + margin_y) * 20)
                    new_img = Image.new("RGB", (img_width, new_height), (255, 255, 255))
                    new_img.paste(img, (0, 0))
                    img = new_img
                    draw = ImageDraw.Draw(img)

                current_x = margin_x
                for i, cell_text in enumerate(row_data):
                    draw.rectangle([current_x, current_y, current_x + col_widths_pixels[i], current_y + row_data_height], fill=row_bg_color, outline=border_color, width=1)
                    
                    wrapped_lines = temp_wrapped_cells[i]
                    text_y = current_y + 5
                    for line in wrapped_lines:
                        draw.text((current_x + 10, text_y), line, font=font_body, fill=text_color)
                        text_y += line_height
                    current_x += col_widths_pixels[i]
                current_y += row_data_height

            current_y += 15
            footer_text = "Report Generated by NregaBot.com"
            footer_font = font_body
            footer_bbox = footer_font.getbbox(footer_text)
            footer_height = footer_bbox[3] - footer_bbox[1]
            footer_y_pos = current_y + 10

            if footer_y_pos + footer_height + margin_y > img.height:
                new_height = int(footer_y_pos + footer_height + margin_y)
                new_img = Image.new("RGB", (img_width, new_height), (255, 255, 255))
                new_img.paste(img, (0, 0))
                img = new_img
                draw = ImageDraw.Draw(img)
            
            draw.text((margin_x, footer_y_pos), footer_text, font=footer_font, fill=text_color)
            current_y = footer_y_pos + footer_height
            final_img = img.crop((0, 0, img_width, current_y + margin_y))
            final_img.save(output_path, "PNG", dpi=(300, 300))
            return True
        except Exception as e:
            messagebox.showerror("PNG Export Error", f"Could not generate PNG report.\nError: {e}", parent=self.app)
            return False

    def _wrap_text(self, text: str, font: Any, max_width: float) -> List[str]:
        """Helper to wrap text for Pillow."""
        if not text: return [""]
        text_lines = text.split('\n')
        final_lines = []
        for text_line in text_lines:
            if not text_line.strip():
                final_lines.append(""); continue
            words = text_line.split(' ')
            lines = []; current_line = []
            for word in words:
                word_too_long = False
                while font.getlength(word) > max_width:
                    word_too_long = True
                    if current_line: lines.append(' '.join(current_line)); current_line = []
                    break_found = False
                    for i in range(len(word) - 1, 0, -1):
                        if font.getlength(word[:i]) <= max_width:
                            lines.append(word[:i]); word = word[i:]; break_found = True; break
                    if not break_found: lines.append(word); word = ""; break
                if not word: continue
                if not word_too_long and font.getlength(' '.join(current_line + [word])) <= max_width:
                    current_line.append(word)
                else:
                    if current_line: lines.append(' '.join(current_line))
                    current_line = [word]
            if current_line: lines.append(' '.join(current_line))
            final_lines.extend(lines)
        return final_lines if final_lines else [""]

    def generate_report_pdf(self, data: List[List[str]], headers: List[str], col_widths: List[float], title: str, date_str: str, file_path: str) -> bool:
        try:
            pdf = ProfessionalPDF(title, date_str, orientation='L', unit='mm', format='A4')
            pdf.alias_nb_pages()
            pdf.add_page()
            
            line_height = 8
            font_size = 9
            pdf.set_font("Arial", size=font_size)
            
            pdf.set_fill_color(44, 62, 80)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Arial", 'B', 10)
            
            for i, h in enumerate(headers):
                width = col_widths[i] if i < len(col_widths) else 40
                pdf.cell(width, 10, str(h), 1, 0, 'C', True)
            pdf.ln()

            pdf.set_font("Arial", size=font_size)
            pdf.set_text_color(0, 0, 0)
            
            fill = False
            
            for row in data:
                max_lines = 1
                for i, cell_data in enumerate(row):
                    text = str(cell_data).encode('latin-1', 'replace').decode('latin-1')
                    width = col_widths[i] if i < len(col_widths) else 40
                    if text:
                        text_width = pdf.get_string_width(text)
                        if text_width > width - 2:
                            lines = int(text_width / (width - 2)) + 1
                            if lines > max_lines: max_lines = lines
                            
                row_height = max_lines * 5
                if row_height < 8: row_height = 8
                
                if pdf.get_y() + row_height > 190:
                    pdf.add_page()
                    pdf.set_fill_color(44, 62, 80)
                    pdf.set_text_color(255, 255, 255)
                    pdf.set_font("Arial", 'B', 10)
                    for i, h in enumerate(headers):
                        width = col_widths[i] if i < len(col_widths) else 40
                        pdf.cell(width, 10, str(h), 1, 0, 'C', True)
                    pdf.ln()
                    pdf.set_font("Arial", size=font_size)
                    pdf.set_text_color(0, 0, 0)
                
                pdf.set_fill_color(240, 240, 240) if fill else pdf.set_fill_color(255, 255, 255)
                
                current_x = pdf.get_x()
                current_y = pdf.get_y()
                
                for i, cell_data in enumerate(row):
                    width = col_widths[i] if i < len(col_widths) else 40
                    text = str(cell_data).encode('latin-1', 'replace').decode('latin-1')
                    
                    if i == 1:
                        if "SUCCESS" in text.upper(): pdf.set_text_color(0, 100, 0)
                        elif "FAIL" in text.upper(): pdf.set_text_color(180, 0, 0)
                        else: pdf.set_text_color(0, 0, 0)
                    else:
                        pdf.set_text_color(0, 0, 0)

                    pdf.rect(current_x, current_y, width, row_height, 'DF' if fill else 'D')
                    pdf.multi_cell(width, 5, text, border=0, align='L')
                    
                    current_x += width
                    pdf.set_xy(current_x, current_y)
                    
                pdf.ln(row_height)
                fill = not fill

            pdf.output(file_path)
            return True
            
        except Exception as e:
            print(f"PDF Gen Error: {e}")
            messagebox.showerror("PDF Export Error", f"Could not generate PDF.\nError: {e}", parent=self.app)
            return False

    def _create_action_buttons(self, parent_frame: Any) -> ctk.CTkFrame:
        """Creates Start, Stop, Reset AND Retry buttons."""
        outer_wrapper = ctk.CTkFrame(parent_frame, fg_color="transparent")
        inner_container = ctk.CTkFrame(outer_wrapper, fg_color="transparent")
        inner_container.pack(expand=True, anchor="center")
        
        self.start_button = ctk.CTkButton(inner_container, text="▶ Start", command=self.start_automation, width=110, height=32, corner_radius=8, fg_color=config.COLORS["btn_start"], hover_color=config.COLORS["btn_start_hover"], font=ctk.CTkFont(size=13, weight="bold"))
        self.start_button.pack(side="left", padx=(0, 8))

        self.stop_button = ctk.CTkButton(inner_container, text="■ Stop", command=self.stop_automation, state="disabled", width=90, height=32, corner_radius=8, fg_color=config.COLORS["btn_stop"], hover_color=config.COLORS["btn_stop_hover"], font=ctk.CTkFont(size=13, weight="bold"))
        self.stop_button.pack(side="left", padx=(0, 8))
        
        # --- NEW RETRY BUTTON ---
        self.retry_btn = ctk.CTkButton(inner_container, text="↻ Retry Failed", command=self.retry_logic_handler, width=110, height=32, corner_radius=8, fg_color=config.COLORS["orange"], hover_color=config.COLORS["orange_hover"], font=ctk.CTkFont(size=13, weight="bold"))
        self.retry_btn.pack(side="left", padx=(0, 8))
        self.retry_btn.configure(state="disabled") # Initially disabled

        self.reset_button = ctk.CTkButton(inner_container, text="↺ Reset", command=self.reset_ui, width=90, height=32, corner_radius=8, fg_color=(config.COLORS["gray70"], config.COLORS["gray40_"]), hover_color=(config.COLORS["gray60"], config.COLORS["gray35_"]), text_color="white", font=ctk.CTkFont(size=13))
        self.reset_button.pack(side="left")
        
        return outer_wrapper

    def _create_log_and_status_area(self, parent_notebook):
        # Remember the tab's main notebook so set_common_ui_state() can
        # auto-switch between "Logs & Status" (while running) and "Results"
        # (when the run finishes).
        self.notebook = parent_notebook
        log_frame = parent_notebook.add("Logs & Status")
        log_frame.grid_columnconfigure(0, weight=1)
        log_frame.grid_rowconfigure(1, weight=1)

        log_actions_frame = ctk.CTkFrame(log_frame, fg_color="transparent")
        log_actions_frame.grid(row=0, column=0, sticky="ew", padx=5, pady=(5, 0))

        def copy_logs_to_clipboard():
            logs = self.log_display.get("1.0", tkinter.END)
            if logs.strip():
                self.app.clipboard_clear(); self.app.clipboard_append(logs)
                messagebox.showinfo("Copied", "Logs copied to clipboard.", parent=self.app)
            else:
                messagebox.showwarning("Empty", "There are no logs to copy.", parent=self.app)

        def clear_log_display():
            self.log_display.configure(state="normal")
            self.log_display.delete("1.0", tkinter.END)
            self.log_display.configure(state="disabled")
            self.update_status("Logs cleared", None)

        copy_button = ctk.CTkButton(log_actions_frame, text="📋 Copy Logs", width=110, command=copy_logs_to_clipboard)
        copy_button.pack(side="right", padx=(0, 5))

        self.clear_logs_button = ctk.CTkButton(log_actions_frame, text="🗑 Clear Logs", width=110, command=clear_log_display,
                                                fg_color=("#DC2626", "#EF4444"), hover_color=("#B91C1C", "#DC2626"))
        self.clear_logs_button.pack(side="right", padx=(0, 5))

        self.log_display = ctk.CTkTextbox(log_frame, state="disabled", font=("Consolas", 12))
        self.log_display.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
        
        status_bar_frame = ctk.CTkFrame(log_frame, height=30)
        status_bar_frame.grid(row=2, column=0, sticky="ew", padx=5, pady=(0, 5))

        self.status_label = ctk.CTkLabel(status_bar_frame, text="Status: Ready", anchor="w")
        self.status_label.pack(side="left", padx=10)
        
        self.progress_bar = ctk.CTkProgressBar(status_bar_frame, mode="determinate")
        self.progress_bar.set(0)
        self.progress_bar.pack(side="right", padx=10, fill="x", expand=True)

    # ────────────────────────────────────────────────────────────────
    # REUSABLE UI BUILDING BLOCKS (P7.2)
    # Shared by automation tabs so every tab gets the same modern
    # card-based look as Pending Bills — without copy-paste.
    # All helpers are pure additions: they never touch automation logic.
    # ────────────────────────────────────────────────────────────────

    def _create_header_card(self, parent: Any, emoji: str, title: str,
                            subtitle: str, row: int = 0,
                            icon_key: Optional[str] = None) -> ctk.CTkFrame:
        """Creates the header/intro card (PNG icon or emoji + bold title + subtitle).

        Mirrors the Pending Bills tab header. Placed at the given grid row
        of `parent` (default row 0). Pass `icon_key` to use the tab's PNG
        icon instead of an emoji (falls back to emoji if the icon fails to
        load). Returns the frame for optional styling.
        """
        header = ctk.CTkFrame(parent, fg_color=("gray95", "gray20"), corner_radius=12)
        header.grid(row=row, column=0, sticky="ew", padx=12, pady=(12, 6))
        icon = None
        if icon_key:
            try:
                icon = self.app.icon_images.get_sized(icon_key, (20, 20))
            except Exception:
                icon = None
        ctk.CTkLabel(
            header, text=f" {title}" if icon is not None else f"{emoji} {title}",
            image=icon, compound="left",
            font=ctk.CTkFont(size=17, weight="bold"),
            text_color=(config.COLORS["blue_dark"], config.COLORS["blue_light"])
        ).pack(anchor="w", padx=14, pady=(10, 0))
        ctk.CTkLabel(
            header, text=subtitle,
            font=ctk.CTkFont(size=12),
            text_color=(config.COLORS["text_dark_alt"], config.COLORS["text_light"])
        ).pack(anchor="w", padx=14, pady=(0, 10))
        return header

    def _create_info_card(self, parent: Any, title: str, text: str,
                          row: int = 2, column: int = 0,
                          columnspan: int = 1) -> ctk.CTkFrame:
        """Creates the 'ℹ️ How it works' info card (bordered, muted fill).

        Placed at the given grid row/column of `parent`. The card stretches
        vertically (sticky nsew) so it fills leftover space in a grid.
        Returns the frame for optional styling.
        """
        info = ctk.CTkFrame(parent, corner_radius=12, border_width=1,
                            border_color=("gray85", "gray30"), fg_color=("gray97", "gray18"))
        info.grid(row=row, column=column, columnspan=columnspan, sticky="nsew", padx=12, pady=6)
        info.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(
            info, text=title,
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=(config.COLORS["blue_dark"], config.COLORS["blue_light"])
        ).grid(row=0, column=0, sticky="w", padx=16, pady=(12, 2))
        ctk.CTkLabel(
            info, text=text, justify="left", anchor="w",
            font=ctk.CTkFont(size=11),
            text_color=(config.COLORS["text_dark_alt"], config.COLORS["text_light"])
        ).grid(row=1, column=0, sticky="w", padx=16, pady=(0, 12))
        return info

    def _create_field(self, parent: Any, key: str, text: str, row: int,
                      col: int = 0, widget_type: str = "entry",
                      values: Optional[List[str]] = None,
                      store: Optional[str] = None,
                      columnspan: Optional[int] = None, **kwargs: Any) -> Any:
        """Generic labelled form field builder shared by all automation tabs.

        Creates a label at (row, col) and a widget at (row, col+1). Returns
        the created widget. Storage is optional and depends on the caller:
          - store='config_vars': StringVar stored in self.config_vars[key]
          - store='ui_fields':   widget stored in self.ui_fields[key]
            (widget_type='combo' additionally stores its var in
             self.dynamic_combo_vars[key], matching the old if_edit helper.)

        Args:
            parent: Parent frame (must have column (col+1) weighted).
            key: Identifier used for config_vars / ui_fields storage.
            text: Label text.
            row: Grid row for label and widget.
            col: Grid column of the label (widget sits at col+1).
            widget_type: 'entry' (default), 'dropdown' or 'combo'.
            values: Dropdown options (ignored for entries).
            store: Optional storage target ('config_vars' | 'ui_fields').
            columnspan: Optional columnspan for the widget's grid placement.
            **kwargs: Passed to the widget constructor.
        """
        ctk.CTkLabel(parent, text=text).grid(row=row, column=col, sticky="w", padx=15, pady=5)
        var = ctk.StringVar()
        if widget_type in ("dropdown", "combo"):
            widget = ctk.CTkOptionMenu(parent, variable=var, values=values or [], **kwargs)
        else:
            widget = ctk.CTkEntry(parent, textvariable=var, **kwargs)
        widget.grid(row=row, column=col + 1, sticky="ew", padx=15, pady=5,
                    columnspan=columnspan if columnspan else 1)

        if store == "config_vars" and hasattr(self, "config_vars") and isinstance(self.config_vars, dict):
            self.config_vars[key] = var
        elif store == "ui_fields" and hasattr(self, "ui_fields") and isinstance(self.ui_fields, dict):
            self.ui_fields[key] = widget
            if widget_type == "combo":
                if not hasattr(self, "dynamic_combo_vars") or not isinstance(self.dynamic_combo_vars, dict):
                    self.dynamic_combo_vars = {}
                self.dynamic_combo_vars[key] = var
        return widget

    def _create_option_field(self, parent: Any, key: str, text: str, row: int,
                             col: int = 0, columnspan: int = 3,
                             store: Optional[str] = "config_vars",
                             **kwargs: Any) -> Any:
        """Label + suggestion dropdown populated from saved history values.

        Convenience wrapper over _create_field() used by the old mb_entry
        helper: reads suggestions for `key` from history_manager and builds
        a dropdown (default columnspan=3 to span label + two field columns).
        """
        values = kwargs.pop("values", None)
        if values is None:
            try:
                values = self.app.history_manager.get_suggestions(key) or [""]
            except Exception:
                values = [""]
        return self._create_field(
            parent, key, text, row, col=col, widget_type="dropdown",
            values=values, columnspan=columnspan, store=store, **kwargs)

    def set_common_ui_state(self, running: bool) -> None:
        """Updates Start/Stop/Reset/Retry buttons based on running state.
        
        Safe to call after tab has been destroyed — checks _is_alive()
        and wraps each configure() in try/except to prevent TclError.
        """
        if not self._is_alive():
            return
        try:
            self.start_button.configure(state="disabled" if running else "normal", text="Running..." if running else "▶ Start")
        except Exception:
            pass
        try:
            self.stop_button.configure(state="normal" if running else "disabled")
        except Exception:
            pass
        try:
            self.reset_button.configure(state="disabled" if running else "normal")
        except Exception:
            pass
        if self.retry_btn:
            try:
                self.retry_btn.configure(state="disabled" if running else "normal")
            except Exception:
                pass

        # Auto-switch the tab's inner notebook: show Logs while the automation
        # runs, then show Results the moment it finishes — so the user never
        # has to hunt for the results / export buttons after a run.
        # (Helper is fully guarded internally — never raises.)
        self._show_automation_tab("running" if running else "finished")

    def _show_automation_tab(self, state: str) -> None:
        """Auto-switch the tab's inner notebook based on automation state.

        state="running"  → show 'Logs & Status' so the user watches progress.
        state="finished" → show 'Results' so results (and their export/report
                           buttons) are immediately visible.

        Safe no-op when the tab has no notebook yet, or the desired tab name
        doesn't exist (tabs use varied names: 'Logs', 'Results Table',
        'MR Summary', ...). Falls back from self.notebook to self.tab_view
        for tabs that build their notebook manually.
        """
        nb = getattr(self, "notebook", None) or getattr(self, "tab_view", None)
        if nb is None:
            return
        try:
            tabs = getattr(nb, "_tab_dict", None)
            if not tabs:
                return
            candidates = ("Logs & Status", "Logs") if state == "running" \
                else ("Results", "Results Table", "MR Summary")
            for name in candidates:
                if name in tabs:
                    nb.set(name)
                    return
        except Exception:
            pass

    def reset_ui(self) -> None:
        self.update_status("Ready", 0)
        self.app.set_status("Ready")
        self.log_display.configure(state="normal")
        self.log_display.delete("1.0", tkinter.END)
        self.log_display.configure(state="disabled")

    def stop_automation(self) -> None:
        self.app.stop_events[self.automation_key].set()
        self.log_warning("Stop signal sent. Finishing current task...")

    def update_status(self, message: str, progress: Optional[float] = None) -> None:
        """Update status label and progress bar.
        
        Safe to call after tab has been destroyed — checks _is_alive()
        and wraps configure() in try/except.
        """
        if not self._is_alive():
            return
        try:
            self.status_label.configure(text=f"Status: {message}")
        except Exception:
            pass
        if progress is not None:
            try:
                self.progress_bar.set(float(progress))
            except Exception:
                pass
            # Footer '%' display — app ko progress report karo (thread-safe)
            try:
                if hasattr(self.app, 'report_automation_progress'):
                    self.app.report_automation_progress(self.automation_key, float(progress))
            except Exception:
                pass
        # --- FIXED: Update Global App Status ---
        if hasattr(self.app, 'set_status'):
            self.app.set_status(message)

    def retry_logic_handler(self) -> None:
        """Override this in child tabs if specific logic is needed, otherwise uses default."""
        # Child tab should define 'self.input_text_widget' (the textbox with codes/jobcards)
        if hasattr(self, 'work_codes_text'):
            self.retry_failed_automation(self.work_codes_text)
        elif hasattr(self, 'jobcards_text'): # For Demand Tab support
            self.retry_failed_automation(self.jobcards_text)
        else:
            messagebox.showinfo("Info", "Retry logic not configured for this tab.")

    def retry_failed_automation(self, input_widget: Any) -> None:
        """
        Generic Logic:
        1. Reads 'Failed' items from Treeview.
        2. Updates the input box with ONLY failed items.
        3. Clears the Treeview.
        4. Auto-starts automation.
        """
        failed_items = []
        all_items = self.results_tree.get_children()
        
        if not all_items:
            messagebox.showinfo("Retry", "No results found to retry.")
            return

        for item_id in all_items:
            values = self.results_tree.item(item_id)['values']
            # Assuming Column 0 is ID (Workcode/Jobcard) and Column 1 is Status
            code = str(values[0])
            status = str(values[1]).lower()
            
            if "success" not in status:
                failed_items.append(code)
        
        if not failed_items:
            messagebox.showinfo("Great!", "No failed items found.")
            return

        # Confirm before action
        if not messagebox.askyesno("Retry Failed", f"Found {len(failed_items)} failed items.\nDo you want to retry them now?"):
            return

        # 1. Update Input Widget
        input_widget.configure(state="normal")
        input_widget.delete("1.0", tkinter.END)
        input_widget.insert("1.0", "\n".join(failed_items))
        input_widget.configure(state="disabled")

        # 2. Clear Previous Results (Optional, but cleaner for retry run)
        for item in all_items:
            self.results_tree.delete(item)

        # 3. Auto Start
        self.log_info(f"Retrying {len(failed_items)} failed items...")
        self.start_automation()

    def _get_style(self) -> ttk.Style:
        """Return a cached ttk.Style singleton to avoid recreation overhead."""
        if not hasattr(self.app, '_cached_style') or self.app._cached_style is None:
            style = ttk.Style()
            style.theme_use("clam")
            self.app._cached_style = style
        return self.app._cached_style

    def style_treeview(self, treeview_widget: Optional[Any] = None) -> None:
        style = self._get_style()

        mode = ctk.get_appearance_mode()

        if mode == "Dark":
            # --- DARK MODE COLORS ---
            bg_color = "#333333"
            text_color = "#e5e7eb"
            row_hover = "#404040"
            selected_bg = "#3B82F6"
            
            header_bg = "#1f2937"       
            header_fg = "#ffffff"       
            header_hover = "#374151"
            
            # Tags Colors (Dark Mode)
            fail_bg = "#5c1e1e"   # Dark Red
            fail_fg = "#ffffff"
            skip_bg = "#5c4e1e"   # Dark Yellow/Brown
            skip_fg = "#ffffff"
            success_bg = "#14532d" # Dark Green
            success_fg = "#ffffff"
            
        else:
            # --- LIGHT MODE COLORS ---
            bg_color = "#ffffff"        
            text_color = "#111827"
            row_hover = "#f3f4f6"       
            selected_bg = "#3B82F6"     
            
            header_bg = "#f9fafb"       
            header_fg = "#111827"       
            header_hover = "#e5e7eb"
            
            # Tags Colors (Light Mode)
            fail_bg = "#fee2e2"   # Light Red
            fail_fg = "#991b1b"
            skip_bg = "#fef9c3"   # Light Yellow
            skip_fg = "#854d0e"
            success_bg = "#dcfce7" # Light Green
            success_fg = "#166534" # Dark Green Text

        # Configure Treeview
        style.configure("Treeview",
                        background=bg_color,
                        foreground=text_color,
                        fieldbackground=bg_color,
                        rowheight=35,             
                        font=("Segoe UI", 11),
                        borderwidth=0)

        style.map("Treeview",
                  background=[('selected', selected_bg), ('active', row_hover)],
                  foreground=[('selected', 'white'), ('active', text_color)])

        style.configure("Treeview.Heading",
                        background=header_bg,
                        foreground=header_fg,
                        relief="flat",
                        font=("Segoe UI", 12, "bold"))

        style.map("Treeview.Heading",
                  background=[('active', header_hover)])

        # Apply Tags
        if treeview_widget:
            treeview_widget.configure(style="Treeview")
            treeview_widget.tag_configure('failed', background=fail_bg, foreground=fail_fg)
            treeview_widget.tag_configure('skipped', background=skip_bg, foreground=skip_fg)
            treeview_widget.tag_configure('success', background=success_bg, foreground=success_fg)
            treeview_widget.tag_configure('warning', background=skip_bg, foreground=skip_fg)

    def _setup_treeview_sorting(self, tree: Any) -> None:
        for col in tree["columns"]:
            tree.heading(col, text=col, command=lambda _col=col: self._treeview_sort_column(tree, _col, False))

    def _treeview_sort_column(self, tv: Any, col: str, reverse: bool) -> None:
        l = [(tv.set(k, col), k) for k in tv.get_children('')]
        try: l.sort(key=lambda t: float(t[0]), reverse=reverse)
        except ValueError: l.sort(reverse=reverse)
        for index, (val, k) in enumerate(l): tv.move(k, '', index)
        tv.heading(col, command=lambda: self._treeview_sort_column(tv, col, not reverse))
        
    # automation_key -> friendly report folder name (kept consistent so every
    # tab's exports land in the same-named folder under Report {fin_year}/).
    _REPORT_CATEGORY_NAMES: Dict[str, str] = {
        "pending_bills": "Pending Bills",
        "mr_tracking": "MR Tracking",
        "issued_mr_report": "Issued MR",
        "fto_gen": "FTO Generation",
        "nmms_attendance": "NMMS Attendance",
        "work_allocation": "Work Allocation",
        "gen": "Wagelist",
        "mr_fill": "MR Fill",
        "emb_verify": "eMB Verify",
        "material_entry": "Material Entry",
        "mis_reports": "MIS",
        "physical_complete": "Physical Complete",
        "sad_update_status": "SAD Update",
        "add_activity": "Add Activity",
        "del_demand": "Delete Demand",
        "sad_auto": "Sarkar Aapke Dwar",
        "mb_entry": "MB Entry",
        "zero_mr": "Zero MR",
        "delete_applicant": "Delete Applicant",
        "demand": "Demand",
        "resend_wg": "Resend Rejected Wagelist",
        "update_estimate": "Update Estimate",
        "wc_gen": "Work Code Generation",
        "send": "Wagelist Send",
        "duplicate_mr": "Duplicate MR",
        "social_audit_respond": "Social Audit",
        "muster": "Muster Roll",
        "mate_mr": "Mate MR",
        "pdf_merger": "PDF Merger",
        "msr": "MSR",
        "dashboard_report": "Dashboard Report",
        "abps_verify": "ABPS Verify",
        "if_edit": "IF Edit",
        "jc_verify": "Jobcard Verify",
        "del_work_alloc": "Delete Work Allocation",
        "macro": "Macro",
        "scheme_closing": "Scheme Closing",
        "ekyc_report": "eKYC Report",
    }

    def _report_category(self) -> str:
        """Human-readable category folder name for this tab's reports."""
        key = self.automation_key or ""
        return self._REPORT_CATEGORY_NAMES.get(
            key, key.replace("_", " ").title() or "Reports")

    def export_treeview_to_csv(self, tree: Any, default_filename: str, category: str = "") -> None:
        """Export treeview contents to CSV inside the standard report folder."""
        reports_dir = self.app.get_report_path(category or self._report_category())
        file_path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")], initialdir=reports_dir, initialfile=default_filename, title="Save CSV Report")
        if not file_path: return
        try:
            with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(tree["columns"])
                for item_id in tree.get_children(): writer.writerow(tree.item(item_id)['values'])
            messagebox.showinfo("Success", f"Report successfully exported to\n{file_path}", parent=self)
        except Exception as e:
            messagebox.showerror("Export Failed", f"An error occurred while saving the CSV file:\n{e}", parent=self)

    def export_treeview_to_excel(self, tree: Any, default_filename: str = "report.xlsx",
                                  filter_mode: str = "Export All",
                                  title_prefix: str = "",
                                  category: str = "") -> Optional[str]:
        """
        Professional Excel export with openpyxl styling.

        Features:
        - ✅ Auto-detects Status column from column headings
        - ✅ Filter support (Export All / Success Only / Failed Only)
        - ✅ Summary statistics row (Total / Success / Failed / Skipped)
        - ✅ Dark blue header row with white text
        - ✅ Alternating row colors (white/gray)
        - ✅ Conditional formatting — Success=Green, Failed=Red
        - ✅ Auto column widths (capped at 50 chars)
        - ✅ Thin borders on all cells
        - ✅ Timestamp footer with NregaBot branding

        Args:
            tree: The ttk.Treeview widget containing data
            default_filename: Suggested filename for the save dialog
            filter_mode: 'Export All', 'Success Only', or 'Failed Only'
            title_prefix: Title text for the report header (e.g. 'MB Entry Report')

        Returns:
            File path if saved, None otherwise
        """
        all_items = tree.get_children()
        if not all_items:
            messagebox.showinfo("No Data", "No records to export.")
            return None

        # ── Auto-detect Status column index ──
        columns = list(tree["columns"])
        status_idx: Optional[int] = None
        for i, col in enumerate(columns):
            if col.lower().strip() == 'status':
                status_idx = i
                break
        if status_idx is None:
            # Heuristic: status is usually second-to-last or at index 1 for 3-col layouts
            n = len(columns)
            if n == 3:
                status_idx = 1
            elif n >= 8:
                status_idx = n - 3  # wide tables (mb_entry=8, demand=10)
            elif n >= 6:
                status_idx = n - 2
            elif n >= 4:
                status_idx = n - 3
            else:
                status_idx = 1 if n > 1 else 0

        # ── Filter data & count statistics ──
        data_to_export: List[List] = []
        total_count = len(all_items)
        success_count = 0
        failed_count = 0

        for item_id in all_items:
            values = tree.item(item_id)['values']
            if not values or len(values) <= status_idx:
                continue

            status_text = str(values[status_idx]).upper()

            if "SUCCESS" in status_text:
                success_count += 1
            elif "FAIL" in status_text or "ERROR" in status_text:
                failed_count += 1

            if filter_mode == "Export All":
                data_to_export.append(values)
            elif filter_mode == "Success Only" and "SUCCESS" in status_text:
                data_to_export.append(values)
            elif filter_mode == "Failed Only" and "SUCCESS" not in status_text and status_text:
                data_to_export.append(values)

        if not data_to_export:
            messagebox.showinfo("Empty", "No data matches the selected filter.")
            return None

        # ── Save file dialog ──
        reports_dir = self.app.get_report_path(category or self._report_category())
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            initialdir=reports_dir,
            initialfile=default_filename,
            title="Save Excel Report"
        )
        if not file_path:
            return None

        try:
            # ── Lazy import openpyxl ──
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Report"

            # ═══════════════════════════════════════════════
            # STYLES
            # ═══════════════════════════════════════════════
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            success_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            failed_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
            summary_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin', color='B0B0B0'),
                right=Side(style='thin', color='B0B0B0'),
                top=Side(style='thin', color='B0B0B0'),
                bottom=Side(style='thin', color='B0B0B0')
            )

            ncols = len(columns)

            # ═══════════════════════════════════════════════
            # ROW 1: Main Title (merged across all columns)
            # ═══════════════════════════════════════════════
            title_text = title_prefix.strip() if title_prefix.strip() else "Automation Report"
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
            c = ws.cell(row=1, column=1, value=title_text)
            c.font = Font(size=14, bold=True, color="FFFFFF")
            c.fill = header_fill
            c.alignment = center_align

            # ═══════════════════════════════════════════════
            # ROW 2: Subtitle — generated by + timestamp
            # ═══════════════════════════════════════════════
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
            c = ws.cell(row=2, column=1,
                       value=f"Generated by NregaBot.com | {datetime.now().strftime('%d-%b-%Y %I:%M %p')}")
            c.font = Font(italic=True, size=9, color="555555")
            c.alignment = center_align
            c.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

            # ═══════════════════════════════════════════════
            # ROW 4-5: Summary statistics
            # ═══════════════════════════════════════════════
            skip_count = total_count - success_count - failed_count
            summary_headers = ["Total", "✅ Success", "❌ Failed", "⏭️ Skipped"]
            summary_values = [total_count, success_count, failed_count, skip_count]

            # Centre summary block under the title
            n_summary = len(summary_headers)
            start_col = max(1, (ncols - n_summary) // 2 + 1)

            for i, (h, v) in enumerate(zip(summary_headers, summary_values)):
                col = start_col + i
                if col > ncols:
                    break

                # Header cells
                h_cell = ws.cell(row=4, column=col, value=h)
                h_cell.font = Font(bold=True, size=10)
                h_cell.fill = summary_fill
                h_cell.alignment = center_align
                h_cell.border = thin_border

                # Value cells
                v_cell = ws.cell(row=5, column=col, value=v)
                v_cell.font = Font(bold=True, size=11)
                v_cell.alignment = center_align
                v_cell.border = thin_border
                if "Failed" in h and v > 0:
                    v_cell.font = Font(color="CC0000", bold=True, size=11)
                elif "Success" in h:
                    v_cell.font = Font(color="006100", bold=True, size=11)

            # ═══════════════════════════════════════════════
            # ROW 7: Data table header
            # ═══════════════════════════════════════════════
            data_start_row = 7
            for i, col_name in enumerate(columns, 1):
                c = ws.cell(row=data_start_row, column=i, value=col_name)
                c.font = header_font
                c.fill = header_fill
                c.alignment = center_align
                c.border = thin_border

            # ═══════════════════════════════════════════════
            # DATA ROWS with conditional formatting
            # ═══════════════════════════════════════════════
            for idx, row_data in enumerate(data_to_export):
                r = data_start_row + 1 + idx
                is_even = idx % 2 == 0

                # Determine row fill based on status
                status_text = str(row_data[status_idx]).upper() if len(row_data) > status_idx else ""
                if "SUCCESS" in status_text:
                    row_fill = success_fill if is_even else PatternFill(start_color="F1F8E9", end_color="F1F8E9", fill_type="solid")
                elif "FAIL" in status_text or "ERROR" in status_text:
                    row_fill = failed_fill if is_even else PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
                else:
                    row_fill = gray_fill if is_even else white_fill

                for j, val in enumerate(row_data):
                    c = ws.cell(row=r, column=j + 1, value=str(val))
                    c.fill = row_fill
                    c.border = thin_border

                    # Alignment — first column & status column centred
                    if j == 0:
                        c.alignment = center_align
                    elif j == status_idx:
                        c.alignment = center_align
                        if "SUCCESS" in status_text:
                            c.font = Font(color="006100", bold=True)
                        elif "FAIL" in status_text or "ERROR" in status_text:
                            c.font = Font(color="CC0000", bold=True)
                    else:
                        c.alignment = left_align

            # ═══════════════════════════════════════════════
            # AUTO COLUMN WIDTHS
            # ═══════════════════════════════════════════════
            for i, col_name in enumerate(columns, 1):
                max_width = len(str(col_name)) + 2
                for row_idx in range(data_start_row + 1, data_start_row + 1 + len(data_to_export)):
                    cell_val = ws.cell(row=row_idx, column=i).value or ""
                    # Unicode/Hindi text needs wider approximation
                    width = len(str(cell_val)) * 1.3
                    if width > max_width:
                        max_width = min(width, 50)  # cap at 50 chars
                ws.column_dimensions[get_column_letter(i)].width = max(max_width, 8)  # min 8

            # ═══════════════════════════════════════════════
            # SAVE
            # ═══════════════════════════════════════════════
            wb.save(file_path)
            messagebox.showinfo("Success", f"✅ Excel report saved successfully!\n{file_path}")

            # Try to open the file automatically
            try:
                if sys.platform == "win32":
                    os.startfile(file_path)
                elif sys.platform == "darwin":
                    import subprocess
                    subprocess.call(['open', file_path])
            except Exception:
                pass

            return file_path

        except ImportError:
            messagebox.showerror("Missing Library",
                                "Excel export requires 'openpyxl'.\n"
                                "Please install it: pip install openpyxl")
            return None
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export Excel:\n{e}")
            return None

    def export_treeview_to_excel_auto(self, tree: Any, default_filename: str = "report.xlsx",
                                          filter_mode: str = "Export All",
                                          title_prefix: str = "") -> Optional[str]:
        """
        Auto-save Excel to a temp directory (no file dialog).
        Same logic as export_treeview_to_excel() but saves to temp path.
        Used by WhatsApp Excel send feature.

        Returns:
            File path if saved, None otherwise
        """
        all_items = tree.get_children()
        if not all_items:
            return None

        # ── Auto-detect Status column index ──
        columns = list(tree["columns"])
        status_idx: Optional[int] = None
        for i, col in enumerate(columns):
            if col.lower().strip() == 'status':
                status_idx = i
                break
        if status_idx is None:
            n = len(columns)
            if n == 3:
                status_idx = 1
            elif n >= 8:
                status_idx = n - 3
            elif n >= 6:
                status_idx = n - 2
            elif n >= 4:
                status_idx = n - 3
            else:
                status_idx = 1 if n > 1 else 0

        # ── Filter data ──
        data_to_export: List[List] = []
        total_count = len(all_items)
        success_count = 0
        failed_count = 0

        for item_id in all_items:
            values = tree.item(item_id)['values']
            if not values or len(values) <= status_idx:
                continue
            status_text = str(values[status_idx]).upper()
            if "SUCCESS" in status_text:
                success_count += 1
            elif "FAIL" in status_text or "ERROR" in status_text:
                failed_count += 1
            if filter_mode == "Export All":
                data_to_export.append(values)
            elif filter_mode == "Success Only" and "SUCCESS" in status_text:
                data_to_export.append(values)
            elif filter_mode == "Failed Only" and "SUCCESS" not in status_text and status_text:
                data_to_export.append(values)

        if not data_to_export:
            return None

        # ── Save to temp file (no dialog) ──
        temp_dir = self.app.get_nregabot_path("Temp")
        os.makedirs(temp_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = os.path.join(temp_dir, f"{timestamp}_{default_filename}")

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Report"

            ncols = len(columns)
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            success_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            failed_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
            summary_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin', color='B0B0B0'),
                right=Side(style='thin', color='B0B0B0'),
                top=Side(style='thin', color='B0B0B0'),
                bottom=Side(style='thin', color='B0B0B0')
            )

            # Title row
            title_text = title_prefix.strip() if title_prefix.strip() else "Automation Report"
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
            c = ws.cell(row=1, column=1, value=title_text)
            c.font = Font(size=14, bold=True, color="FFFFFF")
            c.fill = header_fill
            c.alignment = center_align

            # Subtitle
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
            c = ws.cell(row=2, column=1,
                       value=f"Generated by NregaBot.com | {datetime.now().strftime('%d-%b-%Y %I:%M %p')}")
            c.font = Font(italic=True, size=9, color="555555")
            c.alignment = center_align

            # Summary stats
            skip_count = total_count - success_count - failed_count
            summary_headers = ["Total", "✅ Success", "❌ Failed", "⏭️ Skipped"]
            summary_values = [total_count, success_count, failed_count, skip_count]
            n_summary = len(summary_headers)
            start_col = max(1, (ncols - n_summary) // 2 + 1)
            for i, (h, v) in enumerate(zip(summary_headers, summary_values)):
                col = start_col + i
                if col > ncols:
                    break
                h_cell = ws.cell(row=4, column=col, value=h)
                h_cell.font = Font(bold=True, size=10)
                h_cell.fill = summary_fill
                h_cell.alignment = center_align
                h_cell.border = thin_border
                v_cell = ws.cell(row=5, column=col, value=v)
                v_cell.font = Font(bold=True, size=11)
                v_cell.alignment = center_align
                v_cell.border = thin_border
                if "Failed" in h and v > 0:
                    v_cell.font = Font(color="CC0000", bold=True, size=11)
                elif "Success" in h:
                    v_cell.font = Font(color="006100", bold=True, size=11)

            # Header row
            data_start_row = 7
            for i, col_name in enumerate(columns, 1):
                c = ws.cell(row=data_start_row, column=i, value=col_name)
                c.font = header_font
                c.fill = header_fill
                c.alignment = center_align
                c.border = thin_border

            # Data rows
            for idx, row_data in enumerate(data_to_export):
                r = data_start_row + 1 + idx
                is_even = idx % 2 == 0
                status_text = str(row_data[status_idx]).upper() if len(row_data) > status_idx else ""
                if "SUCCESS" in status_text:
                    row_fill = success_fill if is_even else PatternFill(start_color="F1F8E9", end_color="F1F8E9", fill_type="solid")
                elif "FAIL" in status_text or "ERROR" in status_text:
                    row_fill = failed_fill if is_even else PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
                else:
                    row_fill = gray_fill if is_even else white_fill
                for j, val in enumerate(row_data):
                    c = ws.cell(row=r, column=j + 1, value=str(val))
                    c.fill = row_fill
                    c.border = thin_border
                    if j == 0 or j == status_idx:
                        c.alignment = center_align
                        if j == status_idx:
                            if "SUCCESS" in status_text:
                                c.font = Font(color="006100", bold=True)
                            elif "FAIL" in status_text or "ERROR" in status_text:
                                c.font = Font(color="CC0000", bold=True)
                    else:
                        c.alignment = left_align

            # Auto column widths
            for i, col_name in enumerate(columns, 1):
                max_width = len(str(col_name)) + 2
                for row_idx in range(data_start_row + 1, data_start_row + 1 + len(data_to_export)):
                    cell_val = ws.cell(row=row_idx, column=i).value or ""
                    width = len(str(cell_val)) * 1.3
                    if width > max_width:
                        max_width = min(width, 50)
                ws.column_dimensions[get_column_letter(i)].width = max(max_width, 8)

            wb.save(file_path)
            return file_path

        except Exception as e:
            logger.debug(f"Auto Excel export failed: {e}")
            return None

    def _extract_and_update_workcodes(self, textbox_widget: Any) -> None:
        try:
            input_content = textbox_widget.get("1.0", tkinter.END)
            if not input_content.strip(): return

            work_code_pattern = re.compile(r'\b(34\d{8}(?:/\w+)+/\d+)\b')
            wagelist_pattern = re.compile(r'\b\d+WL\d+\b', re.IGNORECASE)

            found_work_codes = work_code_pattern.findall(input_content)
            found_wagelists = wagelist_pattern.findall(input_content)

            processed_work_codes = [truncate_workcode(code) for code in found_work_codes]
            
            results = processed_work_codes + [wl.upper() for wl in found_wagelists]
            final_results = results 

            if final_results:
                textbox_widget.configure(state="normal")
                textbox_widget.delete("1.0", tkinter.END)
                textbox_widget.insert("1.0", "\n".join(final_results))
                messagebox.showinfo("Extraction Complete", f"Found and extracted {len(final_results)} items.", parent=self)
            else:
                messagebox.showinfo("No Codes Found", "Could not find any matching work codes or wagelist IDs in the text.", parent=self)
        
        except Exception as e:
            messagebox.showerror("Extraction Error", f"An error occurred during extraction: {e}", parent=self)

    # ────────────────────────────────────────────────────────────────
    # DROPDOWN SELECTION HELPERS (P6.2, P6.3)
    # ────────────────────────────────────────────────────────────────

    def select_dropdown(self, driver: Any, element_id: str, value: str,
                        by: Any = None, timeout: int = 15,
                        case_insensitive: bool = True) -> Any:
        """
        Wait for dropdown, select by visible text (case-insensitive by default).

        Before:
            self._select_by_text_case_insensitive(
                Select(wait.until(EC.element_to_be_clickable((By.ID, STATE_ID)))),
                inputs['state'])

        After:
            self.select_dropdown(driver, STATE_ID, inputs['state'])

        Args:
            driver: WebDriver instance
            element_id: ID attribute of the select element
            value: Visible text to select
            by: Selector strategy (default By.ID)
            timeout: WebDriverWait timeout in seconds
            case_insensitive: If True, use case-insensitive matching (default True)

        Returns:
            The Select element, or None if not found
        """
        if by is None:
            by = By.ID
        wait = WebDriverWait(driver, timeout)
        select = Select(wait.until(EC.element_to_be_clickable((by, element_id))))
        if case_insensitive:
            self._select_by_text_case_insensitive(select, value)
        else:
            select.select_by_visible_text(value)
        return select

    def _find(self, driver: Any, by: Any, selector: str) -> Any:
        """Shorthand for driver.find_element(by, selector).

        Before:
            self._find(driver, By.ID, "ctl00_ContentPlaceHolder1_lblmsg")

        After:
            self._find(driver, By.ID, "ctl00_ContentPlaceHolder1_lblmsg")
        """
        return driver.find_element(by, selector)

    @staticmethod
    def _select_by_text_case_insensitive(select_element: Any, target_text: str) -> bool:
        """
        Case-insensitive version of Selenium's select_by_visible_text().
        
        Website par panchayat kabhi "PALOJORI" (UPPERCASE), kabhi "Palojori" (Title Case)
        dikhta hai. User ne jaisa bhi save kiya ho, automation kaam karna chahiye.
        
        Args:
            select_element: A selenium.webdriver.support.ui.Select instance
            target_text: The text to match (case-insensitive)
            
        Returns:
            True if a match was found and selected, False otherwise
        """
        target_lower = target_text.strip().lower()
        for option in select_element.options:
            if option.text.strip().lower() == target_lower:
                select_element.select_by_visible_text(option.text)
                return True
        return False

    @staticmethod
    def _get_select_option_texts(select_element: Any) -> List[str]:
        """Return non-empty visible option texts from a Selenium Select.

        Filters out common placeholder options (---Select---, --Select--, etc.)
        so callers get only real panchayat/agency/village names. Individual tabs
        may apply extra filters on top (e.g. skip 'Total' rows).
        """
        texts: List[str] = []
        for opt in select_element.options:
            t = (opt.text or "").strip()
            low = t.lower()
            if not t or low.startswith("--") or low.startswith("---"):
                continue
            if low in ("select", "select panchayat", "select one", "select all", "choose", "please select"):
                continue
            texts.append(t)
        return texts

    def _with_all_panchayats(self, menu_widget: Any) -> None:
        """Prepend the '🌐 All Panchayats' + '⭐ My Saved Panchayats' options to an option menu's values."""
        try:
            vals = list(menu_widget.cget("values"))
        except Exception:
            return
        clean = [v for v in vals if v and v not in (config.ALL_PANCHAYATS_LABEL, config.MY_PANCHAYATS_LABEL)]
        if not clean:
            clean = [""]
        menu_widget.configure(values=[config.ALL_PANCHAYATS_LABEL, config.MY_PANCHAYATS_LABEL] + clean)

    def _all_panchayat_values(self, raw_vals: Optional[List[str]]) -> List[str]:
        """Build option-menu values = [All Panchayats, My Saved Panchayats] + deduped non-empty values."""
        vals = [v for v in (raw_vals or []) if v and v not in (config.ALL_PANCHAYATS_LABEL, config.MY_PANCHAYATS_LABEL)]
        return [config.ALL_PANCHAYATS_LABEL, config.MY_PANCHAYATS_LABEL] + vals

    def _get_saved_panchayats(self) -> List[str]:
        """All unique panchayat names saved in the app (Settings > Location Data).

        Mirrors SettingsTab._get_panchayat_suggestions() so exactly the panchayats
        the user has set/stored in the app are considered — not the full website list.
        """
        try:
            hm = self.app.history_manager
            keys = ["location_panchayat", "panchayat_name", "panchayat",
                    "dashboard_panchayat", "mr_track_panchayat",
                    "issued_mr_panchayat", "audit_panchayat_respond"]
            vals = set()
            for k in keys:
                for s in (hm.get_suggestions(k) or []):
                    if s and str(s).strip():
                        vals.add(str(s).strip())
            return sorted(vals)
        except Exception:
            return []

    @staticmethod
    def _normalize_panchayat_name(name: str) -> str:
        """Collapse whitespace + uppercase — matches portal names and saved names robustly."""
        return re.sub(r"\s+", " ", str(name or "")).strip().upper()

    def _filter_panchayats_to_saved(self, website_panchayats: List[str]) -> List[str]:
        """Keep only the website's panchayats that the user has saved in the app.

        Case-insensitive match (whitespace-collapsed). Returns [] when nothing is
        saved or nothing matches, so callers can log a clear warning and abort.
        """
        saved = {self._normalize_panchayat_name(p) for p in self._get_saved_panchayats() if p and p.strip()}
        if not saved:
            return []
        return [p for p in (website_panchayats or []) if p and self._normalize_panchayat_name(p) in saved]

    def _reset_ui_state_safe(self) -> None:
        """Reset automation UI state from a background thread (safe no-op if unsupported)."""
        try:
            resetter = getattr(self, "set_ui_state", None) or getattr(self, "set_common_ui_state", None)
            if resetter:
                self.app.after(0, resetter, False)
        except Exception:
            pass

    def _abort_if_no_saved_panchayats(self, panchayats_to_process: List[str]) -> bool:
        """Log a warning + reset UI when the saved-panchayat filter matched nothing.

        Returns True when the caller should abort (return) the automation run.
        """
        if panchayats_to_process:
            return False
        self.log_warning("⚠️ No saved panchayat found on the website. Check Settings > Location Data.")
        self._reset_ui_state_safe()
        return True

    @staticmethod
    def _is_aggregate_panchayat_name(name: str) -> bool:
        """True for header/aggregate/numeric rows that are NOT real panchayats.

        Summary tables often contain rows like 'Total', 'Panchayats' (header)
        or serial numbers ('2', '2.') — these must be excluded when collecting
        panchayat names in 'All Panchayats' mode.
        """
        low = name.strip().lower()
        if low in ("total", "grand total", "panchayats", "panchayat",
                   "panchayat name", "s no", "s no.", "sl no", "sl. no",
                   "sr no", "sr no.", "sno", "sno.", "all"):
            return True
        cleaned = name.strip().replace(".", "").replace(" ", "")
        return cleaned.isdigit()

    # ────────────────────────────────────────────────────────────────
    # LOCATION HIERARCHY HELPERS
    # ────────────────────────────────────────────────────────────────

    # Mapping: location_key → (parent_key, [child_keys_to_clear])
    _HIERARCHY_CLEAR = {
        "location_state":    (None, ["location_district", "location_block", "location_panchayat"]),
        "location_district": ("location_state", ["location_block", "location_panchayat"]),
        "location_block":    ("location_district", ["location_panchayat"]),
        "location_panchayat":("location_block", []),
        "location_village":  ("location_panchayat", []),
    }

    def _make_filter_func(self, child_key: str, parent_key: str, parent_entry: Any):
        """
        Create a filter function for a child dropdown that filters by parent value.
        Usage: filter_func=self._make_filter_func("location_district", "location_state", self.state_entry)
        """
        def _filter():
            parent_val = parent_entry.get().strip() if parent_entry else ""
            return self.app.history_manager.get_filtered_suggestions(
                child_key, parent_key, parent_val
            )
        return _filter

    def _make_parent_callback(self, my_key: str, child_entries: list):
        """
        Create a command callback for when a location value is selected.
        Clears child entries when parent changes so old child values don't
        appear orphaned.
        
        child_entries: list of (entry_widget, child_key) tuples to clear
        
        Note: Hierarchy relationships (parent→child) are built via the
        Settings tab — automation tabs are consumers only.
        """
        def _on_parent_selected(value):
            # Clear all dependent child entries
            for entry, _ in child_entries:
                try:
                    entry.delete(0, 'end')
                except Exception:
                    pass
        return _on_parent_selected

    # ────────────────────────────────────────────────────────────────
    # STANDARDIZED LOG HELPERS
    # ────────────────────────────────────────────────────────────────

    def log_success(self, msg: str) -> None:
        """Log a success message with ✅ prefix."""
        self.app.log_message(self.log_display, f"✅ {msg}", "success")

    def log_error(self, msg: str) -> None:
        """Log an error message with ❌ prefix."""
        self.app.log_message(self.log_display, f"❌ {msg}", "error")

    def log_warning(self, msg: str) -> None:
        """Log a warning message with ⚠️ prefix."""
        self.app.log_message(self.log_display, f"⚠️ {msg}", "warning")

    def log_info(self, msg: str) -> None:
        """Log an info message with ℹ️ prefix."""
        self.app.log_message(self.log_display, f"ℹ️ {msg}", "info")

    # ────────────────────────────────────────────────────────────────
    # STOP EVENT HELPER (P6.1)
    # ────────────────────────────────────────────────────────────────

    def is_stopped(self) -> bool:
        """Check if the stop event has been set for this automation.

        Before (55 chars):
            if self.app.stop_events[self.automation_key].is_set():

        After (18 chars):
            if self.is_stopped():
        """
        return self.app.stop_events[self.automation_key].is_set()

    # ────────────────────────────────────────────────────────────────
    # STANDARDIZED TREE HELPERS (P5.2, P5.3)
    # ────────────────────────────────────────────────────────────────

    def safe_tree_insert(self, values: tuple, tags: tuple = ()) -> None:
        """Thread-safe results_tree insert. Called from background threads.

        Before:
            self.app.after(0, lambda: self.results_tree.insert("", "end",
                values=(work_code, status, details, timestamp), tags=tags))

        After:
            self.safe_tree_insert((work_code, status, details, timestamp), tags)
        """
        if not self._is_alive():
            return
        if not hasattr(self, 'results_tree') or self.results_tree is None:
            return
        self.app.after(0, lambda: self.results_tree.insert("", "end", values=values, tags=tags))

    def _insert_rows_batch(self, batch: List[tuple]) -> None:
        """Insert a batch of rows into the results tree (main thread only).

        Batching avoids flooding Tk with thousands of queued inserts when
        scraping large reports (e.g. 1000+ rows). Schedule with:
            self.app.after(0, lambda b=batch: self._insert_rows_batch(b))
        """
        if not self._is_alive():
            return
        if not hasattr(self, 'results_tree') or self.results_tree is None:
            return
        try:
            for data in batch:
                self.results_tree.insert("", "end", values=data)
        except Exception:
            pass

    def safe_tree_clear(self) -> None:
        """Thread-safe results_tree clear. Called from background threads.

        Before:
            for item in self.results_tree.get_children():
                self.results_tree.delete(item)

        After:
            self.safe_tree_clear()
        """
        if not self._is_alive():
            return
        if not hasattr(self, 'results_tree') or self.results_tree is None:
            return
        self.app.after(0, lambda: [self.results_tree.delete(item)
                                     for item in self.results_tree.get_children()])

    def _apply_appearance_mode(self, theme_color_tuple: Any) -> str:
        if isinstance(theme_color_tuple, (tuple, list)):
            if ctk.get_appearance_mode().lower() == "light": return theme_color_tuple[0]
            else: return theme_color_tuple[1]
        return theme_color_tuple