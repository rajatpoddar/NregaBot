# tabs/wc_gen_tab.py
import tkinter
from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk
import os, csv, time, pyperclip, sys, threading, json, webbrowser, requests
from datetime import datetime
from urllib.parse import urlparse, parse_qs
from collections import defaultdict
from src import config
from .base_tab import BaseAutomationTab
from .date_entry_widget import DateEntry
from .demand_tab import CloudFilePicker 
from src.utils import get_logger, truncate_workcode
from typing import Any, Callable, Dict, List, Optional, Tuple

logger = get_logger()

class WcGenTab(BaseAutomationTab):
    def __init__(self, parent: Any, app_instance: Any) -> None:
        # Lazy imports
        from selenium.webdriver.common.keys import Keys
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, ElementNotInteractableException
        super().__init__(parent, app_instance, automation_key="wc_gen")
        self.csv_path = None
        self.undertaking_pdf_path = None 
        self.ui_fields = {}
        self.profiles = {}
        self.profile_file = self.app.get_data_path("wc_gen_profiles.json")
        self.saved_config = {}
        self.successful_wcs_data = [] 

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self._create_widgets()
        self._load_profiles_from_file()
    def _create_widgets(self) -> None:
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.webdriver.common.keys import Keys
        from selenium.common.exceptions import ElementNotInteractableException
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        import openpyxl
        from selenium import webdriver

        notebook = ctk.CTkTabview(self)
        notebook.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        
        settings_tab = notebook.add("Settings")
        results_tab = notebook.add("Results")

        settings_tab.grid_rowconfigure(0, weight=1)
        settings_tab.grid_columnconfigure(0, weight=1)
        results_tab.grid_rowconfigure(1, weight=1)
        results_tab.grid_columnconfigure(0, weight=1)

        settings_container = ctk.CTkScrollableFrame(settings_tab, label_text="Configuration & Actions")
        settings_container.grid(row=0, column=0, sticky="nsew")
        settings_container.grid_columnconfigure(0, weight=1)
        
        step1_frame = ctk.CTkFrame(settings_container)
        step1_frame.grid(row=0, column=0, sticky='ew', pady=(0, 10))
        step1_frame.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(step1_frame, text="Step 1: Load Panchayat & Profile", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, columnspan=2, padx=15, pady=(10, 5), sticky="w")
        
        profile_frame = ctk.CTkFrame(step1_frame, fg_color="transparent")
        profile_frame.grid(row=1, column=0, columnspan=2, sticky='ew', padx=10)
        profile_frame.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(profile_frame, text="Config Profile:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.profile_var = ctk.StringVar()
        self.profile_combobox = ctk.CTkOptionMenu(profile_frame, variable=self.profile_var, values=[], command=self._load_profile)
        self.profile_combobox.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        self.profile_name_entry = ctk.CTkEntry(profile_frame, placeholder_text="Enter new profile name to save")
        self.profile_name_entry.grid(row=1, column=1, padx=5, pady=(5,10), sticky="ew")
        profile_actions = ctk.CTkFrame(profile_frame, fg_color="transparent")
        profile_actions.grid(row=1, column=0, padx=5, pady=(5,10))
        self.save_profile_button = ctk.CTkButton(profile_actions, text="Save", width=70, command=self._save_profile)
        self.save_profile_button.pack(side="left", padx=(0, 5))
        self.delete_profile_button = ctk.CTkButton(profile_actions, text="Delete", width=70, fg_color="transparent", border_width=1, text_color=("gray10", "#DCE4EE"), command=self._delete_profile)
        self.delete_profile_button.pack(side="left")
        
        panchayat_frame = ctk.CTkFrame(step1_frame, fg_color="transparent")
        panchayat_frame.grid(row=2, column=0, columnspan=2, sticky='ew', padx=10, pady=(0,10))
        panchayat_frame.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(panchayat_frame, text="Panchayat:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        p_vals = self.app.history_manager.get_suggestions("location_panchayat") or [""]
        self.panchayat_var = ctk.StringVar()
        self.panchayat_menu = ctk.CTkOptionMenu(panchayat_frame, variable=self.panchayat_var, values=p_vals)
        self.panchayat_menu.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        self.load_button = ctk.CTkButton(panchayat_frame, text="Load Categories from Website", command=self._start_category_loading_thread)
        self.load_button.grid(row=1, column=0, columnspan=2, padx=5, pady=(5,10), sticky="ew")

        action_frame = self._create_action_buttons(parent_frame=settings_container)
        action_frame.grid(row=1, column=0, sticky="ew", padx=0, pady=10)
        
        integration_frame = ctk.CTkFrame(settings_container)
        integration_frame.grid(row=2, column=0, sticky='ew', pady=(0, 10))
        self.send_to_if_edit_switch = ctk.CTkSwitch(integration_frame, text="Auto-send successful work codes to IF Editor")
        self.send_to_if_edit_switch.grid(row=0, column=0, padx=15, pady=10)

        self.step2_frame = ctk.CTkFrame(settings_container)
        self.step2_frame.grid(row=3, column=0, sticky='ew', pady=(0, 10))
        self.step2_frame.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(self.step2_frame, text="Step 2: Configure Work Details", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, columnspan=2, padx=15, pady=(10, 5), sticky="w")
        self._create_field(self.step2_frame, "master_category", "Master Category", 1, is_dropdown=True)
        self._create_field(self.step2_frame, "work_category", "Work Category", 2, is_dropdown=True)
        self._create_field(self.step2_frame, "beneficiary_type", "Beneficiary Type", 3, is_dropdown=True)
        self._create_field(self.step2_frame, "activity_type", "Activity Type", 4, is_dropdown=True)
        self._create_field(self.step2_frame, "work_type", "Work Type", 5, is_dropdown=True)
        self._create_field(self.step2_frame, "pro_status", "Proposal Status", 6, is_dropdown=True)
        self._create_field(self.step2_frame, "executing_agency", "Executing Agency", 7, is_dropdown=True)
        ctk.CTkLabel(self.step2_frame, text="Proposal Date:").grid(row=8, column=0, sticky="w", padx=15, pady=5)
        self.ui_fields['proposal_date'] = DateEntry(self.step2_frame)
        self.ui_fields['proposal_date'].grid(row=8, column=1, sticky="ew", padx=15, pady=5)
        ctk.CTkLabel(self.step2_frame, text="Work Start Date:").grid(row=9, column=0, sticky="w", padx=15, pady=5)
        self.ui_fields['start_date'] = DateEntry(self.step2_frame)
        self.ui_fields['start_date'].grid(row=9, column=1, sticky="ew", padx=15, pady=5)
        self._create_field(self.step2_frame, "est_labour_cost", "Est. Labour Cost (Lakhs)", 10)
        self._create_field(self.step2_frame, "est_material_cost", "Est. Material Cost (Lakhs)", 11)

        ctk.CTkLabel(self.step2_frame, text="Undertaking PDF (Individual):").grid(row=12, column=0, sticky="w", padx=15, pady=5)
        self.pdf_frame = ctk.CTkFrame(self.step2_frame, fg_color="transparent")
        self.pdf_frame.grid(row=12, column=1, sticky="ew", padx=15, pady=5)
        self.select_pdf_button = ctk.CTkButton(self.pdf_frame, text="Select PDF", width=100, command=self._select_undertaking_pdf)
        self.select_pdf_button.pack(side="left")
        self.pdf_label = ctk.CTkLabel(self.pdf_frame, text="No file selected", text_color="gray", font=("Arial", 11))
        self.pdf_label.pack(side="left", padx=10)

        step3_frame = ctk.CTkFrame(settings_container)
        step3_frame.grid(row=4, column=0, sticky='ew', pady=(0, 10))
        step3_frame.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(step3_frame, text="Step 3: Select Data File", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, columnspan=3, padx=15, pady=(10, 5), sticky="w")
        
        file_buttons_frame = ctk.CTkFrame(step3_frame, fg_color="transparent")
        file_buttons_frame.grid(row=1, column=0, columnspan=2, sticky="w", padx=15, pady=10)

        self.select_button = ctk.CTkButton(file_buttons_frame, text="Select from Computer", command=self.select_csv_file)
        self.select_button.pack(side="left", padx=(0, 10))
        
        self.cloud_csv_button = ctk.CTkButton(file_buttons_frame, text="Select from Cloud", command=self._select_csv_from_cloud, fg_color="teal", hover_color="#00695C")
        self.cloud_csv_button.pack(side="left", padx=(0, 10))

        self.demo_csv_button = ctk.CTkButton(file_buttons_frame, text="Download Demo CSV", command=lambda: self.app.save_demo_csv("wc_gen"), fg_color="#2E8B57", hover_color="#257247")
        self.demo_csv_button.pack(side="left", padx=(0, 10))

        self.online_csv_button = ctk.CTkButton(file_buttons_frame, text="Generate CSV Online", command=self._open_wc_tool_link, fg_color="#1F618D", hover_color="#154360")
        self.online_csv_button.pack(side="left", padx=(0, 10)) 
        
        self.file_label = ctk.CTkLabel(step3_frame, text="No file selected", text_color="gray")
        self.file_label.grid(row=2, column=0, columnspan=2, sticky="w", padx=15, pady=(0, 10))
        
        # Initialize as disabled to prevent editing before loading
        self.set_ui_state(running=False, force_disable_form=True)
        # Ensure PDF button starts disabled until categories are loaded (or you can leave it enabled)
        self.select_pdf_button.configure(state="normal") 

        results_action_frame = ctk.CTkFrame(results_tab, fg_color="transparent")
        results_action_frame.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(5, 10), padx=5)
        self.export_csv_button = ctk.CTkButton(results_action_frame, text="Export for IF Editor", command=self._export_wc_gen_results)
        self.export_csv_button.pack(side="left")
        
        cols = ("Work Code", "Job Card", "Beneficiary Type")
        self.results_tree = ttk.Treeview(results_tab, columns=cols, show='headings')
        for col in cols: self.results_tree.heading(col, text=col)
        self.results_tree.column("Work Code", width=180); self.results_tree.column("Job Card", width=180); self.results_tree.column("Beneficiary Type", width=150)
        self.results_tree.grid(row=1, column=0, sticky='nsew', padx=5, pady=5)
        scrollbar = ctk.CTkScrollbar(results_tab, command=self.results_tree.yview)
        self.results_tree.configure(yscroll=scrollbar.set); scrollbar.grid(row=1, column=1, sticky='ns')
        self.style_treeview(self.results_tree)
        
        self._create_log_and_status_area(notebook)

    def _select_undertaking_pdf(self):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.webdriver.common.keys import Keys
        from selenium.common.exceptions import ElementNotInteractableException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        try:
            path = filedialog.askopenfilename(
                parent=self, 
                title="Select Undertaking PDF",
                filetypes=[("PDF files", "*.pdf")]
            )
            if path:
                self.undertaking_pdf_path = path
                filename = os.path.basename(path)
                display_name = filename if len(filename) < 25 else filename[:22] + "..."
                self.pdf_label.configure(text=display_name, text_color=("black", "white"))
                self.log_info(f"Undertaking PDF selected: {filename}")        except Exception as e:
            messagebox.showerror("Error", f"Could not open file picker: {e}")

    def _open_wc_tool_link(self):
        webbrowser.open_new_tab("https://tools.nregabot.com/work_code_generator")

    def _select_csv_from_cloud(self):
        token = self.app.license_info.get('key')
        if not token:
            messagebox.showerror("Error", "You must be licensed to use cloud storage.")
            return

        picker = CloudFilePicker(parent=self, app_instance=self.app)
        self.wait_window(picker) 

        try:
            self.winfo_toplevel().focus_set()
            self.focus_set()
        except Exception as e:
            print(f"Error setting focus after picker: {e}")
        
        selected_file = picker.selected_file
        
        if selected_file:
            file_id = selected_file['id']
            filename = selected_file['filename']
            
            self.log_info(f"Downloading '{filename}' from cloud...")            temp_path = self._download_file_from_cloud(file_id, filename)
            
            if temp_path:
                self.csv_path = temp_path
                self.file_label.configure(text=os.path.basename(temp_path))
                self.log_info(f"Cloud file '{filename}' selected.")
    def _download_file_from_cloud(self, file_id, filename):
        token = self.app.license_info.get('key')
        if not token:
            self.log_error("Cloud Download Failed: Not licensed.")            return None

        headers = {'Authorization': f'Bearer {token}'}
        url = f"{config.LICENSE_SERVER_URL}/files/api/download/{file_id}"
        
        temp_path = self.app.get_data_path(f"cloud_download_wc_gen_{filename}")
        
        try:
            with self.app.http_session.get(url, headers=headers, stream=True, timeout=30) as r:
                r.raise_for_status() 
                with open(temp_path, 'wb') as f:
                    for chunk in r.iter_content(chunk_size=8192): 
                        f.write(chunk)
            
            self.log_info(f"Successfully downloaded '{filename}'.")            return temp_path
        except Exception as e:
            self.log_error(f"Cloud download failed: {e}")            messagebox.showerror("Download Failed", f"Could not download file: {e}")
            return None

    def _log_result(self, result_data):
        self.app.after(0, lambda: self.results_tree.insert("", "end", values=(
            truncate_workcode(result_data.get('work_code', 'N/A')),
            result_data.get('job_card', 'N/A'),
            result_data.get('beneficiary_type', 'N/A')
        )))

    def _export_wc_gen_results(self):
        if not self.successful_wcs_data:
            messagebox.showinfo("No Data", "There are no successful work codes to export.")
            return
        
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
            initialfile="wc_gen_for_if_edit.csv",
            title="Save Work Code Results for IF Editor"
        )
        if not path:
            return
        
        try:
            headers = ["work_code", "beneficiary_type", "job_card"]
            with open(path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                writer.writerows(self.successful_wcs_data)
            messagebox.showinfo("Success", f"Successfully exported {len(self.successful_wcs_data)} rows to\n{path}")
        except Exception as e:
            messagebox.showerror("Export Error", f"An error occurred while exporting:\n{e}")

    def _create_field(self, parent, key, text, row, is_dropdown=False):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.webdriver.common.keys import Keys
        from selenium.common.exceptions import ElementNotInteractableException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        ctk.CTkLabel(parent, text=text).grid(row=row, column=0, sticky="w", padx=15, pady=5)
        if is_dropdown:
            widget = ctk.CTkOptionMenu(parent, values=[], state="disabled", command=lambda choice, k=key: self._on_dropdown_select(k, choice))
        else:
            widget = ctk.CTkEntry(parent, state="disabled")
        widget.grid(row=row, column=1, sticky="ew", padx=15, pady=5)
        self.ui_fields[key] = widget
        
    def _populate_defaults(self):
        cfg = config.WC_GEN_CONFIG["defaults"]
        current_year = str(datetime.now().year)
        for key, value in cfg.items():
            if key in self.ui_fields:
                field = self.ui_fields[key]
                formatted_value = value.format(year=current_year) if "{year}" in value else value
                
                # We need to temporarily enable if disabled
                try:
                    prev_state = field.cget("state")
                except Exception:
                    prev_state = "normal"
                if prev_state == "disabled":
                     try:
                         field.configure(state="normal")
                     except Exception:
                         pass
                
                if isinstance(field, DateEntry):
                    field.set_date(formatted_value)
                elif isinstance(field, ctk.CTkEntry):
                    field.delete(0, tkinter.END)
                    field.insert(0, formatted_value)
                
                if prev_state == "disabled":
                     try:
                         field.configure(state="disabled")
                     except Exception:
                         pass

    def _load_profiles_from_file(self):
        if not os.path.exists(self.profile_file):
            self.profiles = {}
            self._populate_defaults()
            return
        try:
            with open(self.profile_file, 'r') as f:
                self.profiles = json.load(f)
            profile_names = list(self.profiles.keys())
            self.profile_combobox.configure(values=profile_names)
            last_used = "Last Used Config"
            if last_used in profile_names:
                self.profile_combobox.set(last_used)
                self._load_profile(last_used)
            elif profile_names:
                self.profile_combobox.set(profile_names[0])
                self._load_profile(profile_names[0])
            else:
                self._populate_defaults()
        except Exception as e:
            self.log_warning(f"Could not load profiles: {e}")            self.profiles = {}

    def _save_profile(self, profile_name=None, is_autosave=False):
        if not is_autosave:
            profile_name = self.profile_name_entry.get().strip()
            if not profile_name:
                messagebox.showwarning("Input Error", "Please enter a name for the profile.")
                return
        if not profile_name:
            return

        config_data = {key: field.get() for key, field in self.ui_fields.items()}
        self.profiles[profile_name] = config_data
        
        try:
            with open(self.profile_file, 'w') as f:
                json.dump(self.profiles, f, indent=4)
            profile_names = list(self.profiles.keys())
            if not "Last Used Config" in profile_names:
                profile_names.insert(0, "Last Used Config")
            self.profile_combobox.configure(values=profile_names)
            self.profile_combobox.set(profile_name)
            if not is_autosave:
                self.profile_name_entry.delete(0, tkinter.END)
                messagebox.showinfo("Success", f"Profile '{profile_name}' saved successfully.")
        except Exception as e:
            if not is_autosave:
                messagebox.showerror("Error", f"Failed to save profile: {e}")

    def _load_profile(self, profile_name):
        if not profile_name or not self.profiles:
            return
        self.saved_config = self.profiles.get(profile_name, {})
        if not self.saved_config:
            return
        
        # --- FIX: Handling Disabled State during Load ---
        # This was causing dates and costs not to load because the widgets were disabled.
        for key in ["proposal_date", "start_date", "est_labour_cost", "est_material_cost"]:
            if key in self.saved_config and key in self.ui_fields:
                field = self.ui_fields[key]
                value = self.saved_config[key]
                
                # Check state
                is_disabled = False
                try:
                    if field.cget("state") == "disabled":
                        field.configure(state="normal")
                        is_disabled = True
                except Exception as e: logger.debug("WcGen: Could not check field state: %s", e)
                
                if isinstance(field, DateEntry):
                    field.set_date(value)
                elif isinstance(field, ctk.CTkEntry):
                    field.delete(0, tkinter.END)
                    field.insert(0, value)
                
                # Restore state
                if is_disabled:
                    field.configure(state="disabled")
        
        self.log_info(f"Profile '{profile_name}' loaded. Click 'Load Categories' to continue.")
    def _delete_profile(self):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.webdriver.common.keys import Keys
        from selenium.common.exceptions import ElementNotInteractableException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        profile_name = self.profile_combobox.get()
        if not profile_name or profile_name not in self.profiles or profile_name == "Last Used Config":
            messagebox.showwarning("Selection Error", "Please select a valid, user-saved profile to delete.")
            return
        if not messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete the profile '{profile_name}'?"):
            return
        del self.profiles[profile_name]
        try:
            with open(self.profile_file, 'w') as f:
                json.dump(self.profiles, f, indent=4)
            profile_names = list(self.profiles.keys())
            self.profile_combobox.configure(values=profile_names)
            if profile_names:
                self.profile_combobox.set(profile_names[0])
                self._load_profile(profile_names[0])
            else:
                self.profile_combobox.set("")
                self._populate_defaults()
            messagebox.showinfo("Success", f"Profile '{profile_name}' deleted.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete profile: {e}")

    def _start_category_loading_thread(self):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.webdriver.common.keys import Keys
        from selenium.common.exceptions import ElementNotInteractableException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        panchayat = self.panchayat_var.get().strip()
        if not panchayat:
            messagebox.showwarning("Input Required", "Please enter a Panchayat Name first.")
            return
        self.app.update_history("location_panchayat", panchayat)
        self.load_button.configure(state="disabled", text="Loading...")
        threading.Thread(target=self._load_initial_categories, daemon=True).start()

    def _load_initial_categories(self):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.webdriver.common.keys import Keys
        from selenium.common.exceptions import ElementNotInteractableException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        try:
            driver = self.app.get_driver()
            if not driver:
                self.app.after(0, lambda: self.load_button.configure(state="normal", text="Load Categories from Website"))
                return
            
            target_url = config.WC_GEN_CONFIG["url"]
            self.log_info(f"Navigating to: {target_url}")            driver.get(target_url)
            
            wait = WebDriverWait(driver, 20)
            
            try:
                wait.until(EC.presence_of_element_located((By.ID, "ContentPlaceHolder1_ddlMastercategory")))
            except TimeoutException:
                current_url = driver.current_url
                if "login" in current_url.lower() or "home" in current_url.lower() or "index" in current_url.lower():
                    raise Exception(f"Session Expired or Not Logged In.\nCurrent URL: {current_url}\nPlease log in first.")
                else:
                    raise Exception(f"Element not found (Timeout).\nCurrent URL: {current_url}\nCheck if the page loaded correctly.")

            master_cat_options = self._get_options(driver, "ContentPlaceHolder1_ddlMastercategory")
            agency_options = self._get_options(driver, "ContentPlaceHolder1_ddlExeAgency")
            
            self.app.after(0, self._update_ui_after_load, master_cat_options, agency_options)
            
        except Exception as e:
            import traceback
            full_error = traceback.format_exc()
            print(full_error)
            
            error_msg = str(e)
            if "Message:" in error_msg:
                 error_msg = error_msg.split("Stacktrace:")[0].strip()
            
            self.app.after(0, lambda msg=error_msg: messagebox.showerror("Error", f"Failed to load categories:\n{msg}"))
            self.app.after(0, lambda msg=error_msg: self.app.log_message(self.log_display, f"Error: {msg}", "error"))

        finally:
            self.app.after(0, lambda: self.load_button.configure(state="normal", text="Load Categories from Website"))

    def _update_ui_after_load(self, master_cat_options, agency_options):
        # Optimized: Single pass to enable fields
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.webdriver.common.keys import Keys
        from selenium.common.exceptions import ElementNotInteractableException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        for child in self.step2_frame.winfo_children():
             if isinstance(child, (ctk.CTkEntry, ctk.CTkOptionMenu, DateEntry)):
                child.configure(state="normal")
        
        self.select_pdf_button.configure(state="normal")
        self.ui_fields['master_category'].configure(values=master_cat_options)
        self.ui_fields['executing_agency'].configure(values=agency_options)

        self.log_info("Categories loaded. Restoring selections...")        
        # Restore Dropdowns
        saved_master_cat = self.saved_config.get('master_category')
        if saved_master_cat and saved_master_cat in master_cat_options:
            self.ui_fields['master_category'].set(saved_master_cat)
            self._on_dropdown_select('master_category', saved_master_cat)
        
        saved_agency = self.saved_config.get('executing_agency')
        if saved_agency and saved_agency in agency_options: 
            self.ui_fields['executing_agency'].set(saved_agency)

        # --- FIX: Re-apply Date/Cost fields here to be 100% sure they are populated when enabled ---
        for key in ["proposal_date", "start_date", "est_labour_cost", "est_material_cost"]:
            if key in self.saved_config:
                val = self.saved_config[key]
                field = self.ui_fields[key]
                if isinstance(field, DateEntry): field.set_date(val)
                elif isinstance(field, ctk.CTkEntry):
                    field.delete(0, tkinter.END)
                    field.insert(0, val)

    def _on_dropdown_select(self, dropdown_key, selection):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.webdriver.common.keys import Keys
        from selenium.common.exceptions import ElementNotInteractableException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        if not selection:
            return
        self.log_info(f"Selected {dropdown_key}: '{selection}'. Fetching next options...")        threading.Thread(target=self._update_dependent_dropdown, args=(dropdown_key, selection), daemon=True).start()
    
    def _update_dependent_dropdown(self, dropdown_key, selection):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.webdriver.common.keys import Keys
        from selenium.common.exceptions import ElementNotInteractableException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        dependency_map = {
            'master_category': {'next': 'work_category', 'id': 'ContentPlaceHolder1_ddlMastercategory', 'next_id': 'ContentPlaceHolder1_ddlproposed_work_category'},
            'work_category': {'next': 'beneficiary_type', 'id': 'ContentPlaceHolder1_ddlproposed_work_category', 'next_id': 'ContentPlaceHolder1_ddlbeneficiary_type'},
            'beneficiary_type': {'next': 'activity_type', 'id': 'ContentPlaceHolder1_ddlbeneficiary_type', 'next_id': 'ContentPlaceHolder1_ddlactivity_type'},
            'activity_type': {'next': 'work_type', 'id': 'ContentPlaceHolder1_ddlactivity_type', 'next_id': 'ContentPlaceHolder1_ddlproposed_work_type'},
            'work_type': {'next': 'pro_status', 'id': 'ContentPlaceHolder1_ddlproposed_work_type', 'next_id': 'ContentPlaceHolder1_ddlprostatus'}
        }
        if dropdown_key not in dependency_map:
            return
        current = dependency_map[dropdown_key]
        try:
            driver = self.app.get_driver()
            if not driver:
                self.app.after(0, lambda: self.app.log_message(self.log_display, "Browser not available for dropdown update.", "warning"))
                return
            wait = WebDriverWait(driver, 20)
            
            select_element = wait.until(EC.element_to_be_clickable((By.ID, current['id'])))
            
            try:
                next_element_ref = driver.find_element(By.ID, current['next_id'])
            except:
                next_element_ref = None

            Select(select_element).select_by_visible_text(selection)
            
            if next_element_ref:
                try:
                    wait.until(EC.staleness_of(next_element_ref))
                except TimeoutException:
                    pass
            else:
                time.sleep(1)

            wait.until(EC.presence_of_element_located((By.ID, current['next_id'])))
            time.sleep(0.5) 
            
            new_options = self._get_options(driver, current['next_id'])
            # Use after_idle for smoother UI updates
            self.app.after_idle(self._update_next_combobox, current['next'], new_options, list(dependency_map.keys()))
        except Exception as e:
            error_message = str(e).splitlines()[0]
            self.app.after(0, lambda msg=error_message: self.app.log_message(self.log_display, f"Error updating dropdown: {msg}", "error"))

    def _update_next_combobox(self, next_key, options, all_keys):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.webdriver.common.keys import Keys
        from selenium.common.exceptions import ElementNotInteractableException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        self.ui_fields[next_key].configure(values=options, state="normal")
        
        saved_value = self.saved_config.get(next_key)
        current_selection = ""
        
        if saved_value and saved_value in options:
            self.ui_fields[next_key].set(saved_value)
            current_selection = saved_value
        else:
            current_selection = options[0] if options else ""
            self.ui_fields[next_key].set(current_selection)
        
        if current_selection:
            self._on_dropdown_select(next_key, current_selection)

        # Performance Optimization: Only reset necessary widgets, avoid excessive calls
        start_resetting = False
        for key in all_keys:
            if start_resetting:
                # Use cached state check to avoid redundant Configure calls
                if self.ui_fields[key].cget("state") != "disabled":
                     self.ui_fields[key].configure(values=[], state="disabled")
                     self.ui_fields[key].set("")
            if key == next_key:
                start_resetting = True
    
    def _get_options(self, driver, element_id):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.webdriver.common.keys import Keys
        from selenium.common.exceptions import ElementNotInteractableException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        select_element = Select(driver.find_element(By.ID, element_id))
        return [option.text for option in select_element.options if option.get_attribute("value") not in ["00", "0", ""]]

    def _process_single_row(self, driver, form_config, row_data):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.webdriver.common.keys import Keys
        from selenium.common.exceptions import ElementNotInteractableException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        try:
            (priority, work_name, khata_no, plot_no, village_name, total_plants, covered_area, area_plantation, total_saplings, job_card, beneficiary_type_for_if_edit) = row_data
        except ValueError:
            self.log_error("ERROR: CSV row has incorrect number of columns. Expected 11.")            return None

        driver.get(config.WC_GEN_CONFIG["url"])
        wait = WebDriverWait(driver, 25)

        def select_and_wait(element_id, value):
            # ---- Lazy imports ----
            from selenium.webdriver.common.by import By
            from selenium.webdriver.support.ui import Select, WebDriverWait
            from selenium.webdriver.support import expected_conditions as EC
            from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
            from selenium.webdriver.common.keys import Keys
            from selenium.common.exceptions import ElementNotInteractableException
            from selenium import webdriver
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.page import PageMargins
            from openpyxl.drawing.image import Image as XLImage
            self.log_info(f"  > Selecting '{value}'...")            html_element = driver.find_element(By.TAG_NAME, 'html')
            Select(wait.until(EC.presence_of_element_located((By.ID, element_id)))).select_by_visible_text(value)
            wait.until(EC.staleness_of(html_element))
            self.log_info(f"  > OK.")
        self.log_info("Step 1: Selecting Categories...")        select_and_wait("ContentPlaceHolder1_ddlMastercategory", form_config['master_category'])
        select_and_wait("ContentPlaceHolder1_ddlproposed_work_category", form_config['work_category'])
        select_and_wait("ContentPlaceHolder1_ddlbeneficiary_type", form_config['beneficiary_type'])
        select_and_wait("ContentPlaceHolder1_ddlactivity_type", form_config['activity_type'])
        select_and_wait("ContentPlaceHolder1_ddlproposed_work_type", form_config['work_type'])
        select_and_wait("ContentPlaceHolder1_ddlprostatus", form_config['pro_status'])

        self.log_info("Step 2: Filling Dynamic Fields...")        dynamic_fields = {
            "ContentPlaceHolder1_txtdist": total_plants, 
            "ContentPlaceHolder1_txtAdd_dis": covered_area,
            "ContentPlaceHolder1_txtEst_output": area_plantation, 
            "ContentPlaceHolder1_txtJSA_Inst_unit": total_saplings
        }
        for field_id, value in dynamic_fields.items():
            if value.strip():
                try:
                    field = wait.until(EC.presence_of_element_located((By.ID, field_id)))
                    driver.execute_script("arguments[0].value = arguments[1];", field, value)
                except (NoSuchElementException, TimeoutException): pass 
        
        self.log_info("Step 3: Selecting Location...")        select_and_wait("ContentPlaceHolder1_ddlpanch", form_config['panchayat_name'])
        
        village_select = wait.until(EC.presence_of_element_located((By.ID, "ContentPlaceHolder1_ddlvillage")))
        self._select_by_text_case_insensitive(Select(village_select), village_name)
        
        pdf_path = form_config.get('undertaking_pdf')
        if pdf_path and os.path.exists(pdf_path):
            self.log_info("  > Uploading Undertaking PDF...")            try:
                abs_pdf_path = os.path.abspath(pdf_path)
                file_input = driver.find_element(By.ID, "ContentPlaceHolder1_File_indiv_work_file_pdf")
                driver.execute_script("arguments[0].style.display = 'block'; arguments[0].style.visibility = 'visible';", file_input)
                file_input.send_keys(abs_pdf_path)
            except Exception as e: 
                self.log_warning(f"  > Warning: Could not upload PDF. Error: {e}")
        self.log_info("Step 4: Filling Final Details...")        
        ridge_select = wait.until(EC.presence_of_element_located((By.ID, "ContentPlaceHolder1_ddlridgetype")))
        Select(ridge_select).select_by_value(config.WC_GEN_CONFIG["defaults"]["ridge_type"])
        
        def set_val(eid, val):
            # ---- Lazy imports ----
            from selenium.webdriver.common.by import By
            from selenium.webdriver.support.ui import Select, WebDriverWait
            from selenium.webdriver.support import expected_conditions as EC
            from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
            from selenium.webdriver.common.keys import Keys
            from selenium.common.exceptions import ElementNotInteractableException
            from selenium import webdriver
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.page import PageMargins
            from openpyxl.drawing.image import Image as XLImage
            try:
                el = driver.find_element(By.ID, eid)
                driver.execute_script("arguments[0].value = arguments[1];", el, val)
            except Exception as e: logger.debug("WcGen: Could not set field value: %s", e)

        set_val("ContentPlaceHolder1_txtPriority", priority)
        set_val("ContentPlaceHolder1_txtPropDate", form_config['proposal_date'])
        set_val("ContentPlaceHolder1_txtstartdate", form_config['start_date'])
        
        set_val("ContentPlaceHolder1_TxtEstlb", form_config['est_labour_cost'])
        set_val("ContentPlaceHolder1_txtEstMat", form_config['est_material_cost'])
        
        try:
            driver.execute_script("if(typeof TotEstCostFin === 'function') { TotEstCostFin(); }")
        except Exception as e:
            logger.debug("WcGen: Could not calculate estimate: %s", e)
            try:
                total_c = float(form_config['est_labour_cost']) + float(form_config['est_material_cost'])
                driver.execute_script(f"document.getElementById('ContentPlaceHolder1_Txtestcost').value = '{total_c}';")
            except Exception as e2:
                logger.debug("WcGen: Could not set total estimate: %s", e2)

        set_val("ContentPlaceHolder1_txtkhtano", khata_no)
        set_val("ContentPlaceHolder1_txtPlotNo", plot_no)
        
        work_name_field = driver.find_element(By.ID, "ContentPlaceHolder1_txtworkname")
        driver.execute_script("arguments[0].value = arguments[1];", work_name_field, work_name)
        time.sleep(0.5)

        self.log_info("Step 5: Selecting Agency and Saving...")        
        agency_select = wait.until(EC.presence_of_element_located((By.ID, "ContentPlaceHolder1_ddlExeAgency")))
        self._select_by_text_case_insensitive(Select(agency_select), form_config['executing_agency'])
        
        save_btn = driver.find_element(By.ID, "ContentPlaceHolder1_btSave")
        driver.execute_script("arguments[0].click();", save_btn)

        try:
            wait.until(EC.url_contains("ifedit.aspx?work_code="))
            final_url = driver.current_url
            parsed_url = urlparse(final_url)
            work_code = parse_qs(parsed_url.query).get('work_code', [None])[0]
            if work_code:
                self.log_success(f"SUCCESS! Generated Work Code: {work_code}")                result_data = {
                    "work_code": work_code,
                    "beneficiary_type": beneficiary_type_for_if_edit.strip(),
                    "job_card": job_card.strip()
                }
                self._log_result(result_data)
                return result_data
            else:
                self.log_warning("Row submitted, but could not extract work code from URL.")        except TimeoutException:
            self.log_warning("Row submitted, but URL did not change to the success page.")        
        return None
    def start_automation(self) -> None:
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.webdriver.common.keys import Keys
        from selenium.common.exceptions import ElementNotInteractableException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        if not self.csv_path: messagebox.showwarning("Missing File", "Please select a CSV data file first."); return
        form_config = {key: field.get() for key, field in self.ui_fields.items()}
        form_config["panchayat_name"] = self.panchayat_var.get().strip()
        form_config["undertaking_pdf"] = self.undertaking_pdf_path

        required_fields = ["panchayat_name", "master_category", "work_category", "beneficiary_type", "activity_type", "work_type", "pro_status", "executing_agency", "proposal_date", "start_date", "est_labour_cost", "est_material_cost"]
        if any(not form_config.get(key) for key in required_fields):
            messagebox.showwarning("Input Error", "Please load categories and ensure all configuration fields are filled."); return
        
        self._save_profile(profile_name="Last Used Config", is_autosave=True)
        self.app.start_automation_thread(self.automation_key, self.run_automation_logic, args=(form_config,))
        
    def run_automation_logic(self, form_config):
        self.app.after(0, self.set_ui_state, True)
        self.app.clear_log(self.log_display)
        for item in self.results_tree.get_children(): self.results_tree.delete(item)
        
        self.successful_wcs_data.clear()
        
        self.log_info("--- Starting Workcode Generation ---")        self.app.after(0, self.app.set_status, "Running Workcode Generation...")
        
        local_successful_wcs = [] 
        try:
            driver = self.app.get_driver()
            if not driver: return
            with open(self.csv_path, mode='r', encoding='utf-8') as csvfile:
                rows = list(csv.reader(csvfile))[1:]
                total = len(rows)
                for i, row in enumerate(rows):
                    if self.app.stop_events[self.automation_key].is_set():
                        self.app.log_message(self.log_display, "Automation stopped by user."); break
                    if not any(field.strip() for field in row): continue
                    
                    self.log_info(f"--- Processing Row {i+1}/{total} ---")                    try:
                        result_data = self._process_single_row(driver, form_config, row)
                        if result_data:
                            local_successful_wcs.append(result_data)
                    except Exception as e:
                        self.log_error(f"ERROR processing row {i+1}: {e}")
        except FileNotFoundError: self.log_error("ERROR: CSV file not found.")        except Exception as e: self.log_error(f"An unexpected error occurred: {e}")        finally:
            total_processed = len(local_successful_wcs)
            total_rows = len(rows) if 'rows' in dir() and isinstance(rows, list) else 0
            self.successful_wcs_data = local_successful_wcs 
            self.app.after(0, self.set_ui_state, False)
            self.log_info(f"
{'='*50}")            self.log_info(f"📊 Workcode Generation Complete!")            self.log_info(f"✅ Generated: {total_processed} work codes")            if total_rows > 0:
                failed = total_rows - total_processed
                if failed > 0:
                    self.log_info(f"❌ Failed/Skipped: {failed}")                elif total_processed > 0:
                    self.log_info(f"🎉 All {total_rows} rows processed successfully!")            self.log_info(f"{'='*50}")            self.app.after(0, self.app.set_status, "Automation Finished")

            if self.send_to_if_edit_switch.get() and self.successful_wcs_data:
                self.log_info(f"📤 Sending {len(self.successful_wcs_data)} work codes to IF Editor tab...")                self.app.after(0, self.app.switch_to_if_edit_with_data, self.successful_wcs_data)
            
    def select_csv_file(self):
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.webdriver.common.keys import Keys
        from selenium.common.exceptions import ElementNotInteractableException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        path = filedialog.askopenfilename(title="Select your CSV data file", filetypes=[("CSV files", "*.csv")])
        if path: self.csv_path = path; self.file_label.configure(text=os.path.basename(path))
        
    def set_ui_state(self, running: bool, force_disable_form=False):
        if not self._is_alive():
            return
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.webdriver.common.keys import Keys
        from selenium.common.exceptions import ElementNotInteractableException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        state = "disabled" if running else "normal"
        self.start_button.configure(state=state)
        self.stop_button.configure(state="normal" if running else "disabled")
        self.reset_button.configure(state=state)
        if hasattr(self, 'copy_logs_button'): self.copy_logs_button.configure(state=state)

        self.select_button.configure(state=state)
        self.cloud_csv_button.configure(state=state)
        self.panchayat_menu.configure(state=state)
        self.load_button.configure(state=state)
        self.save_profile_button.configure(state=state)
        self.delete_profile_button.configure(state=state)
        self.profile_combobox.configure(state=state)
        self.send_to_if_edit_switch.configure(state=state)
        self.select_pdf_button.configure(state=state)

        if running or force_disable_form:
            for child in self.step2_frame.winfo_children():
                if isinstance(child, (ctk.CTkEntry, ctk.CTkOptionMenu, DateEntry)): child.configure(state="disabled")
        else:
            if self.ui_fields['master_category'].cget("values"):
                 for child in self.step2_frame.winfo_children():
                    if isinstance(child, (ctk.CTkEntry, ctk.CTkOptionMenu, DateEntry)): child.configure(state="normal")
    def reset_ui(self) -> None:
        # ---- Lazy imports ----
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
        from selenium.webdriver.common.keys import Keys
        from selenium.common.exceptions import ElementNotInteractableException
        from selenium import webdriver
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.drawing.image import Image as XLImage
        if messagebox.askokcancel("Reset Form?", "Are you sure?"):
            self.panchayat_var.set("")
            self.file_label.configure(text="No file selected")
            self.csv_path = None
            
            self.undertaking_pdf_path = None
            self.pdf_label.configure(text="No file selected", text_color="gray")
            
            self.app.clear_log(self.log_display)
            self.send_to_if_edit_switch.deselect()
            self.successful_wcs_data.clear()
            for item in self.results_tree.get_children(): self.results_tree.delete(item)
            
            self.set_ui_state(running=False, force_disable_form=True)

            self._populate_defaults()
            self.log_info("Form has been reset.")            self.app.after(0, self.app.set_status, "Ready")