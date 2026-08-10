# tabs/pending_bills_tab.py
"""
Pending Bills Scraper Tab

Scrapes the MGNREGA public "Liability & Expenditure Report" (pending bills)
website and prepares an Excel sheet of unpaid Muster Rolls & Bills for the
selected State → District → Block → Panchayat and Financial Year.

Behaviour:
- User selects State, District, Block (required), Panchayat & Financial Year.
- If Panchayat is left empty → data is collected for ALL panchayats of the
  selected Block.
- If Financial Year is left empty → data is collected for the last 6 financial
  years (current FY + previous 5).

How the scraping works (public site, plain HTTP — no browser needed):
    page=S  liability_exp_report.aspx  → lists districts (digests from links)
    page=D  liability_exp_report.aspx  → lists blocks
    page=B  liability_exp_report.aspx  → lists panchayats (name + panchayat_code)
    Final   state_html/anticipated_exp.aspx (per panchayat link, carries its own digest)
            → redirects to Netnrega/writereaddata/state_out/{fy2}anticipated_exp{panchayat_code}_{fy2}.html
            → static HTML with 4 sections:
                1) Unpaid Muster Rolls    2) Unpaid Bills
                3) Unpaid MSR (Skilled)   4) Unpaid Voucher (Skilled)

Following the panchayat link (instead of guessing the static file URL) is important:
MGNREGA periodically refreshes its data files, so a hard-coded URL pattern may 404
while the page link always carries a fresh digest + correct redirect target.

Digest handling: every liability report page demands a per-query "Digest" value.
Digests are embedded in the page links, so we only need ONE valid *seed* digest
for page=S (config.PENDING_BILLS_CONFIG['seed_digest']) and then follow links.
Other financial years are reached through the ASP.NET financial-year dropdown
postback, which the server answers with fresh digests for that year.
"""
import os
import re
import sys
import time
from datetime import datetime
from tkinter import filedialog, messagebox, ttk
from typing import Any, Dict, List, Optional, Tuple

import customtkinter as ctk

from .base_tab import BaseAutomationTab
from src import config
from src.utils import get_logger
from src.i18n import tr

logger = get_logger()

# Labels used in the dropdowns for the "empty" (collect everything) options
ALL_PANCHAYATS_LABEL = "🌐 All Panchayats"
ALL_FY_LABEL = "🗓 All (Last 6 Financial Years)"

# Section names inside the final report html (order matters for parsing)
SECTION_MR = "Muster Roll"
SECTION_BILL = "Bill"
SECTION_MSR_SKILLED = "Skilled MSR"
SECTION_VOUCHER_SKILLED = "Skilled Voucher"


class PendingBillsTab(BaseAutomationTab):
    def __init__(self, parent: Any, app_instance: Any) -> None:
        super().__init__(parent, app_instance, automation_key="pending_bills")
        # Detailed scraped rows (used by Excel export). Each row is a dict.
        self.collected_rows: List[Dict[str, Any]] = []
        # Location vars (also enables base-class activity/panchayat extraction)
        self.state_var = ctk.StringVar(value="JHARKHAND")
        self.district_var = ctk.StringVar(value="")
        self.block_var = ctk.StringVar(value="")
        self.panchayat_var = ctk.StringVar(value=ALL_PANCHAYATS_LABEL)
        self.fy_var = ctk.StringVar(value=ALL_FY_LABEL)
        self.config_vars = {
            "location_state": self.state_var,
            "location_district": self.district_var,
            "location_block": self.block_var,
            "location_panchayat": self.panchayat_var,
        }

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self._create_widgets()
        self.load_inputs()

    # ────────────────────────────────────────────────────────────────────────
    # UI BUILDERS
    # ────────────────────────────────────────────────────────────────────────
    def _create_widgets(self) -> None:
        notebook = ctk.CTkTabview(self)
        notebook.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

        settings_tab = notebook.add("Settings")
        results_tab = notebook.add("Results")
        self._create_log_and_status_area(notebook)  # adds "Logs & Status"

        settings_tab.grid_columnconfigure(0, weight=1)
        settings_tab.grid_rowconfigure(2, weight=1)
        results_tab.grid_columnconfigure(0, weight=1)
        results_tab.grid_rowconfigure(1, weight=1)

        # ════════════════════ SETTINGS TAB ════════════════════
        # ── Header / intro card ──
        header = ctk.CTkFrame(settings_tab, fg_color=("gray95", "gray20"), corner_radius=12)
        header.grid(row=0, column=0, sticky="ew", padx=12, pady=(12, 6))
        ctk.CTkLabel(
            header, text=tr("form.pending_bills.scraper_title"),
            font=ctk.CTkFont(size=17, weight="bold"),
            text_color=(config.COLORS["blue_dark"], config.COLORS["blue_light"])
        ).pack(anchor="w", padx=14, pady=(10, 0))
        ctk.CTkLabel(
            header, text="Scrapes unpaid Muster Rolls & Bills from the MGNREGA website and prepares an Excel sheet.",
            font=ctk.CTkFont(size=12),
            text_color=(config.COLORS["text_dark_alt"], config.COLORS["text_light"])
        ).pack(anchor="w", padx=14, pady=(0, 10))

        # ── Inputs card ──
        inputs_card = ctk.CTkFrame(settings_tab, corner_radius=12, border_width=1,
                                   border_color=("gray85", "gray30"))
        inputs_card.grid(row=1, column=0, sticky="ew", padx=12, pady=6)
        inputs_card.grid_columnconfigure(1, weight=1)

        # State
        ctk.CTkLabel(inputs_card, text=tr("form.pending_bills.state_required"), font=ctk.CTkFont(size=12, weight="bold")).grid(
            row=0, column=0, sticky="w", padx=(16, 10), pady=(14, 5))
        self.state_menu = ctk.CTkOptionMenu(inputs_card, variable=self.state_var,
                                            values=self._state_values(), width=240)
        self.state_menu.grid(row=0, column=1, sticky="ew", padx=(0, 16), pady=(14, 5))

        # District
        ctk.CTkLabel(inputs_card, text=tr("form.pending_bills.district_required"), font=ctk.CTkFont(size=12, weight="bold")).grid(
            row=1, column=0, sticky="w", padx=(16, 10), pady=5)
        self.district_menu = ctk.CTkOptionMenu(inputs_card, variable=self.district_var,
                                               values=[], width=240)
        self.district_menu.grid(row=1, column=1, sticky="ew", padx=(0, 16), pady=5)

        # Block
        ctk.CTkLabel(inputs_card, text=tr("form.pending_bills.block_required"), font=ctk.CTkFont(size=12, weight="bold")).grid(
            row=2, column=0, sticky="w", padx=(16, 10), pady=5)
        self.block_menu = ctk.CTkOptionMenu(inputs_card, variable=self.block_var,
                                            values=[], width=240)
        self.block_menu.grid(row=2, column=1, sticky="ew", padx=(0, 16), pady=5)

        # Panchayat
        ctk.CTkLabel(inputs_card, text=tr("common.panchayat_col"), font=ctk.CTkFont(size=12, weight="bold")).grid(
            row=3, column=0, sticky="w", padx=(16, 10), pady=5)
        self.panchayat_menu = ctk.CTkOptionMenu(inputs_card, variable=self.panchayat_var,
                                                values=[ALL_PANCHAYATS_LABEL, config.MY_PANCHAYATS_LABEL], width=240)
        self.panchayat_menu.grid(row=3, column=1, sticky="ew", padx=(0, 16), pady=5)

        # Financial Year
        ctk.CTkLabel(inputs_card, text=tr("form.pending_bills.financial_year"), font=ctk.CTkFont(size=12, weight="bold")).grid(
            row=4, column=0, sticky="w", padx=(16, 10), pady=(5, 14))
        self.fy_menu = ctk.CTkOptionMenu(inputs_card, variable=self.fy_var,
                                         values=self._fy_values(), width=240)
        self.fy_menu.grid(row=4, column=1, sticky="ew", padx=(0, 16), pady=(5, 14))

        # ── Wire up cascading dropdowns ──
        self.state_var.trace_add("write", self._on_state_change)
        self.district_var.trace_add("write", self._on_district_change)
        self.block_var.trace_add("write", self._on_block_change)
        # populate initial values
        self._on_state_change()
        self._on_district_change()
        self._on_block_change()

        # ── Info / rules card ──
        info = ctk.CTkFrame(settings_tab, corner_radius=12, border_width=1,
                            border_color=("gray85", "gray30"), fg_color=("gray97", "gray18"))
        info.grid(row=2, column=0, sticky="nsew", padx=12, pady=6)
        info.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(
            info, text=tr("common.how_it_works"),
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=(config.COLORS["blue_dark"], config.COLORS["blue_light"])
        ).grid(row=0, column=0, sticky="w", padx=16, pady=(12, 2))
        rules_text = (
            "• Data is scraped live from the MGNREGA Liability & Expenditure report website.\n"
            "• District & Block are required. Select them from your saved location data.\n"
            "• Leave Panchayat empty  →  collects data for ALL panchayats of the selected Block.\n"
            "• Leave Financial Year empty  →  collects data for the last 6 financial years.\n"
            "• Each row in Results = one Panchayat. Click 'Export to Excel' for the full item-wise sheet."
        )
        ctk.CTkLabel(
            info, text=rules_text, justify="left", anchor="w",
            font=ctk.CTkFont(size=11),
            text_color=(config.COLORS["text_dark_alt"], config.COLORS["text_light"])
        ).grid(row=1, column=0, sticky="w", padx=16, pady=(0, 12))

        # ── Action buttons ──
        action_frame = self._create_action_buttons(parent_frame=settings_tab)
        action_frame.grid(row=3, column=0, pady=10)

        # ════════════════════ RESULTS TAB ════════════════════
        res_btn_frame = ctk.CTkFrame(results_tab, fg_color="transparent")
        res_btn_frame.grid(row=0, column=0, sticky="ew", padx=5, pady=(5, 0))
        self.export_button = ctk.CTkButton(
            res_btn_frame, text=tr("common.export_excel"), width=150, height=32,
            fg_color=config.COLORS["green_export"], hover_color="#0B5E0B",
            command=self.export_report, font=ctk.CTkFont(size=12, weight="bold"))
        self.export_button.pack(side="right")
        ctk.CTkLabel(
            res_btn_frame, text=tr("form.pending_bills.summary_hint"), font=ctk.CTkFont(size=11),
            text_color=(config.COLORS["text_medium"], config.COLORS["text_light"])
        ).pack(side="left", padx=5)

        cols = ("Financial Year", "Panchayat", "MR Count", "Bill Count", "Amount (₹)", "Status")
        self.results_tree = ttk.Treeview(results_tab, columns=cols, show="headings")
        widths = {"Financial Year": 130, "Panchayat": 160, "MR Count": 90,
                  "Bill Count": 90, "Amount (₹)": 110, "Status": 100}
        for col in cols:
            self.results_tree.heading(col, text=col)
            self.results_tree.column(col, width=widths[col], minwidth=70,
                                     anchor="w" if col == "Panchayat" else "center")
        self.results_tree.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
        scrollbar = ctk.CTkScrollbar(results_tab, command=self.results_tree.yview)
        self.results_tree.configure(yscroll=scrollbar.set)
        scrollbar.grid(row=1, column=1, sticky="ns")
        self.style_treeview(self.results_tree)
        self._setup_treeview_sorting(self.results_tree)

    # ────────────────────────────────────────────────────────────────────────
    # DROPDOWN VALUE HELPERS
    # ────────────────────────────────────────────────────────────────────────
    def _state_values(self) -> List[str]:
        return list(config.PENDING_BILLS_CONFIG.keys()) or ["JHARKHAND"]

    def _district_values(self, state: str) -> List[str]:
        vals: List[str] = []
        try:
            vals = self.app.history_manager.get_filtered_suggestions(
                "location_district", "location_state", state) or []
        except Exception:
            vals = []
        if not vals:
            try:
                from src.location_data import STATE_DISTRICT_MAP
                vals = STATE_DISTRICT_MAP.get(state.title(), [])
            except Exception:
                vals = []
        return [v for v in vals if v] or [""]

    def _block_values(self, district: str) -> List[str]:
        try:
            vals = self.app.history_manager.get_filtered_suggestions(
                "location_block", "location_district", district) or []
        except Exception:
            vals = []
        return [v for v in vals if v] or [""]

    def _panchayat_values(self, block: str) -> List[str]:
        try:
            vals = self.app.history_manager.get_filtered_suggestions(
                "location_panchayat", "location_block", block) or []
        except Exception:
            vals = []
        return [v for v in vals if v] or [""]

    def _fy_values(self) -> List[str]:
        return [ALL_FY_LABEL] + self._last_n_financial_years(6)

    def _on_state_change(self, *_):
        self.district_var.set("")
        self.block_var.set("")
        self.panchayat_var.set(ALL_PANCHAYATS_LABEL)
        self.district_menu.configure(values=self._district_values(self.state_var.get().strip()))

    def _on_district_change(self, *_):
        self.block_var.set("")
        self.panchayat_var.set(ALL_PANCHAYATS_LABEL)
        self.block_menu.configure(values=self._block_values(self.district_var.get().strip()))

    def _on_block_change(self, *_):
        self.panchayat_var.set(ALL_PANCHAYATS_LABEL)
        self.panchayat_menu.configure(
            values=[ALL_PANCHAYATS_LABEL, config.MY_PANCHAYATS_LABEL] + [v for v in self._panchayat_values(self.block_var.get().strip()) if v])

    # ────────────────────────────────────────────────────────────────────────
    # STATE / INPUT MANAGEMENT
    # ────────────────────────────────────────────────────────────────────────
    def set_ui_state(self, running: bool) -> None:
        if not self._is_alive():
            return
        self.set_common_ui_state(running)
        state = "disabled" if running else "normal"
        for w in (self.state_menu, self.district_menu, self.block_menu,
                  self.panchayat_menu, self.fy_menu):
            try:
                w.configure(state=state)
            except Exception:
                pass
        if hasattr(self, "export_button"):
            try:
                self.export_button.configure(state=state)
            except Exception:
                pass

    def save_inputs(self, inputs: Dict[str, str]) -> None:
        try:
            self.app.history_manager.save_tab_inputs_batch("pending_bills", inputs)
        except Exception as e:
            logger.debug("Failed to save pending_bills inputs: %s", e)

    def load_inputs(self) -> None:
        try:
            data = self.app.history_manager.get_tab_inputs("pending_bills") or {}
        except Exception:
            data = {}
        if not data:
            return
        state = (data.get("state") or "JHARKHAND").strip()
        if state not in self._state_values():
            state = "JHARKHAND"
        self.state_var.set(state)
        self.district_var.set(data.get("district", ""))
        self.block_var.set(data.get("block", ""))
        self.panchayat_var.set(data.get("panchayat") or ALL_PANCHAYATS_LABEL)
        fy = (data.get("fy") or "").strip()
        self.fy_var.set(fy if fy in self._fy_values() else ALL_FY_LABEL)

    def retry_logic_handler(self) -> None:
        """Re-run the automation with the same inputs."""
        self.start_automation()

    def reset_ui(self) -> None:
        super().reset_ui()
        self.state_var.set("JHARKHAND")
        self.district_var.set("")
        self.block_var.set("")
        self.panchayat_var.set(ALL_PANCHAYATS_LABEL)
        self.fy_var.set(ALL_FY_LABEL)
        self.safe_tree_clear()
        self.collected_rows = []

    # ────────────────────────────────────────────────────────────────────────
    # START / AUTOMATION ENTRY
    # ────────────────────────────────────────────────────────────────────────
    def start_automation(self) -> None:
        self.safe_tree_clear()
        self.collected_rows = []

        state = self.state_var.get().strip().upper()
        district = self.district_var.get().strip().upper()
        block = self.block_var.get().strip().upper()
        panchayat = self.panchayat_var.get().strip().upper()
        if panchayat == ALL_PANCHAYATS_LABEL.upper():
            panchayat = ""
        elif panchayat == config.MY_PANCHAYATS_LABEL.upper():
            panchayat = config.MY_PANCHAYATS_LABEL  # Keep marker → filtered in scraper
        fy = self.fy_var.get().strip()
        if fy == ALL_FY_LABEL:
            fy = ""

        if not state or not district or not block:
            messagebox.showwarning(tr("errors.input_error"), tr("dialogs.select_state_district_block"),
                                   parent=self)
            return
        if state not in config.PENDING_BILLS_CONFIG:
            messagebox.showwarning(
                "Unsupported State",
                f"Pending Bills scraping is not configured for '{state}'.\n\n"
                "Please add it to config.PENDING_BILLS_CONFIG.",
                parent=self)
            return

        fys: List[str] = [fy] if fy else self._last_n_financial_years(6)

        inputs = {"state": self.state_var.get().strip(), "district": self.district_var.get().strip(),
                  "block": self.block_var.get().strip(), "panchayat": self.panchayat_var.get().strip(),
                  "fy": fy}
        self.save_inputs(inputs)
        for key, val in (("location_state", inputs["state"]),
                         ("location_district", inputs["district"]),
                         ("location_block", inputs["block"])):
            if val:
                try:
                    self.app.update_history(key, val)
                except Exception:
                    pass

        self.app.start_automation_thread(
            self.automation_key, self.run_automation_logic,
            args=(state, district, block, panchayat, fys))

    # ────────────────────────────────────────────────────────────────────────
    # SCRAPING CORE (runs in background thread)
    # ────────────────────────────────────────────────────────────────────────
    def run_automation_logic(self, state: str, district: str, block: str,
                             panchayat: str, fys: List[str]) -> None:
        self.app.after(0, self.set_ui_state, True)
        self.app.after(0, lambda: self.app.clear_log(self.log_display))
        self.log_info("🚀 Starting Pending Bills scraping...")

        try:
            import requests
        except ImportError:
            self.log_error("'requests' is not installed. Run: pip install requests")
            self.app.after(0, self.set_ui_state, False)
            return

        cfg = config.PENDING_BILLS_CONFIG.get(state)
        if not cfg:
            self.log_error(f"No configuration for state '{state}'.")
            self.app.after(0, self.set_ui_state, False)
            return

        session = requests.Session()
        session.headers.update({
            "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                           "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"),
        })

        seed_fy = cfg.get("seed_fin_year") or fys[0]
        seed_digest = cfg.get("seed_digest") or ""

        # ── Open the state page (page=S) once using the seed digest ──
        seed_url = self._state_page_url(cfg, state, seed_fy, seed_digest)
        seed_html = self._fetch(session, seed_url)
        if not seed_html or self._is_error_page(seed_html):
            self.log_error("✗ Could not open the Pending Bills report.")
            if seed_digest:
                self.log_error("The configured seed digest may have expired — please refresh "
                               "'seed_digest' in config.PENDING_BILLS_CONFIG.")
            else:
                self.log_error("No seed digest configured in config.PENDING_BILLS_CONFIG.")
            self.app.after(0, self.set_ui_state, False)
            self.app.after(0, self.update_status, "Failed — seed digest expired", 0.0)
            return

        # current page=S state (url + html) — used to chain FY postbacks
        cur = {"url": seed_url, "html": seed_html}

        total_fys = len(fys)
        grand_mr = grand_bill = 0
        grand_total = 0.0

        try:
            for fi, fy in enumerate(fys, 1):
                if self.is_stopped():
                    self.log_warning("⏹ Stop signal received.")
                    break
                self.log_info(f"── 📅 Financial Year: {fy}  ({fi}/{total_fys}) ──")
                self.app.after(0, self.update_status,
                               f"FY {fy}: opening report...", (fi - 1) / total_fys)

                # Switch to the target FY via the ASP.NET dropdown postback
                if fy != seed_fy:
                    new_html, new_url = self._switch_fy(session, cur["url"], cur["html"], fy)
                    if not new_html or self._is_error_page(new_html):
                        self.log_warning(f"  ⚠ Could not switch to FY {fy}; skipping it.")
                        continue
                    cur = {"url": new_url, "html": new_html}

                s_html = cur["html"]
                self.log_info(f"  ✅ Report opened for FY {fy}.")

                try:
                    # ── page=S → find district page=D link (with its digest) ──
                    d_url = self._find_district_url(cfg, s_html, district)
                    if not d_url:
                        self.log_error(f"✗ District '{district}' not found for FY {fy}.")
                        continue
                    d_html = self._fetch(session, d_url)
                    if not d_html or self._is_error_page(d_html):
                        self.log_error(f"✗ Could not open district page for '{district}' (FY {fy}).")
                        continue
                    self.log_info(f"  ✅ District '{district}' found.")

                    # ── page=D → find block page=B link (with its digest) ──
                    b_url = self._find_block_url(cfg, d_html, block)
                    if not b_url:
                        self.log_error(f"✗ Block '{block}' not found for FY {fy}.")
                        continue
                    b_html = self._fetch(session, b_url)
                    if not b_html or self._is_error_page(b_html):
                        self.log_error(f"✗ Could not open block page for '{block}' (FY {fy}).")
                        continue
                    self.log_info(f"  ✅ Block '{block}' found.")

                    # ── page=B → panchayat codes ──
                    panchayats = self._get_panchayats(cfg, b_html)
                    if panchayat == config.MY_PANCHAYATS_LABEL:
                        saved = {self._normalize_name(p) for p in self._get_saved_panchayats() if p and p.strip()}
                        if not saved:
                            self.log_warning("  ⚠ No saved panchayat found in Settings > Location Data.")
                            self._reset_ui_state_safe()
                            return
                        panchayats = [p for p in panchayats
                                      if self._normalize_name(p["name"]) in saved]
                        self.log_info(f"  ⭐ My Saved Panchayats mode: {len(panchayats)} saved panchayat(s) will be processed.")
                    elif panchayat:
                        panchayats = [p for p in panchayats
                                      if self._normalize_name(p["name"]) == panchayat]
                    if not panchayats:
                        self.log_warning(f"  ⚠ No panchayat found for FY {fy}.")
                        continue

                    self.log_info(f"  🗂 Scraping {len(panchayats)} panchayat(s)...")
                    for pi, panch in enumerate(panchayats, 1):
                        if self.is_stopped():
                            self.log_warning("⏹ Stop signal received.")
                            break
                        progress = (fi - 1 + (pi - 1) / max(len(panchayats), 1)) / total_fys
                        self.app.after(0, self.update_status,
                                       f"{panch['name']} ({pi}/{len(panchayats)}) — FY {fy}", progress)

                        try:
                            rows, ok, msg = self._scrape_panchayat(
                                session, cfg, fy, state, district, block, panch)
                        except Exception as e:
                            rows, ok, msg = [], False, f"Error: {str(e)[:120]}"

                        mr_count = sum(1 for r in rows if r["section"] == SECTION_MR)
                        bill_count = sum(1 for r in rows if r["section"] == SECTION_BILL)
                        amt = sum(r["amount"] for r in rows)
                        grand_mr += mr_count
                        grand_bill += bill_count
                        grand_total += amt
                        self.collected_rows.extend(rows)

                        status = "Success" if rows else "No Data"
                        tags = ("success",) if rows else ("skipped",)
                        self.app.after(0, self.safe_tree_insert,
                                       (fy, panch["name"], mr_count, bill_count,
                                        round(amt, 2), status), tags)

                        if ok and rows:
                            self.log_info(f"  ✓ {panch['name']}: {msg}  (₹{amt:,.2f})")
                        elif ok:
                            self.log_success(f"  ✓ {panch['name']}: {msg}")
                        else:
                            self.log_warning(f"  ⚠ {panch['name']}: {msg}")
                except Exception as e:
                    self.log_error(f"✗ Unexpected error for FY {fy}: {str(e)[:200]}")
        finally:
            self.app.after(0, self.set_ui_state, False)
            self.app.after(0, self.update_status, "Finished", 1.0)
            self.app.after(0, self.app.set_status, "Finished")

        if self.is_stopped():
            self.log_warning("⏹ Automation stopped by user.")
        else:
            self.log_info(f"📊 Summary: {grand_mr} unpaid MRs, {grand_bill} unpaid bills, "
                          f"total ₹{grand_total:,.2f}")
            self.log_success("✅ Pending Bills scraping completed. "
                             "Click 'Export to Excel' to save the sheet.")
            self.app.after(5000, lambda: self.app.set_status("Ready"))

    # ── HTTP helpers ────────────────────────────────────────────────────────
    def _fetch(self, session: Any, url: str, retries: int = 2) -> Optional[str]:
        for attempt in range(retries + 1):
            try:
                r = session.get(url, timeout=30)
                if r.status_code == 200:
                    # Report pages are UTF-8 (Hindi text) — requests often
                    # defaults to ISO-8859-1 when the header lacks a charset.
                    if r.encoding is None or r.encoding.lower() in ("iso-8859-1", "windows-1252"):
                        r.encoding = "utf-8"
                    return r.text
                return None
            except Exception:
                if attempt >= retries:
                    return None
                time.sleep(2)
        return None

    def _state_page_url(self, cfg: Dict[str, str], state: str, fy: str, digest: str) -> str:
        from urllib.parse import urlencode
        return cfg["report_url"] + "?" + urlencode({
            "lflag": "eng", "state_code": cfg["state_code"], "state_name": state,
            "fin_year": fy, "page": "S"}) + (f"&Digest={digest}" if digest else "")

    def _abs_url(self, cfg: Dict[str, str], href: str) -> str:
        """Resolve a (possibly relative) link against the report URL."""
        from urllib.parse import urlsplit
        host = urlsplit(cfg["report_url"]).netloc
        if href.startswith("http://") or href.startswith("https://"):
            return href
        if href.startswith("//"):
            return "https:" + href
        if href.startswith("/"):
            return f"https://{host}{href}"
        base = cfg["report_url"].rsplit("/", 1)[0] + "/"
        return base + href

    def _is_error_page(self, html: str) -> bool:
        if not html:
            return True
        low = html.lower()
        if "url tempered" in low or "url tampered" in low or "fin year not found" in low:
            return True
        # Real report pages are large; error/empty pages are tiny
        if len(html) < 2000:
            return True
        return False

    def _iter_links(self, html: str):
        """Yield dicts for liability/anticipated report links on a page."""
        try:
            from bs4 import BeautifulSoup
            from urllib.parse import parse_qs, urlparse
        except ImportError:
            return
        soup = BeautifulSoup(html, "html.parser")
        for a in soup.find_all("a", href=True):
            href = a.get("href", "")
            if "liability_exp_report.aspx" not in href and "anticipated_exp.aspx" not in href:
                continue
            qs = parse_qs(urlparse(href).query)
            item: Dict[str, str] = {"text": a.get_text(" ", strip=True), "href": href}
            for k in ("page", "state_name", "state_code", "district_name", "district_code",
                      "block_name", "block_code", "panchayat_name", "panchayat_code",
                      "fin_year", "Digest"):
                v = qs.get(k)
                item[k] = v[0] if v else ""
            yield item

    # ── Page traversal (digests travel inside the links) ───────────────────
    def _find_district_url(self, cfg: Dict[str, str], s_html: str, district: str) -> Optional[str]:
        """Return the absolute page=D URL (including its Digest) for the district."""
        target = self._normalize_name(district)
        for a in self._iter_links(s_html):
            if a.get("page") == "D":
                name = a.get("district_name") or a.get("text")
                if name and self._normalize_name(name) == target:
                    return self._abs_url(cfg, a.get("href", ""))
        return None

    def _find_block_url(self, cfg: Dict[str, str], d_html: str, block: str) -> Optional[str]:
        target = self._normalize_name(block)
        for a in self._iter_links(d_html):
            if a.get("page") == "B":
                name = a.get("block_name") or a.get("text")
                if name and self._normalize_name(name) == target:
                    return self._abs_url(cfg, a.get("href", ""))
        return None

    def _get_panchayats(self, cfg: Dict[str, str], b_html: str) -> List[Dict[str, str]]:
        seen = set()
        out: List[Dict[str, str]] = []
        for a in self._iter_links(b_html):
            code = a.get("panchayat_code")
            if not code or code in seen:
                continue
            seen.add(code)
            out.append({
                "name": a.get("panchayat_name") or a.get("text"),
                "code": code,
                "href": self._abs_url(cfg, a.get("href", "")),
            })
        out.sort(key=lambda p: self._normalize_name(p["name"]))
        return out

    def _switch_fy(self, session: Any, current_url: str, current_html: str,
                   new_fy: str) -> Tuple[Optional[str], str]:
        """ASP.NET postback: change the financial-year dropdown to get fresh digests.

        Returns (html_of_new_page, url_to_use_for_next_postback). On failure the
        html is None and the original url is returned unchanged.
        """
        try:
            from bs4 import BeautifulSoup
        except ImportError:
            return None, current_url
        try:
            soup = BeautifulSoup(current_html, "html.parser")
            form = soup.find("form", id="aspnetForm") or soup.find("form")
            if not form:
                return None, current_url

            sel = None
            for s in form.find_all("select"):
                nm = (s.get("name") or "").lower()
                if "finyr" in nm or "fin_year" in nm:
                    sel = s
                    break
            if not sel:
                return None, current_url
            valid = {o.get("value") for o in sel.find_all("option")}
            if new_fy not in valid:
                return None, current_url

            post: Dict[str, str] = {}
            for inp in form.find_all("input"):
                nm = inp.get("name")
                if nm:
                    post[nm] = inp.get("value", "")
            sel_name = sel.get("name")
            post[sel_name] = new_fy
            post["__EVENTTARGET"] = sel_name
            post["__EVENTARGUMENT"] = ""
            post["__LASTFOCUS"] = ""

            r = session.post(current_url, data=post, timeout=40, allow_redirects=True)
            if r.status_code != 200:
                return None, current_url
            if r.encoding is None or r.encoding.lower() in ("iso-8859-1", "windows-1252"):
                r.encoding = "utf-8"
            return r.text, (r.url or current_url)
        except Exception as e:
            logger.debug("FY switch postback failed: %s", e)
            return None, current_url

    def _scrape_panchayat(self, session: Any, cfg: Dict[str, str], fy: str, state: str,
                          district: str, block: str, panch: Dict[str, str]
                          ) -> Tuple[List[Dict[str, Any]], bool, str]:
        """Fetch + parse the final per-panchayat report html. Returns (rows, ok, msg)."""
        # Prefer the page link (fresh digest + current redirect target). Fall back
        # to the guessed static URL only if the link is missing.
        url = panch.get("href")
        if not url:
            fy2 = self._fy_short(fy)
            url = f"{cfg['data_base_url']}{fy2}anticipated_exp{panch['code']}_{fy2}.html"
        html = self._fetch(session, url)
        if html is None:
            return [], False, "Report page not available"
        rows = self._parse_final_page(html, fy, state, district, block, panch["name"])
        if not rows:
            return [], True, "No pending bills/MRs"
        return rows, True, f"{len(rows)} pending item(s)"

    # ── Final page parsing ──────────────────────────────────────────────────
    def _parse_final_page(self, html: str, fy: str, state: str, district: str,
                          block: str, panchayat_name: str) -> List[Dict[str, Any]]:
        try:
            from bs4 import BeautifulSoup
        except ImportError:
            return []
        soup = BeautifulSoup(html, "html.parser")

        table = None
        for t in soup.find_all("table"):
            txt = t.get_text(" ", strip=True)
            if "Work Name" in txt and "Amount" in txt:
                table = t
                break
        if table is None:
            return []

        rows: List[Dict[str, Any]] = []
        section = ""
        for tr in table.find_all("tr"):
            cells = [td.get_text(" ", strip=True) for td in tr.find_all(["td", "th"])]
            if not cells:
                continue
            joined = " ".join(cells).lower()

            # ── Detect section header rows (order matters) ──
            if "muster roll no." in joined:
                section = SECTION_MR
                continue
            if "bill no." in joined:
                section = SECTION_BILL
                continue
            if "msr no." in joined:
                section = SECTION_MSR_SKILLED
                continue
            if "voucher no." in joined:
                section = SECTION_VOUCHER_SKILLED
                continue

            if len(cells) < 4:
                continue
            first = cells[0].strip().lower()
            if first in ("s.no", "s no", "sno", "s.no.", "", "no data") or \
               "state :" in joined or "district:" in joined or "work name" in joined:
                continue
            if "no data" in joined:
                continue

            work_name_raw = cells[1]
            num = cells[2]
            amount_raw = cells[3]

            # Extract work code from the trailing "(...)" of the work name
            wc = ""
            m = re.search(r"\(([^()]+)\)\s*$", work_name_raw)
            if m:
                wc = m.group(1).strip()
                work_name = work_name_raw[:m.start()].strip()
            else:
                work_name = work_name_raw.strip()

            rows.append({
                "fy": fy, "state": state, "district": district, "block": block,
                "panchayat": panchayat_name, "section": section,
                "work_name": work_name, "work_code": wc,
                "number": num, "amount": self._parse_amount(amount_raw),
            })
        return rows

    @staticmethod
    def _parse_amount(raw: str) -> float:
        s = re.sub(r"[^0-9.]", "", raw)
        try:
            return float(s) if s else 0.0
        except ValueError:
            return 0.0

    # ── Misc helpers ────────────────────────────────────────────────────────
    @staticmethod
    def _normalize_name(s: str) -> str:
        return re.sub(r"\s+", " ", str(s or "")).strip().upper()

    @staticmethod
    def _fy_short(fy: str) -> str:
        parts = fy.split("-")
        if len(parts) == 2:
            return parts[0][-2:] + parts[1][-2:]
        return fy.replace("-", "")

    @staticmethod
    def _last_n_financial_years(n: int = 6) -> List[str]:
        """Indian FYs ending on the current year (April → March)."""
        now = datetime.now()
        start_year = now.year if now.month >= 4 else now.year - 1
        return [f"{start_year - i}-{start_year - i + 1}" for i in range(n)]

    # ────────────────────────────────────────────────────────────────────────
    # EXCEL EXPORT
    # ────────────────────────────────────────────────────────────────────────
    def export_report(self) -> None:
        if not self.collected_rows:
            messagebox.showinfo(tr("dialogs.no_data"),
                                tr("dialogs.no_scraped_data_export"),
                                parent=self)
            return
        # Save under ~/Downloads/NregaBot/Report {fin_year}/Pending Bills/
        sel_fy = self.fy_var.get().strip()
        fin_year = sel_fy if (sel_fy and sel_fy != ALL_FY_LABEL) else ""
        reports_dir = self.app.get_report_path("Pending Bills", fin_year)
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            initialdir=reports_dir,
            initialfile=f"Pending_Bills_{datetime.now().strftime('%d-%m-%Y')}.xlsx",
            title=tr("form.pending_bills.save_report"))
        if not file_path:
            return
        try:
            self._write_excel(file_path)
            messagebox.showinfo(tr("dialogs.success"),
                                tr("dialogs.report_saved_success", path=file_path), parent=self)
            try:
                if sys.platform == "win32":
                    os.startfile(file_path)
                elif sys.platform == "darwin":
                    import subprocess
                    subprocess.call(["open", file_path])
            except Exception:
                pass
        except Exception as e:
            messagebox.showerror(tr("dialogs.export_error"), tr("dialogs.failed_export_excel", error=e), parent=self)

    def _write_excel(self, file_path: str) -> None:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        title_font = Font(size=14, bold=True, color="FFFFFF")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        total_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="B0B0B0"), right=Side(style="thin", color="B0B0B0"),
            top=Side(style="thin", color="B0B0B0"), bottom=Side(style="thin", color="B0B0B0"))

        wb = openpyxl.Workbook()

        # ── Color-coded analytics styles ──
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        orange_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_font = Font(bold=True, color="9C0006")
        green_font = Font(bold=True, color="006100")
        section_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")

        # ── Sheet 1: Summary (per FY × panchayat) — color-coded extremes + grand total ──
        ws = wb.active
        ws.title = tr("form.pending_bills.summary_title")
        sum_headers = ["Financial Year", "Panchayat", "Muster Rolls", "Bills",
                       "Skilled MSR", "Skilled Voucher", "Total Amount (Rs)", "Share %"]
        ncols2 = len(sum_headers)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols2)
        c = ws.cell(row=1, column=1, value="Pending Bills Summary (per Panchayat)")
        c.font = title_font
        c.fill = header_fill
        c.alignment = center_align
        for i, h in enumerate(sum_headers, 1):
            cell = ws.cell(row=2, column=i, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Aggregate per (FY, panchayat) — only rows that belong to a known
        # section are counted, so the Summary totals always match the detail sheets.
        summary: Dict[Tuple[str, str], Dict[str, Any]] = {}
        for r in self.collected_rows:
            if r["section"] not in (SECTION_MR, SECTION_BILL,
                                     SECTION_MSR_SKILLED, SECTION_VOUCHER_SKILLED):
                continue
            key = (r["fy"], r["panchayat"])
            s = summary.setdefault(key, {"mr": 0, "bill": 0, "msr": 0, "voucher": 0, "amt": 0.0})
            if r["section"] == SECTION_MR:
                s["mr"] += 1
            elif r["section"] == SECTION_BILL:
                s["bill"] += 1
            elif r["section"] == SECTION_MSR_SKILLED:
                s["msr"] += 1
            elif r["section"] == SECTION_VOUCHER_SKILLED:
                s["voucher"] += 1
            s["amt"] += r["amount"]

        tot_amt = sum(s["amt"] for s in summary.values())
        max_key = max(summary, key=lambda k: summary[k]["amt"]) if summary else None
        min_key = min(summary, key=lambda k: summary[k]["amt"]) if summary else None

        # Sort by amount (highest first) so the biggest pending row is on top
        for idx, ((fy, panch), s) in enumerate(
                sorted(summary.items(), key=lambda kv: kv[1]["amt"], reverse=True), 1):
            row_num = 2 + idx
            share = (s["amt"] / tot_amt * 100) if tot_amt else 0.0
            vals = [fy, panch, s["mr"], s["bill"], s["msr"], s["voucher"],
                    round(s["amt"], 2), round(share, 1)]
            is_max = len(summary) > 1 and (fy, panch) == max_key
            is_min = len(summary) > 1 and (fy, panch) == min_key
            for j, v in enumerate(vals, 1):
                cell = ws.cell(row=row_num, column=j, value=v)
                cell.border = thin_border
                cell.alignment = center_align if j != 2 else left_align
                if is_max:
                    cell.fill = red_fill
                    cell.font = red_font
                elif is_min:
                    cell.fill = green_fill
                    cell.font = green_font
                elif idx % 2 == 0:
                    cell.fill = gray_fill

        # ── GRAND TOTAL row (bottom) ──
        tot_mr = sum(s["mr"] for s in summary.values())
        tot_bill = sum(s["bill"] for s in summary.values())
        tot_msr = sum(s["msr"] for s in summary.values())
        tot_voucher = sum(s["voucher"] for s in summary.values())
        total_row = 2 + len(summary) + 1
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
        cell = ws.cell(row=total_row, column=1, value="GRAND TOTAL")
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.fill = total_fill
        cell.border = thin_border
        for j, v in enumerate([tot_mr, tot_bill, tot_msr, tot_voucher,
                               round(tot_amt, 2), 100.0], 3):
            cell = ws.cell(row=total_row, column=j, value=v)
            cell.font = Font(bold=True, size=12)
            cell.alignment = center_align
            cell.fill = total_fill
            cell.border = thin_border

        for i, w in enumerate([16, 20, 12, 10, 12, 14, 16, 10], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A3"
        ws.sheet_properties.tabColor = "1F4E79"

        # ── Sheet 2: Analytics — color-coded rankings ──
        wa = wb.create_sheet("Analytics")
        a_headers = ["Rank", "Panchayat / Financial Year", "Muster Rolls", "Bills",
                     "Skilled MSR", "Skilled Voucher", "Total Amount (Rs)", "Share %"]
        na = len(a_headers)
        wa.merge_cells(start_row=1, start_column=1, end_row=1, end_column=na)
        c = wa.cell(row=1, column=1, value="📊 Pending Bills Analytics")
        c.font = title_font
        c.fill = header_fill
        c.alignment = center_align
        wa.merge_cells(start_row=2, start_column=1, end_row=2, end_column=na)
        c = wa.cell(row=2, column=1,
                    value=f"Generated by NregaBot.com | {datetime.now().strftime('%d-%b-%Y %I:%M %p')}")
        c.font = Font(italic=True, size=9, color="555555")
        c.alignment = center_align

        def _rank_block(start_row: int, title: str,
                        ranked: List[Tuple[str, int, int, int, int, float]]) -> int:
            """Write a color-coded ranking table (ranked = (label, mr, bill, msr,
            voucher, amt), sorted descending by amount). Returns the next row."""
            n = len(ranked)
            total = sum(item[5] for item in ranked)
            wa.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=na)
            c = wa.cell(row=start_row, column=1, value=title)
            c.font = Font(bold=True, size=12, color="FFFFFF")
            c.fill = section_fill
            c.alignment = Alignment(horizontal="left", vertical="center")
            start_row += 1
            for i, h in enumerate(a_headers, 1):
                cell = wa.cell(row=start_row, column=i, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            start_row += 1
            for pos, (label, mr, bill, msr, vouch, amt) in enumerate(ranked, 1):
                share = (amt / total * 100) if total else 0.0
                fill, font = white_fill, None
                if n > 1:
                    if pos == 1:
                        fill, font = red_fill, red_font
                    elif pos == n:
                        fill, font = green_fill, green_font
                    elif pos <= 3:
                        fill = orange_fill
                row_vals = [pos, label, mr, bill, msr, vouch, round(amt, 2), round(share, 1)]
                for j, v in enumerate(row_vals, 1):
                    cell = wa.cell(row=start_row, column=j, value=v)
                    cell.border = thin_border
                    cell.alignment = center_align if j != 2 else left_align
                    cell.fill = fill
                    if font:
                        cell.font = font
                    elif j in (7, 8):
                        cell.font = Font(bold=True)
                start_row += 1
            # Block total (merge 1-6 so the amount sits outside the merge at col 7)
            wa.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
            c = wa.cell(row=start_row, column=1, value=f"TOTAL ({n})")
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.fill = total_fill
            c.border = thin_border
            c = wa.cell(row=start_row, column=7, value=round(total, 2))
            c.font = Font(bold=True)
            c.alignment = center_align
            c.fill = total_fill
            c.border = thin_border
            return start_row + 2  # blank spacer row

        # Panchayat totals across all FYs
        panch_agg: Dict[str, List[float]] = {}
        for (fy, panch), s in summary.items():
            agg = panch_agg.setdefault(panch, [0, 0, 0, 0, 0.0])
            agg[0] += s["mr"]; agg[1] += s["bill"]; agg[2] += s["msr"]
            agg[3] += s["voucher"]; agg[4] += s["amt"]
        ranked_panch = sorted(
            ((p, int(a[0]), int(a[1]), int(a[2]), int(a[3]), a[4]) for p, a in panch_agg.items()),
            key=lambda t: t[5], reverse=True)

        # FY totals across all panchayats
        fy_agg: Dict[str, List[float]] = {}
        for (fy, panch), s in summary.items():
            agg = fy_agg.setdefault(fy, [0, 0, 0, 0, 0.0])
            agg[0] += s["mr"]; agg[1] += s["bill"]; agg[2] += s["msr"]
            agg[3] += s["voucher"]; agg[4] += s["amt"]
        ranked_fy = sorted(
            ((fy, int(a[0]), int(a[1]), int(a[2]), int(a[3]), a[4]) for fy, a in fy_agg.items()),
            key=lambda t: t[5], reverse=True)

        row = 4
        row = _rank_block(row, "🏆 PANCHAYAT RANKING — Highest Pending First", ranked_panch)
        row = _rank_block(row, "📅 FINANCIAL YEAR ANALYSIS — Highest Pending First", ranked_fy)

        # Quick insights
        wa.merge_cells(start_row=row, start_column=1, end_row=row, end_column=na)
        c = wa.cell(row=row, column=1, value="💡 Quick Insights")
        c.font = Font(bold=True, size=12, color="FFFFFF")
        c.fill = section_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        row += 1
        insights: List[str] = []
        if ranked_panch:
            insights.append(f"• Highest pending panchayat: {ranked_panch[0][0]} — "
                            f"₹{ranked_panch[0][5]:,.2f}")
            insights.append(f"• Lowest pending panchayat: {ranked_panch[-1][0]} — "
                            f"₹{ranked_panch[-1][5]:,.2f}")
        if ranked_fy:
            insights.append(f"• Highest pending financial year: {ranked_fy[0][0]} — "
                            f"₹{ranked_fy[0][5]:,.2f}")
        insights.append(f"• Total pending amount: ₹{tot_amt:,.2f}")
        for line in insights:
            wa.merge_cells(start_row=row, start_column=1, end_row=row, end_column=na)
            c = wa.cell(row=row, column=1, value=line)
            c.font = Font(size=11)
            c.alignment = Alignment(horizontal="left", vertical="center")
            row += 1

        for i, w in enumerate([6, 22, 12, 10, 12, 14, 16, 10], 1):
            wa.column_dimensions[get_column_letter(i)].width = w
        wa.freeze_panes = "A4"
        wa.sheet_properties.tabColor = "2E75B6"

        # ── Section-wise sheets: Muster Roll / Bills / Skilled MSR / Skilled Voucher ──
        detail_headers = ["S.No", "Financial Year", "State", "District", "Block", "Panchayat",
                          "Work Name", "Work Code", "MR/Bill/Voucher No.", "Amount (Rs)"]
        ncols = len(detail_headers)
        sections = [
            (SECTION_MR, "Muster Roll", "1F4E79"),
            (SECTION_BILL, "Bills", "C00000"),
            (SECTION_MSR_SKILLED, "Skilled MSR", "375623"),
            (SECTION_VOUCHER_SKILLED, "Skilled Voucher", "7F6000"),
        ]
        for section_key, sheet_name, tab_color in sections:
            rows = [r for r in self.collected_rows if r["section"] == section_key]
            ws = wb.create_sheet(sheet_name)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
            c = ws.cell(row=1, column=1, value=f"{sheet_name} — Pending Items")
            c.font = title_font
            c.fill = header_fill
            c.alignment = center_align
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
            c = ws.cell(row=2, column=1,
                        value=f"Generated by NregaBot.com | {datetime.now().strftime('%d-%b-%Y %I:%M %p')}")
            c.font = Font(italic=True, size=9, color="555555")
            c.alignment = center_align

            for i, h in enumerate(detail_headers, 1):
                cell = ws.cell(row=3, column=i, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            total_amt = 0.0
            for idx, r in enumerate(rows, 1):
                row_num = 3 + idx
                is_even = idx % 2 == 0
                fill = white_fill if is_even else gray_fill
                vals = [idx, r["fy"], r["state"], r["district"], r["block"], r["panchayat"],
                        r["work_name"], r["work_code"], r["number"], round(r["amount"], 2)]
                total_amt += r["amount"]
                for j, v in enumerate(vals, 1):
                    cell = ws.cell(row=row_num, column=j, value=v)
                    cell.fill = fill
                    cell.border = thin_border
                    cell.alignment = center_align if j in (1, 2, 9, 10) else left_align

            # Per-sheet total row
            total_row = 3 + len(rows) + 1
            ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=9)
            cell = ws.cell(row=total_row, column=1, value=f"TOTAL ({len(rows)} items)")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.fill = total_fill
            cell.border = thin_border
            cell = ws.cell(row=total_row, column=10, value=round(total_amt, 2))
            cell.font = Font(bold=True)
            cell.alignment = center_align
            cell.fill = total_fill
            cell.border = thin_border

            widths = [6, 14, 14, 14, 16, 16, 45, 24, 18, 13]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
            ws.freeze_panes = "A4"
            if rows:
                ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{3 + len(rows)}"
            ws.sheet_properties.tabColor = tab_color

        wb.save(file_path)
