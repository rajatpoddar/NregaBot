# src/tabs/_imports.py
#
# Shared import hub for automation tabs.
# Tabs import only the names they actually use, e.g.:
#     from ._imports import By, Select, WebDriverWait, EC  # noqa: F401
# (previously `from ._imports import *` — migrated by scripts/migrate_explicit_imports.py)
#
# ⚠️ These imports are resolved when the tab module is first loaded.
# Since tab_config.py lazy-imports tab modules via _lazy_import(), this
# happens only when the user first opens that tab — not at app startup.

# --- selenium.webdriver.common ---
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.print_page_options import PrintOptions

# --- selenium.webdriver.support ---
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# --- selenium.common.exceptions ---
from selenium.common.exceptions import (
    InvalidSessionIdException,
    NoAlertPresentException,
    NoSuchElementException,
    NoSuchWindowException,
    StaleElementReferenceException,
    TimeoutException,
    UnexpectedAlertPresentException,
    WebDriverException,
)

# --- selenium itself (for driver creation) ---
from selenium import webdriver

# --- Chrome options (used by mr_tracking and some others) ---
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.chrome.service import Service as ChromeService

# --- openpyxl ---
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

# ────────────────────────────────────────────────────────────────
#  pandas — deliberately NOT imported here.
#
#  Importing pandas used to happen at module level (`import pandas as pd`),
#  which meant EVERY tab import pulled in the (heavy, ~1-3s) pandas import.
#  Worse: pandas is NOT safe to import concurrently — when two threads both
#  trigger `import pandas` at the same time (tab loading, license validation
#  threads, workflow handoffs, frozen-build imports), one thread can grab the
#  partially-initialized module and crash with:
#      AttributeError: partially initialized module 'pandas' has no
#      attribute '_pandas_datetime_CAPI' (most likely due to a circular import)
#  or `cannot import name 'pd' from 'src.tabs._imports'`. That crashed tab
#  opening for affected users.
#
#  Tabs that genuinely need pandas call import_pandas() instead — it loads
#  pandas lazily under a lock so only ONE thread can be mid-import at a time.
# ────────────────────────────────────────────────────────────────
import threading

_PANDAS_IMPORT_LOCK = threading.Lock()
_PANDAS_MODULE = None


def import_pandas():
    """Thread-safe lazy import of the pandas module.

    Returns the fully-initialized `pandas` module. All pandas imports in the
    app must go through this helper so a concurrent first-import can never
    expose a partially-initialized pandas to another thread.
    """
    global _PANDAS_MODULE
    if _PANDAS_MODULE is None:
        with _PANDAS_IMPORT_LOCK:
            if _PANDAS_MODULE is None:
                import pandas as _PANDAS_MODULE
    return _PANDAS_MODULE

# ────────────────────────────────────────────────────────────────
#  __all__  —  explicit list of names exported via import *
# ────────────────────────────────────────────────────────────────
__all__ = [
    # selenium.webdriver.common
    "By",
    "Keys",
    "PrintOptions",
    # selenium.webdriver.support
    "Select",
    "WebDriverWait",
    "EC",
    # selenium.common.exceptions
    "NoAlertPresentException",
    "NoSuchElementException",
    "NoSuchWindowException",
    "InvalidSessionIdException",
    "StaleElementReferenceException",
    "TimeoutException",
    "UnexpectedAlertPresentException",
    "WebDriverException",
    # selenium
    "webdriver",
    # Chrome options
    "ChromeOptions",
    "ChromeService",
    # openpyxl
    "openpyxl",
    "XLImage",
    "Alignment",
    "Border",
    "Font",
    "PatternFill",
    "Side",
    "get_column_letter",
    "PageMargins",
    # thread-safe pandas loader
    "import_pandas",
]
